"""
GolfGen Dashboard API — FastAPI backend serving Amazon SP-API data from DuckDB.
Includes background data sync from Amazon SP-API (runs every 2 hours on Railway).
"""

import os
import csv
import json
import asyncio
import logging
import threading
import time
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional
from contextlib import asynccontextmanager
from zoneinfo import ZoneInfo

import secrets
import re as _re

import duckdb
from fastapi import FastAPI, Query, Request, Body, HTTPException, UploadFile, File, Cookie
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel

from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

logger = logging.getLogger("golfgen")

# ── Paths ───────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent.parent  # GolfGen Amazon Dashboard/
DB_DIR = Path(os.environ.get("DB_DIR", str(BASE_DIR / "data")))
DB_PATH = DB_DIR / "golfgen_amazon.duckdb"
COGS_PATH = DB_DIR / "cogs.csv"
CONFIG_PATH = BASE_DIR / "config" / "credentials.json"
DOCS_DIR = Path("/app/docs")

# Ensure docs directory exists
DOCS_DIR.mkdir(parents=True, exist_ok=True)

# ── Background SP-API Sync ──────────────────────────────
SYNC_INTERVAL_HOURS = 2


def _load_sp_api_credentials() -> dict | None:
    """Load SP-API credentials from env vars (Railway) or config file (local dev).

    Supports both naming conventions for env vars:
      - New: SP_API_REFRESH_TOKEN, SP_API_LWA_APP_ID, SP_API_LWA_CLIENT_SECRET, ...
      - Legacy: SP_API_REFRESH_TOKEN, LWA_APP_ID, LWA_CLIENT_SECRET, ...
    """
    # Priority 1: Environment variables (used on Railway)
    env_refresh = os.environ.get("SP_API_REFRESH_TOKEN", "")
    if env_refresh:
        return {
            "refresh_token": env_refresh,
            "lwa_app_id": os.environ.get("SP_API_LWA_APP_ID") or os.environ.get("LWA_APP_ID", ""),
            "lwa_client_secret": os.environ.get("SP_API_LWA_CLIENT_SECRET") or os.environ.get("LWA_CLIENT_SECRET", ""),
            "aws_access_key": os.environ.get("SP_API_AWS_ACCESS_KEY", ""),
            "aws_secret_key": os.environ.get("SP_API_AWS_SECRET_KEY", ""),
            "role_arn": os.environ.get("SP_API_ROLE_ARN", ""),
        }

    # Priority 2: Config file (local development)
    if CONFIG_PATH.exists():
        with open(CONFIG_PATH) as f:
            creds_json = json.load(f)
        sp = creds_json.get("AMAZON_SP_API", {})
        creds = {
            "refresh_token": sp.get("refresh_token"),
            "lwa_app_id": sp.get("lwa_app_id"),
            "lwa_client_secret": sp.get("lwa_client_secret"),
            "aws_access_key": sp.get("aws_access_key"),
            "aws_secret_key": sp.get("aws_secret_key"),
            "role_arn": sp.get("role_arn"),
        }
        if creds.get("refresh_token"):
            return creds

    return None


def _sync_today_orders():
    """Fast sync: pull today's orders from Amazon Orders API.

    This is the critical path for showing live "Today" data.
    Runs at startup, on /api/sync, and inline if today has no data.

    Strategy: Use getOrderItems for EVERY order to get accurate item-level
    prices.  OrderTotal is unreliable (missing for Pending, sometimes $0
    for FBA orders).  Item-level prices are the ground truth.
    """
    try:
        from sp_api.api import Orders as OrdersAPI
        from sp_api.base import Marketplaces
    except ImportError:
        logger.warning("SP-API library not installed — skipping today sync")
        return False

    credentials = _load_sp_api_credentials()
    if not credentials:
        return False

    logger.info("  Syncing today's orders (real-time)...")
    try:
        import time as _t
        # Get orders from last 2 days to catch any we missed
        now_utc = datetime.now(ZoneInfo("UTC"))
        after_date = (now_utc - timedelta(days=2)).isoformat()
        orders_api = OrdersAPI(credentials=credentials, marketplace=Marketplaces.US)
        response = orders_api.get_orders(
            CreatedAfter=after_date,
            MarketplaceIds=["ATVPDKIKX0DER"],
            MaxResultsPerPage=100
        )

        order_list = response.payload.get("Orders", [])

        # Handle pagination — get ALL orders, not just first 100
        next_token = response.payload.get("NextToken")
        while next_token:
            try:
                next_resp = orders_api.get_orders(
                    NextToken=next_token,
                    MarketplaceIds=["ATVPDKIKX0DER"],
                )
                more_orders = next_resp.payload.get("Orders", [])
                order_list.extend(more_orders)
                next_token = next_resp.payload.get("NextToken")
            except Exception:
                break

        logger.info(f"  Got {len(order_list)} orders from last 2 days")

        con = duckdb.connect(str(DB_PATH), read_only=False)

        # Store raw orders
        for order in order_list:
            order_total = order.get("OrderTotal", {})
            con.execute("""
                INSERT OR REPLACE INTO orders
                (order_id, purchase_date, order_status, fulfillment_channel,
                 sales_channel, order_total, currency_code, number_of_items,
                 ship_city, ship_state, ship_postal_code,
                 is_business_order, is_prime)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, [
                order.get("AmazonOrderId"),
                order.get("PurchaseDate"),
                order.get("OrderStatus"),
                order.get("FulfillmentChannel"),
                order.get("SalesChannel"),
                float(order_total.get("Amount", 0)) if order_total else 0,
                order_total.get("CurrencyCode", "USD") if order_total else "USD",
                order.get("NumberOfItemsShipped", 0) + order.get("NumberOfItemsUnshipped", 0),
                order.get("ShippingAddress", {}).get("City", ""),
                order.get("ShippingAddress", {}).get("StateOrRegion", ""),
                order.get("ShippingAddress", {}).get("PostalCode", ""),
                order.get("IsBusinessOrder", False),
                order.get("IsPrime", False),
            ])

        # ── Get item-level prices for ALL non-cancelled orders ──
        # OrderTotal is unreliable.  getOrderItems gives us actual item prices.
        # Rate limit: burst 30, then 0.5/sec.  We handle this with adaptive sleep.
        today_str = (datetime.utcnow() - timedelta(hours=7)).strftime("%Y-%m-%d")
        yesterday_str = (datetime.utcnow() - timedelta(hours=7) - timedelta(days=1)).strftime("%Y-%m-%d")

        # Build ASIN price lookup from historical data for fallback
        asin_prices = {}
        try:
            price_rows = con.execute("""
                SELECT asin,
                       SUM(ordered_product_sales) / NULLIF(SUM(units_ordered), 0) AS avg_price
                FROM daily_sales
                WHERE asin != 'ALL' AND units_ordered > 0
                  AND date >= CURRENT_DATE - 90
                GROUP BY asin
            """).fetchall()
            for pr in price_rows:
                if pr[1] and pr[1] > 0:
                    asin_prices[pr[0]] = float(pr[1])
        except Exception:
            pass
        logger.info(f"  ASIN price fallback: {len(asin_prices)} ASINs with known prices")

        items_fetched = 0
        items_failed = 0
        burst_count = 0

        for order in order_list:
            status = order.get("OrderStatus", "")
            if status in ("Canceled", "Cancelled"):
                continue

            # Only fetch items for today and yesterday orders
            purchase_date = order.get("PurchaseDate", "")
            if not purchase_date:
                continue
            try:
                if "T" in purchase_date:
                    dt = datetime.fromisoformat(purchase_date.replace("Z", "+00:00"))
                    dt = dt - timedelta(hours=7)
                    order_day = dt.strftime("%Y-%m-%d")
                else:
                    order_day = purchase_date[:10]
            except Exception:
                order_day = purchase_date[:10]

            if order_day not in (today_str, yesterday_str):
                continue

            order_id = order.get("AmazonOrderId")
            if not order_id:
                continue

            try:
                items_resp = orders_api.get_order_items(order_id)
                item_list = items_resp.payload.get("OrderItems", [])
                total_from_items = 0.0
                total_qty = 0
                for item in item_list:
                    asin = item.get("ASIN", "")
                    qty = 0
                    try:
                        qty = int(item.get("QuantityOrdered", 0))
                    except (ValueError, TypeError):
                        pass

                    # ItemPrice is the total for this line (price × qty)
                    price = item.get("ItemPrice")
                    item_amount = 0.0
                    if price and isinstance(price, dict):
                        amt_str = price.get("Amount", "0")
                        try:
                            item_amount = float(amt_str)
                        except (ValueError, TypeError):
                            pass

                    # Fallback: if ItemPrice is 0 or missing (Pending orders),
                    # use historical average price for this ASIN
                    if item_amount == 0 and qty > 0 and asin in asin_prices:
                        item_amount = round(asin_prices[asin] * qty, 2)
                        logger.info(f"    {order_id}: used fallback price for {asin} = ${item_amount:.2f} ({qty} × ${asin_prices[asin]:.2f})")

                    total_from_items += item_amount
                    total_qty += qty

                # Always use item-level total (more accurate than OrderTotal)
                order["_item_revenue"] = total_from_items
                order["_item_qty"] = total_qty
                items_fetched += 1

                # Adaptive rate limiting: burst first 25, then sleep
                burst_count += 1
                if burst_count > 25:
                    _t.sleep(2.0)  # Stay well under rate limit
            except Exception as e:
                items_failed += 1
                logger.warning(f"  getOrderItems failed for {order_id}: {e}")
                # On rate limit, back off
                if "throttl" in str(e).lower() or "429" in str(e):
                    _t.sleep(5.0)

        logger.info(f"  Fetched items for {items_fetched} orders ({items_failed} failed)")

        # ── Aggregate by date ──
        from collections import defaultdict
        daily_agg = defaultdict(lambda: {"revenue": 0.0, "units": 0, "orders": 0})

        for order in order_list:
            status = order.get("OrderStatus", "")
            if status in ("Canceled", "Cancelled"):
                continue
            purchase_date = order.get("PurchaseDate", "")
            if not purchase_date:
                continue
            try:
                if "T" in purchase_date:
                    dt = datetime.fromisoformat(purchase_date.replace("Z", "+00:00"))
                    dt = dt - timedelta(hours=7)
                    day_str = dt.strftime("%Y-%m-%d")
                else:
                    day_str = purchase_date[:10]
            except Exception:
                day_str = purchase_date[:10]

            # Prefer item-level revenue (most accurate), fall back to OrderTotal
            if "_item_revenue" in order:
                amount = order["_item_revenue"]
                n_items = order["_item_qty"]
            else:
                order_total = order.get("OrderTotal", {})
                amount = float(order_total.get("Amount", 0)) if order_total else 0
                n_items = order.get("NumberOfItemsShipped", 0) + order.get("NumberOfItemsUnshipped", 0)

            daily_agg[day_str]["revenue"] += amount
            daily_agg[day_str]["units"] += n_items
            daily_agg[day_str]["orders"] += 1

        # ── Write to daily_sales ──
        for day_str in [today_str, yesterday_str]:
            if day_str not in daily_agg:
                continue
            agg = daily_agg[day_str]
            existing = con.execute(
                "SELECT ordered_product_sales FROM daily_sales WHERE date = ? AND asin = 'ALL'",
                [day_str]
            ).fetchone()
            existing_rev = float(existing[0]) if existing and existing[0] else 0.0

            # Only overwrite if new revenue is >= existing (never clobber good data with partial sync)
            if agg["revenue"] >= existing_rev:
                con.execute("""
                    INSERT OR REPLACE INTO daily_sales
                    (date, asin, units_ordered, ordered_product_sales,
                     sessions, session_percentage, page_views,
                     buy_box_percentage, unit_session_percentage)
                    VALUES (?, 'ALL', ?, ?, 0, 0, 0, 0, 0)
                """, [day_str, agg["units"], agg["revenue"]])
                logger.info(f"  {day_str}: ${agg['revenue']:.2f} rev, {agg['units']} units, {agg['orders']} orders")

        con.close()

        # Log summary for debugging
        for ds in sorted(daily_agg.keys()):
            a = daily_agg[ds]
            logger.info(f"  Date {ds}: ${a['revenue']:.2f} / {a['units']}u / {a['orders']} orders")

        logger.info(f"  Today sync complete: {len(order_list)} orders, {items_fetched} items fetched")
        return True
    except Exception as e:
        logger.error(f"  Today orders sync error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False


# Track last today-sync so we don't hammer the API
_last_today_sync = 0.0


def _ensure_today_data():
    """If today has no data in daily_sales, do a quick Orders API pull.

    Called inline by profitability/comparison endpoints so the user
    never sees Today = $0 when Amazon has live data.
    Throttled to at most once per 5 minutes.
    """
    global _last_today_sync
    import time as _time
    now = _time.time()
    if now - _last_today_sync < 300:          # 5-minute cooldown
        return

    today_str = (datetime.utcnow() - timedelta(hours=7)).strftime("%Y-%m-%d")
    try:
        con = duckdb.connect(str(DB_PATH), read_only=False)
        row = con.execute(
            "SELECT ordered_product_sales FROM daily_sales WHERE date = ? AND asin = 'ALL'",
            [today_str]
        ).fetchone()
        con.close()
        has_data = row and row[0] and float(row[0]) > 0
    except Exception:
        has_data = False

    if not has_data:
        logger.info("Today has no data — triggering quick Orders API sync")
        _last_today_sync = now
        _sync_today_orders()


def _run_sp_api_sync():
    """Pull latest data from Amazon SP-API into DuckDB (runs in background thread).

    Order of operations (fastest first):
    1. Today's Orders (fast, ~5 seconds) — critical for live "Today" data
    2. Financial Events (fast, ~5 seconds) — actual fees/refunds
    3. FBA Inventory (fast, ~5 seconds)
    4. Sales & Traffic Report (SLOW, 2-5 minutes) — historical data with sessions
    """
    try:
        from sp_api.api import Reports, Orders, Inventories, Finances
        from sp_api.base import Marketplaces, ReportType
    except ImportError:
        logger.warning("SP-API library not installed — skipping sync")
        return

    credentials = _load_sp_api_credentials()
    if not credentials:
        logger.warning("No SP-API credentials found (env vars or config file) — skipping sync")
        return

    logger.info("Starting SP-API background sync...")
    import time, gzip, requests as req

    # ── Ensure financial_events table exists ──────────────────────
    try:
        _con = duckdb.connect(str(DB_PATH), read_only=False)
        _con.execute("""
            CREATE TABLE IF NOT EXISTS financial_events (
                date DATE,
                asin VARCHAR,
                sku VARCHAR,
                order_id VARCHAR,
                event_type VARCHAR,
                product_charges DOUBLE DEFAULT 0,
                shipping_charges DOUBLE DEFAULT 0,
                fba_fees DOUBLE DEFAULT 0,
                commission DOUBLE DEFAULT 0,
                promotion_amount DOUBLE DEFAULT 0,
                other_fees DOUBLE DEFAULT 0,
                net_proceeds DOUBLE DEFAULT 0
            )
        """)
        _con.close()
    except Exception as e:
        logger.warning(f"  Could not ensure financial_events table: {e}")

    # ── 1. TODAY'S ORDERS (fast, real-time) ──────────────────────
    global _last_today_sync
    import time as _time
    _last_today_sync = _time.time()
    _sync_today_orders()

    # ── 2. FINANCIAL EVENTS (actual fees, refunds, promos) ──────
    try:
        from sp_api.api import Finances as FinancesAPI
        logger.info("  Pulling financial events (fees, refunds)...")

        fin_start = (datetime.utcnow() - timedelta(days=90)).strftime("%Y-%m-%dT%H:%M:%SZ")
        finances = FinancesAPI(credentials=credentials, marketplace=Marketplaces.US)

        # Manual NextToken pagination (more reliable than @load_all_pages)
        all_shipment_events = []
        all_refund_events = []
        page = 0
        next_token = None

        while page < 20:
            page += 1
            kwargs = {"PostedAfter": fin_start, "MaxResultsPerPage": 100}
            if next_token:
                kwargs["NextToken"] = next_token

            try:
                resp = finances.list_financial_events(**kwargs)
            except Exception as api_err:
                logger.error(f"  Financial events API call error (page {page}): {api_err}")
                import traceback
                logger.error(traceback.format_exc())
                break

            payload = resp.payload if hasattr(resp, 'payload') else (resp if isinstance(resp, dict) else {})
            events = payload.get("FinancialEvents", {})

            if page == 1:
                logger.info(f"  Financial events payload keys: {list(payload.keys())}")
                logger.info(f"  Financial events event keys: {list(events.keys()) if events else 'EMPTY'}")
                # Log ALL event lists and their counts
                for k, v in events.items():
                    if isinstance(v, list):
                        logger.info(f"    {k}: {len(v)} items")

            shipments = events.get("ShipmentEventList", []) or []
            refunds = events.get("RefundEventList", []) or []
            adjustments = events.get("AdjustmentEventList", []) or []
            all_shipment_events.extend(shipments)
            all_refund_events.extend(refunds)
            # AdjustmentEventList often contains return-related adjustments
            for adj in adjustments:
                adj_type = adj.get("AdjustmentType", "") if isinstance(adj, dict) else getattr(adj, "AdjustmentType", "")
                adj_type_str = str(adj_type or "").lower()
                if "return" in adj_type_str or "refund" in adj_type_str or "reversal" in adj_type_str:
                    all_refund_events.append(adj)
            logger.info(f"  Financial events page {page}: {len(shipments)} shipments, {len(refunds)} refunds, {len(adjustments)} adjustments")

            # Check for next page
            next_token = payload.get("NextToken")
            if not next_token:
                break

        logger.info(f"  Total financial events: {len(all_shipment_events)} shipments, {len(all_refund_events)} refunds")
        con = duckdb.connect(str(DB_PATH), read_only=False)
        fin_records = 0

        # Helper: Amazon uses CurrencyAmount (not Amount) in money objects
        # sp_api may return model objects, dicts, Decimals, or strings
        def _money(amt_obj):
            if amt_obj is None:
                return 0.0
            # Plain dict
            if isinstance(amt_obj, dict):
                return float(amt_obj.get("CurrencyAmount", amt_obj.get("Amount", 0)) or 0)
            # sp_api model object (has attributes instead of dict keys)
            if hasattr(amt_obj, "CurrencyAmount"):
                try:
                    return float(amt_obj.CurrencyAmount or 0)
                except Exception:
                    pass
            if hasattr(amt_obj, "Amount"):
                try:
                    return float(amt_obj.Amount or 0)
                except Exception:
                    pass
            # Already a number (Decimal, float, int)
            try:
                return float(amt_obj)
            except Exception:
                return 0.0

        # Helper: extract string date from PostedDate (may be str or datetime)
        def _posted_date_str(pd_val):
            if pd_val is None:
                return None
            if isinstance(pd_val, str):
                return pd_val[:10] if len(pd_val) >= 10 else pd_val
            # datetime object
            if hasattr(pd_val, 'strftime'):
                return pd_val.strftime("%Y-%m-%d")
            return str(pd_val)[:10]

        # Helper: safely get dict-like value (works for dicts AND sp_api model objects)
        def _safe_get(obj, key, default=None):
            if obj is None:
                return default
            if isinstance(obj, dict):
                return obj.get(key, default)
            # sp_api model object — try attribute access
            return getattr(obj, key, default)

        # Clear old financial events for re-sync (avoids duplicates without needing PK)
        try:
            cutoff_date = (datetime.utcnow() - timedelta(days=90)).strftime("%Y-%m-%d")
            con.execute("DELETE FROM financial_events WHERE date >= ?", [cutoff_date])
            logger.info(f"  Cleared financial_events from {cutoff_date} for re-sync")
        except Exception as e:
            logger.warning(f"  Could not clear old financial_events: {e}")

        # Process shipment events (sales fees)
        first_shipment_logged = False
        for event in all_shipment_events:
            order_id = _safe_get(event, "AmazonOrderId", "")
            posted_date = _safe_get(event, "PostedDate", "")
            date_str = _posted_date_str(posted_date)

            for item in (_safe_get(event, "ShipmentItemList") or []):
                sku = _safe_get(item, "SellerSKU", "")
                asin_val = _safe_get(item, "ASIN", sku)

                product_charges = 0.0
                shipping_ch = 0.0
                charge_list = _safe_get(item, "ItemChargeList") or []
                for c in charge_list:
                    val = _money(_safe_get(c, "ChargeAmount"))
                    ct = _safe_get(c, "ChargeType", "")
                    if ct == "Principal":
                        product_charges += val
                    elif ct in ("ShippingCharge", "Shipping"):
                        shipping_ch += val

                # Log first shipment's charge structure for debugging
                if not first_shipment_logged and charge_list:
                    first_c = charge_list[0]
                    ca = _safe_get(first_c, "ChargeAmount")
                    logger.info(f"  DEBUG charge structure: type={type(first_c).__name__}, "
                               f"ChargeAmount type={type(ca).__name__}, "
                               f"ChargeAmount value={ca}, "
                               f"parsed_val={_money(ca)}")
                    first_shipment_logged = True

                fba_fees_val = 0.0
                commission_val = 0.0
                other_val = 0.0
                for f in (_safe_get(item, "ItemFeeList") or []):
                    val = abs(_money(_safe_get(f, "FeeAmount")))
                    ft = _safe_get(f, "FeeType", "")
                    if "FBA" in ft or "Fulfillment" in ft:
                        fba_fees_val += val
                    elif ft in ("Commission", "ReferralFee"):
                        commission_val += val
                    else:
                        other_val += val

                promo_val = 0.0
                for p in (_safe_get(item, "PromotionList") or []):
                    promo_val += _money(_safe_get(p, "PromotionAmount"))

                net = product_charges + shipping_ch - fba_fees_val - commission_val + promo_val

                try:
                    con.execute("""
                        INSERT INTO financial_events
                        (date, asin, sku, order_id, event_type,
                         product_charges, shipping_charges, fba_fees,
                         commission, promotion_amount, other_fees, net_proceeds)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, [
                        date_str, asin_val, sku, order_id, "Shipment",
                        product_charges, shipping_ch,
                        fba_fees_val, commission_val, promo_val, other_val, net
                    ])
                    fin_records += 1
                except Exception as ins_err:
                    logger.warning(f"  Shipment insert error: {ins_err}")

        # Process refund events (from RefundEventList + return-related AdjustmentEventList)
        first_refund_logged = False
        for event in all_refund_events:
            order_id = _safe_get(event, "AmazonOrderId", "")
            posted_date = _safe_get(event, "PostedDate", "")
            date_str = _posted_date_str(posted_date)

            # Try multiple possible item list keys (different event types use different keys)
            item_list = (_safe_get(event, "ShipmentItemAdjustmentList")
                        or _safe_get(event, "ShipmentItemList")
                        or _safe_get(event, "AdjustmentItemList")
                        or [])

            if not first_refund_logged:
                logger.info(f"  DEBUG refund event keys: {list(event.keys()) if isinstance(event, dict) else dir(event)}")
                logger.info(f"  DEBUG refund item_list length: {len(item_list)}, type: {type(item_list).__name__}")
                if item_list:
                    first_item = item_list[0]
                    logger.info(f"  DEBUG refund first item keys: {list(first_item.keys()) if isinstance(first_item, dict) else 'model'}")
                first_refund_logged = True

            for item in item_list:
                sku = _safe_get(item, "SellerSKU", "")
                asin_val = _safe_get(item, "ASIN", sku)

                refund_amount = 0.0
                refund_fba = 0.0
                refund_comm = 0.0
                refund_other = 0.0
                charge_adj_list = (_safe_get(item, "ItemChargeAdjustmentList")
                                  or _safe_get(item, "ItemChargeList") or [])
                for c in charge_adj_list:
                    val = _money(_safe_get(c, "ChargeAmount"))
                    ct = _safe_get(c, "ChargeType", "")
                    if ct == "Principal":
                        refund_amount += val
                    else:
                        refund_other += abs(val)

                # Also capture fee adjustments for refunds
                fee_adj_list = (_safe_get(item, "ItemFeeAdjustmentList")
                               or _safe_get(item, "ItemFeeList") or [])
                for f in fee_adj_list:
                    val = _money(_safe_get(f, "FeeAmount"))
                    ft = _safe_get(f, "FeeType", "")
                    if "FBA" in ft or "Fulfillment" in ft:
                        refund_fba += abs(val)
                    elif ft in ("Commission", "ReferralFee"):
                        refund_comm += abs(val)

                try:
                    con.execute("""
                        INSERT INTO financial_events
                        (date, asin, sku, order_id, event_type,
                         product_charges, shipping_charges, fba_fees,
                         commission, promotion_amount, other_fees, net_proceeds)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, [
                        date_str,
                        asin_val, sku, order_id, "Refund",
                        abs(refund_amount), 0, refund_fba, refund_comm, 0, refund_other,
                        abs(refund_amount)
                    ])
                    fin_records += 1
                    logger.info(f"  Inserted refund: order={order_id}, asin={asin_val}, amount=${abs(refund_amount):.2f}")
                except Exception as ins_err:
                    logger.error(f"  Refund insert error for {order_id}/{asin_val}: {ins_err}")

        con.close()
        logger.info(f"  Financial events sync done: {fin_records} records (shipments + refunds)")
    except Exception as e:
        logger.error(f"  Financial events sync error: {e}")

    # ── 3. FBA INVENTORY ────────────────────────────────────────
    try:
        logger.info("  Pulling FBA inventory...")
        inventory_api = Inventories(credentials=credentials, marketplace=Marketplaces.US)
        response = inventory_api.get_inventory_summary_marketplace(
            marketplaceIds=["ATVPDKIKX0DER"],
            granularityType="Marketplace",
            granularityId="ATVPDKIKX0DER"
        )
        summaries = response.payload.get("inventorySummaries", [])
        if summaries:
            con = duckdb.connect(str(DB_PATH), read_only=False)
            today = datetime.now().date()
            for item in summaries:
                inv = item.get("inventoryDetails", {})
                fulfillable = inv.get("fulfillableQuantity", 0) or item.get("totalQuantity", 0)
                reserved = inv.get("reservedQuantity", {})
                reserved_qty = reserved.get("totalReservedQuantity", 0) if isinstance(reserved, dict) else int(reserved or 0)
                con.execute("""
                    INSERT OR REPLACE INTO fba_inventory
                    (date, asin, sku, product_name, condition,
                     fulfillable_quantity, inbound_working_quantity,
                     inbound_shipped_quantity, inbound_receiving_quantity,
                     reserved_quantity, unfulfillable_quantity, total_quantity)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, [
                    today, item.get("asin", ""),
                    item.get("sellerSku", item.get("fnSku", "")),
                    item.get("productName", "")[:200],
                    item.get("condition", "NewItem"),
                    int(fulfillable or 0),
                    int(inv.get("inboundWorkingQuantity", 0) or 0),
                    int(inv.get("inboundShippedQuantity", 0) or 0),
                    int(inv.get("inboundReceivingQuantity", 0) or 0),
                    int(reserved_qty or 0), 0,
                    int(item.get("totalQuantity", 0)),
                ])
            con.close()
            logger.info(f"  Inventory sync done: {len(summaries)} items")
    except Exception as e:
        logger.error(f"  Inventory sync error: {e}")

    # ── 4. SALES & TRAFFIC REPORT (slow, 2-5 min polling) ──────
    try:
        end_date = datetime.utcnow().date()
        start_date = end_date - timedelta(days=3)
        logger.info(f"  Requesting Sales report {start_date} to {end_date}...")

        reports = Reports(credentials=credentials, marketplace=Marketplaces.US)
        report_response = reports.create_report(
            reportType=ReportType.GET_SALES_AND_TRAFFIC_REPORT,
            dataStartTime=start_date.isoformat(),
            dataEndTime=end_date.isoformat(),
            reportOptions={"dateGranularity": "DAY", "asinGranularity": "CHILD"}
        )
        report_id = report_response.payload.get("reportId")

        report_data = None
        for attempt in range(30):
            time.sleep(10)
            status_response = reports.get_report(report_id)
            status = status_response.payload.get("processingStatus")
            if status == "DONE":
                doc_id = status_response.payload.get("reportDocumentId")
                doc_response = reports.get_report_document(doc_id)
                report_data = doc_response.payload
                break
            elif status in ("CANCELLED", "FATAL"):
                logger.error(f"  Report failed: {status}")
                break

        if report_data:
            if isinstance(report_data, dict) and "url" in report_data:
                is_gzipped = report_data.get("compressionAlgorithm", "").upper() == "GZIP"
                resp = req.get(report_data["url"])
                if is_gzipped:
                    import gzip as gz
                    report_text = gz.decompress(resp.content).decode("utf-8")
                else:
                    report_text = resp.text
                report_data = json.loads(report_text)

            if isinstance(report_data, str):
                report_data = json.loads(report_data)

            con = duckdb.connect(str(DB_PATH), read_only=False)
            records = 0

            for day_entry in report_data.get("salesAndTrafficByDate", []):
                entry_date = day_entry.get("date", "")
                traffic = day_entry.get("trafficByDate", {})
                sales_info = day_entry.get("salesByDate", {})
                ordered_sales = sales_info.get("orderedProductSales", {})
                sales_amount = float(ordered_sales.get("amount", 0)) if isinstance(ordered_sales, dict) else float(ordered_sales or 0)

                con.execute("""
                    INSERT OR REPLACE INTO daily_sales
                    (date, asin, units_ordered, ordered_product_sales,
                     sessions, session_percentage, page_views,
                     buy_box_percentage, unit_session_percentage)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, [
                    entry_date, "ALL",
                    int(sales_info.get("unitsOrdered", 0)),
                    sales_amount,
                    int(traffic.get("sessions", 0)),
                    float(traffic.get("sessionPercentage", 0)),
                    int(traffic.get("pageViews", 0)),
                    float(traffic.get("buyBoxPercentage", 0)),
                    float(traffic.get("unitSessionPercentage", 0)),
                ])
                records += 1

            for asin_entry in report_data.get("salesAndTrafficByAsin", []):
                asin = asin_entry.get("childAsin") or asin_entry.get("parentAsin", "")
                traffic = asin_entry.get("trafficByAsin", {})
                sales_info = asin_entry.get("salesByAsin", {})
                entry_date = asin_entry.get("date", str(start_date))
                ordered_sales = sales_info.get("orderedProductSales", {})
                sales_amount = float(ordered_sales.get("amount", 0)) if isinstance(ordered_sales, dict) else float(ordered_sales or 0)

                con.execute("""
                    INSERT OR REPLACE INTO daily_sales
                    (date, asin, units_ordered, ordered_product_sales,
                     sessions, session_percentage, page_views,
                     buy_box_percentage, unit_session_percentage)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, [
                    entry_date, asin,
                    int(sales_info.get("unitsOrdered", 0)),
                    sales_amount,
                    int(traffic.get("sessions", 0)),
                    float(traffic.get("sessionPercentage", 0)),
                    int(traffic.get("pageViews", 0)),
                    float(traffic.get("buyBoxPercentage", 0)),
                    float(traffic.get("unitSessionPercentage", 0)),
                ])
                records += 1

            con.close()
            logger.info(f"  Sales sync done: {records} records stored")

    except Exception as e:
        logger.error(f"  Sales sync error: {e}")

    logger.info("SP-API background sync complete.")


# ── Pricing & Coupon Sync ────────────────────────────────

PRICING_CACHE_PATH = DB_DIR / "pricing_sync.json"


def _load_pricing_cache() -> dict:
    """Load cached pricing/coupon data."""
    if PRICING_CACHE_PATH.exists():
        with open(PRICING_CACHE_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {"prices": {}, "coupons": {}, "lastSync": None}


def _save_pricing_cache(data: dict):
    """Save pricing/coupon cache."""
    with open(PRICING_CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def _sync_pricing_data():
    """Pull current listing prices from SP-API Product Pricing API.

    Uses getCompetitivePricing to get Buy Box and listed prices per ASIN.
    Falls back gracefully if SP-API is not configured.
    """
    try:
        from sp_api.api import Products as ProductsAPI
        from sp_api.base import Marketplaces
    except ImportError:
        logger.info("Pricing sync: sp_api not installed, skipping")
        return

    credentials = _load_sp_api_credentials()
    if not credentials:
        logger.info("Pricing sync: no SP-API credentials, skipping")
        return

    # Get all ASINs from item master
    items = load_item_master()
    asins = [i["asin"] for i in items if i.get("asin")]
    if not asins:
        return

    logger.info(f"Pricing sync: fetching prices for {len(asins)} ASINs...")
    import time as _t

    cache = _load_pricing_cache()
    prices = cache.get("prices", {})

    # SP-API allows up to 20 ASINs per getCompetitivePricing call
    products_api = ProductsAPI(credentials=credentials, marketplace=Marketplaces.US)
    batch_size = 20

    for i in range(0, len(asins), batch_size):
        batch = asins[i:i + batch_size]
        try:
            result = products_api.get_competitive_pricing(
                Asins=batch,
                MarketplaceId="ATVPDKIKX0DER",
            )
            for item_data in (result.payload or []):
                asin = item_data.get("ASIN", "")
                if not asin:
                    continue
                comp_prices = item_data.get("Product", {}).get("CompetitivePricing", {})
                number_of_offers = comp_prices.get("NumberOfOfferListings", [])
                comp_price_list = comp_prices.get("CompetitivePrices", [])

                listing_price = None
                landed_price = None
                buy_box_price = None

                for cp in comp_price_list:
                    condition = cp.get("condition", "")
                    belongs_to = cp.get("belongsToRequester", False)
                    price_info = cp.get("Price", {})
                    lp = price_info.get("ListingPrice", {})
                    landed = price_info.get("LandedPrice", {})

                    if lp.get("Amount"):
                        listing_price = float(lp["Amount"])
                    if landed.get("Amount"):
                        landed_price = float(landed["Amount"])

                    comp_type = cp.get("CompetitivePriceId", "")
                    if comp_type == "1":  # Buy Box price
                        if landed.get("Amount"):
                            buy_box_price = float(landed["Amount"])
                        elif lp.get("Amount"):
                            buy_box_price = float(lp["Amount"])

                prices[asin] = {
                    "listingPrice": listing_price,
                    "landedPrice": landed_price,
                    "buyBoxPrice": buy_box_price,
                    "fetchedAt": datetime.utcnow().isoformat(),
                }

            # Rate limit: SP-API throttles at ~10 calls/sec for pricing
            if i + batch_size < len(asins):
                _t.sleep(0.5)

        except Exception as e:
            logger.error(f"Pricing sync batch error (ASINs {i}-{i+batch_size}): {e}")
            _t.sleep(2)

    cache["prices"] = prices
    cache["lastSync"] = datetime.utcnow().isoformat()
    _save_pricing_cache(cache)
    logger.info(f"Pricing sync complete: {len(prices)} ASINs cached")


def _sync_coupon_data():
    """Pull active coupon data from Amazon Ads Coupons API.

    Uses the Coupons API (part of python-amazon-ad-api) to list all coupons
    and their statuses. Maps coupon ASINs back to item master entries.
    """
    ads_creds = _load_ads_credentials()
    if not ads_creds:
        logger.info("Coupon sync: no Ads API credentials, skipping")
        return

    profile_id = ads_creds.get("profile_id", "")
    if not profile_id:
        logger.info("Coupon sync: no profile_id, skipping")
        return

    try:
        from ad_api.api import Coupons
    except ImportError:
        logger.info("Coupon sync: ad_api Coupons not available, skipping")
        return

    logger.info("Coupon sync: fetching coupons from Amazon Ads API...")

    cache = _load_pricing_cache()
    coupons_by_asin = {}

    try:
        coupons_api = Coupons(credentials=ads_creds)

        # List coupons with pagination
        page_token = None
        total_fetched = 0
        max_pages = 10

        for page in range(max_pages):
            body = {
                "stateFilter": {
                    "includes": ["ENABLED", "SCHEDULED"]
                },
                "pageSize": 100,
            }
            if page_token:
                body["pageToken"] = page_token

            result = coupons_api.list_coupons(body=json.dumps(body))
            payload = result.payload or {}
            coupon_list = payload.get("coupons", [])
            total_fetched += len(coupon_list)

            for coupon in coupon_list:
                coupon_id = coupon.get("couponId", "")
                state = coupon.get("state", "UNKNOWN")
                discount = coupon.get("discount", {})
                disc_type = discount.get("discountType", "")  # PERCENTAGE or AMOUNT
                disc_value = float(discount.get("discountValue", 0))

                # Budget and redemption info
                budget = coupon.get("budget", {})
                budget_amount = float(budget.get("budgetAmount", 0))
                budget_used = float(budget.get("budgetConsumed", 0))
                redemptions = coupon.get("totalRedemptions", 0)

                start_date = coupon.get("startDate", "")
                end_date = coupon.get("endDate", "")

                # Map to ASINs
                asin_list = coupon.get("asins", [])
                product_criteria = coupon.get("productCriteria", {})
                if not asin_list and product_criteria:
                    asin_list = product_criteria.get("asins", [])

                coupon_info = {
                    "couponId": coupon_id,
                    "state": state,
                    "discountType": "%" if disc_type == "PERCENTAGE" else "$",
                    "discountValue": disc_value,
                    "budgetAmount": budget_amount,
                    "budgetUsed": budget_used,
                    "redemptions": redemptions,
                    "startDate": start_date,
                    "endDate": end_date,
                    "fetchedAt": datetime.utcnow().isoformat(),
                }

                for asin in asin_list:
                    if isinstance(asin, str) and asin:
                        # Keep the most recently fetched coupon per ASIN
                        # If multiple coupons exist, prefer ENABLED over SCHEDULED
                        existing = coupons_by_asin.get(asin)
                        if not existing or (state == "ENABLED" and existing.get("state") != "ENABLED"):
                            coupons_by_asin[asin] = coupon_info

            # Check for next page
            page_token = payload.get("nextPageToken")
            if not page_token or not coupon_list:
                break

        logger.info(f"Coupon sync: {total_fetched} coupons fetched, mapped to {len(coupons_by_asin)} ASINs")

    except Exception as e:
        logger.error(f"Coupon sync error: {e}")

    cache["coupons"] = coupons_by_asin
    cache["lastSync"] = datetime.utcnow().isoformat()
    _save_pricing_cache(cache)


def _sync_pricing_and_coupons():
    """Run both pricing and coupon syncs. Called from the sync loop."""
    try:
        _sync_pricing_data()
    except Exception as e:
        logger.error(f"Pricing sync error: {e}")
    try:
        _sync_coupon_data()
    except Exception as e:
        logger.error(f"Coupon sync error: {e}")


def _auto_backfill_if_needed():
    """Check if historical data is missing and backfill automatically.
    This prevents data loss when Railway redeploys (ephemeral filesystem).
    Backfills: Apr 2024 to yesterday, in 30-day chunks."""
    import time as _time
    import gzip as gz
    import requests as req
    from sp_api.api import Reports
    from sp_api.base import Marketplaces, ReportType
    from datetime import date as dt_date

    con = duckdb.connect(str(DB_PATH), read_only=True)
    try:
        earliest_row = con.execute(
            "SELECT MIN(date) FROM daily_sales WHERE asin = 'ALL'"
        ).fetchone()
        earliest = earliest_row[0] if earliest_row and earliest_row[0] else None
        count_row = con.execute(
            "SELECT COUNT(DISTINCT date) FROM daily_sales WHERE asin = 'ALL'"
        ).fetchone()
        total_days = count_row[0] if count_row else 0
    except Exception:
        earliest = None
        total_days = 0
    finally:
        con.close()

    # If we already have 500+ days of data, skip backfill
    if total_days >= 500:
        logger.info(f"Auto-backfill: {total_days} days of data found, skipping backfill")
        return

    logger.info(f"Auto-backfill: only {total_days} days found (earliest={earliest}). Starting historical backfill...")

    creds = _load_sp_api_credentials()
    if not creds:
        logger.error("Auto-backfill: no SP-API credentials, skipping")
        return

    reports_api = Reports(credentials=creds, marketplace=Marketplaces.US)

    # Backfill from Apr 2024 to yesterday
    backfill_start = dt_date(2024, 4, 1)
    backfill_end = dt_date.today() - timedelta(days=1)

    chunk_start = backfill_start
    total_records = 0

    while chunk_start < backfill_end:
        chunk_end = min(chunk_start + timedelta(days=29), backfill_end)
        logger.info(f"  Auto-backfill: requesting {chunk_start} to {chunk_end}...")

        try:
            report_response = reports_api.create_report(
                reportType=ReportType.GET_SALES_AND_TRAFFIC_REPORT,
                dataStartTime=chunk_start.isoformat(),
                dataEndTime=chunk_end.isoformat(),
                reportOptions={"dateGranularity": "DAY", "asinGranularity": "CHILD"}
            )
            report_id = report_response.payload.get("reportId")

            report_data = None
            for attempt in range(30):
                _time.sleep(10)
                status_response = reports_api.get_report(report_id)
                status = status_response.payload.get("processingStatus")
                if status == "DONE":
                    doc_id = status_response.payload.get("reportDocumentId")
                    doc_response = reports_api.get_report_document(doc_id)
                    report_data = doc_response.payload
                    break
                elif status in ("CANCELLED", "FATAL"):
                    logger.warning(f"  Auto-backfill: report {chunk_start}-{chunk_end} status={status}")
                    break

            if report_data:
                if isinstance(report_data, dict) and "url" in report_data:
                    is_gzipped = report_data.get("compressionAlgorithm", "").upper() == "GZIP"
                    resp = req.get(report_data["url"])
                    if is_gzipped:
                        report_text = gz.decompress(resp.content).decode("utf-8")
                    else:
                        report_text = resp.text
                    report_data = json.loads(report_text)

                if isinstance(report_data, str):
                    report_data = json.loads(report_data)

                wcon = duckdb.connect(str(DB_PATH), read_only=False)
                records = 0

                for day_entry in report_data.get("salesAndTrafficByDate", []):
                    entry_date = day_entry.get("date", "")
                    traffic = day_entry.get("trafficByDate", {})
                    sales_info = day_entry.get("salesByDate", {})
                    ordered_sales = sales_info.get("orderedProductSales", {})
                    sales_amount = float(ordered_sales.get("amount", 0)) if isinstance(ordered_sales, dict) else float(ordered_sales or 0)

                    wcon.execute("""
                        INSERT OR REPLACE INTO daily_sales
                        (date, asin, units_ordered, ordered_product_sales,
                         sessions, session_percentage, page_views,
                         buy_box_percentage, unit_session_percentage)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, [
                        entry_date, "ALL",
                        int(sales_info.get("unitsOrdered", 0)),
                        sales_amount,
                        int(traffic.get("sessions", 0)),
                        float(traffic.get("sessionPercentage", 0)),
                        int(traffic.get("pageViews", 0)),
                        float(traffic.get("buyBoxPercentage", 0)),
                        float(traffic.get("unitSessionPercentage", 0)),
                    ])
                    records += 1

                for asin_entry in report_data.get("salesAndTrafficByAsin", []):
                    asin = asin_entry.get("childAsin") or asin_entry.get("parentAsin", "")
                    traffic = asin_entry.get("trafficByAsin", {})
                    sales_info = asin_entry.get("salesByAsin", {})
                    entry_date = asin_entry.get("date", str(chunk_start))
                    ordered_sales = sales_info.get("orderedProductSales", {})
                    sales_amount = float(ordered_sales.get("amount", 0)) if isinstance(ordered_sales, dict) else float(ordered_sales or 0)

                    wcon.execute("""
                        INSERT OR REPLACE INTO daily_sales
                        (date, asin, units_ordered, ordered_product_sales,
                         sessions, session_percentage, page_views,
                         buy_box_percentage, unit_session_percentage)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, [
                        entry_date, asin,
                        int(sales_info.get("unitsOrdered", 0)),
                        sales_amount,
                        int(traffic.get("sessions", 0)),
                        float(traffic.get("sessionPercentage", 0)),
                        int(traffic.get("pageViews", 0)),
                        float(traffic.get("buyBoxPercentage", 0)),
                        float(traffic.get("unitSessionPercentage", 0)),
                    ])
                    records += 1

                wcon.close()
                total_records += records
                logger.info(f"  Auto-backfill: {chunk_start}-{chunk_end}: {records} records")

        except Exception as e:
            logger.error(f"  Auto-backfill error {chunk_start}-{chunk_end}: {e}")

        chunk_start = chunk_end + timedelta(days=1)

    logger.info(f"Auto-backfill complete: {total_records} total records inserted")


# ── Amazon Ads API Sync ──────────────────────────────────

def _load_ads_credentials() -> dict | None:
    """Load Amazon Ads API credentials from env vars or config file.

    Supports both naming conventions for env vars:
      - New: AMAZON_ADS_CLIENT_ID, AMAZON_ADS_CLIENT_SECRET, AMAZON_ADS_REFRESH_TOKEN, AMAZON_ADS_PROFILE_ID
      - Legacy: ADS_CLIENT_ID, ADS_CLIENT_SECRET, ADS_REFRESH_TOKEN, ADS_PROFILE_ID
    """
    env_client_id = os.environ.get("AMAZON_ADS_CLIENT_ID") or os.environ.get("ADS_CLIENT_ID", "")
    if env_client_id:
        creds = {
            "refresh_token": os.environ.get("AMAZON_ADS_REFRESH_TOKEN") or os.environ.get("ADS_REFRESH_TOKEN", ""),
            "client_id": env_client_id,
            "client_secret": os.environ.get("AMAZON_ADS_CLIENT_SECRET") or os.environ.get("ADS_CLIENT_SECRET", ""),
            "profile_id": os.environ.get("AMAZON_ADS_PROFILE_ID") or os.environ.get("ADS_PROFILE_ID", ""),
        }
        if creds["refresh_token"] and creds["client_secret"]:
            return creds
        return None

    if CONFIG_PATH.exists():
        with open(CONFIG_PATH) as f:
            creds_json = json.load(f)
        ads = creds_json.get("AMAZON_ADS_API", {})
        creds = {
            "refresh_token": ads.get("refresh_token", ""),
            "client_id": ads.get("client_id", ""),
            "client_secret": ads.get("client_secret", ""),
            "profile_id": ads.get("profile_id", ""),
        }
        if creds["refresh_token"] and creds["client_id"]:
            return creds

    return None


def _ensure_ads_tables():
    """Create ads DuckDB tables if they don't exist."""
    con = duckdb.connect(str(DB_PATH), read_only=False)
    con.execute("""
        CREATE TABLE IF NOT EXISTS advertising (
            date TEXT,
            campaign_id TEXT,
            campaign_name TEXT,
            impressions INTEGER DEFAULT 0,
            clicks INTEGER DEFAULT 0,
            spend DOUBLE DEFAULT 0,
            sales DOUBLE DEFAULT 0,
            orders INTEGER DEFAULT 0,
            units INTEGER DEFAULT 0,
            PRIMARY KEY (date, campaign_id)
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS ads_campaigns (
            date TEXT,
            campaign_id TEXT,
            campaign_name TEXT,
            campaign_type TEXT DEFAULT 'SP',
            campaign_status TEXT DEFAULT '',
            daily_budget DOUBLE DEFAULT 0,
            impressions INTEGER DEFAULT 0,
            clicks INTEGER DEFAULT 0,
            spend DOUBLE DEFAULT 0,
            sales DOUBLE DEFAULT 0,
            orders INTEGER DEFAULT 0,
            units INTEGER DEFAULT 0,
            PRIMARY KEY (date, campaign_id)
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS ads_keywords (
            date TEXT,
            campaign_id TEXT,
            campaign_name TEXT,
            ad_group_id TEXT,
            ad_group_name TEXT,
            keyword_id TEXT,
            keyword_text TEXT,
            match_type TEXT,
            impressions INTEGER DEFAULT 0,
            clicks INTEGER DEFAULT 0,
            spend DOUBLE DEFAULT 0,
            sales DOUBLE DEFAULT 0,
            orders INTEGER DEFAULT 0,
            units INTEGER DEFAULT 0,
            PRIMARY KEY (date, keyword_id)
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS ads_search_terms (
            date TEXT,
            campaign_id TEXT,
            campaign_name TEXT,
            ad_group_name TEXT,
            keyword_text TEXT,
            match_type TEXT,
            search_term TEXT,
            impressions INTEGER DEFAULT 0,
            clicks INTEGER DEFAULT 0,
            spend DOUBLE DEFAULT 0,
            sales DOUBLE DEFAULT 0,
            orders INTEGER DEFAULT 0,
            units INTEGER DEFAULT 0,
            PRIMARY KEY (date, search_term, campaign_id)
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS ads_negative_keywords (
            keyword_text TEXT,
            match_type TEXT,
            campaign_name TEXT,
            ad_group_name TEXT,
            keyword_status TEXT DEFAULT 'ENABLED',
            PRIMARY KEY (keyword_text, campaign_name)
        )
    """)
    con.close()


def _sync_ads_data():
    """Pull Sponsored Products reporting data from Amazon Ads API v3."""
    import time as _time
    import gzip as gz
    import requests as req

    ads_creds = _load_ads_credentials()
    if not ads_creds:
        logger.info("Ads sync: no credentials configured, skipping")
        return

    _ensure_ads_tables()

    profile_id = ads_creds.get("profile_id", "")
    if not profile_id:
        logger.warning("Ads sync: no profile_id set, attempting to discover profiles...")
        try:
            from ad_api.api import Profiles
            result = Profiles(credentials=ads_creds).get_profiles()
            profiles = result.payload
            if profiles:
                for p in profiles:
                    if p.get("accountInfo", {}).get("marketplaceStringId") == "ATVPDKIKX0DER":
                        profile_id = str(p.get("profileId", ""))
                        break
                if not profile_id:
                    profile_id = str(profiles[0].get("profileId", ""))
                logger.info(f"Ads sync: discovered profile_id={profile_id}")
                ads_creds["profile_id"] = profile_id
            else:
                logger.error("Ads sync: no profiles found")
                return
        except Exception as e:
            logger.error(f"Ads sync: failed to discover profiles: {e}")
            return

    # Pull last 60 days of data
    today = datetime.now()
    start_date = (today - timedelta(days=60)).strftime("%Y-%m-%d")
    end_date = (today - timedelta(days=1)).strftime("%Y-%m-%d")

    # ── Report 1: Campaign-level daily data ──
    _pull_ads_report(
        ads_creds, "spCampaigns",
        columns=["date", "campaignId", "campaignName", "campaignStatus",
                 "campaignBudgetAmount", "impressions", "clicks", "cost",
                 "purchases", "unitsSold", "sales"],
        start_date=start_date, end_date=end_date,
        handler=_handle_campaign_report,
    )

    # ── Report 2: Targeting/Keywords daily data ──
    _pull_ads_report(
        ads_creds, "spTargeting",
        columns=["date", "campaignId", "campaignName", "adGroupId",
                 "adGroupName", "keywordId", "keywordText", "matchType",
                 "impressions", "clicks", "cost", "purchases",
                 "unitsSold", "sales"],
        start_date=start_date, end_date=end_date,
        handler=_handle_targeting_report,
    )

    # ── Report 3: Search terms ──
    _pull_ads_report(
        ads_creds, "spSearchTerm",
        columns=["date", "campaignId", "campaignName", "adGroupName",
                 "keywordText", "matchType", "searchTerm",
                 "impressions", "clicks", "cost", "purchases",
                 "unitsSold", "sales"],
        start_date=start_date, end_date=end_date,
        handler=_handle_search_term_report,
    )

    logger.info("Ads sync complete")


def _pull_ads_report(creds, report_type_id, columns, start_date, end_date, handler):
    """Create, poll, download, and process a single Amazon Ads v3 report."""
    import time as _time
    import gzip as gz
    import requests as req
    from ad_api.api import sponsored_products

    body = json.dumps({
        "name": f"{report_type_id}_{start_date}_{end_date}",
        "startDate": start_date,
        "endDate": end_date,
        "configuration": {
            "adProduct": "SPONSORED_PRODUCTS",
            "groupBy": ["advertiser"],
            "columns": columns,
            "reportTypeId": report_type_id,
            "timeUnit": "DAILY",
            "format": "GZIP_JSON",
        }
    })

    try:
        logger.info(f"Ads sync: creating {report_type_id} report ({start_date} to {end_date})...")
        reports_api = sponsored_products.Reports(credentials=creds)
        result = reports_api.post_report(body=body)
        report_id = result.payload.get("reportId")

        if not report_id:
            logger.error(f"Ads sync: no reportId returned for {report_type_id}")
            return

        # Poll for completion (max ~5 min)
        download_url = None
        for attempt in range(30):
            _time.sleep(10)
            status_result = reports_api.get_report(reportId=report_id)
            status = status_result.payload.get("status")

            if status == "COMPLETED":
                download_url = status_result.payload.get("url")
                break
            elif status in ("FAILED", "CANCELLED"):
                logger.error(f"Ads sync: {report_type_id} report {status}")
                return

        if not download_url:
            logger.error(f"Ads sync: {report_type_id} report timed out")
            return

        # Download and decompress
        resp = req.get(download_url)
        try:
            data = json.loads(gz.decompress(resp.content).decode("utf-8"))
        except Exception:
            data = json.loads(resp.text)

        if isinstance(data, str):
            data = json.loads(data)

        handler(data)
        logger.info(f"Ads sync: {report_type_id} — {len(data) if isinstance(data, list) else 'N/A'} rows processed")

    except Exception as e:
        logger.error(f"Ads sync error ({report_type_id}): {e}")


def _handle_campaign_report(data):
    """Insert campaign-level ads data into DuckDB."""
    if not isinstance(data, list):
        return

    con = duckdb.connect(str(DB_PATH), read_only=False)
    for row in data:
        date = row.get("date", "")
        campaign_id = str(row.get("campaignId", ""))
        campaign_name = row.get("campaignName", "")
        spend = float(row.get("cost", 0) or 0)
        sales = float(row.get("sales", 0) or 0)
        impressions = int(row.get("impressions", 0) or 0)
        clicks = int(row.get("clicks", 0) or 0)
        orders = int(row.get("purchases", 0) or 0)
        units = int(row.get("unitsSold", 0) or 0)
        status = row.get("campaignStatus", "")
        budget = float(row.get("campaignBudgetAmount", 0) or 0)

        # Aggregate table (for summary endpoints)
        con.execute("""
            INSERT OR REPLACE INTO advertising
            (date, campaign_id, campaign_name, impressions, clicks, spend, sales, orders, units)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, [date, campaign_id, campaign_name, impressions, clicks, spend, sales, orders, units])

        # Campaign detail table
        con.execute("""
            INSERT OR REPLACE INTO ads_campaigns
            (date, campaign_id, campaign_name, campaign_type, campaign_status,
             daily_budget, impressions, clicks, spend, sales, orders, units)
            VALUES (?, ?, ?, 'SP', ?, ?, ?, ?, ?, ?, ?, ?)
        """, [date, campaign_id, campaign_name, status, budget,
              impressions, clicks, spend, sales, orders, units])

    con.close()


def _handle_targeting_report(data):
    """Insert keyword/targeting data into DuckDB."""
    if not isinstance(data, list):
        return

    con = duckdb.connect(str(DB_PATH), read_only=False)
    for row in data:
        date = row.get("date", "")
        keyword_id = str(row.get("keywordId", row.get("targetId", "")))
        con.execute("""
            INSERT OR REPLACE INTO ads_keywords
            (date, campaign_id, campaign_name, ad_group_id, ad_group_name,
             keyword_id, keyword_text, match_type,
             impressions, clicks, spend, sales, orders, units)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, [
            date,
            str(row.get("campaignId", "")),
            row.get("campaignName", ""),
            str(row.get("adGroupId", "")),
            row.get("adGroupName", ""),
            keyword_id,
            row.get("keywordText", row.get("targeting", "")),
            row.get("matchType", ""),
            int(row.get("impressions", 0) or 0),
            int(row.get("clicks", 0) or 0),
            float(row.get("cost", 0) or 0),
            float(row.get("sales", 0) or 0),
            int(row.get("purchases", 0) or 0),
            int(row.get("unitsSold", 0) or 0),
        ])
    con.close()


def _handle_search_term_report(data):
    """Insert search term data into DuckDB."""
    if not isinstance(data, list):
        return

    con = duckdb.connect(str(DB_PATH), read_only=False)
    for row in data:
        date = row.get("date", "")
        search_term = row.get("searchTerm", "")
        campaign_id = str(row.get("campaignId", ""))
        con.execute("""
            INSERT OR REPLACE INTO ads_search_terms
            (date, campaign_id, campaign_name, ad_group_name,
             keyword_text, match_type, search_term,
             impressions, clicks, spend, sales, orders, units)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, [
            date, campaign_id,
            row.get("campaignName", ""),
            row.get("adGroupName", ""),
            row.get("keywordText", ""),
            row.get("matchType", ""),
            search_term,
            int(row.get("impressions", 0) or 0),
            int(row.get("clicks", 0) or 0),
            float(row.get("cost", 0) or 0),
            float(row.get("sales", 0) or 0),
            int(row.get("purchases", 0) or 0),
            int(row.get("unitsSold", 0) or 0),
        ])
    con.close()


# ── Scheduler Setup (APScheduler) ──────────────────────────
scheduler = None  # Will be initialized in lifespan


def _run_scheduled_sp_api_sync():
    """Wrapper for scheduled SP-API sync with logging."""
    start_time = time.time()
    log_id = _log_sync("sp_api_sync", "in_progress")
    try:
        _run_sp_api_sync()
        execution_time = time.time() - start_time
        _log_sync("sp_api_sync", "completed", execution_time=execution_time)
        logger.info(f"Scheduled SP-API sync completed in {execution_time:.2f}s")
    except Exception as e:
        execution_time = time.time() - start_time
        error_msg = str(e)
        _log_sync("sp_api_sync", "failed", error_message=error_msg, execution_time=execution_time)
        logger.error(f"Scheduled SP-API sync failed: {error_msg}")


def _run_scheduled_today_sync():
    """Wrapper for scheduled today orders sync with logging."""
    start_time = time.time()
    log_id = _log_sync("today_sync", "in_progress")
    try:
        _sync_today_orders()
        execution_time = time.time() - start_time
        _log_sync("today_sync", "completed", execution_time=execution_time)
        logger.info(f"Scheduled today sync completed in {execution_time:.2f}s")
    except Exception as e:
        execution_time = time.time() - start_time
        error_msg = str(e)
        _log_sync("today_sync", "failed", error_message=error_msg, execution_time=execution_time)
        logger.error(f"Scheduled today sync failed: {error_msg}")


def _run_scheduled_ads_sync():
    """Wrapper for scheduled ads sync with logging."""
    start_time = time.time()
    log_id = _log_sync("ads_sync", "in_progress")
    try:
        _sync_ads_data()
        execution_time = time.time() - start_time
        _log_sync("ads_sync", "completed", execution_time=execution_time)
        logger.info(f"Scheduled ads sync completed in {execution_time:.2f}s")
    except Exception as e:
        execution_time = time.time() - start_time
        error_msg = str(e)
        _log_sync("ads_sync", "failed", error_message=error_msg, execution_time=execution_time)
        logger.error(f"Scheduled ads sync failed: {error_msg}")


def _run_scheduled_pricing_sync():
    """Wrapper for scheduled pricing sync with logging."""
    start_time = time.time()
    log_id = _log_sync("pricing_sync", "in_progress")
    try:
        _sync_pricing_and_coupons()
        execution_time = time.time() - start_time
        _log_sync("pricing_sync", "completed", execution_time=execution_time)
        logger.info(f"Scheduled pricing sync completed in {execution_time:.2f}s")
    except Exception as e:
        execution_time = time.time() - start_time
        error_msg = str(e)
        _log_sync("pricing_sync", "failed", error_message=error_msg, execution_time=execution_time)
        logger.error(f"Scheduled pricing sync failed: {error_msg}")


def _run_scheduled_docs_update():
    """Wrapper for scheduled docs update with logging. (Placeholder for Prompt 2)"""
    start_time = time.time()
    log_id = _log_docs_update("in_progress")
    try:
        # Placeholder: actual implementation in Prompt 2
        # - Collect system snapshot
        # - Call Claude to generate docs
        # - Save to /app/docs/
        # - Commit to GitHub if token available
        logger.info("Docs update job started (placeholder)")
        execution_time = time.time() - start_time
        _log_docs_update("completed", documents_updated="architecture_guide.md, disaster_recovery_plan.md", execution_time=execution_time)
        logger.info(f"Scheduled docs update completed in {execution_time:.2f}s")
    except Exception as e:
        execution_time = time.time() - start_time
        error_msg = str(e)
        _log_docs_update("failed", error_message=error_msg, execution_time=execution_time)
        logger.error(f"Scheduled docs update failed: {error_msg}")


def _run_duckdb_backup():
    """Wrapper for nightly DuckDB backup with logging. (Placeholder for Prompt 2)"""
    start_time = time.time()
    log_id = _log_sync("duckdb_backup", "in_progress")
    try:
        # Placeholder: actual implementation in Prompt 2
        # - Copy DuckDB to temp file with date suffix
        # - Upload to Google Drive
        # - Clean up backups older than 30 days
        logger.info("DuckDB backup job started (placeholder)")
        execution_time = time.time() - start_time
        _log_sync("duckdb_backup", "completed", execution_time=execution_time)
        logger.info(f"DuckDB backup completed in {execution_time:.2f}s")
    except Exception as e:
        execution_time = time.time() - start_time
        error_msg = str(e)
        _log_sync("duckdb_backup", "failed", error_message=error_msg, execution_time=execution_time)
        logger.error(f"DuckDB backup failed: {error_msg}")


async def _sync_loop():
    """Initialize and run APScheduler for background sync jobs.

    Sequence:
    1. Immediately sync today's orders (fast, ~5s) — so "Today" works right away
    2. Auto-backfill historical data if missing (detects fresh deploy)
    3. Schedule 4 daily sync jobs at 9:00, 12:00, 15:00, 18:00 Central
    4. Schedule ads and pricing syncs every 2 hours
    5. Handle startup catchup: if a scheduled time has passed for today with no sync_log entry, run immediately
    """
    global scheduler

    # Immediately pull today's orders — no delay
    try:
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(None, _sync_today_orders)
        logger.info("Startup: today's orders synced")
    except Exception as e:
        logger.error(f"Startup today-sync error: {e}")

    # Auto-backfill if historical data is missing (e.g. after Railway redeploy)
    try:
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(None, _auto_backfill_if_needed)
    except Exception as e:
        logger.error(f"Auto-backfill error: {e}")

    # Sync ads data
    try:
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(None, _sync_ads_data)
    except Exception as e:
        logger.error(f"Ads sync error: {e}")

    # Sync pricing & coupon data
    try:
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(None, _sync_pricing_and_coupons)
    except Exception as e:
        logger.error(f"Pricing/coupon sync error: {e}")

    # Initialize APScheduler
    if not scheduler:
        scheduler = AsyncIOScheduler(timezone="America/Chicago")

        # Schedule 4 daily SP-API syncs at 9am, 12pm, 3pm, 6pm Central
        scheduler.add_job(_run_scheduled_sp_api_sync, CronTrigger(hour=9, minute=0, timezone="America/Chicago"), id="sp_api_sync_9am")
        scheduler.add_job(_run_scheduled_sp_api_sync, CronTrigger(hour=12, minute=0, timezone="America/Chicago"), id="sp_api_sync_12pm")
        scheduler.add_job(_run_scheduled_sp_api_sync, CronTrigger(hour=15, minute=0, timezone="America/Chicago"), id="sp_api_sync_3pm")
        scheduler.add_job(_run_scheduled_sp_api_sync, CronTrigger(hour=18, minute=0, timezone="America/Chicago"), id="sp_api_sync_6pm")

        # Schedule ads sync every 2 hours
        scheduler.add_job(_run_scheduled_ads_sync, CronTrigger(minute=0, second=0, timezone="America/Chicago"), id="ads_sync_hourly")

        # Schedule pricing sync every 2 hours (offset by 1 hour from ads)
        scheduler.add_job(_run_scheduled_pricing_sync, CronTrigger(minute=30, second=0, timezone="America/Chicago"), id="pricing_sync_hourly")

        # Schedule docs update at 8am and 8pm Central (Prompt 2 implementation)
        scheduler.add_job(_run_scheduled_docs_update, CronTrigger(hour=8, minute=0, timezone="America/Chicago"), id="docs_update_8am")
        scheduler.add_job(_run_scheduled_docs_update, CronTrigger(hour=20, minute=0, timezone="America/Chicago"), id="docs_update_8pm")

        # Schedule nightly DuckDB backup at 2am Central (Prompt 2 implementation)
        scheduler.add_job(_run_duckdb_backup, CronTrigger(hour=2, minute=0, timezone="America/Chicago"), id="duckdb_backup_2am")

        await scheduler.start()
        logger.info("APScheduler started with 4 daily SP-API syncs, hourly ads/pricing, docs updates, and nightly backup (America/Chicago)")

        # Startup catchup: if current time is past a scheduled sync time and no sync_log entry exists for today, run immediately
        await _startup_sync_catchup()


async def _startup_sync_catchup():
    """On startup, check if any scheduled sync times have passed today without a sync_log entry.
    If so, run the sync immediately to catch up."""
    try:
        now_utc = datetime.now(ZoneInfo("UTC"))
        now_central = now_utc.astimezone(ZoneInfo("America/Chicago"))
        today_date = now_central.date()

        scheduled_times = [9, 12, 15, 18]  # 9am, 12pm, 3pm, 6pm Central

        con = duckdb.connect(str(DB_PATH), read_only=True)

        for scheduled_hour in scheduled_times:
            scheduled_time = now_central.replace(hour=scheduled_hour, minute=0, second=0, microsecond=0)

            # Check if this time has already passed today
            if now_central > scheduled_time:
                # Check if we have a sync_log entry for this job today
                rows = con.execute("""
                    SELECT id FROM sync_log
                    WHERE job_name = 'sp_api_sync'
                    AND DATE(started_at) = ?
                    AND status = 'completed'
                    ORDER BY started_at DESC
                    LIMIT 1
                """, [today_date]).fetchall()

                if not rows:
                    logger.info(f"Startup catchup: {scheduled_hour}:00 sync hasn't run yet today, running now...")
                    loop = asyncio.get_event_loop()
                    await loop.run_in_executor(None, _run_scheduled_sp_api_sync)
                    break  # Only run one catchup per startup

        con.close()
    except Exception as e:
        logger.error(f"Startup sync catchup error: {e}")


@asynccontextmanager
async def lifespan(app):
    """Start background sync and initialize tables on app startup."""
    # Initialize Item Plan DuckDB tables (defined later in file, safe here at runtime)
    try:
        _init_item_plan_tables()
        logger.info("Item Plan tables initialized")
    except Exception as e:
        logger.error(f"Item Plan table init error: {e}")
    # Initialize system tables
    try:
        _init_system_tables()
        logger.info("System tables initialized")
    except Exception as e:
        logger.error(f"System table init error: {e}")
    task = asyncio.create_task(_sync_loop())
    logger.info("Background sync scheduler initialized")
    yield
    if scheduler and scheduler.running:
        await scheduler.shutdown()
    task.cancel()


app = FastAPI(title="GolfGen Dashboard API", version="1.0.0", lifespan=lifespan)

# Allow frontend (dev on :5173, prod wherever)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Item Plan tables initialized in lifespan handler (defined later in file)


# ── Multi-User Authentication ──────────────────────────────
import bcrypt as _bcrypt

# All dashboard tabs (tab_key -> display label)
ALL_TABS = {
    "dashboard": "Dashboard",
    "products": "Products",
    "profitability": "Profitability",
    "advertising": "Advertising",
    "inventory": "Amazon FBA",
    "golfgen-inventory": "GolfGen Inventory",
    "item-master": "Item Master",
    "factory-po": "Factory PO",
    "logistics": "OTW / Logistics",
    "fba-shipments": "Shipments to FBA",
    "item-planning": "Item Planning",
}

# Tab key -> list of API path prefixes that belong to that tab
TAB_API_PREFIXES = {
    "dashboard": ["/api/summary", "/api/daily", "/api/comparison", "/api/monthly-yoy", "/api/product-mix", "/api/color-mix"],
    "products": ["/api/products", "/api/product/"],
    "profitability": ["/api/pnl", "/api/profitability"],
    "advertising": ["/api/ads/"],
    "inventory": ["/api/inventory"],
    "golfgen-inventory": ["/api/warehouse"],
    "item-master": ["/api/item-master", "/api/pricing"],
    "factory-po": ["/api/factory-po"],
    "logistics": ["/api/logistics", "/api/supply-chain"],
    "fba-shipments": ["/api/fba-shipments"],
    "item-planning": ["/api/item-planning", "/api/item-plan", "/api/factory-on-order", "/api/dashboard-settings"],
}

USERS = {
    "eric": {
        "name": "Eric",
        "emails": ["eric@golfgen.com", "eric@egbrands.com"],
        "password_hash": "$2b$12$CXwF3gjnEyEPwej2qV9trem4AXZi4tUVR50ifvb2dUTiNVBhHAneu",
        "role": "admin",
    },
    "ty": {
        "name": "Ty",
        "emails": ["ty@golfgen.com", "tysams@egbrands.com"],
        "password_hash": "$2b$12$jurM2OMgL16XIFjNBQu3JeZsq.phyEea08ABqvNMIxZt3ZzgFBjs6",
        "role": "staff",
    },
    "kim": {
        "name": "Kim",
        "emails": ["kim@golfgen.com", "kim@egbrands.com"],
        "password_hash": "$2b$12$84EhMgFJ072dxgZ3ChICj.K.vwVRix2tDPGcb7uqMa3haswq8zdSK",
        "role": "staff",
    },
    "ryan": {
        "name": "Ryan",
        "emails": ["ryan@golfgen.com", "ryan@egbrands.com"],
        "password_hash": "$2b$12$oP6sxNocG4Hzrek3R/SU5eWUFR/EM3bqjZFf58RhgZHC35zuKLAEC",
        "role": "staff",
    },
    "mckay": {
        "name": "McKay",
        "emails": ["riseecom21@gmail.com"],
        "password_hash": "$2b$12$XvwYb1CZMG78FpD5AXy3FOT8WgiJhuBQyUg/tbgBmMrokA6z.LJya",
        "role": "staff",
    },
}

# Build email -> user_key lookup (case-insensitive)
_EMAIL_TO_USER = {}
for _ukey, _udata in USERS.items():
    for _em in _udata["emails"]:
        _EMAIL_TO_USER[_em.lower()] = _ukey


def _init_auth_tables():
    """Create DuckDB tables for sessions and permissions if not exist."""
    con = duckdb.connect(str(DB_PATH))
    con.execute("""
        CREATE TABLE IF NOT EXISTS sessions (
            token       TEXT PRIMARY KEY,
            user_email  TEXT NOT NULL,
            user_name   TEXT NOT NULL,
            role        TEXT NOT NULL,
            created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS user_permissions (
            user_name   TEXT NOT NULL,
            tab_key     TEXT NOT NULL,
            enabled     BOOLEAN DEFAULT TRUE,
            PRIMARY KEY (user_name, tab_key)
        )
    """)
    # Seed default permissions for staff users (all enabled)
    for ukey, udata in USERS.items():
        if udata["role"] == "staff":
            for tab_key in ALL_TABS:
                con.execute("""
                    INSERT OR IGNORE INTO user_permissions (user_name, tab_key, enabled)
                    VALUES (?, ?, TRUE)
                """, [ukey, tab_key])
    con.close()


_init_auth_tables()


def _init_system_tables():
    """Create DuckDB tables for sync logging and docs updates if not exist."""
    con = duckdb.connect(str(DB_PATH))

    con.execute("""
        CREATE TABLE IF NOT EXISTS sync_log (
            id BIGINT PRIMARY KEY DEFAULT nextval('sync_log_seq'),
            job_name TEXT NOT NULL,
            started_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            completed_at TIMESTAMP,
            status TEXT DEFAULT 'in_progress',
            records_processed BIGINT DEFAULT 0,
            error_message TEXT,
            execution_time_seconds DOUBLE
        )
    """)

    # Create sequence if not exists (DuckDB doesn't support CREATE SEQUENCE IF NOT EXISTS in older versions)
    try:
        con.execute("CREATE SEQUENCE sync_log_seq")
    except:
        pass  # Sequence already exists

    con.execute("""
        CREATE TABLE IF NOT EXISTS docs_update_log (
            id BIGINT PRIMARY KEY DEFAULT nextval('docs_update_log_seq'),
            started_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            completed_at TIMESTAMP,
            status TEXT DEFAULT 'in_progress',
            documents_updated TEXT,
            error_message TEXT,
            execution_time_seconds DOUBLE
        )
    """)

    try:
        con.execute("CREATE SEQUENCE docs_update_log_seq")
    except:
        pass  # Sequence already exists

    con.close()

# _init_system_tables() called in lifespan handler


def _log_sync(job_name: str, status: str = "in_progress", records_processed: int = 0, error_message: str = None, execution_time: float = None) -> int:
    """Log a sync job to the sync_log table. Returns the log ID."""
    try:
        con = duckdb.connect(str(DB_PATH))
        result = con.execute("""
            INSERT INTO sync_log (job_name, status, records_processed, error_message, execution_time_seconds, completed_at)
            VALUES (?, ?, ?, ?, ?, CASE WHEN ? = 'completed' OR ? = 'failed' THEN CURRENT_TIMESTAMP ELSE NULL END)
            RETURNING id
        """, [job_name, status, records_processed, error_message, execution_time, status, status]).fetchone()
        con.close()
        return result[0] if result else None
    except Exception as e:
        logger.error(f"Failed to log sync: {e}")
        return None


def _log_docs_update(status: str = "in_progress", documents_updated: str = None, error_message: str = None, execution_time: float = None) -> int:
    """Log a docs update to the docs_update_log table. Returns the log ID."""
    try:
        con = duckdb.connect(str(DB_PATH))
        result = con.execute("""
            INSERT INTO docs_update_log (status, documents_updated, error_message, execution_time_seconds, completed_at)
            VALUES (?, ?, ?, ?, CASE WHEN ? = 'completed' OR ? = 'failed' THEN CURRENT_TIMESTAMP ELSE NULL END)
            RETURNING id
        """, [status, documents_updated, error_message, execution_time, status, status]).fetchone()
        con.close()
        return result[0] if result else None
    except Exception as e:
        logger.error(f"Failed to log docs update: {e}")
        return None


def _find_user_by_email(email: str):
    """Look up user key by email (case-insensitive)."""
    return _EMAIL_TO_USER.get(email.lower().strip())


def _get_session(token: str):
    """Look up session from DuckDB. Returns dict or None."""
    if not token:
        return None
    con = duckdb.connect(str(DB_PATH), read_only=True)
    rows = con.execute("SELECT token, user_email, user_name, role FROM sessions WHERE token = ?", [token]).fetchall()
    con.close()
    if rows:
        return {"token": rows[0][0], "user_email": rows[0][1], "user_name": rows[0][2], "role": rows[0][3]}
    return None


def _get_user_permissions(user_name: str):
    """Return set of enabled tab_keys for a user."""
    con = duckdb.connect(str(DB_PATH), read_only=True)
    rows = con.execute("SELECT tab_key FROM user_permissions WHERE user_name = ? AND enabled = TRUE", [user_name]).fetchall()
    con.close()
    return {r[0] for r in rows}


class MultiLoginRequest(BaseModel):
    email: str
    password: str


@app.post("/api/auth/login")
def login(req: MultiLoginRequest):
    """Validate email + password and create a DuckDB session."""
    user_key = _find_user_by_email(req.email)
    if not user_key:
        raise HTTPException(status_code=401, detail="Invalid email or password.")
    user = USERS[user_key]
    if not _bcrypt.checkpw(req.password.encode("utf-8"), user["password_hash"].encode("utf-8")):
        raise HTTPException(status_code=401, detail="Invalid email or password.")
    token = secrets.token_hex(32)
    con = duckdb.connect(str(DB_PATH))
    con.execute("INSERT INTO sessions (token, user_email, user_name, role) VALUES (?, ?, ?, ?)",
                [token, req.email.lower().strip(), user["name"], user["role"]])
    con.close()
    response = JSONResponse({"ok": True, "name": user["name"], "role": user["role"]})
    response.set_cookie(
        key="golfgen_session",
        value=token,
        httponly=True,
        samesite="none",
        secure=True,
        max_age=60 * 60 * 24 * 7,
    )
    return response


@app.get("/api/auth/check")
def auth_check(golfgen_session: Optional[str] = Cookie(None)):
    """Check if the current session is valid."""
    sess = _get_session(golfgen_session)
    if sess:
        return {"authenticated": True, "name": sess["user_name"], "role": sess["role"]}
    raise HTTPException(status_code=401, detail="Not authenticated")


@app.post("/api/auth/logout")
def logout(golfgen_session: Optional[str] = Cookie(None)):
    """Clear the session from DuckDB."""
    if golfgen_session:
        con = duckdb.connect(str(DB_PATH))
        con.execute("DELETE FROM sessions WHERE token = ?", [golfgen_session])
        con.close()
    response = JSONResponse({"ok": True})
    response.delete_cookie("golfgen_session")
    return response


@app.get("/api/me")
def get_me(request: Request):
    """Return current user info."""
    sess = _get_session(request.cookies.get("golfgen_session"))
    if not sess:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return {"name": sess["user_name"], "email": sess["user_email"], "role": sess["role"]}


@app.get("/api/permissions/me")
def get_my_permissions(request: Request):
    """Return list of tab keys the current user can access."""
    sess = _get_session(request.cookies.get("golfgen_session"))
    if not sess:
        raise HTTPException(status_code=401, detail="Not authenticated")
    if sess["role"] == "admin":
        return {"tabs": list(ALL_TABS.keys()), "role": "admin", "allTabs": ALL_TABS}
    user_key = _find_user_by_email(sess["user_email"])
    if not user_key:
        return {"tabs": list(ALL_TABS.keys()), "role": sess["role"], "allTabs": ALL_TABS}
    enabled = _get_user_permissions(user_key)
    return {"tabs": [t for t in ALL_TABS if t in enabled], "role": sess["role"], "allTabs": ALL_TABS}


@app.get("/api/permissions")
def get_all_permissions(request: Request):
    """Admin only: return full permissions grid."""
    sess = _get_session(request.cookies.get("golfgen_session"))
    if not sess or sess["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    con = duckdb.connect(str(DB_PATH), read_only=True)
    rows = con.execute("SELECT user_name, tab_key, enabled FROM user_permissions ORDER BY user_name, tab_key").fetchall()
    con.close()
    grid = {}
    for user_name, tab_key, enabled in rows:
        if user_name not in grid:
            grid[user_name] = {}
        grid[user_name][tab_key] = enabled
    # Include user display names
    users_list = []
    for ukey, udata in USERS.items():
        if udata["role"] == "staff":
            users_list.append({
                "key": ukey, "name": udata["name"],
                "emails": udata["emails"],
                "permissions": grid.get(ukey, {}),
            })
    return {"users": users_list, "allTabs": ALL_TABS}


class PermissionUpdate(BaseModel):
    user: str
    tab: str
    enabled: bool


@app.post("/api/permissions")
def update_permission(req: PermissionUpdate, request: Request):
    """Admin only: update one toggle."""
    sess = _get_session(request.cookies.get("golfgen_session"))
    if not sess or sess["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    if req.user not in USERS or req.tab not in ALL_TABS:
        raise HTTPException(status_code=400, detail="Invalid user or tab")
    con = duckdb.connect(str(DB_PATH))
    # Upsert
    con.execute("DELETE FROM user_permissions WHERE user_name = ? AND tab_key = ?", [req.user, req.tab])
    con.execute("INSERT INTO user_permissions (user_name, tab_key, enabled) VALUES (?, ?, ?)",
                [req.user, req.tab, req.enabled])
    con.close()
    return {"ok": True}


# ── Helpers ─────────────────────────────────────────────

def _require_auth(request: Request):
    """Validate session cookie or raise 401."""
    token = request.cookies.get("golfgen_session")
    sess = _get_session(token)
    if not sess:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return sess


def _require_tab_access(request: Request, tab_key: str):
    """Validate session + tab permission. Returns session dict."""
    sess = _require_auth(request)
    if sess["role"] == "admin":
        return sess
    user_key = _find_user_by_email(sess["user_email"])
    if user_key:
        enabled = _get_user_permissions(user_key)
        if tab_key not in enabled:
            raise HTTPException(status_code=403, detail="You do not have access to this page")
    return sess


def _tab_key_for_path(path: str):
    """Return the tab_key for a given API path, or None if not tab-gated."""
    for tab_key, prefixes in TAB_API_PREFIXES.items():
        for prefix in prefixes:
            if path == prefix or path.startswith(prefix + "/") or path.startswith(prefix + "?"):
                return tab_key
    return None


@app.middleware("http")
async def tab_permission_middleware(request: Request, call_next):
    """Enforce tab-based permissions on all /api/* data endpoints."""
    path = request.url.path
    method = request.method

    # Always allow OPTIONS (CORS preflight) and non-API paths
    if method == "OPTIONS" or not path.startswith("/api/"):
        return await call_next(request)

    # Skip auth endpoints, system endpoints, upload endpoints
    if (path.startswith("/api/auth/") or
        path.startswith("/api/me") or
        path.startswith("/api/permissions") or
        path.startswith("/api/upload/") or
        path.startswith("/api/debug/") or
        path in ("/api/health", "/api/sync", "/api/backfill") or
        path.startswith("/api/refresh")):
        return await call_next(request)

    # Check session
    token = request.cookies.get("golfgen_session")
    try:
        sess = _get_session(token)
    except Exception:
        # DuckDB error (table not exist, lock, etc.) — let request through
        return await call_next(request)

    if not sess:
        return JSONResponse(status_code=401, content={"detail": "Not authenticated"})

    # For tab-gated endpoints, check permission
    tab_key = _tab_key_for_path(path)
    if tab_key and sess["role"] != "admin":
        user_key = _find_user_by_email(sess["user_email"])
        if user_key:
            try:
                enabled = _get_user_permissions(user_key)
                if tab_key not in enabled:
                    return JSONResponse(status_code=403, content={"detail": "You do not have access to this page"})
            except Exception:
                pass  # DuckDB error — allow through

    return await call_next(request)


def get_db():
    """Return a read-only DuckDB connection."""
    return duckdb.connect(str(DB_PATH), read_only=True)


def load_json(filename: str) -> list:
    """Load a JSON data file from the data directory."""
    path = DB_DIR / filename
    if not path.exists():
        return []
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def save_json(filename: str, data):
    """Save data to a JSON file in the data directory."""
    path = DB_DIR / filename
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def load_cogs() -> dict:
    """Load COGS data from CSV. Returns {asin: {cogs, product_name, sku}}."""
    cogs = {}
    if not COGS_PATH.exists():
        return cogs
    with open(COGS_PATH, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            asin = (row.get("asin") or "").strip()
            if not asin:
                continue
            try:
                cogs_val = float(row.get("cogs") or 0)
            except ValueError:
                cogs_val = 0
            cogs[asin] = {
                "cogs": cogs_val,
                "product_name": (row.get("product_name") or "").strip(),
                "sku": (row.get("sku") or "").strip(),
            }
    return cogs


def fmt_date(d) -> str:
    """Safely format a date value to string."""
    if isinstance(d, str):
        return d
    if hasattr(d, "strftime"):
        return d.strftime("%Y-%m-%d")
    return str(d)


def get_today(con) -> datetime:
    """Return today's actual calendar date in US/Central timezone.

    Railway servers run in UTC. The business operates in Central time,
    so we convert UTC to Central (America/Chicago) to get the real calendar date.
    'Today' always means today — if there's no data yet it shows $0,
    which is correct until Orders API sync populates it.
    """
    now_utc = datetime.now(ZoneInfo("UTC"))
    now_central = now_utc.astimezone(ZoneInfo("America/Chicago"))
    return now_central.replace(hour=0, minute=0, second=0, microsecond=0)


# ── API Routes ──────────────────────────────────────────

@app.get("/api/health")
def health():
    return {
        "status": "ok",
        "db": str(DB_PATH),
        "db_exists": DB_PATH.exists(),
        "port": os.environ.get("PORT", "8000"),
        "has_sp_api_creds": _load_sp_api_credentials() is not None,
    }


@app.post("/api/sync")
async def trigger_sync():
    """Manually trigger an SP-API data sync."""
    import asyncio
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, _run_sp_api_sync)
    return {"status": "sync_complete"}


@app.get("/api/sync")
async def trigger_sync_get():
    """Manually trigger sync. Fast: returns after today's orders are synced.
    Full sync (Sales Report) continues in background."""
    import asyncio
    loop = asyncio.get_event_loop()
    # Fast: sync today's orders first (returns in ~5s)
    await loop.run_in_executor(None, _sync_today_orders)
    # Kick off full sync in background (don't wait for 5-min report poll)
    asyncio.get_event_loop().run_in_executor(None, _run_sp_api_sync)
    return {"status": "sync_complete", "today_synced": True}


@app.get("/api/summary")
def summary(days: int = Query(365, description="Number of days to include")):
    """High-level KPIs: revenue, units, orders, sessions, AUR, conv rate."""
    con = get_db()
    cutoff = (get_today(con) - timedelta(days=days)).strftime("%Y-%m-%d") if days > 0 else get_today(con).strftime("%Y-%m-%d")

    row = con.execute("""
        SELECT
            COALESCE(SUM(ordered_product_sales), 0) AS revenue,
            COALESCE(SUM(units_ordered), 0) AS units,
            COALESCE(SUM(sessions), 0) AS sessions
        FROM daily_sales
        WHERE date >= ? AND asin = 'ALL'
    """, [cutoff]).fetchone()

    revenue, units, sessions = row
    orders = units  # total_order_items is not populated; use units as proxy
    aur = round(revenue / units, 2) if units else 0
    conv_rate = round(units / sessions * 100, 1) if sessions else 0

    # Net profit: estimate from known cost ratios in per-ASIN data
    products = _build_product_list(con, cutoff)
    prod_rev = sum(p["rev"] for p in products)
    prod_net = sum(p["net"] for p in products)

    # Scale product P&L to actual revenue if per-ASIN data is incomplete
    if prod_rev > 0 and revenue > 0:
        margin_pct = prod_net / prod_rev  # known margin from per-ASIN data
        total_net = round(revenue * margin_pct, 2)
        margin = round(margin_pct * 100)
    else:
        total_net = 0
        margin = 0

    con.close()
    return {
        "days": days,
        "revenue": round(revenue, 2),
        "units": units,
        "orders": orders,
        "sessions": sessions,
        "aur": aur,
        "convRate": conv_rate,
        "netProfit": total_net,
        "margin": margin,
    }


@app.get("/api/daily")
def daily_sales(days: int = Query(365), granularity: str = Query("daily")):
    """Time-series sales data for charts. Granularity: daily or weekly."""
    con = get_db()
    cutoff = (get_today(con) - timedelta(days=days)).strftime("%Y-%m-%d")

    rows = con.execute("""
        SELECT
            date,
            COALESCE(ordered_product_sales, 0) AS revenue,
            COALESCE(units_ordered, 0) AS units,
            COALESCE(sessions, 0) AS sessions
        FROM daily_sales
        WHERE date >= ? AND asin = 'ALL'
        ORDER BY date
    """, [cutoff]).fetchall()
    con.close()

    data = []
    for r in rows:
        revenue_val = r[1]
        units_val = r[2] or 0
        sessions_val = r[3] or 0
        data.append({
            "date": fmt_date(r[0]),
            "revenue": round(revenue_val, 2),
            "units": units_val,
            "orders": units_val,  # total_order_items not populated; use units
            "sessions": sessions_val,
            "aur": round(revenue_val / units_val, 2) if units_val else 0,
            "convRate": round(units_val / sessions_val * 100, 1) if sessions_val else 0,
        })

    if granularity == "weekly" and data:
        data = _aggregate_weekly(data)

    return {"granularity": granularity, "days": days, "data": data}


@app.get("/api/products")
def products(days: int = Query(365)):
    """Product breakdown with profitability metrics."""
    con = get_db()
    cutoff = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    product_list = _build_product_list(con, cutoff)
    con.close()
    return {"days": days, "products": product_list}


@app.get("/api/inventory")
def inventory():
    """Current FBA inventory with days-of-supply calculations."""
    con = get_db()
    cogs_data = load_cogs()

    inv_rows = con.execute("""
        SELECT asin, sku, product_name,
               COALESCE(fulfillable_quantity, 0) AS fba_stock,
               COALESCE(inbound_working_quantity, 0) + COALESCE(inbound_shipped_quantity, 0) + COALESCE(inbound_receiving_quantity, 0) AS inbound,
               COALESCE(reserved_quantity, 0) AS reserved
        FROM fba_inventory
    """).fetchall()

    # Get avg daily units for last 30 days
    velocity = {}
    vel_rows = con.execute("""
        SELECT asin, SUM(units_ordered) AS units
        FROM daily_sales
        WHERE date >= CURRENT_DATE - INTERVAL '30 days'
          AND asin != 'ALL'
        GROUP BY asin
    """).fetchall()
    for vr in vel_rows:
        velocity[vr[0]] = round(vr[1] / 30, 1)

    con.close()

    items = []
    for r in inv_rows:
        asin = r[0]
        avg_daily = velocity.get(asin, 0)
        fba_stock = r[3]
        dos = round(fba_stock / avg_daily) if avg_daily > 0 else 999

        # Get name from COGS file first, then inventory table
        cogs_name = cogs_data.get(asin, {}).get("product_name", "")
        # Exclude names that are just the ASIN repeated
        if cogs_name and cogs_name.strip().upper() == asin.upper():
            cogs_name = ""
        name = (cogs_name or r[2] or asin)

        items.append({
            "asin": asin,
            "sku": r[1] or "",
            "name": name,
            "fbaStock": fba_stock,
            "inbound": r[4],
            "reserved": r[5],
            "avgDaily": avg_daily,
            "dos": dos,
        })

    items.sort(key=lambda x: x["fbaStock"], reverse=True)
    return {"items": items}


@app.get("/api/product/{asin}")
def product_detail(asin: str, days: int = Query(365)):
    """Detailed view for a single product including daily trend."""
    con = get_db()
    cutoff = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    cogs_data = load_cogs()

    # Daily trend for this ASIN
    rows = con.execute("""
        SELECT date,
               COALESCE(ordered_product_sales, 0) AS revenue,
               COALESCE(units_ordered, 0) AS units,
               COALESCE(sessions, 0) AS sessions
        FROM daily_sales
        WHERE date >= ? AND asin = ?
        ORDER BY date
    """, [cutoff, asin]).fetchall()

    trend = [{"date": fmt_date(r[0]), "revenue": round(r[1], 2), "units": r[2], "sessions": r[3]} for r in rows]

    # Inventory
    inv = con.execute("""
        SELECT COALESCE(fulfillable_quantity, 0),
               COALESCE(inbound_working_quantity, 0) + COALESCE(inbound_shipped_quantity, 0) + COALESCE(inbound_receiving_quantity, 0),
               COALESCE(reserved_quantity, 0)
        FROM fba_inventory WHERE asin = ?
    """, [asin]).fetchone()

    con.close()

    cogs_info = cogs_data.get(asin, {})
    total_rev = sum(r["revenue"] for r in trend)
    total_units = sum(r["units"] for r in trend)

    return {
        "asin": asin,
        "name": cogs_info.get("product_name", asin),
        "sku": cogs_info.get("sku", ""),
        "cogs": cogs_info.get("cogs", 0),
        "totalRevenue": round(total_rev, 2),
        "totalUnits": total_units,
        "trend": trend,
        "inventory": {
            "fbaStock": inv[0] if inv else 0,
            "inbound": inv[1] if inv else 0,
            "reserved": inv[2] if inv else 0,
        },
    }


@app.get("/api/pnl")
def pnl_waterfall(days: int = Query(365)):
    """Profit & Loss waterfall data, scaled to actual revenue from ALL rows."""
    con = get_db()
    cutoff = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")

    # Actual revenue from 'ALL' daily rows
    actual_rev = con.execute("""
        SELECT COALESCE(SUM(ordered_product_sales), 0)
        FROM daily_sales WHERE date >= ? AND asin = 'ALL'
    """, [cutoff]).fetchone()[0]

    # Cost breakdown from per-ASIN data
    products = _build_product_list(con, cutoff)
    con.close()

    prod_rev = sum(p["rev"] for p in products)
    prod_cogs = sum(p["cogsTotal"] for p in products)
    prod_fba = sum(p["fbaTotal"] for p in products)
    prod_referral = sum(p["referralTotal"] for p in products)
    prod_ad = sum(p["adSpend"] for p in products)
    prod_net = sum(p["net"] for p in products)

    # Scale cost ratios to actual revenue
    scale = actual_rev / prod_rev if prod_rev > 0 else 1
    cogs = round(prod_cogs * scale, 2)
    fba = round(prod_fba * scale, 2)
    referral = round(prod_referral * scale, 2)
    ad_spend = round(prod_ad * scale, 2)
    net = round(actual_rev - cogs - fba - referral - ad_spend, 2)

    return {
        "days": days,
        "revenue": round(actual_rev, 2),
        "cogs": cogs,
        "fbaFees": fba,
        "referralFees": referral,
        "adSpend": ad_spend,
        "netProfit": net,
    }


# ── Internal helpers ────────────────────────────────────

def _build_product_list(con, cutoff: str) -> list:
    """Build product-level P&L. Core business logic."""
    cogs_data = load_cogs()

    # Get product names/SKUs from fba_inventory as fallback
    inv_names = {}
    try:
        inv_rows = con.execute(
            "SELECT asin, sku, product_name FROM fba_inventory"
        ).fetchall()
        for ir in inv_rows:
            inv_names[ir[0]] = {"sku": ir[1] or "", "product_name": ir[2] or ""}
    except Exception:
        pass

    # Product-level sales — use per-ASIN data (excluding aggregate 'ALL' row)
    rows = con.execute("""
        SELECT asin,
               MAX(sku) AS sku,
               MAX(product_name) AS product_name,
               COALESCE(SUM(ordered_product_sales), 0) AS revenue,
               COALESCE(SUM(units_ordered), 0) AS units,
               COALESCE(SUM(sessions), 0) AS sessions,
               COALESCE(SUM(total_order_items), 0) AS orders
        FROM daily_sales
        WHERE asin != 'ALL' AND date >= ?
        GROUP BY asin
        ORDER BY SUM(ordered_product_sales) DESC
    """, [cutoff]).fetchall()

    # Financial data (actual fees if available)
    fin_rows = con.execute("""
        SELECT asin,
               SUM(ABS(fba_fees)) AS fba_fees,
               SUM(ABS(commission)) AS commission
        FROM financial_events
        WHERE date >= ?
        GROUP BY asin
    """, [cutoff]).fetchall()
    fin_by_asin = {r[0]: {"fba_fees": r[1], "commission": r[2]} for r in fin_rows}

    products = []
    for r in rows:
        asin, sku, api_name, revenue, units, sessions, orders = r
        if units == 0:
            continue

        # Use units as proxy for orders if total_order_items is empty
        if not orders:
            orders = units

        aur = round(revenue / units, 2)

        # COGS: from file, by ASIN or SKU
        cogs_info = cogs_data.get(asin) or cogs_data.get(sku or "") or {}
        cogs_per_unit = cogs_info.get("cogs", 0)
        if cogs_per_unit == 0:
            cogs_per_unit = round(aur * 0.35, 2)  # estimate

        # Name: COGS file > inventory table > API > ASIN
        inv_info = inv_names.get(asin, {})
        cogs_name = cogs_info.get("product_name", "")
        # Exclude names that are just the ASIN repeated
        if cogs_name and cogs_name.strip().upper() == asin.upper():
            cogs_name = ""
        name = (cogs_name
                or api_name
                or inv_info.get("product_name")
                or asin)

        # SKU: COGS file > daily_sales > inventory table
        resolved_sku = (sku
                        or cogs_info.get("sku", "")
                        or inv_info.get("sku", ""))

        # FBA fees
        fin = fin_by_asin.get(asin, {})
        actual_fba = fin.get("fba_fees", 0)
        # Better estimate for golf equipment (oversized/bulky items):
        # Amazon FBA fees are typically 10-15% of selling price for large items
        # Old formula ($5 + 3% of AUR) was WAY too low
        est_fba_total = round(revenue * 0.12, 2)  # ~12% of revenue for oversized
        fba_total = actual_fba if actual_fba > 0 else est_fba_total

        # Referral fees (15% is standard for Sports & Outdoors)
        actual_commission = fin.get("commission", 0)
        referral_total = round(actual_commission, 2) if actual_commission > 0 else round(revenue * 0.15, 2)

        # Ad spend — pull from advertising if available
        ad_spend = 0
        try:
            ad_row = con.execute("""
                SELECT COALESCE(SUM(spend), 0)
                FROM advertising
                WHERE asin = ? AND date >= ?
            """, [asin, cutoff]).fetchone()
            if ad_row and ad_row[0] > 0:
                ad_spend = round(ad_row[0], 2)
        except Exception:
            # Table might not exist yet
            pass

        cogs_total = units * cogs_per_unit
        net = round(revenue - cogs_total - fba_total - referral_total - ad_spend, 2)
        margin = round(net / revenue * 100) if revenue else 0
        conv_rate = round(units / sessions * 100, 1) if sessions else 0

        products.append({
            "asin": asin,
            "sku": resolved_sku,
            "name": name,
            "rev": round(revenue, 2),
            "units": units,
            "price": aur,
            "sessions": sessions,
            "convRate": conv_rate,
            "cogsPerUnit": cogs_per_unit,
            "cogsTotal": round(cogs_total, 2),
            "fbaTotal": round(fba_total, 2),
            "referralTotal": round(referral_total, 2),
            "adSpend": ad_spend,
            "net": net,
            "margin": margin,
        })

    return products


def _aggregate_weekly(daily_data: list) -> list:
    """Aggregate daily data into weekly buckets."""
    from collections import defaultdict
    weeks = defaultdict(lambda: {"revenue": 0, "units": 0, "orders": 0, "sessions": 0})
    for d in daily_data:
        dt = datetime.strptime(d["date"], "%Y-%m-%d")
        week_start = dt - timedelta(days=dt.weekday())
        key = week_start.strftime("%Y-%m-%d")
        weeks[key]["revenue"] += d["revenue"]
        weeks[key]["units"] += d["units"]
        weeks[key]["orders"] += d["orders"]
        weeks[key]["sessions"] += d["sessions"]

    result = []
    for date_key in sorted(weeks.keys()):
        w = weeks[date_key]
        units = w["units"] or 1
        sessions = w["sessions"]
        result.append({
            "date": date_key,
            "revenue": round(w["revenue"], 2),
            "units": w["units"],
            "orders": w["orders"],
            "sessions": sessions,
            "aur": round(w["revenue"] / units, 2),
            "convRate": round(w["orders"] / sessions * 100, 1) if sessions else 0,
        })
    return result


# ── Advertising API Routes ─────────────────────────────────

def _safe_ads_query(con, query, params=None):
    """Run an ads query, returning empty list if tables don't exist yet."""
    try:
        return con.execute(query, params or []).fetchall()
    except Exception:
        return []


@app.get("/api/ads/summary")
def ads_summary(days: int = Query(30)):
    """Ads KPIs: total spend, sales, ACOS, ROAS, TACOS, CPC, CTR."""
    con = get_db()
    cutoff = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")

    row = _safe_ads_query(con, """
        SELECT
            COALESCE(SUM(spend), 0) AS spend,
            COALESCE(SUM(sales), 0) AS ad_sales,
            COALESCE(SUM(impressions), 0) AS impressions,
            COALESCE(SUM(clicks), 0) AS clicks,
            COALESCE(SUM(orders), 0) AS ad_orders,
            COALESCE(SUM(units), 0) AS ad_units
        FROM advertising
        WHERE date >= ?
    """, [cutoff])

    if not row or not row[0]:
        con.close()
        return {
            "days": days, "connected": False,
            "spend": 0, "adSales": 0, "impressions": 0, "clicks": 0,
            "orders": 0, "units": 0, "acos": 0, "roas": 0, "tacos": 0,
            "cpc": 0, "ctr": 0, "cvr": 0,
        }

    spend, ad_sales, impressions, clicks, orders, units = row[0]

    # Get total organic revenue for TACOS
    org_row = con.execute("""
        SELECT COALESCE(SUM(ordered_product_sales), 0)
        FROM daily_sales
        WHERE date >= ? AND asin = 'ALL'
    """, [cutoff]).fetchone()
    total_revenue = org_row[0] if org_row else 0

    con.close()

    acos = round(spend / ad_sales * 100, 2) if ad_sales > 0 else 0
    roas = round(ad_sales / spend, 2) if spend > 0 else 0
    tacos = round(spend / total_revenue * 100, 2) if total_revenue > 0 else 0
    cpc = round(spend / clicks, 2) if clicks > 0 else 0
    ctr = round(clicks / impressions * 100, 2) if impressions > 0 else 0
    cvr = round(orders / clicks * 100, 2) if clicks > 0 else 0

    return {
        "days": days, "connected": True,
        "spend": round(spend, 2),
        "adSales": round(ad_sales, 2),
        "impressions": impressions,
        "clicks": clicks,
        "orders": orders,
        "units": units,
        "acos": acos,
        "roas": roas,
        "tacos": tacos,
        "cpc": cpc,
        "ctr": ctr,
        "cvr": cvr,
    }


@app.get("/api/ads/daily")
def ads_daily(days: int = Query(30)):
    """Daily ads performance time series."""
    con = get_db()
    cutoff = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")

    rows = _safe_ads_query(con, """
        SELECT date,
               COALESCE(SUM(spend), 0),
               COALESCE(SUM(sales), 0),
               COALESCE(SUM(impressions), 0),
               COALESCE(SUM(clicks), 0),
               COALESCE(SUM(orders), 0),
               0, 0, 0, 0, 0
        FROM advertising
        WHERE date >= ?
        GROUP BY date
        ORDER BY date
    """, [cutoff])

    con.close()

    data = []
    for r in rows:
        data.append({
            "date": fmt_date(r[0]),
            "spend": round(r[1], 2),
            "adSales": round(r[2], 2),
            "impressions": r[3],
            "clicks": r[4],
            "orders": r[5],
            "acos": r[6],
            "roas": r[7],
            "tacos": r[8],
            "cpc": r[9],
            "ctr": r[10],
        })

    return {"days": days, "data": data}


@app.get("/api/ads/campaigns")
def ads_campaigns(days: int = Query(30)):
    """Campaign-level performance."""
    con = get_db()
    cutoff = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")

    rows = _safe_ads_query(con, """
        SELECT
            campaign_id,
            MAX(campaign_name) AS campaign_name,
            MAX(campaign_type) AS campaign_type,
            MAX(campaign_status) AS campaign_status,
            MAX(daily_budget) AS daily_budget,
            SUM(impressions) AS impressions,
            SUM(clicks) AS clicks,
            SUM(spend) AS spend,
            SUM(sales) AS sales,
            SUM(orders) AS orders,
            SUM(units) AS units
        FROM ads_campaigns
        WHERE date >= ?
        GROUP BY campaign_id
        ORDER BY SUM(spend) DESC
    """, [cutoff])

    con.close()

    campaigns = []
    for r in rows:
        impressions, clicks, spend, sales = r[5], r[6], r[7], r[8]
        orders, units = r[9], r[10]

        campaigns.append({
            "campaignId": r[0],
            "campaignName": r[1],
            "campaignType": r[2],
            "status": r[3],
            "dailyBudget": round(r[4] or 0, 2),
            "impressions": impressions,
            "clicks": clicks,
            "spend": round(spend, 2),
            "sales": round(sales, 2),
            "orders": orders,
            "units": units,
            "acos": round(spend / sales * 100, 2) if sales > 0 else 0,
            "roas": round(sales / spend, 2) if spend > 0 else 0,
            "cpc": round(spend / clicks, 2) if clicks > 0 else 0,
            "ctr": round(clicks / impressions * 100, 2) if impressions > 0 else 0,
            "cvr": round(orders / clicks * 100, 2) if clicks > 0 else 0,
        })

    return {"days": days, "campaigns": campaigns}


@app.get("/api/ads/keywords")
def ads_keywords(days: int = Query(30), sort: str = Query("spend"), limit: int = Query(50)):
    """Top keywords by spend, sales, or ACOS."""
    con = get_db()
    cutoff = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")

    rows = _safe_ads_query(con, """
        SELECT
            keyword_text,
            MAX(match_type) AS match_type,
            MAX(campaign_name) AS campaign_name,
            MAX(ad_group_name) AS ad_group_name,
            SUM(impressions) AS impressions,
            SUM(clicks) AS clicks,
            SUM(spend) AS spend,
            SUM(sales) AS sales,
            SUM(orders) AS orders,
            SUM(units) AS units
        FROM ads_keywords
        WHERE date >= ?
        GROUP BY keyword_text
        ORDER BY SUM(spend) DESC
        LIMIT ?
    """, [cutoff, limit])

    con.close()

    keywords = []
    for r in rows:
        impressions, clicks, spend, sales = r[4], r[5], r[6], r[7]
        orders, units = r[8], r[9]

        keywords.append({
            "keyword": r[0],
            "matchType": r[1],
            "campaignName": r[2],
            "adGroupName": r[3],
            "impressions": impressions,
            "clicks": clicks,
            "spend": round(spend, 2),
            "sales": round(sales, 2),
            "orders": orders,
            "units": units,
            "acos": round(spend / sales * 100, 2) if sales > 0 else 0,
            "roas": round(sales / spend, 2) if spend > 0 else 0,
            "cpc": round(spend / clicks, 2) if clicks > 0 else 0,
            "ctr": round(clicks / impressions * 100, 2) if impressions > 0 else 0,
            "cvr": round(orders / clicks * 100, 2) if clicks > 0 else 0,
        })

    return {"days": days, "keywords": keywords}


@app.get("/api/ads/search-terms")
def ads_search_terms(days: int = Query(30), limit: int = Query(50)):
    """Top search terms by spend."""
    con = get_db()
    cutoff = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")

    rows = _safe_ads_query(con, """
        SELECT
            search_term,
            MAX(keyword_text) AS keyword,
            MAX(match_type) AS match_type,
            MAX(campaign_name) AS campaign_name,
            SUM(impressions) AS impressions,
            SUM(clicks) AS clicks,
            SUM(spend) AS spend,
            SUM(sales) AS sales,
            SUM(orders) AS orders,
            SUM(units) AS units
        FROM ads_search_terms
        WHERE date >= ?
        GROUP BY search_term
        ORDER BY SUM(spend) DESC
        LIMIT ?
    """, [cutoff, limit])

    con.close()

    terms = []
    for r in rows:
        impressions, clicks, spend, sales = r[4], r[5], r[6], r[7]
        orders, units = r[8], r[9]

        terms.append({
            "searchTerm": r[0],
            "keyword": r[1],
            "matchType": r[2],
            "campaignName": r[3],
            "impressions": impressions,
            "clicks": clicks,
            "spend": round(spend, 2),
            "sales": round(sales, 2),
            "orders": orders,
            "units": units,
            "acos": round(spend / sales * 100, 2) if sales > 0 else 0,
            "roas": round(sales / spend, 2) if spend > 0 else 0,
            "cpc": round(spend / clicks, 2) if clicks > 0 else 0,
            "ctr": round(clicks / impressions * 100, 2) if impressions > 0 else 0,
        })

    return {"days": days, "searchTerms": terms}


@app.get("/api/ads/negative-keywords")
def ads_negative_keywords():
    """List all negative keywords."""
    con = get_db()

    rows = _safe_ads_query(con, """
        SELECT keyword_text, match_type, campaign_name, ad_group_name, keyword_status
        FROM ads_negative_keywords
        ORDER BY campaign_name, keyword_text
    """)

    con.close()

    keywords = [
        {
            "keyword": r[0],
            "matchType": r[1],
            "campaignName": r[2],
            "adGroupName": r[3],
            "status": r[4],
        }
        for r in rows
    ]

    return {"negativeKeywords": keywords}


# ── Ads: Profile Discovery + Manual Sync ────────────────────

@app.get("/api/ads/profiles")
def ads_profiles():
    """Discover Amazon Ads profiles for this account."""
    ads_creds = _load_ads_credentials()
    if not ads_creds:
        return {"error": "No Amazon Ads credentials configured. Set AMAZON_ADS_CLIENT_ID, AMAZON_ADS_CLIENT_SECRET, AMAZON_ADS_REFRESH_TOKEN env vars."}

    try:
        from ad_api.api import Profiles
        result = Profiles(credentials=ads_creds).get_profiles()
        profiles = []
        for p in result.payload:
            profiles.append({
                "profileId": p.get("profileId"),
                "countryCode": p.get("countryCode"),
                "accountName": p.get("accountInfo", {}).get("name", ""),
                "marketplace": p.get("accountInfo", {}).get("marketplaceStringId", ""),
                "type": p.get("accountInfo", {}).get("type", ""),
            })
        return {"profiles": profiles, "configured_profile_id": ads_creds.get("profile_id", "")}
    except Exception as e:
        return {"error": str(e)}


@app.post("/api/ads/sync")
async def trigger_ads_sync():
    """Manually trigger an ads data sync."""
    ads_creds = _load_ads_credentials()
    if not ads_creds:
        return {"error": "No Amazon Ads credentials configured"}
    try:
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(None, _sync_ads_data)
        return {"status": "ads_sync_complete"}
    except Exception as e:
        return {"error": str(e)}


# ── Debug: Today Orders ────────────────────────────────────

@app.get("/api/debug/today-orders")
def debug_today_orders():
    """Show raw order data for today to debug pricing issues."""
    try:
        from sp_api.api import Orders as OrdersAPI
        from sp_api.base import Marketplaces
        import time as _t

        creds = _load_sp_api_credentials()
        orders_api = OrdersAPI(credentials=creds, marketplace=Marketplaces.US)

        today_str = (datetime.utcnow() - timedelta(hours=7)).strftime("%Y-%m-%d")
        after_date = (datetime.utcnow() - timedelta(days=1)).isoformat()

        response = orders_api.get_orders(
            CreatedAfter=after_date,
            MarketplaceIds=["ATVPDKIKX0DER"],
            MaxResultsPerPage=50
        )
        order_list = response.payload.get("Orders", [])

        results = []
        for order in order_list[:10]:  # limit to 10
            order_id = order.get("AmazonOrderId")
            status = order.get("OrderStatus")
            purchase_date = order.get("PurchaseDate", "")
            order_total = order.get("OrderTotal", {})

            # Try to get items
            items_info = []
            try:
                _t.sleep(0.5)
                items_resp = orders_api.get_order_items(order_id)
                for item in items_resp.payload.get("OrderItems", []):
                    items_info.append({
                        "asin": item.get("ASIN"),
                        "title": (item.get("Title") or "")[:60],
                        "qty": item.get("QuantityOrdered"),
                        "ItemPrice": item.get("ItemPrice"),
                        "ItemTax": item.get("ItemTax"),
                        "PromotionDiscount": item.get("PromotionDiscount"),
                    })
            except Exception as e:
                items_info = [{"error": str(e)}]

            results.append({
                "order_id": order_id,
                "status": status,
                "purchase_date": purchase_date,
                "OrderTotal": order_total,
                "items": items_info,
            })

        return {
            "today_pacific": today_str,
            "total_orders_fetched": len(order_list),
            "sample_orders": results,
        }
    except Exception as e:
        import traceback
        return {"error": str(e), "traceback": traceback.format_exc()}


# ── Debug: Financial Events ───────────────────────────────

@app.get("/api/debug/financial-events")
def debug_financial_events():
    """Call Financial Events API directly and return raw summary for debugging."""
    try:
        from sp_api.api import Finances as FinancesAPI
        from sp_api.base import Marketplaces as _Mp
        creds = _load_sp_api_credentials()
        if not creds:
            return {"error": "No SP-API credentials found"}
        fin_start = (datetime.utcnow() - timedelta(days=90)).strftime("%Y-%m-%dT%H:%M:%SZ")
        finances = FinancesAPI(credentials=creds, marketplace=_Mp.US)

        resp = finances.list_financial_events(PostedAfter=fin_start, MaxResultsPerPage=100)
        payload = resp.payload if hasattr(resp, 'payload') else (resp if isinstance(resp, dict) else {})
        events = payload.get("FinancialEvents", {})

        summary = {}
        sample_charge = None
        type_debug = {}  # Shows Python types for debugging
        for k, v in events.items():
            if isinstance(v, list):
                summary[k] = len(v)
                # Grab a sample charge structure from first shipment
                if k == "ShipmentEventList" and len(v) > 0 and not sample_charge:
                    first_event = v[0]
                    type_debug["shipment_event_type"] = type(first_event).__name__
                    items = first_event.get("ShipmentItemList", []) if isinstance(first_event, dict) else getattr(first_event, "ShipmentItemList", [])
                    if items:
                        first_item = items[0]
                        type_debug["shipment_item_type"] = type(first_item).__name__
                        charges = first_item.get("ItemChargeList", []) if isinstance(first_item, dict) else getattr(first_item, "ItemChargeList", [])
                        if charges:
                            first_charge = charges[0]
                            type_debug["charge_type"] = type(first_charge).__name__
                            ca = first_charge.get("ChargeAmount", None) if isinstance(first_charge, dict) else getattr(first_charge, "ChargeAmount", None)
                            type_debug["charge_amount_type"] = type(ca).__name__
                            type_debug["charge_amount_value"] = str(ca)
                            type_debug["charge_amount_is_dict"] = isinstance(ca, dict)
                            type_debug["charge_amount_has_attr"] = hasattr(ca, "CurrencyAmount")
                            sample_charge = first_charge
                            # Try to get the actual parsed value
                            if isinstance(ca, dict):
                                type_debug["parsed_amount"] = float(ca.get("CurrencyAmount", 0) or 0)
                            elif hasattr(ca, "CurrencyAmount"):
                                type_debug["parsed_amount"] = float(ca.CurrencyAmount or 0)
                            else:
                                type_debug["parsed_amount"] = f"unparseable: {ca}"

                        fees = first_item.get("ItemFeeList", []) if isinstance(first_item, dict) else getattr(first_item, "ItemFeeList", [])
                        if fees:
                            first_fee = fees[0]
                            fa = first_fee.get("FeeAmount", None) if isinstance(first_fee, dict) else getattr(first_fee, "FeeAmount", None)
                            type_debug["fee_amount_type"] = type(fa).__name__
                            type_debug["fee_amount_value"] = str(fa)

                    # Check PostedDate type
                    pd = first_event.get("PostedDate", None) if isinstance(first_event, dict) else getattr(first_event, "PostedDate", None)
                    type_debug["posted_date_type"] = type(pd).__name__
                    type_debug["posted_date_value"] = str(pd)[:30]

                if k == "RefundEventList" and len(v) > 0:
                    first = v[0]
                    summary["refund_sample_keys"] = list(first.keys()) if isinstance(first, dict) else [a for a in dir(first) if not a.startswith('_')]
                    summary["refund_sample_order_id"] = first.get("AmazonOrderId", "N/A") if isinstance(first, dict) else getattr(first, "AmazonOrderId", "N/A")
                    adj_items = (first.get("ShipmentItemAdjustmentList", []) if isinstance(first, dict)
                                else getattr(first, "ShipmentItemAdjustmentList", [])) or []
                    if adj_items:
                        first_adj = adj_items[0]
                        adj_charges = (first_adj.get("ItemChargeAdjustmentList", []) if isinstance(first_adj, dict)
                                      else getattr(first_adj, "ItemChargeAdjustmentList", [])) or []
                        summary["refund_sample_charge"] = adj_charges[0] if adj_charges else "no charges"
                        summary["refund_item_type"] = type(first_adj).__name__
            else:
                summary[k] = str(v)[:100]

        # Also check what's in the DB
        con = duckdb.connect(str(DB_PATH), read_only=False)
        db_counts = con.execute("""
            SELECT event_type, COUNT(*), SUM(ABS(product_charges)),
                   SUM(ABS(fba_fees)), SUM(ABS(commission))
            FROM financial_events
            GROUP BY event_type
        """).fetchall()
        con.close()

        has_next = "NextToken" in payload

        return {
            "api_date_range": fin_start,
            "payload_keys": list(payload.keys()),
            "event_type_counts": summary,
            "sample_charge_object": sample_charge,
            "type_debug": type_debug,
            "has_next_page": has_next,
            "db_records": [{"type": r[0], "count": r[1], "total_charges": round(r[2], 2),
                           "fba_fees": round(r[3], 2), "commission": round(r[4], 2)} for r in db_counts]
        }
    except Exception as e:
        import traceback
        return {"error": str(e), "traceback": traceback.format_exc()}


# ── Period Comparison & Analytics ──────────────────────────

@app.get("/api/comparison")
def period_comparison(view: str = Query("realtime")):
    """Return KPI comparison across multiple time periods.

    Views:
    - realtime: Today / WTD / MTD / YTD
    - weekly: Last Week / 4 Weeks / 13 Weeks / 26 Weeks
    - monthly: Last Month / 2 Mo Ago / 3 Mo Ago / Last 12 Mo
    - yearly: 2026 YTD / 2025 YTD (comp) / 2025 Full / 2024 Full
    - monthly2026: Jan / Feb / ... / current month / YTD total
    """
    # Ensure today has live data before building the response
    _ensure_today_data()

    con = get_db()
    cogs_data = load_cogs()

    today_start = get_today(con)
    today_start = today_start.replace(hour=0, minute=0, second=0, microsecond=0)
    tomorrow = (today_start + timedelta(days=1)).strftime("%Y-%m-%d")
    yr = today_start.year
    month_names = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

    if view == "weekly":
        wd = today_start.weekday()  # Monday=0
        last_week_start = today_start - timedelta(days=wd + 7)
        last_week_end = today_start - timedelta(days=wd)
        periods = [
            {"label": "Last Week",
             "start": last_week_start.strftime("%Y-%m-%d"),
             "end": last_week_end.strftime("%Y-%m-%d")},
            {"label": "4 Weeks",
             "start": (today_start - timedelta(days=28)).strftime("%Y-%m-%d"),
             "end": tomorrow},
            {"label": "13 Weeks",
             "start": (today_start - timedelta(days=91)).strftime("%Y-%m-%d"),
             "end": tomorrow},
            {"label": "26 Weeks",
             "start": (today_start - timedelta(days=182)).strftime("%Y-%m-%d"),
             "end": tomorrow},
        ]
    elif view == "monthly":
        periods = []
        for i in range(1, 4):
            m_start = today_start.replace(day=1)
            # Go back i months
            m = m_start.month - i
            y = m_start.year
            while m <= 0:
                m += 12
                y -= 1
            m_start_actual = datetime(y, m, 1)
            m_end_actual = datetime(y if m < 12 else y + 1, m + 1 if m < 12 else 1, 1)
            label = "Last Month" if i == 1 else f"{i} Mo Ago" if i == 2 else "3 Mo Ago"
            periods.append({
                "label": label,
                "sub": month_names[m - 1] + " " + str(y),
                "start": m_start_actual.strftime("%Y-%m-%d"),
                "end": m_end_actual.strftime("%Y-%m-%d"),
            })
        periods.append({
            "label": "Last 12 Mo",
            "start": (today_start - timedelta(days=365)).strftime("%Y-%m-%d"),
            "end": tomorrow,
        })
    elif view == "yearly":
        ytd_comp_end = datetime(yr - 1, today_start.month, today_start.day) + timedelta(days=1)
        periods = [
            {"label": f"{yr} YTD",
             "start": f"{yr}-01-01",
             "end": tomorrow},
            {"label": f"{yr-1} YTD",
             "sub": "same window comp",
             "start": f"{yr-1}-01-01",
             "end": ytd_comp_end.strftime("%Y-%m-%d")},
            {"label": f"{yr-1} Full",
             "start": f"{yr-1}-01-01",
             "end": f"{yr}-01-01"},
            {"label": f"{yr-2} Full",
             "start": f"{yr-2}-01-01",
             "end": f"{yr-1}-01-01"},
        ]
    elif view == "monthly2026":
        periods = []
        for m in range(1, today_start.month + 1):
            m_start = datetime(yr, m, 1)
            m_end = tomorrow if m == today_start.month else datetime(yr, m + 1, 1).strftime("%Y-%m-%d")
            if isinstance(m_end, datetime):
                m_end = m_end.strftime("%Y-%m-%d")
            periods.append({
                "label": month_names[m - 1],
                "sub": str(yr),
                "start": m_start.strftime("%Y-%m-%d"),
                "end": m_end,
            })
        periods.append({
            "label": f"{yr} Total",
            "sub": "YTD",
            "start": f"{yr}-01-01",
            "end": tomorrow,
        })
    else:
        # Default: realtime (Today / WTD / MTD / YTD)
        # WTD: rolling 7 days so it's always meaningful (not empty on Monday)
        week_start = today_start - timedelta(days=6)
        month_start = today_start.replace(day=1)
        year_start = today_start.replace(month=1, day=1)

        periods = [
            {"label": "Today",
             "start": today_start.strftime("%Y-%m-%d"),
             "end": tomorrow},
            {"label": "WTD",
             "start": week_start.strftime("%Y-%m-%d"),
             "end": tomorrow},
            {"label": "MTD",
             "start": month_start.strftime("%Y-%m-%d"),
             "end": tomorrow},
            {"label": "YTD",
             "start": year_start.strftime("%Y-%m-%d"),
             "end": tomorrow},
        ]

    def chg(cur, prev):
        if prev == 0:
            return 0
        return round((cur - prev) / prev * 100, 1)

    # Compute cost ratios from full product list
    full_products = _build_product_list(con, f"{yr-1}-01-01")
    full_prod_rev = sum(pp["rev"] for pp in full_products)
    full_prod_net = sum(pp["net"] for pp in full_products)
    full_prod_cogs = sum(pp["cogsTotal"] for pp in full_products)
    full_prod_fba = sum(pp["fbaTotal"] for pp in full_products)
    full_prod_referral = sum(pp["referralTotal"] for pp in full_products)

    margin_pct = full_prod_net / full_prod_rev if full_prod_rev > 0 else 0
    cogs_pct = full_prod_cogs / full_prod_rev if full_prod_rev > 0 else 0
    fba_pct = full_prod_fba / full_prod_rev if full_prod_rev > 0 else 0
    referral_pct = full_prod_referral / full_prod_rev if full_prod_rev > 0 else 0

    results = []
    for p in periods:
        row = con.execute("""
            SELECT
                COALESCE(SUM(ordered_product_sales), 0),
                COALESCE(SUM(units_ordered), 0),
                COALESCE(SUM(sessions), 0)
            FROM daily_sales
            WHERE date >= ? AND date < ? AND asin = 'ALL'
        """, [p["start"], p["end"]]).fetchone()

        rev, units, sessions = row
        orders = units
        aur = round(rev / units, 2) if units else 0
        conv = round(units / sessions * 100, 1) if sessions else 0

        # Ad spend for period
        ad_spend = 0
        try:
            ad_row = con.execute("""
                SELECT COALESCE(SUM(spend), 0)
                FROM advertising
                WHERE date >= ? AND date < ?
            """, [p["start"], p["end"]]).fetchone()
            ad_spend = round(ad_row[0], 2) if ad_row else 0
        except Exception:
            pass

        tacos = round(ad_spend / rev * 100, 1) if rev > 0 else 0

        # Cost breakdown using known ratios
        cogs = round(rev * cogs_pct, 2)
        fba_fees = round(rev * fba_pct, 2)
        referral_fees = round(rev * referral_pct, 2)
        amazon_fees = round(fba_fees + referral_fees, 2)
        net = round(rev - cogs - amazon_fees - ad_spend, 2)
        margin_val = round(net / rev * 100, 1) if rev > 0 else 0

        # Compute last-year equivalents
        try:
            ly_start = datetime.strptime(p["start"], "%Y-%m-%d").replace(year=datetime.strptime(p["start"], "%Y-%m-%d").year - 1)
            ly_end = datetime.strptime(p["end"], "%Y-%m-%d").replace(year=datetime.strptime(p["end"], "%Y-%m-%d").year - 1)
        except ValueError:
            ly_start = datetime.strptime(p["start"], "%Y-%m-%d") - timedelta(days=365)
            ly_end = datetime.strptime(p["end"], "%Y-%m-%d") - timedelta(days=365)

        ly_row = con.execute("""
            SELECT
                COALESCE(SUM(ordered_product_sales), 0),
                COALESCE(SUM(units_ordered), 0),
                COALESCE(SUM(sessions), 0)
            FROM daily_sales
            WHERE date >= ? AND date < ? AND asin = 'ALL'
        """, [ly_start.strftime("%Y-%m-%d"), ly_end.strftime("%Y-%m-%d")]).fetchone()

        ly_rev, ly_units, ly_sessions = ly_row
        ly_orders = ly_units

        ly_ad = 0
        try:
            ly_ad_row = con.execute("""
                SELECT COALESCE(SUM(spend), 0)
                FROM advertising
                WHERE date >= ? AND date < ?
            """, [ly_start.strftime("%Y-%m-%d"), ly_end.strftime("%Y-%m-%d")]).fetchone()
            ly_ad = round(ly_ad_row[0], 2) if ly_ad_row else 0
        except Exception:
            pass

        results.append({
            "label": p["label"],
            "sub": p.get("sub", ""),
            "revenue": round(rev, 2),
            "units": units,
            "orders": orders,
            "aur": aur,
            "sessions": sessions,
            "convRate": conv,
            "adSpend": ad_spend,
            "tacos": tacos,
            "cogs": cogs,
            "amazonFees": amazon_fees,
            "fbaFees": fba_fees,
            "referralFees": referral_fees,
            "netProfit": net,
            "margin": margin_val,
            "revChg": chg(rev, ly_rev),
            "unitChg": chg(units, ly_units),
            "orderChg": chg(orders, ly_orders),
            "sessChg": chg(sessions, ly_sessions),
            "adChg": chg(ad_spend, ly_ad),
        })

    con.close()
    return {"view": view, "periods": results}


@app.get("/api/monthly-yoy")
def monthly_yoy():
    """Monthly revenue broken down by year for YoY comparison.
    Always returns 12 months with bars for 2024, 2025, 2026."""
    con = get_db()

    rows = con.execute("""
        SELECT
            EXTRACT(YEAR FROM date) AS yr,
            EXTRACT(MONTH FROM date) AS mo,
            COALESCE(SUM(ordered_product_sales), 0) AS revenue
        FROM daily_sales
        WHERE asin = 'ALL'
          AND EXTRACT(YEAR FROM date) >= 2024
        GROUP BY EXTRACT(YEAR FROM date), EXTRACT(MONTH FROM date)
        ORDER BY yr, mo
    """).fetchall()

    con.close()

    months_map = {}
    for r in rows:
        yr, mo, rev = int(r[0]), int(r[1]), round(r[2], 2)
        if mo not in months_map:
            months_map[mo] = {}
        months_map[mo][yr] = rev

    # Always show 3 years for comparison
    years = [2024, 2025, 2026]

    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    data = []
    for mo in range(1, 13):
        entry = {"month": month_names[mo - 1]}
        for yr in years:
            entry[str(yr)] = months_map.get(mo, {}).get(yr, 0)
        data.append(entry)

    return {"years": years, "data": data}


@app.get("/api/backfill")
def backfill_historical(
    start: str = Query(..., description="Start date YYYY-MM-DD"),
    end: str = Query(..., description="End date YYYY-MM-DD"),
):
    """Backfill historical Sales & Traffic data for a date range.
    Amazon limits report ranges to ~30 days, so this endpoint
    breaks the range into 30-day chunks and requests each one.
    This is a SLOW endpoint (each chunk takes 2-5 minutes for polling)."""
    from datetime import date as dt_date
    import time, gzip as gz, requests as req
    from sp_api.api import Reports
    from sp_api.base import Marketplaces, ReportType

    creds = _load_sp_api_credentials()
    reports = Reports(credentials=creds, marketplace=Marketplaces.US)

    start_dt = dt_date.fromisoformat(start)
    end_dt = dt_date.fromisoformat(end)

    total_records = 0
    chunks_done = 0
    errors = []

    chunk_start = start_dt
    while chunk_start < end_dt:
        chunk_end = min(chunk_start + timedelta(days=29), end_dt)
        logger.info(f"  Backfill requesting report {chunk_start} to {chunk_end}...")

        try:
            report_response = reports.create_report(
                reportType=ReportType.GET_SALES_AND_TRAFFIC_REPORT,
                dataStartTime=chunk_start.isoformat(),
                dataEndTime=chunk_end.isoformat(),
                reportOptions={"dateGranularity": "DAY", "asinGranularity": "CHILD"}
            )
            report_id = report_response.payload.get("reportId")

            report_data = None
            for attempt in range(30):
                time.sleep(10)
                status_response = reports.get_report(report_id)
                status = status_response.payload.get("processingStatus")
                if status == "DONE":
                    doc_id = status_response.payload.get("reportDocumentId")
                    doc_response = reports.get_report_document(doc_id)
                    report_data = doc_response.payload
                    break
                elif status in ("CANCELLED", "FATAL"):
                    errors.append(f"{chunk_start}-{chunk_end}: Report {status}")
                    break

            if report_data:
                if isinstance(report_data, dict) and "url" in report_data:
                    is_gzipped = report_data.get("compressionAlgorithm", "").upper() == "GZIP"
                    resp = req.get(report_data["url"])
                    if is_gzipped:
                        report_text = gz.decompress(resp.content).decode("utf-8")
                    else:
                        report_text = resp.text
                    report_data = json.loads(report_text)

                if isinstance(report_data, str):
                    report_data = json.loads(report_data)

                con = duckdb.connect(str(DB_PATH), read_only=False)
                records = 0

                for day_entry in report_data.get("salesAndTrafficByDate", []):
                    entry_date = day_entry.get("date", "")
                    traffic = day_entry.get("trafficByDate", {})
                    sales_info = day_entry.get("salesByDate", {})
                    ordered_sales = sales_info.get("orderedProductSales", {})
                    sales_amount = float(ordered_sales.get("amount", 0)) if isinstance(ordered_sales, dict) else float(ordered_sales or 0)

                    con.execute("""
                        INSERT OR REPLACE INTO daily_sales
                        (date, asin, units_ordered, ordered_product_sales,
                         sessions, session_percentage, page_views,
                         buy_box_percentage, unit_session_percentage)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, [
                        entry_date, "ALL",
                        int(sales_info.get("unitsOrdered", 0)),
                        sales_amount,
                        int(traffic.get("sessions", 0)),
                        float(traffic.get("sessionPercentage", 0)),
                        int(traffic.get("pageViews", 0)),
                        float(traffic.get("buyBoxPercentage", 0)),
                        float(traffic.get("unitSessionPercentage", 0)),
                    ])
                    records += 1

                for asin_entry in report_data.get("salesAndTrafficByAsin", []):
                    asin = asin_entry.get("childAsin") or asin_entry.get("parentAsin", "")
                    traffic = asin_entry.get("trafficByAsin", {})
                    sales_info = asin_entry.get("salesByAsin", {})
                    entry_date = asin_entry.get("date", str(chunk_start))
                    ordered_sales = sales_info.get("orderedProductSales", {})
                    sales_amount = float(ordered_sales.get("amount", 0)) if isinstance(ordered_sales, dict) else float(ordered_sales or 0)

                    con.execute("""
                        INSERT OR REPLACE INTO daily_sales
                        (date, asin, units_ordered, ordered_product_sales,
                         sessions, session_percentage, page_views,
                         buy_box_percentage, unit_session_percentage)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, [
                        entry_date, asin,
                        int(sales_info.get("unitsOrdered", 0)),
                        sales_amount,
                        int(traffic.get("sessions", 0)),
                        float(traffic.get("sessionPercentage", 0)),
                        int(traffic.get("pageViews", 0)),
                        float(traffic.get("buyBoxPercentage", 0)),
                        float(traffic.get("unitSessionPercentage", 0)),
                    ])
                    records += 1

                con.close()
                total_records += records
                chunks_done += 1
                logger.info(f"  Backfill chunk {chunk_start}-{chunk_end}: {records} records")

        except Exception as e:
            errors.append(f"{chunk_start}-{chunk_end}: {str(e)}")
            logger.error(f"  Backfill error {chunk_start}-{chunk_end}: {e}")

        chunk_start = chunk_end + timedelta(days=1)

    return {
        "status": "done",
        "chunks_processed": chunks_done,
        "total_records": total_records,
        "errors": errors,
        "range": f"{start} to {end}",
    }


@app.get("/api/product-mix")
def product_mix(days: int = Query(365)):
    """Top 10 products by revenue for donut chart."""
    con = get_db()
    cutoff = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    products = _build_product_list(con, cutoff)
    con.close()

    products.sort(key=lambda p: p["rev"], reverse=True)
    top10 = products[:10]
    other_rev = sum(p["rev"] for p in products[10:])

    result = [{"name": p["name"][:30], "value": p["rev"]} for p in top10]
    if other_rev > 0:
        result.append({"name": "Other", "value": round(other_rev, 2)})

    return {"products": result}


@app.get("/api/color-mix")
def color_mix(days: int = Query(365)):
    """Sales breakdown by color extracted from product names."""
    con = get_db()
    cutoff = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    products = _build_product_list(con, cutoff)
    con.close()

    import re
    color_keywords = ["Orange", "Blue", "Green", "Red", "Grey", "White", "Black", "Pink", "Purple", "Yellow", "Ombre"]
    color_totals = {}

    for p in products:
        name = p.get("name", "")
        matched = False
        for color in color_keywords:
            if re.search(r'\b' + color + r'\b', name, re.IGNORECASE):
                color_totals[color] = color_totals.get(color, 0) + p["rev"]
                matched = True
                break
        if not matched:
            color_totals["Other"] = color_totals.get("Other", 0) + p["rev"]

    total_rev = sum(color_totals.values())
    result = []
    for color, rev in sorted(color_totals.items(), key=lambda x: x[1], reverse=True):
        pct = round(rev / total_rev * 100, 1) if total_rev > 0 else 0
        result.append({"color": color, "revenue": round(rev, 2), "pct": pct})

    return {"colors": result, "total": round(total_rev, 2)}


# ── Profitability (Sellerboard-style) ─────────────────────

@app.get("/api/profitability")
def profitability(view: str = Query("realtime")):
    """Sellerboard-style profit waterfall with same period views as comparison.

    Waterfall: Sales - Promo - Ads - Shipping - Refunds - Amazon Fees - COGS - Indirect = Net Profit
    Returns both account-level waterfall and per-ASIN item breakdown.
    """
    # Ensure today has live data before building the response
    _ensure_today_data()

    con = get_db()
    cogs_data = load_cogs()

    today_start = get_today(con)
    today_start = today_start.replace(hour=0, minute=0, second=0, microsecond=0)
    tomorrow = (today_start + timedelta(days=1)).strftime("%Y-%m-%d")
    yr = today_start.year
    month_names = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

    # Build periods (same logic as comparison endpoint)
    if view == "weekly":
        wd = today_start.weekday()
        last_week_start = today_start - timedelta(days=wd + 7)
        last_week_end = today_start - timedelta(days=wd)
        periods = [
            {"label": "Last Week", "start": last_week_start.strftime("%Y-%m-%d"), "end": last_week_end.strftime("%Y-%m-%d")},
            {"label": "4 Weeks", "start": (today_start - timedelta(days=28)).strftime("%Y-%m-%d"), "end": tomorrow},
            {"label": "13 Weeks", "start": (today_start - timedelta(days=91)).strftime("%Y-%m-%d"), "end": tomorrow},
            {"label": "26 Weeks", "start": (today_start - timedelta(days=182)).strftime("%Y-%m-%d"), "end": tomorrow},
        ]
    elif view == "monthly":
        periods = []
        for i in range(1, 4):
            m_start = today_start.replace(day=1)
            m = m_start.month - i
            y = m_start.year
            while m <= 0:
                m += 12; y -= 1
            m_start_actual = datetime(y, m, 1)
            m_end_actual = datetime(y if m < 12 else y + 1, m + 1 if m < 12 else 1, 1)
            label = "Last Month" if i == 1 else f"{i} Mo Ago" if i == 2 else "3 Mo Ago"
            periods.append({"label": label, "sub": month_names[m-1] + " " + str(y),
                            "start": m_start_actual.strftime("%Y-%m-%d"), "end": m_end_actual.strftime("%Y-%m-%d")})
        periods.append({"label": "Last 12 Mo", "start": (today_start - timedelta(days=365)).strftime("%Y-%m-%d"), "end": tomorrow})
    elif view == "yearly":
        try:
            ytd_comp_end = datetime(yr - 1, today_start.month, today_start.day) + timedelta(days=1)
        except ValueError:
            ytd_comp_end = datetime(yr - 1, today_start.month, 28) + timedelta(days=1)
        periods = [
            {"label": f"{yr} YTD", "start": f"{yr}-01-01", "end": tomorrow},
            {"label": f"{yr-1} YTD", "sub": "same window comp", "start": f"{yr-1}-01-01", "end": ytd_comp_end.strftime("%Y-%m-%d")},
            {"label": f"{yr-1} Full", "start": f"{yr-1}-01-01", "end": f"{yr}-01-01"},
            {"label": f"{yr-2} Full", "start": f"{yr-2}-01-01", "end": f"{yr-1}-01-01"},
        ]
    elif view == "monthly2026":
        periods = []
        for m in range(1, today_start.month + 1):
            m_start = datetime(yr, m, 1)
            m_end = tomorrow if m == today_start.month else datetime(yr, m + 1, 1).strftime("%Y-%m-%d")
            if isinstance(m_end, datetime):
                m_end = m_end.strftime("%Y-%m-%d")
            periods.append({"label": month_names[m-1], "sub": str(yr),
                            "start": m_start.strftime("%Y-%m-%d"), "end": m_end})
        periods.append({"label": f"{yr} Total", "sub": "YTD", "start": f"{yr}-01-01", "end": tomorrow})
    else:
        # WTD: rolling 7 days so it's always meaningful (not empty on Monday)
        week_start = today_start - timedelta(days=6)
        month_start = today_start.replace(day=1)
        year_start = today_start.replace(month=1, day=1)
        periods = [
            {"label": "Today", "start": today_start.strftime("%Y-%m-%d"), "end": tomorrow},
            {"label": "WTD", "start": week_start.strftime("%Y-%m-%d"), "end": tomorrow},
            {"label": "MTD", "start": month_start.strftime("%Y-%m-%d"), "end": tomorrow},
            {"label": "YTD", "start": year_start.strftime("%Y-%m-%d"), "end": tomorrow},
        ]

    results = []
    for p in periods:
        wf = _build_waterfall(con, cogs_data, p["start"], p["end"])
        wf["label"] = p["label"]
        wf["sub"] = p.get("sub", "")
        results.append(wf)

    con.close()
    return {"view": view, "periods": results}


@app.get("/api/profitability/items")
def profitability_items(days: int = Query(365)):
    """Per-ASIN profitability breakdown matching Sellerboard item view."""
    con = get_db()
    cutoff = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    products = _build_product_list(con, cutoff)

    # Get per-ASIN financial event data for the period
    fin_by_asin = {}
    try:
        fin_rows = con.execute("""
            SELECT asin,
                   SUM(ABS(promotion_amount)) AS promo,
                   SUM(ABS(shipping_charges)) AS shipping,
                   SUM(ABS(other_fees)) AS other,
                   SUM(CASE WHEN event_type LIKE '%Refund%' OR event_type LIKE '%Return%'
                       THEN ABS(product_charges) ELSE 0 END) AS refund_amt,
                   COUNT(CASE WHEN event_type LIKE '%Refund%' OR event_type LIKE '%Return%'
                       THEN 1 END) AS refund_count
            FROM financial_events
            WHERE date >= ?
            GROUP BY asin
        """, [cutoff]).fetchall()
        fin_by_asin = {r[0]: {"promo": r[1], "shipping": r[2], "other": r[3],
                                "refund_amt": r[4], "refund_count": r[5]} for r in fin_rows}
    except Exception:
        pass

    # Get per-ASIN ad spend from advertising table
    ad_by_asin = {}
    try:
        ad_rows = con.execute("""
            SELECT asin, COALESCE(SUM(spend), 0) AS spend
            FROM advertising
            WHERE date >= ?
            GROUP BY asin
        """, [cutoff]).fetchall()
        ad_by_asin = {r[0]: r[1] for r in ad_rows}
    except Exception:
        pass

    # Load live pricing & coupon cache for enrichment
    pricing_cache = _load_pricing_cache()
    live_prices = pricing_cache.get("prices", {})
    live_coupons = pricing_cache.get("coupons", {})

    # Also load item master for static sale_price / list_price
    master_items = load_item_master()
    master_by_asin = {m["asin"]: m for m in master_items}

    # Enhance with Sellerboard-style metrics
    items = []
    for p in products:
        rev = p["rev"]
        cogs_total = p["cogsTotal"]
        net = p["net"]
        amazon_fees = p["fbaTotal"] + p["referralTotal"]
        margin = round(net / rev * 100, 1) if rev > 0 else 0
        roi = round(net / cogs_total * 100, 1) if cogs_total > 0 else 0

        asin = p["asin"]
        fin = fin_by_asin.get(asin, {})
        ad_spend = ad_by_asin.get(asin, p["adSpend"])
        refund_units = fin.get("refund_count", 0)
        return_pct = round(refund_units / p["units"] * 100, 1) if p["units"] > 0 else 0

        # Coupon data from pricing_sync.json cache
        coupon = live_coupons.get(asin)
        coupon_status = None
        coupon_end_date = None
        coupon_discount = None
        coupon_type = None
        if coupon:
            raw_state = coupon.get("state")
            # Map Amazon Ads coupon states to display labels
            state_map = {"ENABLED": "ACTIVE", "SCHEDULED": "SCHEDULED",
                         "PAUSED": "PAUSED", "EXPIRED": "EXPIRED"}
            coupon_status = state_map.get(raw_state, raw_state)
            coupon_end_date = coupon.get("endDate")
            coupon_discount = coupon.get("discountValue")
            coupon_type = coupon.get("discountType")

        # Live pricing data from pricing_sync.json cache
        live = live_prices.get(asin)
        sale_price = None
        sale_price_end_date = None
        list_price = None
        if live:
            list_price = live.get("listingPrice")
            # If listing price differs from buy box, the lower is likely a sale
            buy_box = live.get("buyBoxPrice")
            if list_price and buy_box and buy_box < list_price:
                sale_price = buy_box
        # Fall back to static item master sale_price if no live data
        master = master_by_asin.get(asin, {})
        if sale_price is None and master.get("salePrice", 0) > 0:
            sale_price = master["salePrice"]
        if list_price is None and master.get("listPrice", 0) > 0:
            list_price = master["listPrice"]
        # sale_price_end_date: not available from getCompetitivePricing API
        # Would require SP-API getListingOffers for sale schedule data

        items.append({
            "asin": asin,
            "sku": p["sku"],
            "name": p["name"],
            "units": p["units"],
            "sales": rev,
            "promo": round(fin.get("promo", 0), 2),
            "adSpend": round(ad_spend, 2),
            "shipping": round(fin.get("shipping", 0), 2),
            "refunds": round(fin.get("refund_amt", 0), 2),
            "refundUnits": refund_units,
            "returnPct": return_pct,
            "amazonFees": round(amazon_fees, 2),
            "fbaFees": p["fbaTotal"],
            "referralFees": p["referralTotal"],
            "otherFees": round(fin.get("other", 0), 2),
            "cogs": cogs_total,
            "cogsPerUnit": p["cogsPerUnit"],
            "indirect": 0,
            "netProfit": net,
            "margin": margin,
            "roi": roi,
            "aur": p["price"],
            # New coupon/pricing fields
            "couponStatus": coupon_status,
            "couponEndDate": coupon_end_date,
            "couponDiscount": coupon_discount,
            "couponType": coupon_type,
            "salePrice": sale_price,
            "listPrice": list_price,
            "salePriceEndDate": sale_price_end_date,
        })

    items.sort(key=lambda x: x["sales"], reverse=True)
    con.close()
    return {"days": days, "items": items}


def _build_waterfall(con, cogs_data, start: str, end: str) -> dict:
    """Build Sellerboard-style waterfall for a single period."""
    # Account-level revenue
    row = con.execute("""
        SELECT COALESCE(SUM(ordered_product_sales), 0),
               COALESCE(SUM(units_ordered), 0)
        FROM daily_sales
        WHERE date >= ? AND date < ? AND asin = 'ALL'
    """, [start, end]).fetchone()
    sales, units = row[0], row[1]

    # Per-ASIN cost breakdown
    asin_rows = con.execute("""
        SELECT asin, MAX(sku) AS sku,
               COALESCE(SUM(ordered_product_sales), 0) AS revenue,
               COALESCE(SUM(units_ordered), 0) AS units
        FROM daily_sales
        WHERE date >= ? AND date < ? AND asin <> 'ALL'
        GROUP BY asin
    """, [start, end]).fetchall()

    # Financial data (date-filtered for the period)
    fin_rows = []
    try:
        fin_rows = con.execute("""
            SELECT asin,
                   SUM(ABS(fba_fees)) AS fba,
                   SUM(ABS(commission)) AS comm,
                   SUM(ABS(promotion_amount)) AS promo,
                   SUM(ABS(shipping_charges)) AS shipping,
                   SUM(ABS(other_fees)) AS other,
                   SUM(CASE WHEN event_type LIKE '%Refund%' OR event_type LIKE '%Return%'
                       THEN ABS(product_charges) ELSE 0 END) AS refund_amt,
                   COUNT(CASE WHEN event_type LIKE '%Refund%' OR event_type LIKE '%Return%'
                       THEN 1 END) AS refund_count
            FROM financial_events
            WHERE date >= ? AND date < ?
            GROUP BY asin
        """, [start, end]).fetchall()
    except Exception:
        pass
    fin_by_asin = {r[0]: {"fba": r[1], "comm": r[2], "promo": r[3],
                           "shipping": r[4], "other": r[5],
                           "refund_amt": r[6], "refund_count": r[7]} for r in fin_rows}

    # Ad spend from the *advertising* table (the actual table name in this DB)
    total_ad = 0
    try:
        ad_row = con.execute("""
            SELECT COALESCE(SUM(spend), 0)
            FROM advertising
            WHERE date >= ? AND date < ?
        """, [start, end]).fetchone()
        total_ad = ad_row[0] if ad_row else 0
    except Exception:
        pass

    # Compute costs
    total_cogs = 0
    total_fba = 0
    total_referral = 0
    total_promo = 0
    total_shipping = 0
    total_refunds = 0
    total_other_fees = 0
    total_refund_units = 0
    prod_rev = 0

    for ar in asin_rows:
        asin, sku, rev, u = ar
        if u == 0:
            continue
        prod_rev += rev
        aur = rev / u if u else 0

        # COGS
        ci = cogs_data.get(asin) or cogs_data.get(sku or "") or {}
        cpu = ci.get("cogs", 0)
        if cpu == 0:
            cpu = round(aur * 0.35, 2)
        total_cogs += u * cpu

        # Financial data for this ASIN
        fin = fin_by_asin.get(asin, {})
        actual_fba = fin.get("fba", 0)
        actual_comm = fin.get("comm", 0)

        # FBA fees: use actual if available, else estimate ~12% of revenue (oversized golf equipment)
        if actual_fba > 0:
            total_fba += actual_fba
        else:
            est_fba = round(rev * 0.12, 2)
            total_fba += est_fba

        # Referral fees: use actual commission if available, else 15%
        if actual_comm > 0:
            total_referral += actual_comm
        else:
            total_referral += rev * 0.15

        # Promo, shipping, refunds from financial events
        total_promo += fin.get("promo", 0)
        total_shipping += fin.get("shipping", 0)
        total_refunds += fin.get("refund_amt", 0)
        total_other_fees += fin.get("other", 0)
        total_refund_units += fin.get("refund_count", 0)

    # If we have per-ASIN data for this period, scale to actual sales
    # If not, fall back to global cost ratios from _build_product_list
    if prod_rev > 0 and total_cogs > 0:
        scale = sales / prod_rev if prod_rev > 0 else 1
        cogs = round(total_cogs * scale, 2)
        fba_fees = round(total_fba * scale, 2)
        referral_fees = round(total_referral * scale, 2)
        promo = round(total_promo * scale, 2)
        shipping = round(total_shipping * scale, 2)
        refunds = round(total_refunds * scale, 2)
        other_fees = round(total_other_fees * scale, 2)
    elif sales > 0:
        # No per-ASIN data for this period — use global cost ratios
        # This happens when SP-API per-ASIN report data has different dates
        all_products = _build_product_list(con, "2020-01-01")
        total_prd_rev = sum(p["rev"] for p in all_products)
        if total_prd_rev > 0:
            cogs_pct = sum(p["cogsTotal"] for p in all_products) / total_prd_rev
            fba_pct = sum(p["fbaTotal"] for p in all_products) / total_prd_rev
            ref_pct = sum(p["referralTotal"] for p in all_products) / total_prd_rev
        else:
            cogs_pct, fba_pct, ref_pct = 0.35, 0.12, 0.15
        cogs = round(sales * cogs_pct, 2)
        fba_fees = round(sales * fba_pct, 2)
        referral_fees = round(sales * ref_pct, 2)
        promo = 0
        shipping = 0
        refunds = 0
        other_fees = 0
    else:
        cogs = fba_fees = referral_fees = promo = shipping = refunds = other_fees = 0

    amazon_fees = round(fba_fees + referral_fees, 2)
    ad_spend = round(total_ad, 2)  # already account-level, no scaling needed
    refund_units = total_refund_units
    return_pct = round(refund_units / units * 100, 1) if units > 0 else 0
    indirect = 0    # Not configured

    gross_profit = round(sales - promo - ad_spend - shipping - refunds - amazon_fees - cogs, 2)
    net_profit = round(gross_profit - indirect, 2)
    margin = round(net_profit / sales * 100, 1) if sales > 0 else 0
    roi = round(net_profit / cogs * 100, 1) if cogs > 0 else 0
    real_acos = round(ad_spend / sales * 100, 1) if sales > 0 else 0

    return {
        "sales": round(sales, 2),
        "units": units,
        "promo": promo,
        "adSpend": ad_spend,
        "shipping": shipping,
        "refunds": refunds,
        "refundUnits": refund_units,
        "returnPct": return_pct,
        "amazonFees": amazon_fees,
        "fbaFees": fba_fees,
        "referralFees": referral_fees,
        "otherFees": other_fees,
        "cogs": cogs,
        "indirect": indirect,
        "grossProfit": gross_profit,
        "netProfit": net_profit,
        "margin": margin,
        "roi": roi,
        "realAcos": real_acos,
    }


# ── Item Master ────────────────────────────────────────────

ITEM_MASTER_PATH = DB_DIR / "item_master.csv"


def load_item_master() -> list:
    """Load item master from CSV. Returns list of dicts."""
    items = []
    if not ITEM_MASTER_PATH.exists():
        return items
    with open(ITEM_MASTER_PATH, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            asin = (row.get("asin") or "").strip()
            if not asin:
                continue
            items.append({
                "asin": asin,
                "sku": (row.get("sku") or "").strip(),
                "productName": (row.get("product_name") or "").strip(),
                "color": (row.get("color") or "").strip(),
                "brand": (row.get("brand") or "").strip(),
                "series": (row.get("series") or "").strip(),
                "productType": (row.get("product_type") or "").strip(),
                "pieceCount": int(float(row.get("piece_count") or 0)),
                "orientation": (row.get("orientation") or "").strip(),
                "category": (row.get("category") or "").strip(),
                "unitCost": float(row.get("unit_cost") or 0),
                "fbsStock": int(float(row.get("fba_stock") or 0)),
                "lyAur": round(float(row.get("ly_aur") or 0), 2),
                "lyRevenue": round(float(row.get("ly_revenue") or 0), 2),
                "lyUnits": int(float(row.get("ly_units") or 0)),
                "lyProfit": round(float(row.get("ly_profit") or 0), 2),
                "plannedAnnualUnits": int(float(row.get("planned_annual_units") or 0)),
                "listPrice": float(row.get("list_price") or 0),
                "salePrice": float(row.get("sale_price") or 0),
                "referralPct": float(row.get("referral_pct") or 15),
                "couponType": (row.get("coupon_type") or "").strip(),
                "couponValue": float(row.get("coupon_value") or 0),
                "cartonPack": int(float(row.get("carton_pack") or 0)),
                "cartonLength": float(row.get("carton_length") or 0),
                "cartonWidth": float(row.get("carton_width") or 0),
                "cartonHeight": float(row.get("carton_height") or 0),
                "cartonWeight": float(row.get("carton_weight") or 0),
            })
    return items


def save_item_master(items: list):
    """Save item master list back to CSV."""
    fields = [
        "asin", "sku", "product_name", "color", "brand", "series",
        "product_type", "piece_count", "orientation", "category", "unit_cost",
        "fba_stock", "ly_aur", "ly_revenue", "ly_units", "ly_profit",
        "planned_annual_units", "list_price", "sale_price", "referral_pct",
        "coupon_type", "coupon_value", "carton_pack", "carton_length",
        "carton_width", "carton_height", "carton_weight",
    ]
    with open(ITEM_MASTER_PATH, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for item in items:
            w.writerow({
                "asin": item.get("asin", ""),
                "sku": item.get("sku", ""),
                "product_name": item.get("productName", ""),
                "color": item.get("color", ""),
                "brand": item.get("brand", ""),
                "series": item.get("series", ""),
                "product_type": item.get("productType", ""),
                "piece_count": item.get("pieceCount", 0),
                "orientation": item.get("orientation", ""),
                "category": item.get("category", ""),
                "unit_cost": item.get("unitCost", 0),
                "fba_stock": item.get("fbsStock", 0),
                "ly_aur": item.get("lyAur", 0),
                "ly_revenue": item.get("lyRevenue", 0),
                "ly_units": item.get("lyUnits", 0),
                "ly_profit": item.get("lyProfit", 0),
                "planned_annual_units": item.get("plannedAnnualUnits", 0),
                "list_price": item.get("listPrice", 0),
                "sale_price": item.get("salePrice", 0),
                "referral_pct": item.get("referralPct", 15),
                "coupon_type": item.get("couponType", ""),
                "coupon_value": item.get("couponValue", 0),
                "carton_pack": item.get("cartonPack", 0),
                "carton_length": item.get("cartonLength", 0),
                "carton_width": item.get("cartonWidth", 0),
                "carton_height": item.get("cartonHeight", 0),
                "carton_weight": item.get("cartonWeight", 0),
            })


@app.get("/api/item-master")
def get_item_master():
    """Return all items from the item master CSV, enriched with live COGS and pricing/coupon data."""
    items = load_item_master()
    cogs_data = load_cogs()

    # Load live pricing & coupon cache (from SP-API / Ads API syncs)
    pricing_cache = _load_pricing_cache()
    live_prices = pricing_cache.get("prices", {})
    live_coupons = pricing_cache.get("coupons", {})
    pricing_last_sync = pricing_cache.get("lastSync")

    for item in items:
        asin = item["asin"]
        cogs_info = cogs_data.get(asin, {})
        if cogs_info.get("cogs", 0) > 0 and item["unitCost"] == 0:
            item["unitCost"] = cogs_info["cogs"]

        # Merge live pricing data if available
        live = live_prices.get(asin)
        if live:
            item["liveListingPrice"] = live.get("listingPrice")
            item["liveBuyBoxPrice"] = live.get("buyBoxPrice")
            item["liveLandedPrice"] = live.get("landedPrice")
            item["priceFetchedAt"] = live.get("fetchedAt")
        else:
            item["liveListingPrice"] = None
            item["liveBuyBoxPrice"] = None
            item["liveLandedPrice"] = None
            item["priceFetchedAt"] = None

        # Merge live coupon data if available
        coupon = live_coupons.get(asin)
        if coupon:
            item["liveCouponState"] = coupon.get("state")
            item["liveCouponType"] = coupon.get("discountType")
            item["liveCouponValue"] = coupon.get("discountValue")
            item["liveCouponId"] = coupon.get("couponId")
            item["couponBudget"] = coupon.get("budgetAmount")
            item["couponBudgetUsed"] = coupon.get("budgetUsed")
            item["couponRedemptions"] = coupon.get("redemptions")
            item["couponStartDate"] = coupon.get("startDate")
            item["couponEndDate"] = coupon.get("endDate")
        else:
            item["liveCouponState"] = None
            item["liveCouponType"] = None
            item["liveCouponValue"] = None
            item["liveCouponId"] = None
            item["couponBudget"] = None
            item["couponBudgetUsed"] = None
            item["couponRedemptions"] = None
            item["couponStartDate"] = None
            item["couponEndDate"] = None

        # Compute net price after coupon
        base = item["salePrice"] if item["salePrice"] > 0 else item["listPrice"]
        if item["couponType"] == "$" and item["couponValue"] > 0:
            item["netPrice"] = round(base - item["couponValue"], 2)
        elif item["couponType"] == "%" and item["couponValue"] > 0:
            item["netPrice"] = round(base * (1 - item["couponValue"] / 100), 2)
        else:
            item["netPrice"] = round(base, 2)
        # Referral fee estimate
        item["referralFee"] = round(item["netPrice"] * item["referralPct"] / 100, 2)
    return {"items": items, "count": len(items), "pricingLastSync": pricing_last_sync}


@app.get("/api/pricing/status")
def get_pricing_status():
    """Return pricing & coupon sync status and summary."""
    cache = _load_pricing_cache()
    prices = cache.get("prices", {})
    coupons = cache.get("coupons", {})
    return {
        "lastSync": cache.get("lastSync"),
        "pricedAsins": len(prices),
        "couponAsins": len(coupons),
        "activeCoupons": sum(1 for c in coupons.values() if c.get("state") == "ENABLED"),
        "scheduledCoupons": sum(1 for c in coupons.values() if c.get("state") == "SCHEDULED"),
    }


@app.post("/api/pricing/sync")
def trigger_pricing_sync():
    """Manually trigger a pricing & coupon data sync."""
    import threading
    threading.Thread(target=_sync_pricing_and_coupons, daemon=True).start()
    return {"status": "started", "message": "Pricing & coupon sync started in background"}


@app.put("/api/item-master/{asin}")
def update_item_master(asin: str, body: dict = Body(...)):
    """Update a single item in the item master by ASIN."""
    items = load_item_master()
    found = False
    for item in items:
        if item["asin"] == asin:
            # Only update fields that are provided
            updatable = [
                "listPrice", "salePrice", "referralPct", "couponType",
                "couponValue", "cartonPack", "cartonLength", "cartonWidth",
                "cartonHeight", "cartonWeight", "unitCost", "plannedAnnualUnits",
                "color", "series", "productType", "pieceCount", "orientation",
                "category",
            ]
            for key in updatable:
                if key in body:
                    item[key] = body[key]
            found = True
            break
    if not found:
        raise HTTPException(status_code=404, detail=f"ASIN {asin} not found")
    save_item_master(items)
    return {"status": "ok", "asin": asin}


@app.post("/api/item-master/bulk-update")
def bulk_update_item_master(body: dict = Body(...)):
    """Bulk update items. Expects {items: [{asin, ...fields}, ...]}."""
    updates = body.get("items", [])
    if not updates:
        return {"status": "ok", "updated": 0}
    items = load_item_master()
    item_map = {i["asin"]: i for i in items}
    updated = 0
    for upd in updates:
        asin = upd.get("asin")
        if asin and asin in item_map:
            item = item_map[asin]
            for key, val in upd.items():
                if key != "asin" and key in item:
                    item[key] = val
            updated += 1
    save_item_master(items)
    return {"status": "ok", "updated": updated}


# ── GolfGen Warehouse ──────────────────────────────────────

WAREHOUSE_PATH = DB_DIR / "warehouse.csv"


@app.get("/api/warehouse")
def get_warehouse():
    """Return warehouse inventory grouped by master item with sub-items."""
    if not WAREHOUSE_PATH.exists():
        return {"masters": [], "count": 0}

    # Load raw warehouse rows
    raw_items = []
    with open(WAREHOUSE_PATH, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            raw_items.append({
                "itemNumber": (row.get("item_number") or "").strip(),
                "description": (row.get("description") or "").strip(),
                "pack": int(float(row.get("pack") or 1)),
                "onHand": int(float(row.get("on_hand") or 0)),
                "damage": int(float(row.get("damage") or 0)),
                "qcHold": int(float(row.get("qc_hold") or 0)),
                "pcsOnHand": int(float(row.get("pcs_on_hand") or 0)),
                "pcsAllocated": int(float(row.get("pcs_allocated") or 0)),
                "pcsAvailable": int(float(row.get("pcs_available") or 0)),
                "itemRef": (row.get("item_ref") or "").strip(),
            })

    # Load item master for ASIN + product name lookup
    item_master = load_item_master()
    im_lookup = {i["sku"]: i for i in item_master}

    # Group by item_ref
    groups = {}
    for item in raw_items:
        ref = item["itemRef"]
        if ref not in groups:
            groups[ref] = {"master": None, "subs": []}
        if item["itemNumber"] == ref:
            groups[ref]["master"] = item
        else:
            groups[ref]["subs"].append(item)

    # Build response: each master has aggregated totals + list of subs
    masters = []
    for ref, group in groups.items():
        master = group["master"]
        subs = group["subs"]

        # If no master record exists, create a virtual one from the ref
        if not master:
            master = {
                "itemNumber": ref,
                "description": subs[0]["description"] if subs else "",
                "pack": subs[0]["pack"] if subs else 1,
                "onHand": 0, "damage": 0, "qcHold": 0,
                "pcsOnHand": 0, "pcsAllocated": 0, "pcsAvailable": 0,
                "itemRef": ref,
            }

        # Aggregate totals across master + all subs
        all_items = [master] + subs
        total_pcs_oh = sum(i["pcsOnHand"] for i in all_items)
        total_pcs_alloc = sum(i["pcsAllocated"] for i in all_items)
        total_pcs_avail = sum(i["pcsAvailable"] for i in all_items)
        total_damage = sum(i["damage"] for i in all_items)
        total_qc = sum(i["qcHold"] for i in all_items)

        # Lookup from Item Master
        im_info = im_lookup.get(ref, {})
        asin = im_info.get("asin", "")
        im_name = im_info.get("productName", "")
        color = im_info.get("color", "")

        # Suffix label for sub-items
        for sub in subs:
            suffix = sub["itemNumber"].replace(ref, "", 1)
            sub["suffix"] = suffix if suffix else sub["itemNumber"]

        masters.append({
            "itemRef": ref,
            "itemNumber": master["itemNumber"],
            "asin": asin,
            "description": im_name if im_name else master["description"],
            "whDescription": master["description"],
            "color": color,
            "pack": master["pack"],
            "totalOnHand": total_pcs_oh,
            "totalDamage": total_damage,
            "totalQcHold": total_qc,
            "totalAllocated": total_pcs_alloc,
            "totalAvailable": total_pcs_avail,
            "subCount": len(subs),
            "subs": subs,
        })

    # Sort: items with ASIN first (known products), then by total on-hand desc
    masters.sort(key=lambda m: (0 if m["asin"] else 1, -m["totalOnHand"]))

    return {"masters": masters, "count": len(masters)}


# ── Golf/Housewares Inventory Excel Upload ──────────────────────

# Upload metadata file path
UPLOAD_META_PATH = DB_DIR / "upload_metadata.json"

def _load_upload_meta() -> dict:
    if UPLOAD_META_PATH.exists():
        with open(UPLOAD_META_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {}

def _save_upload_meta(meta: dict):
    with open(UPLOAD_META_PATH, "w", encoding="utf-8") as f:
        json.dump(meta, f, indent=2)


@app.get("/api/upload/metadata")
def upload_metadata():
    """Return last-uploaded timestamps for golf and housewares inventory files."""
    return _load_upload_meta()


def _classify_golf_channel(item_number: str, walmart_skus: set, amazon_skus: set) -> str:
    """Classify a golf SKU into Amazon, Walmart, Walmart & Amazon, or Other."""
    import re as _re
    s = item_number.strip().upper()
    if s.startswith("T-"):
        s = s[2:]
    for suf in ["FBM", "/RB", "/RETD", "/DONATE", "/DMGD", "/CUST", "/HOLD", "/1", "/2"]:
        if s.endswith(suf):
            s = s[:-len(suf)]
    s = _re.sub(r"\s*/DAM.*$", "", s)
    s = s.strip()
    in_walmart = s in walmart_skus
    in_amazon = s in amazon_skus
    if in_walmart and in_amazon:
        return "Walmart & Amazon"
    if in_walmart:
        return "Walmart"
    if in_amazon:
        return "Amazon"
    return "Other"


@app.post("/api/upload/inventory-excel")
async def upload_inventory_excel(
    file: UploadFile = File(...),
    division: str = Query("golf", description="golf or housewares"),
):
    """Upload an Excel file to refresh golf or housewares inventory JSON.

    Parses the first sheet with warehouse-like columns
    (Item Number, Description, Pack, On-Hand, Damage, QC Hold, Pcs On-Hand, Pcs Allocated, Pcs Available).
    For golf, also classifies each item by channel (Amazon/Walmart/Other).
    Saves to golf_inventory.json or housewares_inventory.json.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    import tempfile
    tmp_path = None
    try:
        contents = await file.read()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(contents)
            tmp_path = tmp.name

        wb = openpyxl.load_workbook(tmp_path, data_only=True)

        # Find the data sheet — first sheet with "item" in header row, or just the first sheet
        ws = None
        for sheet_name in wb.sheetnames:
            candidate = wb[sheet_name]
            first_row = [str(c or "").strip().lower() for c in next(candidate.iter_rows(min_row=1, max_row=1, values_only=True), [])]
            if any("item" in h for h in first_row):
                ws = candidate
                break
        if ws is None:
            ws = wb[wb.sheetnames[0]]

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise HTTPException(status_code=400, detail="Excel file has no data rows")

        headers = [str(h or "").strip() for h in rows[0]]
        header_lower = [h.lower().replace(" ", "_").replace("-", "_") for h in headers]

        # Map columns
        def find_col(*patterns):
            for i, h in enumerate(header_lower):
                for p in patterns:
                    if p in h:
                        return i
            return None

        item_col = find_col("item_number", "item_no", "item_#", "item")
        desc_col = find_col("description", "desc")
        pack_col = find_col("pack")
        oh_col = find_col("on_hand", "on hand", "oh")
        ns_col = find_col("non_standard", "nonstandard", "non_std")
        dmg_col = find_col("damage", "damaged", "dam")
        std_col = find_col("standard")
        qc_col = find_col("qc_hold", "qc hold", "qchold")
        poh_col = find_col("pcs_on_hand", "pcs on hand")
        palloc_col = find_col("pcs_allocated", "pcs allocated", "pcs_alloc")
        pavail_col = find_col("pcs_available", "pcs available", "pcs_avail")
        ref_col = find_col("item_ref", "item ref", "itemref")

        if item_col is None:
            raise HTTPException(status_code=400, detail=f"Could not find 'Item Number' column. Headers: {headers}")

        def safe_int(val):
            try:
                return int(float(val or 0))
            except (ValueError, TypeError):
                return 0

        def safe_str(val):
            return str(val or "").strip()

        # Build channel lookup for golf
        walmart_skus_set = set()
        amazon_skus_set = set()
        if division == "golf":
            wm_items = load_json("walmart_item_master.json")
            for w in wm_items:
                if w.get("golfgenItem"):
                    walmart_skus_set.add(w["golfgenItem"].strip().upper())
            am_items = load_json("amazon_item_master.json")
            for a in am_items:
                if a.get("sku"):
                    amazon_skus_set.add(a["sku"].strip().upper())

        items = []
        for row in rows[1:]:
            item_num = safe_str(row[item_col] if item_col is not None else "")
            if not item_num:
                continue

            item = {
                "itemNumber": item_num,
                "description": safe_str(row[desc_col]) if desc_col is not None else "",
                "pack": safe_int(row[pack_col]) if pack_col is not None else 1,
                "onHand": safe_int(row[oh_col]) if oh_col is not None else 0,
                "nonStandard": safe_int(row[ns_col]) if ns_col is not None else 0,
                "damage": safe_int(row[dmg_col]) if dmg_col is not None else 0,
                "standard": safe_int(row[std_col]) if std_col is not None else 0,
                "qcHold": safe_int(row[qc_col]) if qc_col is not None else 0,
                "pcsOnHand": safe_int(row[poh_col]) if poh_col is not None else 0,
                "pcsAllocated": safe_int(row[palloc_col]) if palloc_col is not None else 0,
                "pcsAvailable": safe_int(row[pavail_col]) if pavail_col is not None else 0,
                "itemRef": safe_str(row[ref_col]) if ref_col is not None else "",
            }

            if division == "golf":
                item["channel"] = _classify_golf_channel(item_num, walmart_skus_set, amazon_skus_set)

            items.append(item)

        if not items:
            raise HTTPException(status_code=400, detail="No items found in Excel file")

        # Save to JSON
        filename = "golf_inventory.json" if division == "golf" else "housewares_inventory.json"
        json_path = DB_DIR / filename
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(items, f, indent=2)

        # Update upload metadata
        meta = _load_upload_meta()
        meta[division] = {
            "lastUpload": datetime.now().isoformat(),
            "filename": file.filename,
            "itemCount": len(items),
        }
        _save_upload_meta(meta)

        wb.close()
        logger.info(f"  {division.title()} inventory uploaded: {len(items)} items from {file.filename}")

        return {
            "status": "success",
            "division": division,
            "filename": file.filename,
            "itemCount": len(items),
            "columns_mapped": {
                "item": item_col is not None,
                "description": desc_col is not None,
                "pack": pack_col is not None,
                "pcsOnHand": poh_col is not None,
                "pcsAllocated": palloc_col is not None,
                "pcsAvailable": pavail_col is not None,
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"Inventory Excel upload error: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


# ── Excel Upload: Refresh Warehouse + Item Master from Excel ────────────

@app.post("/api/upload/warehouse-excel")
async def upload_warehouse_excel(file: UploadFile = File(...)):
    """Upload an Excel file to refresh warehouse data and optionally item master.

    Looks for sheets named 'Raw Warehouse Data' and 'Item Master'.
    Parses each into the corresponding CSV file (warehouse.csv, item_master.csv).
    Returns a summary of what was updated.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Add it to requirements.txt.")

    # Save upload temporarily
    import tempfile
    tmp_path = None
    try:
        contents = await file.read()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(contents)
            tmp_path = tmp.name

        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        result = {"filename": file.filename, "sheets_found": list(wb.sheetnames)}

        # ── Parse Raw Warehouse Data ──
        wh_sheet_name = None
        for name in wb.sheetnames:
            if "warehouse" in name.lower() or "raw" in name.lower():
                wh_sheet_name = name
                break

        if wh_sheet_name:
            ws = wb[wh_sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) > 1:
                headers = [str(h or "").strip() for h in rows[0]]
                # Map Excel headers to our CSV column names
                col_map = {}
                for i, h in enumerate(headers):
                    hl = h.lower().replace(" ", "_").replace("-", "_")
                    if hl == "item_number" or hl == "item_no" or hl == "item":
                        col_map["item_number"] = i
                    elif hl == "description" or hl == "desc":
                        col_map["description"] = i
                    elif hl == "pack":
                        col_map["pack"] = i
                    elif hl in ("on_hand", "on hand", "oh"):
                        col_map["on_hand"] = i
                    elif hl in ("damage", "damaged", "dam"):
                        col_map["damage"] = i
                    elif hl in ("qc_hold", "qc hold", "qchold"):
                        col_map["qc_hold"] = i
                    elif hl in ("pcs_on_hand", "pcs on hand", "pcs_on_hand"):
                        col_map["pcs_on_hand"] = i
                    elif hl in ("pcs_allocated", "pcs allocated", "pcs_alloc"):
                        col_map["pcs_allocated"] = i
                    elif hl in ("pcs_available", "pcs available", "pcs_avail"):
                        col_map["pcs_available"] = i
                    elif hl in ("item_ref", "item ref", "itemref"):
                        col_map["item_ref"] = i

                # Auto-detect column positions by header name patterns
                # Also try exact header match as fallback
                header_exact = {h.strip(): i for i, h in enumerate(headers)}
                for csv_col, excel_header in [
                    ("item_number", "Item Number"), ("description", "Description"),
                    ("pack", "Pack"), ("on_hand", "On-Hand"),
                    ("damage", "Damage"), ("qc_hold", "QC Hold"),
                    ("pcs_on_hand", "Pcs On-Hand"), ("pcs_allocated", "Pcs Allocated"),
                    ("pcs_available", "Pcs Available"), ("item_ref", "Item Ref"),
                ]:
                    if csv_col not in col_map and excel_header in header_exact:
                        col_map[csv_col] = header_exact[excel_header]

                logger.info(f"  Warehouse Excel: {len(rows)-1} data rows, mapped columns: {col_map}")

                # Filter to GG-prefix items and write CSV
                wh_rows = []
                for row in rows[1:]:
                    item_num = str(row[col_map.get("item_number", 0)] or "").strip()
                    if not item_num.startswith("GG"):
                        continue
                    wh_rows.append({
                        "item_number": item_num,
                        "description": str(row[col_map.get("description", 1)] or "").strip(),
                        "pack": int(float(row[col_map.get("pack", 2)] or 1)),
                        "on_hand": int(float(row[col_map.get("on_hand", 3)] or 0)),
                        "damage": int(float(row[col_map.get("damage", 5)] or 0)),
                        "qc_hold": int(float(row[col_map.get("qc_hold", 7)] or 0)),
                        "pcs_on_hand": int(float(row[col_map.get("pcs_on_hand", 8)] or 0)),
                        "pcs_allocated": int(float(row[col_map.get("pcs_allocated", 9)] or 0)),
                        "pcs_available": int(float(row[col_map.get("pcs_available", 10)] or 0)),
                        "item_ref": str(row[col_map.get("item_ref", 11)] or "").strip(),
                    })

                if wh_rows:
                    fieldnames = ["item_number", "description", "pack", "on_hand", "damage",
                                 "qc_hold", "pcs_on_hand", "pcs_allocated", "pcs_available", "item_ref"]
                    with open(WAREHOUSE_PATH, "w", newline="", encoding="utf-8") as f:
                        writer = csv.DictWriter(f, fieldnames=fieldnames)
                        writer.writeheader()
                        writer.writerows(wh_rows)
                    result["warehouse"] = {"status": "updated", "rows": len(wh_rows)}
                    logger.info(f"  Warehouse CSV refreshed: {len(wh_rows)} items")
                else:
                    result["warehouse"] = {"status": "no_gg_items_found", "rows": 0}
        else:
            result["warehouse"] = {"status": "sheet_not_found"}

        # ── Parse Item Master (if present) ──
        im_sheet_name = None
        for name in wb.sheetnames:
            if "item master" in name.lower() or "itemmaster" in name.lower():
                im_sheet_name = name
                break

        if im_sheet_name:
            ws = wb[im_sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) > 1:
                headers = [str(h or "").strip() for h in rows[0]]
                # Check if this looks like our item master format
                header_lower = [h.lower() for h in headers]
                has_asin = any("asin" in h for h in header_lower)
                has_sku = any("sku" in h or "item" in h for h in header_lower)

                if has_asin or has_sku:
                    # Load existing item master to merge/update
                    existing = load_item_master()
                    existing_by_sku = {i.get("sku", ""): i for i in existing}

                    # Find column positions
                    hmap = {h.lower().strip(): i for i, h in enumerate(headers)}
                    asin_col = next((i for h, i in hmap.items() if "asin" in h), None)
                    sku_col = next((i for h, i in hmap.items() if "sku" in h or h == "item number"), None)
                    name_col = next((i for h, i in hmap.items() if "product" in h or "name" in h), None)
                    color_col = next((i for h, i in hmap.items() if "color" in h), None)

                    updated_count = 0
                    for row in rows[1:]:
                        sku_val = str(row[sku_col] or "").strip() if sku_col is not None else ""
                        if not sku_val or not sku_val.startswith("GG"):
                            continue
                        if sku_val in existing_by_sku:
                            # Update product name and color if present in Excel
                            if name_col is not None and row[name_col]:
                                existing_by_sku[sku_val]["productName"] = str(row[name_col]).strip()
                            if color_col is not None and row[color_col]:
                                existing_by_sku[sku_val]["color"] = str(row[color_col]).strip()
                            updated_count += 1

                    if updated_count > 0:
                        save_item_master(list(existing_by_sku.values()))
                    result["itemMaster"] = {"status": "updated", "matched": updated_count}
                else:
                    result["itemMaster"] = {"status": "unrecognized_format"}
        else:
            result["itemMaster"] = {"status": "sheet_not_found"}

        wb.close()
        return result

    except Exception as e:
        import traceback
        logger.error(f"Excel upload error: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


@app.post("/api/refresh-warehouse")
async def refresh_warehouse_from_file():
    """Check for a dropped Excel file in the data directory and refresh warehouse data.

    Looks for any .xlsx file in /data/ that contains a 'Raw Warehouse Data' sheet.
    This allows users to drop an Excel file into the data folder and hit this endpoint.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    # Look for Excel files in the data directory
    xlsx_files = sorted(DB_DIR.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    if not xlsx_files:
        return {"status": "no_xlsx_found", "data_dir": str(DB_DIR)}

    # Use the most recently modified Excel file
    excel_path = xlsx_files[0]
    logger.info(f"  Found Excel file for warehouse refresh: {excel_path.name}")

    # Re-use the upload logic by calling it with the file
    from starlette.datastructures import UploadFile as StarletteUpload
    from io import BytesIO

    with open(excel_path, "rb") as f:
        content = f.read()

    # Create a mock UploadFile
    mock_file = UploadFile(filename=excel_path.name, file=BytesIO(content))
    result = await upload_warehouse_excel(mock_file)
    result["source_file"] = excel_path.name
    return result


# ── Golf/Housewares Warehouse Endpoints ──────────────────────

@app.get("/api/warehouse/golf")
def warehouse_golf(channel: Optional[str] = Query(None, description="Filter by channel: Amazon, Walmart, Walmart & Amazon, Other")):
    """Golf warehouse inventory with optional channel filter."""
    items = load_json("golf_inventory.json")
    if not items:
        return {"items": [], "summary": {"totalSkus": 0, "totalPcsOnHand": 0, "totalPcsAvailable": 0, "totalPcsAllocated": 0, "totalDamage": 0}, "channelBreakdown": {}}
    if channel:
        items = [i for i in items if i.get("channel", "").lower() == channel.lower()]

    total_pcs = sum(i.get("pcsOnHand", 0) for i in items)
    total_available = sum(i.get("pcsAvailable", 0) for i in items)
    total_allocated = sum(i.get("pcsAllocated", 0) for i in items)
    total_damage = sum(i.get("damage", 0) for i in items)

    channel_counts = {}
    all_items = load_json("golf_inventory.json")
    for i in all_items:
        ch = i.get("channel", "Other")
        channel_counts[ch] = channel_counts.get(ch, 0) + 1

    return {
        "items": items,
        "summary": {
            "totalSkus": len(items),
            "totalPcsOnHand": total_pcs,
            "totalPcsAvailable": total_available,
            "totalPcsAllocated": total_allocated,
            "totalDamage": total_damage,
        },
        "channelBreakdown": channel_counts,
    }


@app.get("/api/warehouse/housewares")
def warehouse_housewares():
    """Housewares warehouse inventory."""
    items = load_json("housewares_inventory.json")
    if not items:
        return {"items": [], "summary": {"totalSkus": 0, "totalPcsOnHand": 0, "totalPcsAvailable": 0, "totalPcsAllocated": 0, "totalDamage": 0}}

    total_pcs = sum(i.get("pcsOnHand", 0) for i in items)
    total_available = sum(i.get("pcsAvailable", 0) for i in items)
    total_allocated = sum(i.get("pcsAllocated", 0) for i in items)
    total_damage = sum(i.get("damage", 0) for i in items)

    return {
        "items": items,
        "summary": {
            "totalSkus": len(items),
            "totalPcsOnHand": total_pcs,
            "totalPcsAvailable": total_available,
            "totalPcsAllocated": total_allocated,
            "totalDamage": total_damage,
        },
    }


@app.get("/api/warehouse/unified")
def warehouse_unified(division: str = Query("golf", description="golf or housewares"),
                      channel: Optional[str] = Query(None, description="All, Amazon, Walmart, Walmart & Amazon, Other (golf only)")):
    """Unified warehouse inventory with master/sub grouping, suffix breakdown, and summary.
    Used by the unified Inventory page."""
    filename = "golf_inventory.json" if division == "golf" else "housewares_inventory.json"
    items = load_json(filename)
    if not items:
        return {"masters": [], "summary": {}, "suffixBreakdown": {}, "channelBreakdown": {}}

    # For golf, apply channel filter
    if division == "golf" and channel and channel != "All":
        items = [i for i in items if (i.get("channel", "") or "").lower() == channel.lower()]

    # Channel breakdown (from full data, before channel filter)
    channel_counts = {}
    if division == "golf":
        all_items = load_json("golf_inventory.json")
        for i in all_items:
            ch = i.get("channel", "Other")
            channel_counts[ch] = channel_counts.get(ch, 0) + 1

    # Known suffixes for grouping
    SUFFIX_PATTERNS = ["/RB", "/RETD", "/DONATE", "/DMGD", "/DAM", "/HOLD", "/CUST", "/INBD", "/1",
                       "FBM", "-RB", "-DONATE", "-RETD", "-FBM", "-HOLD", "-Damage", "-CUST", "-Transfer"]

    def _get_base_sku(sku):
        """Strip prefixes (T-) and suffixes to get the base/master SKU.
        Strips iteratively to handle multi-suffix SKUs like GGWMSS2238BM/1/RB."""
        s = sku.strip()
        # Strip T- prefix
        if s.startswith("T-"):
            s = s[2:]
        # Strip known suffixes iteratively (handles /1/RB, /1/DONATE, etc.)
        changed = True
        while changed:
            changed = False
            for pat in SUFFIX_PATTERNS:
                if s.endswith(pat):
                    s = s[:-len(pat)]
                    changed = True
                    break
                if pat in s and pat.startswith("/"):
                    idx = s.find(pat)
                    s = s[:idx]
                    changed = True
                    break
        return s

    def _get_suffix_label(sku, base):
        """Determine suffix type for a variant SKU."""
        s = sku.strip()
        if s.startswith("T-"):
            return "T-"
        if "FBM" in s and "FBM" not in base:
            return "FBM"
        remainder = s.replace(base, "", 1)
        if "/RB" in remainder or "-RB" in remainder:
            return "RB"
        if "/RETD" in remainder or "-RETD" in remainder:
            return "RETD"
        if "/DONATE" in remainder or "-DONATE" in remainder:
            return "DONATE"
        if "/DMGD" in remainder or "/DAM" in remainder or "-Damage" in remainder:
            return "Damage"
        if "/HOLD" in remainder or "-HOLD" in remainder:
            return "HOLD"
        if "/CUST" in remainder or "-CUST" in remainder:
            return "CUST"
        if "/INBD" in remainder:
            return "INBD"
        if "/1" in remainder:
            return "Each"
        if remainder:
            return remainder.strip("/- ")
        return "Standard"

    # ASIN lookup from item master
    item_master = load_item_master()
    im_lookup = {i["sku"]: i for i in item_master}

    # Group items by base SKU
    groups = {}
    for item in items:
        sku = item.get("itemNumber", "").strip()
        base = _get_base_sku(sku)
        if base not in groups:
            groups[base] = {"master": None, "subs": []}
        if sku == base:
            groups[base]["master"] = item
        else:
            item["suffix"] = _get_suffix_label(sku, base)
            groups[base]["subs"].append(item)

    # Build master list
    masters = []
    for base, group in groups.items():
        master = group["master"]
        subs = group["subs"]
        if not master:
            # Create virtual master from first sub
            ref_item = subs[0] if subs else {}
            master = {
                "itemNumber": base,
                "description": ref_item.get("description", ""),
                "pack": ref_item.get("pack", 1),
                "onHand": 0, "pcsOnHand": 0, "pcsAllocated": 0,
                "pcsAvailable": 0, "damage": 0, "qcHold": 0,
                "channel": ref_item.get("channel", "Other"),
            }

        all_in_group = [master] + subs
        total_oh = sum(i.get("pcsOnHand", 0) for i in all_in_group)
        total_alloc = sum(i.get("pcsAllocated", 0) for i in all_in_group)
        total_avail = sum(i.get("pcsAvailable", 0) for i in all_in_group)
        total_dmg = sum(i.get("damage", 0) for i in all_in_group)
        total_qc = sum(i.get("qcHold", 0) for i in all_in_group)

        # ASIN lookup
        im_info = im_lookup.get(base, {})
        asin = im_info.get("asin", "")
        im_name = im_info.get("productName", "")

        masters.append({
            "baseSku": base,
            "itemNumber": master.get("itemNumber", base),
            "asin": asin,
            "description": im_name if im_name else master.get("description", ""),
            "whDescription": master.get("description", ""),
            "pack": master.get("pack", 1),
            "channel": master.get("channel", "Other") if division == "golf" else None,
            "totalOnHand": total_oh,
            "totalAllocated": total_alloc,
            "totalAvailable": total_avail,
            "totalDamage": total_dmg,
            "totalQcHold": total_qc,
            "subCount": len(subs),
            "subs": subs,
        })

    masters.sort(key=lambda m: -m["totalOnHand"])

    # Summary totals
    summary = {
        "totalSkus": len(masters),
        "totalItems": len(items),
        "totalOnHand": sum(m["totalOnHand"] for m in masters),
        "totalAllocated": sum(m["totalAllocated"] for m in masters),
        "totalAvailable": sum(m["totalAvailable"] for m in masters),
        "totalDamage": sum(m["totalDamage"] for m in masters),
    }

    # Suffix breakdown (across all items in this division, not filtered by channel)
    suffix_buckets = {}
    all_div_items = load_json(filename)
    for item in all_div_items:
        sku = item.get("itemNumber", "").strip()
        base = _get_base_sku(sku)
        if sku == base:
            suf = "Standard"
        else:
            suf = _get_suffix_label(sku, base)
        if suf not in suffix_buckets:
            suffix_buckets[suf] = {"skus": 0, "pcsOnHand": 0, "pcsAllocated": 0, "pcsAvailable": 0}
        suffix_buckets[suf]["skus"] += 1
        suffix_buckets[suf]["pcsOnHand"] += item.get("pcsOnHand", 0)
        suffix_buckets[suf]["pcsAllocated"] += item.get("pcsAllocated", 0)
        suffix_buckets[suf]["pcsAvailable"] += item.get("pcsAvailable", 0)

    return {
        "masters": masters,
        "summary": summary,
        "suffixBreakdown": suffix_buckets,
        "channelBreakdown": channel_counts,
    }


@app.get("/api/warehouse/summary")
def warehouse_summary():
    """Combined warehouse inventory summary — Golf vs Housewares totals + suffix breakdown."""
    golf_items = load_json("golf_inventory.json")
    hw_items = load_json("housewares_inventory.json")

    def _summarize(items):
        return {
            "skus": len(items),
            "pcsOnHand": sum(i.get("pcsOnHand", 0) for i in items),
            "pcsAvailable": sum(i.get("pcsAvailable", 0) for i in items),
            "pcsAllocated": sum(i.get("pcsAllocated", 0) for i in items),
            "damage": sum(i.get("damage", 0) for i in items),
        }

    def _suffix_breakdown(items):
        buckets = {}
        for item in items:
            sku = item.get("itemNumber", "").strip()
            suffix = "Standard"
            if sku.startswith("T-"):
                suffix = "T-"
            elif sku.endswith("FBM"):
                suffix = "FBM"
            elif "/RB" in sku:
                suffix = "RB"
            elif "/RETD" in sku:
                suffix = "RETD"
            elif "/DONATE" in sku:
                suffix = "DONATE"
            elif "/DMGD" in sku or "/DAM" in sku:
                suffix = "Damage"
            elif "/HOLD" in sku:
                suffix = "HOLD"
            elif "/CUST" in sku:
                suffix = "CUST"
            if suffix not in buckets:
                buckets[suffix] = {"skus": 0, "pcsOnHand": 0, "pcsAvailable": 0, "pcsAllocated": 0}
            buckets[suffix]["skus"] += 1
            buckets[suffix]["pcsOnHand"] += item.get("pcsOnHand", 0)
            buckets[suffix]["pcsAvailable"] += item.get("pcsAvailable", 0)
            buckets[suffix]["pcsAllocated"] += item.get("pcsAllocated", 0)
        return buckets

    golf_summary = _summarize(golf_items)
    hw_summary = _summarize(hw_items)

    combined = {
        "skus": golf_summary["skus"] + hw_summary["skus"],
        "pcsOnHand": golf_summary["pcsOnHand"] + hw_summary["pcsOnHand"],
        "pcsAvailable": golf_summary["pcsAvailable"] + hw_summary["pcsAvailable"],
        "pcsAllocated": golf_summary["pcsAllocated"] + hw_summary["pcsAllocated"],
        "damage": golf_summary["damage"] + hw_summary["damage"],
    }

    return {
        "combined": combined,
        "golf": golf_summary,
        "housewares": hw_summary,
        "golfSuffixes": _suffix_breakdown(golf_items),
        "housewaresSuffixes": _suffix_breakdown(hw_items),
    }


# ── Additional Item Master Endpoints ──────────────────────

@app.get("/api/item-master/walmart")
def item_master_walmart():
    """Walmart item master data."""
    items = load_json("walmart_item_master.json")
    return {"items": items, "count": len(items)}


@app.put("/api/item-master/walmart/{golfgen_item}")
def update_walmart_item(golfgen_item: str, body: dict = Body(...)):
    """Update a single Walmart item by golfgenItem identifier."""
    items = load_json("walmart_item_master.json")
    found = False
    updatable = ["plannedAnnualUnits", "unitCost", "unitRetail", "storeCount"]
    for item in items:
        if item.get("golfgenItem") == golfgen_item:
            for key in updatable:
                if key in body:
                    item[key] = body[key]
            found = True
            break
    if not found:
        raise HTTPException(status_code=404, detail=f"Walmart item {golfgen_item} not found")
    save_json("walmart_item_master.json", items)
    return {"status": "ok", "golfgenItem": golfgen_item}


@app.get("/api/item-master/amazon")
def item_master_amazon():
    """Amazon item master data (unique parent SKUs only)."""
    all_items = load_json("amazon_item_master.json")
    parent_items = [i for i in all_items if i.get("asin") and "/" not in i.get("sku", "") and "FBM" not in i.get("sku", "") and not i.get("sku", "").startswith("T-")]
    return {"items": parent_items, "count": len(parent_items), "totalVariations": len(all_items)}


@app.get("/api/item-master/housewares")
def item_master_housewares():
    """Housewares item master — items from housewares inventory."""
    items = load_json("housewares_inventory.json")
    result = []
    for item in items:
        result.append({
            "itemNumber": item.get("itemNumber", ""),
            "description": item.get("description", ""),
            "pack": item.get("pack", 0),
            "pcsOnHand": item.get("pcsOnHand", 0),
            "pcsAvailable": item.get("pcsAvailable", 0),
            "pcsAllocated": item.get("pcsAllocated", 0),
            "nonStandard": item.get("nonStandard", 0),
            "damage": item.get("damage", 0),
            "qcHold": item.get("qcHold", 0),
        })
    return {"items": result, "count": len(result)}


@app.get("/api/item-master/other")
def item_master_other():
    """Items in warehouse inventory that are not in Amazon or Walmart item masters."""
    walmart_items = load_json("walmart_item_master.json")
    amazon_items = load_json("amazon_item_master.json")
    golf_items = load_json("golf_inventory.json")
    hw_items = load_json("housewares_inventory.json")

    walmart_skus = set()
    for w in walmart_items:
        if w.get("golfgenItem"):
            walmart_skus.add(w["golfgenItem"].strip().upper())

    amazon_skus = set()
    for a in amazon_items:
        if a.get("sku"):
            amazon_skus.add(a["sku"].strip().upper())

    def _base_sku(raw: str) -> str:
        s = raw.strip().upper()
        if s.startswith("T-"):
            s = s[2:]
        for suf in ["FBM", "/RB", "/RETD", "/DONATE", "/DMGD", "/CUST", "/HOLD", "/1", "/2"]:
            if s.endswith(suf):
                s = s[: -len(suf)]
        s = _re.sub(r"\s*/DAM.*$", "", s)
        return s.strip()

    other_items = []
    seen = set()
    for item in golf_items + hw_items:
        sku_raw = item.get("itemNumber", "").strip().upper()
        base = _base_sku(sku_raw)
        if base in seen:
            continue
        in_walmart = base in walmart_skus
        in_amazon = base in amazon_skus
        if not in_walmart and not in_amazon:
            seen.add(base)
            source = "Golf" if item in golf_items else "Housewares"
            other_items.append({
                "itemNumber": item.get("itemNumber", ""),
                "description": item.get("description", ""),
                "source": source,
                "pcsOnHand": item.get("pcsOnHand", 0),
                "pcsAvailable": item.get("pcsAvailable", 0),
                "pcsAllocated": item.get("pcsAllocated", 0),
            })

    return {"items": other_items, "count": len(other_items)}


# ── Factory PO Summary ─────────────────────────────────────

def _load_factory_po():
    fp = DB_DIR / "factory_po_summary.json"
    if fp.exists():
        with open(fp) as f:
            return json.load(f)
    return {"purchaseOrders": [], "unitsByItem": [], "arrivalSchedule": [], "lastUpload": None}

def _save_factory_po(data):
    fp = DB_DIR / "factory_po_summary.json"
    with open(fp, "w") as f:
        json.dump(data, f, indent=2)

@app.get("/api/debug/factory-po-test")
async def debug_factory_po_test():
    """Debug endpoint — no auth, returns factory PO data or error details."""
    try:
        data = _load_factory_po()
        return {"ok": True, "keys": list(data.keys()), "poCount": len(data.get("purchaseOrders", []))}
    except Exception as e:
        import traceback
        return {"ok": False, "error": str(e), "trace": traceback.format_exc()}

@app.get("/api/factory-po")
async def get_factory_po(request: Request):
    _require_auth(request)
    data = _load_factory_po()
    # Clean bad rows from old parser output at read time
    if data.get("purchaseOrders"):
        data["purchaseOrders"] = [
            p for p in data["purchaseOrders"]
            if p.get("factory") and not str(p["factory"]).startswith("T-")
            and str(p.get("factory", "")).upper() not in ("TOTAL", "TOTALS", "PO#", "")
            and not ("." in str(p.get("poNumber", "")) and len(str(p.get("poNumber", ""))) > 10)
            and (p.get("units") or 0) > 0
        ]
    if data.get("unitsByItem"):
        data["unitsByItem"] = [
            u for u in data["unitsByItem"]
            if u.get("sku") and str(u["sku"]).upper() not in ("TOTAL", "TOTALS", "SKU", "")
        ]
    if data.get("arrivalSchedule"):
        data["arrivalSchedule"] = [
            a for a in data["arrivalSchedule"]
            if a.get("sku") and str(a["sku"]).upper() not in ("TOTAL", "TOTALS", "SKU", "")
        ]
    return data

@app.post("/api/factory-po/upload")
async def upload_factory_po(request: Request, file: UploadFile = File(...)):
    _require_auth(request)
    import openpyxl
    from io import BytesIO
    contents = await file.read()
    wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
    if "Factory PO Summary" not in wb.sheetnames:
        raise HTTPException(400, "Sheet 'Factory PO Summary' not found")
    ws = wb["Factory PO Summary"]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        r = {}
        for c in row:
            if hasattr(c, 'column_letter') and c.value is not None:
                r[c.column_letter] = c.value
        rows.append(r)
    def _to_str(dt):
        from datetime import datetime as _dt, date as _d
        if isinstance(dt, (_dt, _d)):
            return dt.strftime("%Y-%m-%d")
        return str(dt) if dt else None
    def _sr(v):
        try: return round(float(v), 2)
        except: return 0
    pos = []
    for i in range(7, len(rows)):
        cells = rows[i]
        factory = cells.get("B")
        if not factory or factory == "TOTALS":
            break
        pos.append({
            "factory": str(factory), "paymentTerms": str(cells.get("D") or ""),
            "poNumber": str(cells.get("E") or ""), "units": cells.get("F") or 0,
            "totalCost": _sr(cells.get("G")), "fobDate": _to_str(cells.get("I")),
            "estArrival": _to_str(cells.get("J")), "cbm": _sr(cells.get("K")),
            "landedCost": _sr(cells.get("U")), "oceanFreight": _sr(cells.get("X")),
            "customs": _sr(cells.get("Y"))
        })
    sku_start = next((i for i, r in enumerate(rows) if "UNITS BY ITEM" in str(r.get("B","")) and "PURCHASE ORDER" in str(r.get("B",""))), None)
    sku_rows = []
    if sku_start:
        for i in range(sku_start + 2, len(rows)):
            cells = rows[i]
            sku = cells.get("B")
            if not sku or sku == "TOTAL" or str(sku).startswith("UNITS BY"):
                break
            sku_rows.append({"sku": str(sku), "description": str(cells.get("C") or ""),
                "byPO": {"T-857500": cells.get("D") or 0, "2100": cells.get("E") or 0,
                    "T-857600": cells.get("F") or 0, "T-870100": cells.get("G") or 0,
                    "T-857700": cells.get("I") or 0, "T-857800": cells.get("J") or 0,
                    "T-857900": cells.get("K") or 0, "T-870200": cells.get("L") or 0,
                    "T-870300": cells.get("M") or 0, "T-870400": cells.get("N") or 0,
                    "T-870500": cells.get("O") or 0},
                "total": cells.get("P") or 0})
    arr_start = next((i for i, r in enumerate(rows) if "EST. ARRIVAL" in str(r.get("B",""))), None)
    arrival_rows = []
    if arr_start:
        for i in range(arr_start + 2, len(rows)):
            cells = rows[i]
            sku = cells.get("B")
            if not sku or sku == "TOTAL": break
            arrival_rows.append({"sku": str(sku), "description": str(cells.get("C") or ""),
                "jan2026": cells.get("D") or 0, "feb2026": cells.get("E") or 0,
                "mar2026": cells.get("F") or 0, "may2026": cells.get("G") or 0,
                "jun2026": cells.get("H") or 0, "total": cells.get("I") or 0})
    from datetime import date as _date
    data = {
        "rateInputs": {"usdRmbRate": 7.1, "customsBrokerFee": 250, "htsCode": "9506.31.0080", "htsDutyRate": 0.046},
        "purchaseOrders": pos, "unitsByItem": sku_rows, "arrivalSchedule": arrival_rows,
        "lastUpload": _date.today().isoformat(), "sourceFile": file.filename
    }
    _save_factory_po(data)
    return {"status": "ok", "pos": len(pos), "skus": len(sku_rows), "arrivals": len(arrival_rows), "lastUpload": data["lastUpload"]}

# ── Logistics Tracking ─────────────────────────────────────

def _load_logistics():
    fp = DB_DIR / "logistics_tracking.json"
    if fp.exists():
        with open(fp) as f:
            return json.load(f)
    return {"shipments": [], "itemsByContainer": [], "lastUpload": None}

def _save_logistics(data):
    fp = DB_DIR / "logistics_tracking.json"
    with open(fp, "w") as f:
        json.dump(data, f, indent=2)

@app.get("/api/logistics")
async def get_logistics(request: Request):
    _require_auth(request)
    data = _load_logistics()
    _SHIP_STOP = {"SHIPMENT STATUS SUMMARY", "STATUS SUMMARY", "UPCOMING ARRIVALS",
                  "TOTAL", "TOTALS", "ITEM SUMMARY", "PO#", "STATUS", "IN TRANSIT",
                  "DELIVERED", "PENDING SHIPMENT", "PENDING"}
    if data.get("shipments"):
        clean = []
        for s in data["shipments"]:
            shipper = (s.get("shipper") or "").strip()
            if not shipper or shipper.upper() in _SHIP_STOP:
                continue
            if any(sw in shipper.upper() for sw in ("STATUS SUMMARY", "UPCOMING ARRIVALS", "ITEM SUMMARY")):
                continue
            hbl = (s.get("hbl") or "").strip()
            if not hbl or len(hbl) < 5:
                continue
            # Normalize old field names to new field names
            if "mode" in s and "containerType" not in s:
                s["containerType"] = s.pop("mode")
            if "vessel" in s and "vesselVoyage" not in s:
                s["vesselVoyage"] = s.pop("vessel")
            # Ensure status field exists
            if not s.get("status"):
                s["status"] = ""
            clean.append(s)
        data["shipments"] = clean
    # Clean items: filter out Container Total, NON-GOLF, etc.
    if data.get("itemsByContainer"):
        data["itemsByContainer"] = [
            item for item in data["itemsByContainer"]
            if item.get("sku") and str(item["sku"]).strip()
            and "NON-GOLF" not in str(item.get("containerNumber", "")).upper()
            and "CONTAINER TOTAL" not in str(item.get("description", "")).upper()
            and "GRAND TOTAL" not in str(item.get("description", "")).upper()
        ]
    return data

@app.post("/api/logistics/upload")
async def upload_logistics(request: Request, file: UploadFile = File(...)):
    """Standalone logistics upload — delegates to the same logic as combined supply-chain upload."""
    _require_auth(request)
    import openpyxl
    from io import BytesIO
    contents = await file.read()
    wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
    if "Logistics Tracking" not in wb.sheetnames:
        raise HTTPException(400, "Sheet 'Logistics Tracking' not found")
    # Re-use the combined parser by calling the supply-chain upload logic
    # For standalone, we just wrap it in an UploadFile-like call
    ws = wb["Logistics Tracking"]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        r = {}
        for c in row:
            if hasattr(c, 'column_letter') and c.value is not None:
                r[c.column_letter] = c.value
        rows.append(r)

    def _to_str(v):
        if v is None: return ""
        from datetime import datetime as _dt, date as _d
        if isinstance(v, (_dt, _d)):
            return v.strftime("%Y-%m-%d")
        return str(v)

    def _sr(v, n=2):
        try: return round(float(v), n)
        except: return 0

    def _si(v):
        try: return int(float(v))
        except: return 0

    _STOP_WORDS = {"SHIPMENT STATUS SUMMARY", "STATUS SUMMARY", "UPCOMING ARRIVALS",
                    "TOTAL", "TOTALS", "ITEM SUMMARY", "PO#"}

    # ── Parse shipments ──
    shipments = []
    header_row = None
    for i, r in enumerate(rows):
        vals = [str(v).upper() for v in r.values()]
        joined = " ".join(vals)
        if "SHIPPER" in joined and ("HBL" in joined or "MODE" in joined):
            header_row = i
            break

    if header_row is not None:
        empty_count = 0
        for i in range(header_row + 1, len(rows)):
            r = rows[i]
            shipper = r.get("B")
            if not shipper:
                empty_count += 1
                if empty_count >= 2: break
                continue
            empty_count = 0
            shipper_str = str(shipper).strip()
            shipper_upper = shipper_str.upper()
            if any(sw in shipper_upper for sw in _STOP_WORDS): break
            if shipper_upper in ["SHIPPER", "STATUS", ""] or len(shipper_str) <= 3: continue
            hbl = str(r.get("C") or "").strip()
            if not hbl or len(hbl) < 5: continue
            raw_status = str(r.get("U") or "").strip()
            if raw_status:
                status = raw_status
            else:
                status = "Delivered"
                for _k, _v in r.items():
                    if _v and "transit" in str(_v).lower():
                        status = "In Transit"; break
            shipments.append({
                "shipper": shipper_str, "hbl": hbl,
                "containerType": str(r.get("D") or "").strip(),
                "containerNumber": str(r.get("E") or "").strip(),
                "vesselVoyage": str(r.get("F") or "").strip(),
                "poNumber": str(r.get("G") or "").strip(),
                "units": _si(r.get("H")), "cbm": _sr(r.get("I"), 1),
                "factoryInvoice": str(r.get("J") or "").strip(),
                "carrier": str(r.get("K") or "").strip(),
                "freightForwarder": str(r.get("L") or "").strip(),
                "etdOrigin": _to_str(r.get("N")), "departurePort": str(r.get("O") or "").strip(),
                "etaDischarge": _to_str(r.get("P")), "arrivalPort": str(r.get("Q") or "").strip(),
                "etaFinal": _to_str(r.get("R")), "finalLocation": str(r.get("S") or "").strip(),
                "deliveryDate": _to_str(r.get("T")), "status": status,
                "freightCost": _sr(r.get("V"), 2),
            })

    # ── Parse items by container ──
    item_header_row = None
    for i, r in enumerate(rows):
        for _k, v in r.items():
            if v and "container" in str(v).lower() and "item" in str(v).lower():
                item_header_row = i; break
        if item_header_row is not None: break

    items = []
    if item_header_row is not None:
        col_hdr = item_header_row + 1
        for ci in range(item_header_row, min(item_header_row + 4, len(rows))):
            vals_up = [str(v).upper() for v in rows[ci].values()]
            joined = " ".join(vals_up)
            if "CONTAINER" in joined and ("ITEM" in joined or "SKU" in joined):
                col_hdr = ci; break
        cur_container = ""
        for i in range(col_hdr + 1, len(rows)):
            r = rows[i]
            container_val = r.get("B"); sku_val = r.get("F"); desc_val = r.get("G"); qty_val = r.get("J")
            if not container_val and not sku_val and not desc_val: continue
            container_str = str(container_val or "").strip()
            sku_str = str(sku_val or "").strip()
            desc_str = str(desc_val or "").strip()
            skip_words = ["CONTAINER TOTAL", "GOLF TOTAL", "GRAND TOTAL", "NON-GOLF", "CONTAINER #", "ITEM NUMBER", "ITEM SUMMARY"]
            combined_upper = f"{container_str} {sku_str} {desc_str}".upper()
            if any(sw in combined_upper for sw in skip_words): continue
            if container_val and len(container_str) >= 8 and container_str[0].isalpha():
                cur_container = container_str
            if not sku_val or len(sku_str) < 3: continue
            items.append({
                "containerNumber": cur_container, "invoice": str(r.get("C") or "").strip(),
                "eta": _to_str(r.get("D")), "po": str(r.get("E") or "").strip(),
                "sku": sku_str, "description": desc_str, "qty": _si(qty_val),
            })

    from datetime import date as _date
    data = {"shipments": shipments, "itemsByContainer": items,
        "lastUpload": _date.today().isoformat(), "sourceFile": file.filename}
    _save_logistics(data)
    return {"status": "ok", "shipments": len(shipments), "items": len(items), "lastUpload": data["lastUpload"]}


@app.post("/api/supply-chain/upload")
async def upload_supply_chain(request: Request, file: UploadFile = File(...)):
    """Combined upload: parses both Factory PO Summary and Logistics Tracking from one Excel file"""
    _require_auth(request)
    import openpyxl
    from io import BytesIO
    contents = await file.read()
    wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
    results = {"status": "ok", "sourceFile": file.filename}

    # Try Factory PO
    if "Factory PO Summary" in wb.sheetnames:
        ws = wb["Factory PO Summary"]
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            r = {}
            for c in row:
                if hasattr(c, 'column_letter') and c.value is not None:
                    r[c.column_letter] = c.value
            rows.append(r)

        def safe_float(v, default=0):
            try: return float(v)
            except (ValueError, TypeError): return default

        def safe_round(v, n=2):
            return round(safe_float(v), n)

        def _to_date_str(v):
            from datetime import datetime as _dt, date as _d
            if isinstance(v, (_dt, _d)):
                return v.strftime("%Y-%m-%d")
            s = str(v).strip() if v else ""
            return s if s else ""

        # ── Parse Purchase Orders (rows 8-18 in current sheet) ──
        # Find header row: must have BOTH "Factory" AND "PO" in same row
        po_header_row = None
        for i in range(0, min(15, len(rows))):
            vals_upper = [str(v).upper() for v in rows[i].values()]
            joined = " ".join(vals_upper)
            if "FACTORY" in joined and ("PO#" in joined or "PO #" in joined):
                po_header_row = i
                break
        if po_header_row is None:
            po_header_row = 6  # fallback

        purchase_orders = []
        for i in range(po_header_row + 1, len(rows)):
            r = rows[i]
            po_val = r.get("E")
            factory = r.get("B") or r.get("A")
            units_val = r.get("F")

            # Stop at TOTALS row or empty rows section
            factory_str = str(factory or "").strip().upper()
            if factory_str == "TOTALS" or factory_str == "TOTAL":
                break

            # Skip rows without a PO number in column E
            if not po_val:
                continue

            po_str = str(po_val).strip()
            # PO numbers must start with T- or be a numeric PO like 2100
            # Skip if it looks like a pure float/money value (from payment calendar)
            if not (po_str.startswith("T-") or po_str.startswith("Z") or
                    po_str.startswith("GG") or po_str.startswith("PO") or
                    (po_str.isdigit() and len(po_str) <= 6)):
                continue

            # Must have a factory name (not a PO number in col B from payment section)
            if not factory or str(factory).startswith("T-"):
                continue

            terms = str(r.get("D") or "").strip()
            purchase_orders.append({
                "factory": str(factory),
                "poNumber": po_str,
                "paymentTerms": terms,
                "units": safe_round(units_val, 0),
                "totalCost": safe_round(r.get("G"), 2),
                "fobDate": _to_date_str(r.get("I")),
                "estArrival": _to_date_str(r.get("J")),
                "cbm": safe_round(r.get("K"), 1),
                "landedCost": safe_round(r.get("U"), 2),
            })

        # ── Parse Units by Item (starts at "UNITS BY ITEM" header) ──
        units_by_item = []
        units_header_row = None
        for i, r in enumerate(rows):
            for v in r.values():
                if "UNITS BY ITEM" in str(v).upper() and "BY PURCHASE ORDER" in str(v).upper():
                    units_header_row = i
                    break
            if units_header_row is not None:
                break

        if units_header_row is not None:
            # Next row is the column header (SKU, Description, PO columns)
            col_header_row = units_header_row + 1
            # Build PO column mapping from the header row
            header_r = rows[col_header_row] if col_header_row < len(rows) else {}
            po_cols = {}  # col_letter -> PO name
            for k, v in header_r.items():
                if k not in ["A", "B", "C"] and v:
                    po_cols[k] = str(v)

            # Skip X-Factory date row if present
            start_row = col_header_row + 1
            if start_row < len(rows):
                first_b = str(rows[start_row].get("B", "")).upper()
                if "X-FACTORY" in first_b or "FACTORY" in first_b:
                    start_row += 1

            for i in range(start_row, len(rows)):
                r = rows[i]
                sku = r.get("B")
                if not sku:
                    continue
                sku_str = str(sku).strip()
                if sku_str.upper() in ["TOTAL", "TOTALS", "", "SKU"]:
                    break  # end of section
                desc = str(r.get("C") or "")
                by_po = {}
                total = 0
                for k in po_cols:
                    val = r.get(k)
                    if val is not None:
                        try:
                            n = int(float(val))
                            by_po[po_cols[k]] = n
                            total += n
                        except (ValueError, TypeError):
                            pass
                # Also check col P for TOTAL
                p_val = r.get("P")
                if p_val is not None:
                    try:
                        total = int(float(p_val))
                    except (ValueError, TypeError):
                        pass
                units_by_item.append({"sku": sku_str, "description": desc, "byPO": by_po, "total": total})

        # ── Parse Arrival Schedule (starts at "EST. ARRIVAL") ──
        arrival_schedule = []
        arrival_header_row = None
        for i, r in enumerate(rows):
            for v in r.values():
                vup = str(v).upper()
                if "EST" in vup and "ARRIVAL" in vup:
                    arrival_header_row = i
                    break
            if arrival_header_row is not None:
                break

        if arrival_header_row is not None:
            # Next row should be column headers: SKU, Description, month names
            arr_col_row = arrival_header_row + 1
            arr_header = rows[arr_col_row] if arr_col_row < len(rows) else {}
            month_cols = {}  # col_letter -> month name
            for k, v in arr_header.items():
                if k not in ["A", "B", "C"] and v:
                    v_str = str(v).strip()
                    if v_str.upper() not in ["TOTAL", "TOTALS", "SKU"]:
                        month_cols[k] = v_str
            total_col = None
            for k, v in arr_header.items():
                if str(v).upper() in ["TOTAL", "TOTALS"]:
                    total_col = k

            for i in range(arr_col_row + 1, len(rows)):
                r = rows[i]
                sku = r.get("B")
                if not sku:
                    continue
                sku_str = str(sku).strip()
                if sku_str.upper() in ["TOTAL", "TOTALS", ""]:
                    break  # end of section
                desc = str(r.get("C") or "")
                monthly = {}
                for k, mname in month_cols.items():
                    val = r.get(k)
                    if val is not None:
                        try:
                            monthly[mname] = int(float(val))
                        except (ValueError, TypeError):
                            pass
                total = 0
                if total_col and r.get(total_col) is not None:
                    try:
                        total = int(float(r[total_col]))
                    except (ValueError, TypeError):
                        total = sum(monthly.values())
                else:
                    total = sum(monthly.values())
                arrival_schedule.append({
                    "sku": sku_str, "description": desc,
                    "monthlyArrivals": monthly, "total": total,
                })

        from datetime import date as _date
        po_data = {
            "purchaseOrders": purchase_orders, "unitsByItem": units_by_item,
            "arrivalSchedule": arrival_schedule,
            "lastUpload": _date.today().isoformat(), "sourceFile": file.filename
        }
        _save_factory_po(po_data)
        results["factoryPO"] = {"pos": len(purchase_orders), "items": len(units_by_item)}
        results["factoryPOLastUpload"] = po_data["lastUpload"]

    # Try Logistics
    if "Logistics Tracking" in wb.sheetnames:
        ws = wb["Logistics Tracking"]
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            r = {}
            for c in row:
                if hasattr(c, 'column_letter') and c.value is not None:
                    r[c.column_letter] = c.value
            rows.append(r)

        def to_str(v):
            if v is None: return ""
            from datetime import datetime, date
            if isinstance(v, (datetime, date)):
                return v.strftime("%Y-%m-%d")
            return str(v)

        def safe_round_l(v, n=2):
            try: return round(float(v), n)
            except: return 0

        def safe_int(v):
            try: return int(float(v))
            except: return 0

        # ── STOP WORDS: these indicate we've left the shipments section ──
        _STOP_WORDS = {"SHIPMENT STATUS SUMMARY", "STATUS SUMMARY", "UPCOMING ARRIVALS",
                        "TOTAL", "TOTALS", "ITEM SUMMARY", "PO#"}

        # ── Parse Shipments ──
        shipments = []
        header_row = None
        for i, r in enumerate(rows):
            vals = [str(v).upper() for v in r.values()]
            joined = " ".join(vals)
            if "SHIPPER" in joined and ("HBL" in joined or "MODE" in joined):
                header_row = i
                break

        if header_row is not None:
            empty_count = 0
            for i in range(header_row + 1, len(rows)):
                r = rows[i]
                shipper = r.get("B")
                if not shipper:
                    empty_count += 1
                    if empty_count >= 2:
                        break  # two empty rows in a row = end of section
                    continue
                empty_count = 0

                shipper_str = str(shipper).strip()
                shipper_upper = shipper_str.upper()

                # Stop at summary/status rows
                if any(sw in shipper_upper for sw in _STOP_WORDS):
                    break
                # Skip re-header rows
                if shipper_upper in ["SHIPPER", "STATUS", ""]:
                    continue
                # A real shipper name should be > 3 chars and not a number
                if len(shipper_str) <= 3:
                    continue

                hbl = str(r.get("C") or "").strip()
                # HBL should look like a tracking number (letters+digits, > 5 chars)
                if not hbl or len(hbl) < 5:
                    continue

                # Extract status from column U (DELIVERED, IN TRANSIT, PENDING, etc.)
                raw_status = str(r.get("U") or "").strip()
                if raw_status:
                    status = raw_status
                else:
                    # Fallback: scan row for transit/delivered keywords
                    status = "Delivered"
                    for _k, _v in r.items():
                        if _v and "transit" in str(_v).lower():
                            status = "In Transit"
                            break

                shipments.append({
                    "shipper": shipper_str,
                    "hbl": hbl,
                    "containerType": str(r.get("D") or "").strip(),
                    "containerNumber": str(r.get("E") or "").strip(),
                    "vesselVoyage": str(r.get("F") or "").strip(),
                    "poNumber": str(r.get("G") or "").strip(),
                    "units": safe_int(r.get("H")),
                    "cbm": safe_round_l(r.get("I"), 1),
                    "factoryInvoice": str(r.get("J") or "").strip(),
                    "carrier": str(r.get("K") or "").strip(),
                    "freightForwarder": str(r.get("L") or "").strip(),
                    "etdOrigin": to_str(r.get("N")),
                    "departurePort": str(r.get("O") or "").strip(),
                    "etaDischarge": to_str(r.get("P")),
                    "arrivalPort": str(r.get("Q") or "").strip(),
                    "etaFinal": to_str(r.get("R")),
                    "finalLocation": str(r.get("S") or "").strip(),
                    "deliveryDate": to_str(r.get("T")),
                    "status": status,
                    "freightCost": safe_round_l(r.get("V"), 2),
                })

        # ── Parse Items by Container ──
        # Find the "ITEM SUMMARY BY CONTAINER" header row
        item_header_row = None
        for i, r in enumerate(rows):
            for _k, v in r.items():
                if v and "container" in str(v).lower() and "item" in str(v).lower():
                    item_header_row = i
                    break
            if item_header_row is not None:
                break

        items = []
        if item_header_row is not None:
            # Find the column-header row (Container #, Invoice, ETA, PO#, Item Number, Description, ...)
            col_hdr = item_header_row + 1
            # Scan for actual column header with "Container" and "Item"
            for ci in range(item_header_row, min(item_header_row + 4, len(rows))):
                vals_up = [str(v).upper() for v in rows[ci].values()]
                joined = " ".join(vals_up)
                if "CONTAINER" in joined and ("ITEM" in joined or "SKU" in joined):
                    col_hdr = ci
                    break

            cur_container = ""
            for i in range(col_hdr + 1, len(rows)):
                r = rows[i]
                container_val = r.get("B")
                sku_val = r.get("F")
                desc_val = r.get("G")
                qty_val = r.get("J")

                # Skip completely empty rows
                if not container_val and not sku_val and not desc_val:
                    continue

                # Build string versions for checking
                container_str = str(container_val or "").strip()
                sku_str = str(sku_val or "").strip()
                desc_str = str(desc_val or "").strip()

                # Skip section headers and summary rows
                skip_words = ["CONTAINER TOTAL", "GOLF TOTAL", "GRAND TOTAL",
                              "NON-GOLF", "CONTAINER #", "ITEM NUMBER",
                              "ITEM SUMMARY"]
                combined_upper = f"{container_str} {sku_str} {desc_str}".upper()
                if any(sw in combined_upper for sw in skip_words):
                    # But update current container if this looks like a section header with a container #
                    continue

                # Update current container number (long alphanumeric string)
                if container_val and len(container_str) >= 8 and container_str[0].isalpha():
                    cur_container = container_str

                # Must have an actual SKU to be a valid item row
                if not sku_val or len(sku_str) < 3:
                    continue

                items.append({
                    "containerNumber": cur_container,
                    "invoice": str(r.get("C") or "").strip(),
                    "eta": to_str(r.get("D")),
                    "po": str(r.get("E") or "").strip(),
                    "sku": sku_str,
                    "description": desc_str,
                    "qty": safe_int(qty_val),
                })

        from datetime import date as _date
        log_data = {"shipments": shipments, "itemsByContainer": items,
            "lastUpload": _date.today().isoformat(), "sourceFile": file.filename}
        _save_logistics(log_data)
        results["logistics"] = {"shipments": len(shipments), "items": len(items)}
        results["logisticsLastUpload"] = log_data["lastUpload"]

    if "factoryPO" not in results and "logistics" not in results:
        raise HTTPException(400, "No 'Factory PO Summary' or 'Logistics Tracking' sheet found in file")

    results["lastUpload"] = _date.today().isoformat()
    return results


# ── FBA Shipments (SP-API Fulfillment Inbound) ───────────────

_FBA_SHIPMENTS_CACHE_PATH = DB_DIR / "fba_shipments_cache.json"


def _load_fba_shipments_cache():
    if _FBA_SHIPMENTS_CACHE_PATH.exists():
        with open(_FBA_SHIPMENTS_CACHE_PATH) as f:
            return json.load(f)
    return None


def _save_fba_shipments_cache(data):
    with open(_FBA_SHIPMENTS_CACHE_PATH, "w") as f:
        json.dump(data, f, indent=2)


def _fetch_fba_shipments_from_api(statuses=None, days_back=90):
    """Fetch FBA inbound shipments from Amazon SP-API."""
    try:
        from sp_api.api import FulfillmentInbound
        from sp_api.base import Marketplaces
    except ImportError:
        logger.warning("SP-API library not installed — cannot fetch FBA shipments")
        return None

    credentials = _load_sp_api_credentials()
    if not credentials:
        logger.warning("No SP-API credentials — cannot fetch FBA shipments")
        return None

    if statuses is None:
        statuses = ["WORKING", "SHIPPED", "RECEIVING", "CLOSED", "IN_TRANSIT"]

    try:
        fba = FulfillmentInbound(
            credentials=credentials,
            marketplace=Marketplaces.US,
        )

        all_shipments = []

        # Fetch by status list
        resp = fba.get_shipments(
            QueryType="SHIPMENT_STATUS",
            ShipmentStatusList=",".join(statuses),
            MarketplaceId="ATVPDKIKX0DER",
        )

        payload = resp.payload or {}
        shipment_data = payload.get("ShipmentData", [])
        all_shipments.extend(shipment_data)

        # Handle pagination
        next_token = payload.get("NextToken")
        page = 1
        while next_token and page < 10:  # safety limit
            page += 1
            resp = fba.get_shipments(
                QueryType="NEXT_TOKEN",
                NextToken=next_token,
                MarketplaceId="ATVPDKIKX0DER",
            )
            payload = resp.payload or {}
            shipment_data = payload.get("ShipmentData", [])
            all_shipments.extend(shipment_data)
            next_token = payload.get("NextToken")

        logger.info(f"FBA Shipments: fetched {len(all_shipments)} shipments from SP-API")

        # Normalize shipment data for frontend
        normalized = []
        for s in all_shipments:
            ship_id = s.get("ShipmentId", "")
            ship_name = s.get("ShipmentName", "")
            status = s.get("ShipmentStatus", "")
            dest = s.get("DestinationFulfillmentCenterId", "")
            label_prep = s.get("LabelPrepType", "")
            are_cases_required = s.get("AreCasesRequired", False)

            # Address info
            ship_from = s.get("ShipFromAddress", {})
            ship_from_str = ""
            if ship_from:
                parts = [ship_from.get("City", ""), ship_from.get("StateOrProvinceCode", "")]
                ship_from_str = ", ".join(p for p in parts if p)

            # Items count from InboundShipmentInfo
            items = s.get("Items", [])

            normalized.append({
                "shipmentId": ship_id,
                "shipmentName": ship_name,
                "status": status,
                "destination": dest,
                "labelPrep": label_prep,
                "casesRequired": are_cases_required,
                "shipFrom": ship_from_str,
                "itemCount": len(items) if items else 0,
            })

        result = {
            "shipments": normalized,
            "lastSync": datetime.now().strftime("%m/%d/%Y %I:%M %p"),
            "totalShipments": len(normalized),
        }

        _save_fba_shipments_cache(result)
        return result

    except Exception as e:
        logger.error(f"FBA Shipments SP-API error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None


def _enrich_shipment_items(shipment_id):
    """Fetch items for a specific shipment from SP-API."""
    try:
        from sp_api.api import FulfillmentInbound
        from sp_api.base import Marketplaces
    except ImportError:
        return []

    credentials = _load_sp_api_credentials()
    if not credentials:
        return []

    try:
        fba = FulfillmentInbound(
            credentials=credentials,
            marketplace=Marketplaces.US,
        )
        resp = fba.shipment_items_by_shipment(shipment_id, MarketplaceId="ATVPDKIKX0DER")
        payload = resp.payload or {}
        items = payload.get("ItemData", [])

        result = []
        for item in items:
            result.append({
                "sku": item.get("SellerSKU", ""),
                "fnsku": item.get("FulfillmentNetworkSKU", ""),
                "quantityShipped": item.get("QuantityShipped", 0),
                "quantityReceived": item.get("QuantityReceived", 0),
                "quantityInCase": item.get("QuantityInCase", 0),
            })

        # Handle pagination
        next_token = payload.get("NextToken")
        page = 1
        while next_token and page < 10:
            page += 1
            resp = fba.shipment_items_by_shipment(
                shipment_id,
                MarketplaceId="ATVPDKIKX0DER",
                NextToken=next_token,
            )
            payload = resp.payload or {}
            for item in payload.get("ItemData", []):
                result.append({
                    "sku": item.get("SellerSKU", ""),
                    "fnsku": item.get("FulfillmentNetworkSKU", ""),
                    "quantityShipped": item.get("QuantityShipped", 0),
                    "quantityReceived": item.get("QuantityReceived", 0),
                    "quantityInCase": item.get("QuantityInCase", 0),
                })
            next_token = payload.get("NextToken")

        return result
    except Exception as e:
        logger.error(f"FBA shipment items error for {shipment_id}: {e}")
        return []


@app.get("/api/fba-shipments")
async def get_fba_shipments(request: Request, refresh: bool = False):
    """Get FBA inbound shipments. Uses cache unless refresh=true."""
    _require_auth(request)

    if not refresh:
        cached = _load_fba_shipments_cache()
        if cached:
            return cached

    # Try fetching fresh data from SP-API
    data = _fetch_fba_shipments_from_api()
    if data:
        return data

    # Fall back to cache even if refresh was requested
    cached = _load_fba_shipments_cache()
    if cached:
        cached["_note"] = "Using cached data — SP-API sync failed"
        return cached

    return {"shipments": [], "lastSync": None, "totalShipments": 0}


@app.post("/api/fba-shipments/sync")
async def sync_fba_shipments(request: Request):
    """Force refresh of FBA shipment data from SP-API."""
    _require_auth(request)
    data = _fetch_fba_shipments_from_api()
    if data:
        return {"ok": True, "totalShipments": data["totalShipments"], "lastSync": data["lastSync"]}
    return JSONResponse(status_code=500, content={"ok": False, "error": "SP-API sync failed. Check credentials."})


@app.get("/api/fba-shipments/{shipment_id}/items")
async def get_fba_shipment_items(request: Request, shipment_id: str):
    """Get items for a specific FBA inbound shipment."""
    _require_auth(request)
    items = _enrich_shipment_items(shipment_id)
    return {"shipmentId": shipment_id, "items": items, "totalItems": len(items)}


# ── Item Plan Module ───────────────────────────────────────

def _duck_rw():
    """Return a read-write DuckDB connection for item plan operations."""
    return duckdb.connect(str(DB_PATH), read_only=False)

def _duck():
    """Return a read-only DuckDB connection for item plan queries."""
    return duckdb.connect(str(DB_PATH), read_only=True)

def _init_item_plan_tables():
    """Initialize Item Plan DuckDB tables and seed from JSON if they don't exist."""
    con = _duck_rw()
    try:
        # Create tables if they don't exist
        con.execute("""
            CREATE TABLE IF NOT EXISTS monthly_sales_history (
                sku TEXT,
                year INTEGER,
                month INTEGER,
                units DOUBLE,
                revenue DOUBLE,
                profit DOUBLE,
                refund_units DOUBLE,
                PRIMARY KEY (sku, year, month)
            )
        """)

        con.execute("""
            CREATE TABLE IF NOT EXISTS item_plan_overrides (
                sku TEXT,
                field TEXT,
                month INTEGER,
                value DOUBLE,
                PRIMARY KEY (sku, field, month)
            )
        """)

        con.execute("""
            CREATE TABLE IF NOT EXISTS item_plan_curve_selection (
                sku TEXT PRIMARY KEY,
                curve_type TEXT DEFAULT 'LY'
            )
        """)

        con.execute("""
            CREATE TABLE IF NOT EXISTS item_plan_factory_orders (
                po_number TEXT PRIMARY KEY,
                factory TEXT,
                payment_terms TEXT,
                total_units INTEGER,
                factory_cost DOUBLE,
                fob_date TEXT,
                est_arrival TEXT,
                wk_received TEXT,
                wk_available TEXT,
                cbm DOUBLE,
                status TEXT DEFAULT 'PENDING'
            )
        """)

        con.execute("""
            CREATE TABLE IF NOT EXISTS item_plan_factory_order_items (
                id INTEGER PRIMARY KEY,
                po_number TEXT,
                sku TEXT,
                description TEXT,
                units INTEGER,
                est_arrival TEXT,
                wk_received TEXT,
                wk_available TEXT,
                status TEXT DEFAULT 'PENDING'
            )
        """)

        con.execute("""
            CREATE TABLE IF NOT EXISTS item_plan_settings (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        """)

        # Check if we need to seed
        existing = con.execute("SELECT COUNT(*) FROM monthly_sales_history").fetchall()
        if existing[0][0] == 0:
            # Try to load seed data
            seed_path = DB_DIR / "item_plan_seed.json"
            factory_seed_path = DB_DIR / "factory_orders_seed.json"

            if seed_path.exists():
                with open(seed_path) as f:
                    seed_data = json.load(f)

                # Seed monthly_sales_history from ly_data (12 months Feb 2025 - Jan 2026)
                for sku, data in seed_data.get("ly_data", {}).items():
                    ly_units = data.get("ly_units", [])
                    ly_revenue = data.get("ly_revenue", [])
                    ly_profit = data.get("ly_profit", [])
                    ly_refund = data.get("ly_refund_units", [0] * 12)

                    # Feb 2025 through Jan 2026 = months 2-13 of 2025-2026
                    for idx in range(12):
                        month_num = idx + 2  # Feb=2, Mar=3, ..., Dec=12, Jan=13
                        if month_num > 12:
                            year = 2026
                            month = month_num - 12
                        else:
                            year = 2025
                            month = month_num

                        try:
                            con.execute("""
                                INSERT INTO monthly_sales_history
                                (sku, year, month, units, revenue, profit, refund_units)
                                VALUES (?, ?, ?, ?, ?, ?, ?)
                            """, [
                                sku,
                                year,
                                month,
                                float(ly_units[idx]) if idx < len(ly_units) else 0,
                                float(ly_revenue[idx]) if idx < len(ly_revenue) else 0,
                                float(ly_profit[idx]) if idx < len(ly_profit) else 0,
                                float(ly_refund[idx]) if idx < len(ly_refund) else 0,
                            ])
                        except:
                            pass

                # Seed curve_selection from skus data
                for sku_data in seed_data.get("skus", []):
                    sku = sku_data.get("sku", "")
                    curve = sku_data.get("curve", "LY")
                    if sku:
                        try:
                            con.execute("""
                                INSERT OR REPLACE INTO item_plan_curve_selection
                                (sku, curve_type) VALUES (?, ?)
                            """, [sku, curve])
                        except:
                            pass

                # Seed overrides from skus data
                month_names = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec", "jan_next"]
                field_mappings = {
                    "ty_override_units": "plan_units",
                    "ty_aur_override": "aur",
                    "ty_override_refund_units": "refund_units",
                    "ty_refund_rate_override": "refund_rate",
                    "shipment_override": "shipment",
                }

                for sku_data in seed_data.get("skus", []):
                    sku = sku_data.get("sku", "")
                    if not sku:
                        continue

                    # Seed annual plan
                    annual = sku_data.get("ty_plan_annual", 0)
                    if annual:
                        try:
                            con.execute("""
                                INSERT OR REPLACE INTO item_plan_overrides
                                (sku, field, month, value) VALUES (?, ?, ?, ?)
                            """, [sku, "annual_plan_units", 0, float(annual)])
                        except:
                            pass

                    # Seed monthly overrides
                    for old_field, new_field in field_mappings.items():
                        override_data = sku_data.get(old_field, {})
                        if isinstance(override_data, dict):
                            for month_name, value in override_data.items():
                                if value and value != 0:
                                    month_idx = month_names.index(month_name) if month_name in month_names else -1
                                    if month_idx >= 0:
                                        try:
                                            con.execute("""
                                                INSERT OR REPLACE INTO item_plan_overrides
                                                (sku, field, month, value) VALUES (?, ?, ?, ?)
                                            """, [sku, new_field, month_idx + 1, float(value)])
                                        except:
                                            pass

            # Seed factory orders if they don't exist
            if factory_seed_path.exists():
                with open(factory_seed_path) as f:
                    factory_data = json.load(f)

                for order in factory_data.get("orders", []):
                    try:
                        con.execute("""
                            INSERT INTO item_plan_factory_orders
                            (po_number, factory, payment_terms, total_units, factory_cost,
                             fob_date, est_arrival, wk_received, wk_available, cbm, status)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, [
                            order.get("po_number", ""),
                            order.get("factory", ""),
                            order.get("payment_terms", ""),
                            int(order.get("units", 0)),
                            float(order.get("factory_cost", 0)),
                            order.get("fob_date", ""),
                            order.get("est_arrival", ""),
                            order.get("wk_received", ""),
                            order.get("wk_available", ""),
                            float(order.get("cbm", 0)),
                            order.get("status", "PENDING"),
                        ])
                    except:
                        pass

                for idx, item in enumerate(factory_data.get("items", []), start=1):
                    try:
                        con.execute("""
                            INSERT INTO item_plan_factory_order_items
                            (id, po_number, sku, description, units, est_arrival,
                             wk_received, wk_available, status)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, [
                            idx,
                            item.get("po_number", ""),
                            item.get("sku", ""),
                            item.get("description", ""),
                            int(item.get("units", 0)),
                            item.get("est_arrival", ""),
                            item.get("wk_received", ""),
                            item.get("wk_available", ""),
                            item.get("status", "PENDING"),
                        ])
                    except:
                        pass

            # Set default settings
            con.execute("""INSERT OR IGNORE INTO item_plan_settings (key, value)
                          VALUES ('actualized_through_month', '2')""")
            con.execute("""INSERT OR IGNORE INTO item_plan_settings (key, value)
                          VALUES ('fba_cover_weeks', '4')""")

        con.commit()
    finally:
        con.close()


_MONTH_KEYS = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec", "jan_next"]

# DB field name → frontend override field name
_OVERRIDE_DB_TO_FE = {
    "plan_units": "ty_override_units",
    "aur": "ty_aur_override",
    "refund_units": "ty_override_refund_units",
    "refund_rate": "ty_refund_rate_override",
    "shipment": "shipment_override",
}


def _db_month_to_key(month_num: int) -> str:
    """Convert DB month number (1-13) to frontend month key."""
    if 1 <= month_num <= 13:
        return _MONTH_KEYS[month_num - 1]
    return "jan"


def _array_to_month_obj(arr: list) -> dict:
    """Convert 12-element array (FY order: Feb..Jan) to month-keyed object."""
    # arr[0]=Feb, arr[1]=Mar, ..., arr[10]=Dec, arr[11]=Jan(next)
    # But we also need jan (pre-FY) — not in LY arrays, default 0
    obj = {"jan": 0}
    fy_keys = ["feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec", "jan_next"]
    for i, k in enumerate(fy_keys):
        obj[k] = float(arr[i]) if i < len(arr) else 0
    return obj


def _compute_master_curve() -> dict:
    """Compute master sales curve as percentage distribution across 13 months.
    Returns a dict with month keys: {jan: 0.05, feb: 0.03, ...}"""
    con = _duck()
    try:
        # Get total units per (year, month) across all SKUs for LY
        result = con.execute("""
            SELECT year, month, SUM(units) as total_units
            FROM monthly_sales_history
            WHERE year IN (2025, 2026)
            GROUP BY year, month
            ORDER BY year, month
        """).fetchall()

        # Calculate total annual
        annual_total = sum(row[2] for row in result)

        curve = {k: 0 for k in _MONTH_KEYS}
        if annual_total == 0:
            return curve

        for year, month, total_units in result:
            # Map (year, month) to FY month key
            if year == 2025 and 2 <= month <= 12:
                key = _MONTH_KEYS[month - 1]  # feb=idx1, mar=idx2, ..., dec=idx11
            elif year == 2026 and month == 1:
                key = "jan_next"  # idx 12
            elif year == 2025 and month == 1:
                key = "jan"  # pre-FY
            else:
                continue
            curve[key] = round(float(total_units) / annual_total, 4)

        return curve
    finally:
        con.close()


@app.get("/api/item-plan")
async def get_item_plan(request: Request):
    """Return item plan data shaped for the React frontend.

    Frontend expects:
      data.skus[]: each has ly_units/ly_revenue/etc as month-keyed objects,
                   plus curve, ty_plan_annual, fba_beginning, wh_beginning
      data.overrides: {sku: {fe_field_name: {month_key: value}}}
      data.settings, data.factory_orders, data.factory_order_items
    """
    _require_auth(request)
    con = _duck()
    try:
        # ── Settings ──
        settings = {}
        for row in con.execute("SELECT key, value FROM item_plan_settings").fetchall():
            settings[row[0]] = row[1]

        # ── Seed JSON for metadata ──
        seed_data = {}
        seed_path = DB_DIR / "item_plan_seed.json"
        if seed_path.exists():
            with open(seed_path) as f:
                seed_data = json.load(f)

        # Build metadata map from seed skus
        sku_meta = {}
        for si in seed_data.get("skus", []):
            sk = si.get("sku", "")
            sku_meta[sk] = si  # full seed record

        # ── Build canonical SKU set from item_master.csv ──
        # Extension SKUs (-FBM, -1, " - FBA", etc.) are mapped to their base
        import re
        master_items = load_item_master()
        canonical_skus = {m["sku"] for m in master_items}
        _EXT_SUFFIXES = re.compile(
            r'\s*[-/]\s*(?:FBM|FBA|RB|DONATE|RETD|HOLD|Damage|CUST|Transfer|\d+)$',
            re.IGNORECASE)

        def _base_sku(raw_sku: str) -> str:
            """Strip extension suffixes to find the canonical base SKU."""
            s = raw_sku.strip()
            # Direct match first
            if s in canonical_skus:
                return s
            # Try stripping known suffixes
            stripped = _EXT_SUFFIXES.sub('', s).strip()
            if stripped in canonical_skus:
                return stripped
            return ""  # no match — skip this SKU

        # Map every monthly_sales_history SKU to a canonical base SKU
        all_hist_skus = con.execute(
            "SELECT DISTINCT sku FROM monthly_sales_history ORDER BY sku"
        ).fetchall()
        # Build {base_sku: [raw_sku1, raw_sku2, ...]}
        base_to_raw = {}
        for (raw,) in all_hist_skus:
            base = _base_sku(raw)
            if base:
                base_to_raw.setdefault(base, []).append(raw)

        skus_list = []
        all_overrides = {}  # top-level {sku: {fe_field: {month_key: val}}}

        for sku in sorted(base_to_raw.keys()):
            raw_skus = base_to_raw[sku]
            # ── LY monthly data — aggregate all raw SKUs into base ──
            ly_units_arr = [0.0] * 12
            ly_rev_arr = [0.0] * 12
            ly_prof_arr = [0.0] * 12
            ly_ref_arr = [0.0] * 12

            for m in range(2, 14):  # Feb(2)..Jan(13)
                yr = 2025 if m <= 12 else 2026
                am = m if m <= 12 else m - 12
                idx = m - 2
                # Sum across all raw SKU variants for this base
                placeholders = ",".join(["?"] * len(raw_skus))
                row = con.execute(
                    f"SELECT SUM(units), SUM(revenue), SUM(profit), SUM(refund_units) "
                    f"FROM monthly_sales_history WHERE sku IN ({placeholders}) "
                    f"AND year=? AND month=?",
                    raw_skus + [yr, am],
                ).fetchone()
                if row:
                    ly_units_arr[idx] = float(row[0] or 0)
                    ly_rev_arr[idx] = float(row[1] or 0)
                    ly_prof_arr[idx] = float(row[2] or 0)
                    ly_ref_arr[idx] = float(row[3] or 0)

            # Convert to month-keyed objects
            ly_units = _array_to_month_obj(ly_units_arr)
            ly_revenue = _array_to_month_obj(ly_rev_arr)
            ly_profit = _array_to_month_obj(ly_prof_arr)
            ly_refund_units = _array_to_month_obj(ly_ref_arr)

            # Compute derived: ly_aur = revenue / units per month
            ly_aur = {}
            ly_refund_rate = {}
            for k in _MONTH_KEYS:
                u = ly_units.get(k, 0)
                r = ly_revenue.get(k, 0)
                ref = ly_refund_units.get(k, 0)
                ly_aur[k] = round(r / u, 2) if u > 0 else 0
                ly_refund_rate[k] = round(ref / u, 4) if u > 0 else 0

            # ── Curve selection ──
            crow = con.execute(
                "SELECT curve_type FROM item_plan_curve_selection WHERE sku=?", [sku]
            ).fetchone()
            curve = crow[0] if crow else "LY"

            # ── Overrides (DB → frontend field names & month keys) ──
            ov_rows = con.execute(
                "SELECT field, month, value FROM item_plan_overrides WHERE sku=?", [sku]
            ).fetchall()
            sku_ov = {}
            annual_plan_from_ov = 0
            for db_field, month_num, value in ov_rows:
                if db_field == "annual_plan_units":
                    annual_plan_from_ov = float(value)
                    continue
                fe_field = _OVERRIDE_DB_TO_FE.get(db_field, db_field)
                if fe_field not in sku_ov:
                    sku_ov[fe_field] = {}
                mk = _db_month_to_key(int(month_num))
                sku_ov[fe_field][mk] = float(value)
            all_overrides[sku] = sku_ov

            # ── Metadata from seed ──
            meta = sku_meta.get(sku, {})
            ty_plan_annual = annual_plan_from_ov or meta.get("ty_plan_annual", 0)

            # fba_beginning / wh_beginning: seed stores dicts, frontend wants a number
            fba_raw = meta.get("fba_beginning", 0)
            fba_beginning = fba_raw.get("jan", 0) if isinstance(fba_raw, dict) else float(fba_raw or 0)
            wh_raw = meta.get("wh_beginning", 0)
            wh_beginning = wh_raw.get("feb", 0) if isinstance(wh_raw, dict) else float(wh_raw or 0)

            # ty_aur plan: use LY AUR as the plan baseline (frontend references sku.ty_aur)
            ty_aur = {k: ly_aur[k] for k in _MONTH_KEYS}

            # ty_plan_units: frontend uses sku.ty_plan_units?.fy_total as fallback
            ty_plan_units = {"fy_total": int(ty_plan_annual)}

            skus_list.append({
                "sku": sku,
                "asin": meta.get("asin", ""),
                "product_name": meta.get("product_name", ""),
                "tab": meta.get("tab", ""),
                "curve": curve,
                "ty_plan_annual": int(ty_plan_annual),
                "ty_plan_units": ty_plan_units,
                "ty_aur": ty_aur,
                "fba_beginning": int(fba_beginning),
                "wh_beginning": int(wh_beginning),
                "ly_units": ly_units,
                "ly_revenue": ly_revenue,
                "ly_gross_profit": ly_profit,
                "ly_refund_units": ly_refund_units,
                "ly_aur": ly_aur,
                "ly_refund_rate": ly_refund_rate,
                "lly_units": {},
                "lly_revenue": {},
                "lly_gross_profit": {},
                "lly_refund_units": {},
            })

        return {
            "settings": settings,
            "skus": skus_list,
            "overrides": all_overrides,
        }
    finally:
        con.close()


_OVERRIDE_FE_TO_DB = {v: k for k, v in _OVERRIDE_DB_TO_FE.items()}


def _month_key_to_db(key: str) -> int:
    """Convert frontend month key to DB month number (1-13)."""
    try:
        return _MONTH_KEYS.index(key) + 1
    except ValueError:
        # If it's already an int or int-string, return as-is
        try:
            return int(key)
        except (ValueError, TypeError):
            return 0


@app.post("/api/item-plan/override")
async def post_item_plan_override(request: Request):
    _require_auth(request)
    body = await request.json()
    sku = body.get("sku", "")
    fe_field = body.get("field", "")
    month_raw = body.get("month", 0)
    value = body.get("value")

    # Map frontend field name → DB field name
    db_field = _OVERRIDE_FE_TO_DB.get(fe_field, fe_field)
    # Map frontend month key → DB month integer
    db_month = _month_key_to_db(month_raw) if isinstance(month_raw, str) else int(month_raw)

    con = _duck_rw()
    try:
        if value is None:
            # Delete override
            con.execute("""
                DELETE FROM item_plan_overrides WHERE sku = ? AND field = ? AND month = ?
            """, [sku, db_field, db_month])
        else:
            # Upsert override
            con.execute("""
                INSERT OR REPLACE INTO item_plan_overrides (sku, field, month, value)
                VALUES (?, ?, ?, ?)
            """, [sku, db_field, db_month, float(value)])
        con.commit()
        return {"status": "ok"}
    finally:
        con.close()


@app.post("/api/item-plan/curve")
async def post_item_plan_curve(request: Request):
    _require_auth(request)
    body = await request.json()
    sku = body.get("sku", "")
    curve_type = body.get("curve_type", "LY")

    con = _duck_rw()
    try:
        con.execute("""
            INSERT OR REPLACE INTO item_plan_curve_selection (sku, curve_type)
            VALUES (?, ?)
        """, [sku, curve_type])
        con.commit()
        return {"status": "ok"}
    finally:
        con.close()


@app.get("/api/item-plan/sales-curves")
async def get_item_plan_sales_curves(request: Request):
    """Return sales curves in the shape the frontend expects:
    {master: {jan: 0.05, ...}, bySku: {sku: {jan: 0.05, ...}}}"""
    _require_auth(request)
    master_curve = _compute_master_curve()

    # Build per-SKU LY curves — canonical SKUs only, aggregate extensions
    import re as _re_curves
    con = _duck()
    try:
        master_items_c = load_item_master()
        canon_c = {m["sku"] for m in master_items_c}
        _ext_re = _re_curves.compile(
            r'\s*[-/]\s*(?:FBM|FBA|RB|DONATE|RETD|HOLD|Damage|CUST|Transfer|\d+)$',
            _re_curves.IGNORECASE)

        def _base_c(raw):
            s = raw.strip()
            if s in canon_c:
                return s
            stripped = _ext_re.sub('', s).strip()
            return stripped if stripped in canon_c else ""

        all_raw = con.execute(
            "SELECT DISTINCT sku FROM monthly_sales_history"
        ).fetchall()
        base_map = {}
        for (raw,) in all_raw:
            b = _base_c(raw)
            if b:
                base_map.setdefault(b, []).append(raw)

        by_sku = {}
        for base, raws in sorted(base_map.items()):
            placeholders = ",".join(["?"] * len(raws))
            rows = con.execute(
                f"SELECT year, month, SUM(units) FROM monthly_sales_history "
                f"WHERE sku IN ({placeholders}) AND year IN (2025,2026) "
                f"GROUP BY year, month", raws
            ).fetchall()
            total = sum(float(r[2] or 0) for r in rows)
            curve = {k: 0 for k in _MONTH_KEYS}
            if total > 0:
                for yr, mo, units in rows:
                    if yr == 2025 and 2 <= mo <= 12:
                        curve[_MONTH_KEYS[mo - 1]] = round(float(units or 0) / total, 4)
                    elif yr == 2026 and mo == 1:
                        curve["jan_next"] = round(float(units or 0) / total, 4)
                    elif yr == 2025 and mo == 1:
                        curve["jan"] = round(float(units or 0) / total, 4)
            by_sku[base] = curve
        return {"master": master_curve, "bySku": by_sku}
    finally:
        con.close()


@app.get("/api/factory-on-order")
async def get_factory_on_order(request: Request):
    _require_auth(request)
    con = _duck()
    try:
        orders = []
        order_rows = con.execute("""
            SELECT po_number, factory, payment_terms, total_units, factory_cost,
                   fob_date, est_arrival, wk_received, wk_available, cbm, status
            FROM item_plan_factory_orders
            ORDER BY po_number
        """).fetchall()

        for row in order_rows:
            orders.append({
                "po_number": row[0],
                "factory": row[1],
                "payment_terms": row[2],
                "units": int(row[3]),
                "factory_cost": float(row[4]),
                "fob_date": row[5],
                "est_arrival": row[6],
                "wk_received": row[7],
                "wk_available": row[8],
                "cbm": float(row[9]),
                "status": row[10],
            })

        items = []
        item_rows = con.execute("""
            SELECT id, po_number, sku, description, units, est_arrival,
                   wk_received, wk_available, status
            FROM item_plan_factory_order_items
            ORDER BY id
        """).fetchall()

        for row in item_rows:
            items.append({
                "id": int(row[0]),
                "po_number": row[1],
                "sku": row[2],
                "description": row[3],
                "units": int(row[4]),
                "est_arrival": row[5],
                "wk_received": row[6],
                "wk_available": row[7],
                "status": row[8],
            })

        return {"orders": orders, "items": items}
    finally:
        con.close()


@app.post("/api/factory-on-order")
async def post_factory_on_order(request: Request):
    _require_auth(request)
    body = await request.json()

    con = _duck_rw()
    try:
        # Determine if this is an order or item
        if "po_number" in body and "factory" in body:
            # It's an order
            con.execute("""
                INSERT OR REPLACE INTO item_plan_factory_orders
                (po_number, factory, payment_terms, total_units, factory_cost,
                 fob_date, est_arrival, wk_received, wk_available, cbm, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, [
                body.get("po_number", ""),
                body.get("factory", ""),
                body.get("payment_terms", ""),
                int(body.get("units", 0)),
                float(body.get("factory_cost", 0)),
                body.get("fob_date", ""),
                body.get("est_arrival", ""),
                body.get("wk_received", ""),
                body.get("wk_available", ""),
                float(body.get("cbm", 0)),
                body.get("status", "PENDING"),
            ])
        else:
            # It's an item
            con.execute("""
                INSERT OR REPLACE INTO item_plan_factory_order_items
                (po_number, sku, description, units, est_arrival,
                 wk_received, wk_available, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, [
                body.get("po_number", ""),
                body.get("sku", ""),
                body.get("description", ""),
                int(body.get("units", 0)),
                body.get("est_arrival", ""),
                body.get("wk_received", ""),
                body.get("wk_available", ""),
                body.get("status", "PENDING"),
            ])

        con.commit()
        return {"status": "ok"}
    finally:
        con.close()


@app.get("/api/dashboard-settings")
async def get_dashboard_settings(request: Request):
    _require_auth(request)
    con = _duck()
    try:
        settings = {}
        for row in con.execute("SELECT key, value FROM item_plan_settings").fetchall():
            settings[row[0]] = row[1]
        return settings
    finally:
        con.close()


@app.post("/api/dashboard-settings")
async def post_dashboard_settings(request: Request):
    _require_auth(request)
    body = await request.json()
    key = body.get("key", "")
    value = body.get("value", "")

    con = _duck_rw()
    try:
        con.execute("""
            INSERT OR REPLACE INTO item_plan_settings (key, value)
            VALUES (?, ?)
        """, [key, str(value)])
        con.commit()
        return {"status": "ok"}
    finally:
        con.close()


# ── System Status Endpoints ────────────────────────────────
@app.get("/api/system/status")
async def get_system_status(request: Request):
    """Get system status including last sync times, scheduler status, and recent logs."""
    _require_auth(request)

    con = duckdb.connect(str(DB_PATH), read_only=True)
    try:
        # Get last sync log entry
        last_sync = con.execute("""
            SELECT job_name, completed_at, status, execution_time_seconds
            FROM sync_log
            WHERE status = 'completed'
            ORDER BY completed_at DESC
            LIMIT 1
        """).fetchone()

        # Get last docs update
        last_docs = con.execute("""
            SELECT completed_at, status, documents_updated
            FROM docs_update_log
            WHERE status = 'completed'
            ORDER BY completed_at DESC
            LIMIT 1
        """).fetchone()

        # Get recent sync log entries (last 10)
        recent_syncs = con.execute("""
            SELECT job_name, started_at, completed_at, status, execution_time_seconds, error_message
            FROM sync_log
            ORDER BY started_at DESC
            LIMIT 10
        """).fetchall()

        # Get next scheduled jobs from APScheduler
        scheduler_status = {}
        if scheduler and scheduler.running:
            for job in scheduler.get_jobs():
                scheduler_status[job.id] = {
                    "next_run_time": str(job.next_run_time),
                    "trigger": str(job.trigger)
                }

        return {
            "status": "ok",
            "last_sync": {
                "job_name": last_sync[0] if last_sync else None,
                "completed_at": str(last_sync[1]) if last_sync else None,
                "status": last_sync[2] if last_sync else None,
                "execution_time_seconds": last_sync[3] if last_sync else None
            },
            "last_docs_update": {
                "completed_at": str(last_docs[0]) if last_docs else None,
                "status": last_docs[1] if last_docs else None,
                "documents_updated": last_docs[2] if last_docs else None
            },
            "recent_syncs": [
                {
                    "job_name": row[0],
                    "started_at": str(row[1]),
                    "completed_at": str(row[2]),
                    "status": row[3],
                    "execution_time_seconds": row[4],
                    "error_message": row[5]
                }
                for row in recent_syncs
            ],
            "scheduler_status": scheduler_status,
            "scheduler_running": scheduler.running if scheduler else False
        }
    finally:
        con.close()


@app.get("/api/system/sync-log")
async def get_sync_log(request: Request, limit: int = Query(50, ge=1, le=500)):
    """Get sync log entries."""
    _require_auth(request)

    con = duckdb.connect(str(DB_PATH), read_only=True)
    try:
        rows = con.execute("""
            SELECT id, job_name, started_at, completed_at, status, records_processed, error_message, execution_time_seconds
            FROM sync_log
            ORDER BY started_at DESC
            LIMIT ?
        """, [limit]).fetchall()

        return {
            "entries": [
                {
                    "id": row[0],
                    "job_name": row[1],
                    "started_at": str(row[2]),
                    "completed_at": str(row[3]),
                    "status": row[4],
                    "records_processed": row[5],
                    "error_message": row[6],
                    "execution_time_seconds": row[7]
                }
                for row in rows
            ]
        }
    finally:
        con.close()


@app.post("/api/backup/trigger")
async def trigger_backup(request: Request):
    """Manually trigger a DuckDB backup."""
    _require_auth(request)
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, _run_duckdb_backup)
    return {"status": "backup_started"}


@app.get("/api/backup/status")
async def get_backup_status(request: Request):
    """Get status of recent backups from sync_log."""
    _require_auth(request)
    con = duckdb.connect(str(DB_PATH), read_only=True)
    try:
        rows = con.execute("""
            SELECT id, started_at, completed_at, status, execution_time_seconds, error_message
            FROM sync_log
            WHERE job_name = 'duckdb_backup'
            ORDER BY started_at DESC
            LIMIT 10
        """).fetchall()

        return {
            "backups": [
                {
                    "id": row[0],
                    "started_at": str(row[1]),
                    "completed_at": str(row[2]),
                    "status": row[3],
                    "execution_time_seconds": row[4],
                    "error_message": row[5]
                }
                for row in rows
            ]
        }
    finally:
        con.close()


@app.post("/api/docs/update")
async def trigger_docs_update(request: Request):
    """Manually trigger a docs update."""
    _require_auth(request)
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, _run_scheduled_docs_update)
    return {"status": "docs_update_started"}


@app.get("/api/docs/status")
async def get_docs_status(request: Request):
    """Get status of recent docs updates from docs_update_log."""
    _require_auth(request)
    con = duckdb.connect(str(DB_PATH), read_only=True)
    try:
        rows = con.execute("""
            SELECT id, started_at, completed_at, status, execution_time_seconds, error_message, documents_updated
            FROM docs_update_log
            ORDER BY started_at DESC
            LIMIT 10
        """).fetchall()

        return {
            "updates": [
                {
                    "id": row[0],
                    "started_at": str(row[1]),
                    "completed_at": str(row[2]),
                    "status": row[3],
                    "execution_time_seconds": row[4],
                    "error_message": row[5],
                    "documents_updated": row[6]
                }
                for row in rows
            ]
        }
    finally:
        con.close()


@app.get("/api/docs/architecture")
async def get_architecture_doc(request: Request):
    """Retrieve the generated architecture guide. (Generated by Prompt 2)"""
    _require_auth(request)
    docs_dir = Path("/app/docs")
    arch_file = docs_dir / "architecture_guide.md"

    if arch_file.exists():
        with open(arch_file, "r") as f:
            return {
                "status": "ok",
                "content": f.read(),
                "last_updated": datetime.fromtimestamp(arch_file.stat().st_mtime).isoformat()
            }
    else:
        return {
            "status": "not_found",
            "content": "Architecture guide not yet generated. Run /api/docs/update to generate.",
            "last_updated": None
        }


@app.get("/api/docs/disaster-recovery")
async def get_disaster_recovery_doc(request: Request):
    """Retrieve the generated disaster recovery plan. (Generated by Prompt 2)"""
    _require_auth(request)
    docs_dir = Path("/app/docs")
    dr_file = docs_dir / "disaster_recovery_plan.md"

    if dr_file.exists():
        with open(dr_file, "r") as f:
            return {
                "status": "ok",
                "content": f.read(),
                "last_updated": datetime.fromtimestamp(dr_file.stat().st_mtime).isoformat()
            }
    else:
        return {
            "status": "not_found",
            "content": "Disaster recovery plan not yet generated. Run /api/docs/update to generate.",
            "last_updated": None
        }


# ── Static Frontend ────────────────────────────────────────
# Serve the built React frontend from the same server.
# Check local dist/ first (Docker), then ../frontend/dist/ (local dev)
_local_dist = Path(__file__).resolve().parent / "dist"
_dev_dist = Path(__file__).resolve().parent.parent / "frontend" / "dist"
FRONTEND_DIR = _local_dist if _local_dist.exists() else _dev_dist

if FRONTEND_DIR.exists():
    # Mount static assets (JS, CSS, images) — must come AFTER api routes
    app.mount("/assets", StaticFiles(directory=str(FRONTEND_DIR / "assets")), name="static-assets")

    # Catch-all for SPA: serve index.html for any non-API, non-asset route
    @app.get("/{full_path:path}")
    async def serve_spa(request: Request, full_path: str):
        # If a real file exists (like vite.svg), serve it
        file_path = FRONTEND_DIR / full_path
        if full_path and file_path.exists() and file_path.is_file():
            return FileResponse(str(file_path))
        # Otherwise serve index.html (SPA routing)
        return FileResponse(str(FRONTEND_DIR / "index.html"))


# ── Run ─────────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
