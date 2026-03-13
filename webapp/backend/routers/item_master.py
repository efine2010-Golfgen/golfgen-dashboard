"""Item Master and pricing routes."""
import csv
import json
import logging
import threading
from datetime import datetime
from pathlib import Path
from typing import Optional
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Query, Body, Request, HTTPException, UploadFile, File
import duckdb

from core.config import DB_PATH, DB_DIR, TIMEZONE

logger = logging.getLogger("golfgen")
router = APIRouter()

# ── Configuration ──────────────────────────────────────────────
ITEM_MASTER_PATH = DB_DIR / "item_master.csv"
PRICING_CACHE_PATH = DB_DIR / "pricing_sync.json"
COGS_PATH = DB_DIR / "cogs.csv"
UPLOAD_META_PATH = DB_DIR / "upload_metadata.json"


# ── Helper Functions ───────────────────────────────────────────

def load_item_master() -> list:
    """Load item master from CSV. Returns list of dicts."""
    items = []
    if not ITEM_MASTER_PATH.exists():
        return items
    with open(ITEM_MASTER_PATH, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            asin = (row.get("asin") or "").strip()
            if not asin:
                continue
            items.append({
                "asin": asin,
                "sku": (row.get("sku") or "").strip(),
                "productName": (row.get("product_name") or "").strip(),
                "color": (row.get("color") or "").strip(),
                "brand": (row.get("brand") or "").strip(),
                "series": (row.get("series") or "").strip(),
                "productType": (row.get("product_type") or "").strip(),
                "pieceCount": int(float(row.get("piece_count") or 0)),
                "orientation": (row.get("orientation") or "").strip(),
                "category": (row.get("category") or "").strip(),
                "unitCost": float(row.get("unit_cost") or 0),
                "fbsStock": int(float(row.get("fba_stock") or 0)),
                "lyAur": round(float(row.get("ly_aur") or 0), 2),
                "lyRevenue": round(float(row.get("ly_revenue") or 0), 2),
                "lyUnits": int(float(row.get("ly_units") or 0)),
                "lyProfit": round(float(row.get("ly_profit") or 0), 2),
                "plannedAnnualUnits": int(float(row.get("planned_annual_units") or 0)),
                "listPrice": float(row.get("list_price") or 0),
                "salePrice": float(row.get("sale_price") or 0),
                "salePriceStartDate": (row.get("sale_price_start_date") or "").strip(),
                "salePriceEndDate": (row.get("sale_price_end_date") or "").strip(),
                "referralPct": float(row.get("referral_pct") or 15),
                "couponType": (row.get("coupon_type") or "").strip(),
                "couponValue": float(row.get("coupon_value") or 0),
                "cartonPack": int(float(row.get("carton_pack") or 0)),
                "cartonLength": float(row.get("carton_length") or 0),
                "cartonWidth": float(row.get("carton_width") or 0),
                "cartonHeight": float(row.get("carton_height") or 0),
                "cartonWeight": float(row.get("carton_weight") or 0),
            })
    return items


def save_item_master(items: list):
    """Save item master list back to CSV."""
    fields = [
        "asin", "sku", "product_name", "color", "brand", "series",
        "product_type", "piece_count", "orientation", "category", "unit_cost",
        "fba_stock", "ly_aur", "ly_revenue", "ly_units", "ly_profit",
        "planned_annual_units", "list_price", "sale_price",
        "sale_price_start_date", "sale_price_end_date", "referral_pct",
        "coupon_type", "coupon_value", "carton_pack", "carton_length",
        "carton_width", "carton_height", "carton_weight",
    ]
    with open(ITEM_MASTER_PATH, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for item in items:
            w.writerow({
                "asin": item.get("asin", ""),
                "sku": item.get("sku", ""),
                "product_name": item.get("productName", ""),
                "color": item.get("color", ""),
                "brand": item.get("brand", ""),
                "series": item.get("series", ""),
                "product_type": item.get("productType", ""),
                "piece_count": item.get("pieceCount", 0),
                "orientation": item.get("orientation", ""),
                "category": item.get("category", ""),
                "unit_cost": item.get("unitCost", 0),
                "fba_stock": item.get("fbsStock", 0),
                "ly_aur": item.get("lyAur", 0),
                "ly_revenue": item.get("lyRevenue", 0),
                "ly_units": item.get("lyUnits", 0),
                "ly_profit": item.get("lyProfit", 0),
                "planned_annual_units": item.get("plannedAnnualUnits", 0),
                "list_price": item.get("listPrice", 0),
                "sale_price": item.get("salePrice", 0),
                "sale_price_start_date": item.get("salePriceStartDate", ""),
                "sale_price_end_date": item.get("salePriceEndDate", ""),
                "referral_pct": item.get("referralPct", 15),
                "coupon_type": item.get("couponType", ""),
                "coupon_value": item.get("couponValue", 0),
                "carton_pack": item.get("cartonPack", 0),
                "carton_length": item.get("cartonLength", 0),
                "carton_width": item.get("cartonWidth", 0),
                "carton_height": item.get("cartonHeight", 0),
                "carton_weight": item.get("cartonWeight", 0),
            })


def _load_pricing_cache() -> dict:
    """Load cached pricing/coupon data."""
    if PRICING_CACHE_PATH.exists():
        with open(PRICING_CACHE_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {"prices": {}, "coupons": {}, "lastSync": None}


def load_json(filename: str) -> list:
    """Load a JSON data file from the data directory."""
    path = DB_DIR / filename
    if not path.exists():
        return []
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def save_json(filename: str, data):
    """Save data to a JSON file in the data directory."""
    path = DB_DIR / filename
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def load_cogs() -> dict:
    """Load COGS data from CSV. Returns {asin: {cogs, product_name, sku}}."""
    cogs = {}
    if not COGS_PATH.exists():
        return cogs
    with open(COGS_PATH, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            asin = (row.get("asin") or "").strip()
            if not asin:
                continue
            try:
                cogs_val = float(row.get("cogs") or 0)
            except ValueError:
                cogs_val = 0
            cogs[asin] = {
                "cogs": cogs_val,
                "product_name": (row.get("product_name") or "").strip(),
                "sku": (row.get("sku") or "").strip(),
            }
    return cogs


def _load_upload_meta() -> dict:
    if UPLOAD_META_PATH.exists():
        with open(UPLOAD_META_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {}


def _save_upload_meta(meta: dict):
    with open(UPLOAD_META_PATH, "w", encoding="utf-8") as f:
        json.dump(meta, f, indent=2)


def _classify_golf_channel(item_number: str, walmart_skus: set, amazon_skus: set) -> str:
    """Classify a golf SKU into Amazon, Walmart, Walmart & Amazon, or Other."""
    import re as _re
    s = item_number.strip().upper()
    if s.startswith("T-"):
        s = s[2:]
    for suf in ["FBM", "/RB", "/RETD", "/DONATE", "/DMGD", "/CUST", "/HOLD", "/1", "/2"]:
        if s.endswith(suf):
            s = s[:-len(suf)]
    s = _re.sub(r"\s*/DAM.*$", "", s)
    s = s.strip()
    in_walmart = s in walmart_skus
    in_amazon = s in amazon_skus
    if in_walmart and in_amazon:
        return "Walmart & Amazon"
    if in_walmart:
        return "Walmart"
    if in_amazon:
        return "Amazon"
    return "Other"


# ── Routes ─────────────────────────────────────────────────────

@router.get("/api/item-master")
def get_item_master():
    """Return all items from the item master CSV, enriched with live COGS and pricing/coupon data."""
    items = load_item_master()
    cogs_data = load_cogs()

    # Load live pricing & coupon cache (from SP-API / Ads API syncs)
    pricing_cache = _load_pricing_cache()
    live_prices = pricing_cache.get("prices", {})
    live_coupons = pricing_cache.get("coupons", {})
    pricing_last_sync = pricing_cache.get("lastSync")

    for item in items:
        asin = item["asin"]
        cogs_info = cogs_data.get(asin, {})
        if cogs_info.get("cogs", 0) > 0 and item["unitCost"] == 0:
            item["unitCost"] = cogs_info["cogs"]

        # Merge live pricing data if available
        live = live_prices.get(asin)
        if live:
            item["liveListingPrice"] = live.get("listingPrice")
            item["liveBuyBoxPrice"] = live.get("buyBoxPrice")
            item["liveLandedPrice"] = live.get("landedPrice")
            item["priceFetchedAt"] = live.get("fetchedAt")
        else:
            item["liveListingPrice"] = None
            item["liveBuyBoxPrice"] = None
            item["liveLandedPrice"] = None
            item["priceFetchedAt"] = None

        # Merge live coupon data if available
        coupon = live_coupons.get(asin)
        if coupon:
            item["liveCouponState"] = coupon.get("state")
            item["liveCouponType"] = coupon.get("discountType")
            item["liveCouponValue"] = coupon.get("discountValue")
            item["liveCouponId"] = coupon.get("couponId")
            item["couponBudget"] = coupon.get("budgetAmount")
            item["couponBudgetUsed"] = coupon.get("budgetUsed")
            item["couponRedemptions"] = coupon.get("redemptions")
            item["couponStartDate"] = coupon.get("startDate")
            item["couponEndDate"] = coupon.get("endDate")
        else:
            item["liveCouponState"] = None
            item["liveCouponType"] = None
            item["liveCouponValue"] = None
            item["liveCouponId"] = None
            item["couponBudget"] = None
            item["couponBudgetUsed"] = None
            item["couponRedemptions"] = None
            item["couponStartDate"] = None
            item["couponEndDate"] = None

        # Compute net price after coupon
        base = item["salePrice"] if item["salePrice"] > 0 else item["listPrice"]
        if item["couponType"] == "$" and item["couponValue"] > 0:
            item["netPrice"] = round(base - item["couponValue"], 2)
        elif item["couponType"] == "%" and item["couponValue"] > 0:
            item["netPrice"] = round(base * (1 - item["couponValue"] / 100), 2)
        else:
            item["netPrice"] = round(base, 2)
        # Referral fee estimate
        item["referralFee"] = round(item["netPrice"] * item["referralPct"] / 100, 2)
    return {"items": items, "count": len(items), "pricingLastSync": pricing_last_sync}


@router.get("/api/pricing/status")
def get_pricing_status():
    """Return pricing & coupon sync status and summary."""
    cache = _load_pricing_cache()
    prices = cache.get("prices", {})
    coupons = cache.get("coupons", {})
    return {
        "lastSync": cache.get("lastSync"),
        "pricedAsins": len(prices),
        "couponAsins": len(coupons),
        "activeCoupons": sum(1 for c in coupons.values() if c.get("state") == "ENABLED"),
        "scheduledCoupons": sum(1 for c in coupons.values() if c.get("state") == "SCHEDULED"),
    }


@router.post("/api/pricing/sync")
def trigger_pricing_sync():
    """Manually trigger a pricing & coupon data sync."""
    from services.ads_api import _sync_pricing_and_coupons
    threading.Thread(target=_sync_pricing_and_coupons, daemon=True).start()
    return {"status": "started", "message": "Pricing & coupon sync started in background"}


@router.put("/api/item-master/{asin}")
def update_item_master(asin: str, body: dict = Body(...)):
    """Update a single item in the item master by ASIN."""
    items = load_item_master()
    found = False
    for item in items:
        if item["asin"] == asin:
            # Only update fields that are provided
            updatable = [
                "listPrice", "salePrice", "salePriceStartDate", "salePriceEndDate",
                "referralPct", "couponType",
                "couponValue", "cartonPack", "cartonLength", "cartonWidth",
                "cartonHeight", "cartonWeight", "unitCost", "plannedAnnualUnits",
                "color", "series", "productType", "pieceCount", "orientation",
                "category",
            ]
            for key in updatable:
                if key in body:
                    item[key] = body[key]
            found = True
            break
    if not found:
        raise HTTPException(status_code=404, detail=f"ASIN {asin} not found")
    save_item_master(items)
    return {"status": "ok", "asin": asin}


@router.post("/api/item-master/bulk-update")
def bulk_update_item_master(body: dict = Body(...)):
    """Bulk update items. Expects {items: [{asin, ...fields}, ...]}."""
    updates = body.get("items", [])
    if not updates:
        return {"status": "ok", "updated": 0}
    items = load_item_master()
    item_map = {i["asin"]: i for i in items}
    updated = 0
    for upd in updates:
        asin = upd.get("asin")
        if asin and asin in item_map:
            item = item_map[asin]
            for key, val in upd.items():
                if key != "asin" and key in item:
                    item[key] = val
            updated += 1
    save_item_master(items)
    return {"status": "ok", "updated": updated}


@router.get("/api/upload/metadata")
def upload_metadata():
    """Return last-uploaded timestamps for golf and housewares inventory files."""
    return _load_upload_meta()


@router.post("/api/upload/inventory-excel")
async def upload_inventory_excel(
    file: UploadFile = File(...),
    division: str = Query("golf", description="golf or housewares"),
):
    """Upload an Excel file to refresh golf or housewares inventory JSON.

    Parses the first sheet with warehouse-like columns
    (Item Number, Description, Pack, On-Hand, Damage, QC Hold, Pcs On-Hand, Pcs Allocated, Pcs Available).
    For golf, also classifies each item by channel (Amazon/Walmart/Other).
    Saves to golf_inventory.json or housewares_inventory.json.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    import tempfile
    import os
    tmp_path = None
    try:
        contents = await file.read()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(contents)
            tmp_path = tmp.name

        wb = openpyxl.load_workbook(tmp_path, data_only=True)

        # Find the data sheet — first sheet with "item" in header row, or just the first sheet
        ws = None
        for sheet_name in wb.sheetnames:
            candidate = wb[sheet_name]
            first_row = [str(c or "").strip().lower() for c in next(candidate.iter_rows(min_row=1, max_row=1, values_only=True), [])]
            if any("item" in h for h in first_row):
                ws = candidate
                break
        if ws is None:
            ws = wb[wb.sheetnames[0]]

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise HTTPException(status_code=400, detail="Excel file has no data rows")

        headers = [str(h or "").strip() for h in rows[0]]
        header_lower = [h.lower().replace(" ", "_").replace("-", "_") for h in headers]

        # Map columns
        def find_col(*patterns):
            for i, h in enumerate(header_lower):
                for p in patterns:
                    if p in h:
                        return i
            return None

        item_col = find_col("item_number", "item_no", "item_#", "item")
        desc_col = find_col("description", "desc")
        pack_col = find_col("pack")
        oh_col = find_col("on_hand", "on hand", "oh")
        ns_col = find_col("non_standard", "nonstandard", "non_std")
        dmg_col = find_col("damage", "damaged", "dam")
        std_col = find_col("standard")
        qc_col = find_col("qc_hold", "qc hold", "qchold")
        poh_col = find_col("pcs_on_hand", "pcs on hand")
        palloc_col = find_col("pcs_allocated", "pcs allocated", "pcs_alloc")
        pavail_col = find_col("pcs_available", "pcs available", "pcs_avail")
        ref_col = find_col("item_ref", "item ref", "itemref")

        if item_col is None:
            raise HTTPException(status_code=400, detail=f"Could not find 'Item Number' column. Headers: {headers}")

        def safe_int(val):
            try:
                return int(float(val or 0))
            except (ValueError, TypeError):
                return 0

        def safe_str(val):
            return str(val or "").strip()

        # Build channel lookup for golf
        walmart_skus_set = set()
        amazon_skus_set = set()
        if division == "golf":
            wm_items = load_json("walmart_item_master.json")
            for w in wm_items:
                if w.get("golfgenItem"):
                    walmart_skus_set.add(w["golfgenItem"].strip().upper())
            am_items = load_json("amazon_item_master.json")
            for a in am_items:
                if a.get("sku"):
                    amazon_skus_set.add(a["sku"].strip().upper())

        items = []
        for row in rows[1:]:
            item_num = safe_str(row[item_col] if item_col is not None else "")
            if not item_num:
                continue

            item = {
                "itemNumber": item_num,
                "description": safe_str(row[desc_col]) if desc_col is not None else "",
                "pack": safe_int(row[pack_col]) if pack_col is not None else 1,
                "onHand": safe_int(row[oh_col]) if oh_col is not None else 0,
                "nonStandard": safe_int(row[ns_col]) if ns_col is not None else 0,
                "damage": safe_int(row[dmg_col]) if dmg_col is not None else 0,
                "standard": safe_int(row[std_col]) if std_col is not None else 0,
                "qcHold": safe_int(row[qc_col]) if qc_col is not None else 0,
                "pcsOnHand": safe_int(row[poh_col]) if poh_col is not None else 0,
                "pcsAllocated": safe_int(row[palloc_col]) if palloc_col is not None else 0,
                "pcsAvailable": safe_int(row[pavail_col]) if pavail_col is not None else 0,
                "itemRef": safe_str(row[ref_col]) if ref_col is not None else "",
            }

            if division == "golf":
                item["channel"] = _classify_golf_channel(item_num, walmart_skus_set, amazon_skus_set)

            items.append(item)

        if not items:
            raise HTTPException(status_code=400, detail="No items found in Excel file")

        # Save to JSON
        filename = "golf_inventory.json" if division == "golf" else "housewares_inventory.json"
        json_path = DB_DIR / filename
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(items, f, indent=2)

        # Update upload metadata
        meta = _load_upload_meta()
        meta[division] = {
            "lastUpload": datetime.now(TIMEZONE).isoformat(),
            "filename": file.filename,
            "itemCount": len(items),
        }
        _save_upload_meta(meta)

        wb.close()
        logger.info(f"  {division.title()} inventory uploaded: {len(items)} items from {file.filename}")

        return {
            "status": "success",
            "division": division,
            "filename": file.filename,
            "itemCount": len(items),
            "columns_mapped": {
                "item": item_col is not None,
                "description": desc_col is not None,
                "pack": pack_col is not None,
                "pcsOnHand": poh_col is not None,
                "pcsAllocated": palloc_col is not None,
                "pcsAvailable": pavail_col is not None,
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"Inventory Excel upload error: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


@router.get("/api/item-master/walmart")
def item_master_walmart():
    """Walmart item master data."""
    items = load_json("walmart_item_master.json")
    return {"items": items, "count": len(items)}


@router.put("/api/item-master/walmart/{golfgen_item}")
def update_walmart_item(golfgen_item: str, body: dict = Body(...)):
    """Update a single Walmart item by golfgenItem identifier."""
    items = load_json("walmart_item_master.json")
    found = False
    updatable = ["plannedAnnualUnits", "unitCost", "unitRetail", "storeCount"]
    for item in items:
        if item.get("golfgenItem") == golfgen_item:
            for key in updatable:
                if key in body:
                    item[key] = body[key]
            found = True
            break
    if not found:
        raise HTTPException(status_code=404, detail=f"Walmart item {golfgen_item} not found")
    save_json("walmart_item_master.json", items)
    return {"status": "ok", "golfgenItem": golfgen_item}


@router.get("/api/item-master/amazon")
def item_master_amazon():
    """Amazon item master data (unique parent SKUs only)."""
    all_items = load_json("amazon_item_master.json")
    parent_items = [i for i in all_items if i.get("asin") and "/" not in i.get("sku", "") and "FBM" not in i.get("sku", "") and not i.get("sku", "").startswith("T-")]
    return {"items": parent_items, "count": len(parent_items), "totalVariations": len(all_items)}


@router.get("/api/item-master/housewares")
def item_master_housewares():
    """Housewares item master — items from housewares inventory."""
    items = load_json("housewares_inventory.json")
    result = []
    for item in items:
        result.append({
            "itemNumber": item.get("itemNumber", ""),
            "description": item.get("description", ""),
            "pack": item.get("pack", 0),
            "pcsOnHand": item.get("pcsOnHand", 0),
            "pcsAvailable": item.get("pcsAvailable", 0),
            "pcsAllocated": item.get("pcsAllocated", 0),
            "nonStandard": item.get("nonStandard", 0),
            "damage": item.get("damage", 0),
            "qcHold": item.get("qcHold", 0),
        })
    return {"items": result, "count": len(result)}


@router.get("/api/item-master/other")
def item_master_other():
    """Items in warehouse inventory that are not in Amazon or Walmart item masters."""
    import re as _re
    walmart_items = load_json("walmart_item_master.json")
    amazon_items = load_json("amazon_item_master.json")
    golf_items = load_json("golf_inventory.json")
    hw_items = load_json("housewares_inventory.json")

    walmart_skus = set()
    for w in walmart_items:
        if w.get("golfgenItem"):
            walmart_skus.add(w["golfgenItem"].strip().upper())

    amazon_skus = set()
    for a in amazon_items:
        if a.get("sku"):
            amazon_skus.add(a["sku"].strip().upper())

    def _base_sku(raw: str) -> str:
        s = raw.strip().upper()
        if s.startswith("T-"):
            s = s[2:]
        for suf in ["FBM", "/RB", "/RETD", "/DONATE", "/DMGD", "/CUST", "/HOLD", "/1", "/2"]:
            if s.endswith(suf):
                s = s[: -len(suf)]
        s = _re.sub(r"\s*/DAM.*$", "", s)
        return s.strip()

    other_items = []
    seen = set()
    for item in golf_items + hw_items:
        sku_raw = item.get("itemNumber", "").strip().upper()
        base = _base_sku(sku_raw)
        if base in seen:
            continue
        in_walmart = base in walmart_skus
        in_amazon = base in amazon_skus
        if not in_walmart and not in_amazon:
            seen.add(base)
            source = "Golf" if item in golf_items else "Housewares"
            other_items.append({
                "itemNumber": item.get("itemNumber", ""),
                "description": item.get("description", ""),
                "source": source,
                "pcsOnHand": item.get("pcsOnHand", 0),
                "pcsAvailable": item.get("pcsAvailable", 0),
                "pcsAllocated": item.get("pcsAllocated", 0),
            })

    return {"items": other_items, "count": len(other_items)}
