"""OTW / Logistics routes."""
import json
import logging
from pathlib import Path
from datetime import datetime
from zoneinfo import ZoneInfo
from fastapi import APIRouter, Request, UploadFile, File, HTTPException

from core.config import DB_DIR, TIMEZONE
from core.auth import require_auth

logger = logging.getLogger("golfgen")
router = APIRouter()


def _load_logistics():
    fp = DB_DIR / "logistics_tracking.json"
    if fp.exists():
        with open(fp) as f:
            return json.load(f)
    return {"shipments": [], "itemsByContainer": [], "lastUpload": None}

def _save_logistics(data):
    fp = DB_DIR / "logistics_tracking.json"
    with open(fp, "w") as f:
        json.dump(data, f, indent=2)

@router.get("/api/logistics")
async def get_logistics(request: Request):
    require_auth(request)
    data = _load_logistics()
    _SHIP_STOP = {"SHIPMENT STATUS SUMMARY", "STATUS SUMMARY", "UPCOMING ARRIVALS",
                  "TOTAL", "TOTALS", "ITEM SUMMARY", "PO#", "STATUS", "IN TRANSIT",
                  "DELIVERED", "PENDING SHIPMENT", "PENDING"}
    if data.get("shipments"):
        clean = []
        for s in data["shipments"]:
            shipper = (s.get("shipper") or "").strip()
            if not shipper or shipper.upper() in _SHIP_STOP:
                continue
            if any(sw in shipper.upper() for sw in ("STATUS SUMMARY", "UPCOMING ARRIVALS", "ITEM SUMMARY")):
                continue
            hbl = (s.get("hbl") or "").strip()
            if not hbl or len(hbl) < 5:
                continue
            # Normalize old field names to new field names
            if "mode" in s and "containerType" not in s:
                s["containerType"] = s.pop("mode")
            if "vessel" in s and "vesselVoyage" not in s:
                s["vesselVoyage"] = s.pop("vessel")
            # Ensure status field exists
            if not s.get("status"):
                s["status"] = ""
            clean.append(s)
        data["shipments"] = clean
    # Clean items: filter out Container Total, NON-GOLF, TOTAL rows, etc.
    if data.get("itemsByContainer"):
        data["itemsByContainer"] = [
            item for item in data["itemsByContainer"]
            if item.get("sku") and str(item["sku"]).strip()
            and str(item["sku"]).strip().upper() not in ("TOTAL", "TOTALS", "GRAND TOTAL")
            and "NON-GOLF" not in str(item.get("containerNumber", "")).upper()
            and "CONTAINER TOTAL" not in str(item.get("description", "")).upper()
            and "GRAND TOTAL" not in str(item.get("description", "")).upper()
        ]

    # ── Enrich items with container status from shipments ──
    container_status = {}
    container_eta = {}
    for s in data.get("shipments", []):
        cn = (s.get("containerNumber") or "").strip()
        if cn:
            container_status[cn] = s.get("status") or ""
            container_eta[cn] = s.get("deliveryDate") or s.get("etaFinal") or s.get("etaDischarge") or ""
    for item in data.get("itemsByContainer", []):
        cn = (item.get("containerNumber") or "").strip()
        item["status"] = container_status.get(cn, "")

    # ── If no itemsByContainer from logistics sheet, build from Factory PO Summary ──
    if not data.get("itemsByContainer"):
        try:
            fp = DB_DIR / "factory_po_summary.json"
            if fp.exists():
                with open(fp) as f:
                    po_data = json.load(f)
                # Build PO → container mapping from shipments (by matching HBL/PO)
                # Use logistics items if available in PO data
                logistics_items = po_data.get("logisticsItems", [])
                if logistics_items:
                    for item in logistics_items:
                        cn = (item.get("containerNumber") or "").strip()
                        item["status"] = container_status.get(cn, "")
                        item["arrivalDate"] = container_eta.get(cn, item.get("eta", ""))
                    data["itemsByContainer"] = logistics_items
                else:
                    # Build from unitsByItem + container mapping
                    units = po_data.get("unitsByItem", [])
                    # Build SKU → containers from PO containers
                    sku_po_map = {}  # sku -> list of PO numbers from byPO
                    for u in units:
                        sku = (u.get("sku") or "").strip()
                        if not sku:
                            continue
                        for po_num in (u.get("byPO") or {}).keys():
                            sku_po_map.setdefault(sku, set()).add(po_num)

                    # Build PO → containers from purchaseOrders
                    po_container_map = {}
                    for po in po_data.get("purchaseOrders", []):
                        po_num = (po.get("poNumber") or "").strip()
                        for c in po.get("containers", []):
                            cn = (c.get("containerNumber") or "").strip()
                            if cn:
                                po_container_map.setdefault(po_num, []).append(cn)

                    items = []
                    for u in units:
                        sku = (u.get("sku") or "").strip()
                        desc = u.get("description", "")
                        total_qty = u.get("total") or u.get("quantity") or 0
                        # Find containers for this SKU via PO mapping
                        containers_for_sku = set()
                        for po_num in sku_po_map.get(sku, set()):
                            for cn in po_container_map.get(po_num, []):
                                containers_for_sku.add(cn)
                        if containers_for_sku:
                            for cn in sorted(containers_for_sku):
                                items.append({
                                    "containerNumber": cn,
                                    "sku": sku,
                                    "description": desc,
                                    "qty": total_qty,
                                    "eta": container_eta.get(cn, ""),
                                    "status": container_status.get(cn, ""),
                                    "arrivalDate": container_eta.get(cn, ""),
                                })
                        else:
                            items.append({
                                "containerNumber": "",
                                "sku": sku,
                                "description": desc,
                                "qty": total_qty,
                                "eta": "",
                                "status": "",
                                "arrivalDate": "",
                            })
                    data["itemsByContainer"] = items
        except Exception as e:
            logger.error(f"Error building items from Factory PO: {e}")

    return data

@router.post("/api/logistics/upload")
async def upload_logistics(request: Request, file: UploadFile = File(...)):
    """Standalone logistics upload — delegates to the same logic as combined supply-chain upload."""
    require_auth(request)
    import openpyxl
    from io import BytesIO
    contents = await file.read()
    if not contents:
        raise HTTPException(400, "Empty file uploaded")
    try:
        wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
    except Exception as e:
        logger.error(f"Failed to open Excel file: {e}")
        raise HTTPException(400, f"Could not read Excel file: {e}. Make sure it is a valid .xlsx file.")
    # Case-insensitive sheet name matching
    logistics_sheet = None
    for sn in wb.sheetnames:
        if sn.strip().lower() == "logistics tracking":
            logistics_sheet = sn
            break
    if not logistics_sheet:
        raise HTTPException(400, f"Sheet 'Logistics Tracking' not found. Available sheets: {wb.sheetnames}")
    ws = wb[logistics_sheet]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        r = {}
        for c in row:
            if hasattr(c, 'column_letter') and c.value is not None:
                r[c.column_letter] = c.value
        rows.append(r)

    def _to_str(v):
        if v is None: return ""
        from datetime import datetime as _dt, date as _d
        if isinstance(v, (_dt, _d)):
            return v.strftime("%Y-%m-%d")
        return str(v)

    def _sr(v, n=2):
        try: return round(float(v), n)
        except (ValueError, TypeError): return 0

    def _si(v):
        try: return int(float(v))
        except (ValueError, TypeError): return 0

    _STOP_WORDS = {"SHIPMENT STATUS SUMMARY", "STATUS SUMMARY", "UPCOMING ARRIVALS",
                    "TOTAL", "TOTALS", "ITEM SUMMARY", "PO#"}

    # ── Parse shipments ──
    shipments = []
    header_row = None
    for i, r in enumerate(rows):
        vals = [str(v).upper() for v in r.values()]
        joined = " ".join(vals)
        if ("SHIPPER" in joined or "FACTORY" in joined) and ("HBL" in joined or "HB" in joined or "MODE" in joined):
            header_row = i
            break

    if header_row is not None:
        empty_count = 0
        for i in range(header_row + 1, len(rows)):
            r = rows[i]
            shipper = r.get("B")
            if not shipper:
                empty_count += 1
                if empty_count >= 2: break
                continue
            empty_count = 0
            shipper_str = str(shipper).strip()
            shipper_upper = shipper_str.upper()
            if any(sw in shipper_upper for sw in _STOP_WORDS): break
            if shipper_upper in ["SHIPPER", "FACTORY", "STATUS", ""] or len(shipper_str) <= 3: continue
            hbl = str(r.get("D") or "").strip()
            if not hbl or len(hbl) < 5: continue
            # Infer status from delivery date
            delivery_raw = r.get("N")
            if delivery_raw and str(delivery_raw).strip():
                status = "Delivered"
            else:
                eta_raw = r.get("J")
                if eta_raw:
                    status = "In Transit"
                else:
                    status = "Pending"
            shipments.append({
                "shipper": shipper_str,                            # B = Factory/Shipper
                "factory": str(r.get("C") or "").strip(),          # C = Shipped (brand)
                "hbl": hbl,                                        # D = HB#
                "containerType": str(r.get("E") or "").strip(),    # E = Type
                "containerNumber": str(r.get("F") or "").strip(),  # F = CNTR#
                "vesselVoyage": str(r.get("G") or "").strip(),     # G = Vessel
                "etdOrigin": _to_str(r.get("H")),                  # H = ETD
                "departurePort": str(r.get("I") or "").strip(),    # I = ETD PORT
                "etaDischarge": _to_str(r.get("J")),               # J = ETA
                "arrivalPort": str(r.get("K") or "").strip(),      # K = ETA PORT
                "etaFinal": _to_str(r.get("L")),                   # L = ETA Dest. (date)
                "finalLocation": str(r.get("M") or "").strip(),    # M = ETA Dest. (location)
                "deliveryDate": _to_str(r.get("N")),               # N = Actual Delivery
                "comments": str(r.get("O") or "").strip(),         # O = Comments
                "status": status,
            })
    # ── Parse items by container ──
    item_header_row = None
    for i, r in enumerate(rows):
        for _k, v in r.items():
            if v and "container" in str(v).lower() and "item" in str(v).lower():
                item_header_row = i; break
        if item_header_row is not None: break

    items = []
    if item_header_row is not None:
        col_hdr = item_header_row + 1
        for ci in range(item_header_row, min(item_header_row + 4, len(rows))):
            vals_up = [str(v).upper() for v in rows[ci].values()]
            joined = " ".join(vals_up)
            if "CONTAINER" in joined and ("ITEM" in joined or "SKU" in joined):
                col_hdr = ci; break
        cur_container = ""
        for i in range(col_hdr + 1, len(rows)):
            r = rows[i]
            container_val = r.get("B"); sku_val = r.get("F"); desc_val = r.get("G"); qty_val = r.get("J")
            if not container_val and not sku_val and not desc_val: continue
            container_str = str(container_val or "").strip()
            sku_str = str(sku_val or "").strip()
            desc_str = str(desc_val or "").strip()
            skip_words = ["CONTAINER TOTAL", "GOLF TOTAL", "GRAND TOTAL", "NON-GOLF", "CONTAINER #", "ITEM NUMBER", "ITEM SUMMARY"]
            combined_upper = f"{container_str} {sku_str} {desc_str}".upper()
            if any(sw in combined_upper for sw in skip_words): continue
            if container_val and len(container_str) >= 8 and container_str[0].isalpha():
                cur_container = container_str
            if not sku_val or len(sku_str) < 3: continue
            items.append({
                "containerNumber": cur_container, "invoice": str(r.get("C") or "").strip(),
                "eta": _to_str(r.get("D")), "po": str(r.get("E") or "").strip(),
                "sku": sku_str, "description": desc_str, "qty": _si(qty_val),
            })

    from datetime import date as _date
    data = {"shipments": shipments, "itemsByContainer": items,
        "lastUpload": _date.today().isoformat(), "sourceFile": file.filename}
    _save_logistics(data)
    return {"status": "ok", "shipments": len(shipments), "items": len(items), "lastUpload": data["lastUpload"]}


# NOTE: The /api/supply-chain/upload endpoint has been moved to routers/supply_chain.py
# The old handler was removed to avoid route conflicts.
def _save_factory_po(data):
    fp = DB_DIR / "factory_po_summary.json"
    with open(fp, "w") as f:
        json.dump(data, f, indent=2)
