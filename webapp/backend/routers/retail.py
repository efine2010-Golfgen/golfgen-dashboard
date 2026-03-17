"""
Retail Reporting router — Walmart POS data from Scintilla reports.

Upload endpoints parse Excel files and load into walmart_* tables.
Query endpoints serve data to the Retail Reporting frontend page.
"""
import logging
import uuid
from datetime import datetime
from zoneinfo import ZoneInfo
from io import BytesIO

from fastapi import APIRouter, UploadFile, File, Query, HTTPException, Body
from core.database import get_db, get_db_rw

logger = logging.getLogger("golfgen")
router = APIRouter()

CT = ZoneInfo("America/Chicago")

# ── Brand → division mapping ────────────────────────────────────────────
BRAND_DIVISION = {
    "pga tour": "golf",
    "lpga": "golf",
    "holiday time": "housewares",
    "way to celebrate": "housewares",
}

def _brand_to_division(brand_name: str) -> str:
    if not brand_name:
        return "golf"
    return BRAND_DIVISION.get(brand_name.strip().lower(), "golf")

def _n(v):
    """Coerce value to float, returning 0 for None/empty."""
    if v is None or v == "":
        return 0.0
    try:
        from decimal import Decimal
        if isinstance(v, Decimal):
            return float(v)
        return float(v)
    except (TypeError, ValueError):
        return 0.0

def _safe_int(v):
    if v is None or v == "":
        return 0
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0

def _safe_str(v, max_len=None):
    if v is None:
        return None
    s = str(v).strip()
    if max_len:
        s = s[:max_len]
    return s or None


# ═══════════════════════════════════════════════════════════════════════
#  UPLOAD ENDPOINT — auto-detect report type and parse
# ═══════════════════════════════════════════════════════════════════════

@router.post("/api/retail/upload")
async def upload_retail_report(file: UploadFile = File(...)):
    """Upload a Scintilla Excel report. Auto-detects report type."""
    import openpyxl

    upload_id = str(uuid.uuid4())[:8]
    fname = file.filename or "unknown.xlsx"
    data = await file.read()
    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    ws = wb.active

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_set = set(h for h in headers if h)

    # Auto-detect report type from headers
    report_type = None
    if "store_number" in header_set and "walmart_calendar_week" in header_set:
        report_type = "store_weekly"
    elif "prime_item_number" in header_set and "walmart_calendar_week" in header_set:
        report_type = "item_weekly"
    elif "prime_item_number" in header_set and "wm_time_window_week" in header_set:
        report_type = "item_period"
    elif "all_links_item_number" in header_set:
        report_type = "ecomm"
    elif "store_dc_nbr" in header_set:
        report_type = "order_forecast"
    else:
        # Check for scorecard (pivoted format — row 3 has TY/LY/Diff)
        row3 = [c.value for c in list(ws.iter_rows(min_row=3, max_row=3))[0]]
        if "TY" in row3 and "LY" in row3:
            report_type = "scorecard"

    wb.close()

    if not report_type:
        raise HTTPException(400, "Could not auto-detect report type from headers")

    # Re-open for parsing
    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    ws = wb.active

    try:
        if report_type == "store_weekly":
            rows = _parse_store_weekly(ws, headers)
        elif report_type == "item_weekly":
            rows = _parse_item_weekly(ws, headers, period_type="weekly")
        elif report_type == "item_period":
            rows = _parse_item_weekly(ws, headers, period_type=None)  # read from data
        elif report_type == "scorecard":
            rows = _parse_scorecard(ws, upload_id)
        elif report_type == "ecomm":
            rows = _parse_ecomm(ws, headers)
        elif report_type == "order_forecast":
            rows = _parse_order_forecast(ws, headers)
        else:
            rows = 0
    except Exception as e:
        wb.close()
        _log_upload(upload_id, fname, report_type, 0, "ERROR", str(e))
        raise HTTPException(500, f"Parse error: {e}")

    wb.close()
    _log_upload(upload_id, fname, report_type, rows, "SUCCESS", None)

    return {
        "status": "ok",
        "upload_id": upload_id,
        "report_type": report_type,
        "rows_loaded": rows,
        "filename": fname,
    }


def _log_upload(upload_id, fname, report_type, rows, status, error):
    try:
        con = get_db_rw()
        con.execute("""
            INSERT INTO retail_upload_log
                (upload_id, filename, report_type, channel, rows_loaded, uploaded_by, status, error)
            VALUES (?, ?, ?, 'walmart_stores', ?, 'upload', ?, ?)
        """, [upload_id, fname, report_type, rows, status, error])
        con.close()
    except Exception as e:
        logger.error(f"Upload log error: {e}")


# ═══════════════════════════════════════════════════════════════════════
#  PARSERS
# ═══════════════════════════════════════════════════════════════════════

def _parse_store_weekly(ws, headers):
    """Parse Store Level Detail → walmart_store_weekly."""
    h = {v: i for i, v in enumerate(headers) if v}
    con = get_db_rw()
    count = 0
    batch = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        wk = _safe_str(row[h.get("walmart_calendar_week", 0)])
        yr = _safe_str(row[h.get("walmart_calendar_year", 1)])
        sn = _safe_str(row[h.get("store_number", 3)])
        if not wk or not sn:
            continue
        batch.append((
            wk, yr,
            _safe_str(row[h.get("store_name", 2)], 120),
            sn,
            _safe_str(row[h.get("vendor_department_number", 4)]),
            _n(row[h["pos_quantity_this_year"]]) if "pos_quantity_this_year" in h else 0,
            _n(row[h["pos_quantity_last_year"]]) if "pos_quantity_last_year" in h else 0,
            _n(row[h["pos_sales_this_year"]]) if "pos_sales_this_year" in h else 0,
            _n(row[h["pos_sales_last_year"]]) if "pos_sales_last_year" in h else 0,
            _n(row[h["gross_receipt_quantity_this_year"]]) if "gross_receipt_quantity_this_year" in h else 0,
            _n(row[h["gross_receipt_quantity_last_year"]]) if "gross_receipt_quantity_last_year" in h else 0,
            _n(row[h["net_receipt_quantity_this_year"]]) if "net_receipt_quantity_this_year" in h else 0,
            _n(row[h["net_receipt_quantity_last_year"]]) if "net_receipt_quantity_last_year" in h else 0,
            _n(row[h["total_store_customer_returns_quantity_this_year"]]) if "total_store_customer_returns_quantity_this_year" in h else 0,
            _n(row[h["total_store_customer_returns_quantity_last_year"]]) if "total_store_customer_returns_quantity_last_year" in h else 0,
            _n(row[h["total_store_customer_returns_quantity_defective_this_year"]]) if "total_store_customer_returns_quantity_defective_this_year" in h else 0,
            _n(row[h["total_store_customer_returns_quantity_defective_last_year"]]) if "total_store_customer_returns_quantity_defective_last_year" in h else 0,
            _n(row[h["store_on_hand_quantity_this_year"]]) if "store_on_hand_quantity_this_year" in h else 0,
            _n(row[h["store_on_hand_quantity_last_year"]]) if "store_on_hand_quantity_last_year" in h else 0,
            _n(row[h["current_clearance_on_hand_quantity_this_year"]]) if "current_clearance_on_hand_quantity_this_year" in h else 0,
            _n(row[h["current_clearance_on_hand_quantity_last_year"]]) if "current_clearance_on_hand_quantity_last_year" in h else 0,
            _n(row[h["store_in_transit_quantity_this_year"]]) if "store_in_transit_quantity_this_year" in h else 0,
            _n(row[h["store_in_transit_quantity_last_year"]]) if "store_in_transit_quantity_last_year" in h else 0,
            _n(row[h["store_in_warehouse_quantity_this_year"]]) if "store_in_warehouse_quantity_this_year" in h else 0,
            _n(row[h["store_in_warehouse_quantity_last_year"]]) if "store_in_warehouse_quantity_last_year" in h else 0,
            _n(row[h["instock_percentage_this_year"]]) if "instock_percentage_this_year" in h else None,
            _n(row[h["instock_percentage_last_year"]]) if "instock_percentage_last_year" in h else None,
            _n(row[h["repl_instock_percentage_this_year"]]) if "repl_instock_percentage_this_year" in h else None,
            _n(row[h["repl_instock_percentage_last_year"]]) if "repl_instock_percentage_last_year" in h else None,
            _n(row[h["store_returns_quantity_to_vendor_this_year"]]) if "store_returns_quantity_to_vendor_this_year" in h else 0,
            _n(row[h["store_returns_quantity_to_vendor_last_year"]]) if "store_returns_quantity_to_vendor_last_year" in h else 0,
            _n(row[h["store_returns_quantity_to_return_center_this_year"]]) if "store_returns_quantity_to_return_center_this_year" in h else 0,
            _n(row[h["store_returns_quantity_to_return_center_last_year"]]) if "store_returns_quantity_to_return_center_last_year" in h else 0,
            _n(row[h["store_returns_quantity_to_dc_this_year"]]) if "store_returns_quantity_to_dc_this_year" in h else 0,
            _n(row[h["store_returns_quantity_to_dc_last_year"]]) if "store_returns_quantity_to_dc_last_year" in h else 0,
            _n(row[h["store_returns_quantity_recall_this_year"]]) if "store_returns_quantity_recall_this_year" in h else 0,
            _n(row[h["store_returns_quantity_recall_last_year"]]) if "store_returns_quantity_recall_last_year" in h else 0,
            _n(row[h["inventory_adjustment_quantity_this_year"]]) if "inventory_adjustment_quantity_this_year" in h else 0,
            _n(row[h["inventory_adjustment_quantity_last_year"]]) if "inventory_adjustment_quantity_last_year" in h else 0,
            _safe_int(row[h["traited_store_count_this_year"]]) if "traited_store_count_this_year" in h else 0,
            _safe_int(row[h["traited_store_count_last_year"]]) if "traited_store_count_last_year" in h else 0,
        ))
        count += 1
        if len(batch) >= 500:
            _insert_store_weekly_batch(con, batch)
            batch = []

    if batch:
        _insert_store_weekly_batch(con, batch)
    con.close()
    return count


def _insert_store_weekly_batch(con, batch):
    for r in batch:
        con.execute("""
            INSERT OR IGNORE INTO walmart_store_weekly
                (walmart_week, walmart_year, store_name, store_number, vendor_dept_number,
                 pos_qty_ty, pos_qty_ly, pos_sales_ty, pos_sales_ly,
                 gross_receipt_qty_ty, gross_receipt_qty_ly,
                 net_receipt_qty_ty, net_receipt_qty_ly,
                 returns_qty_ty, returns_qty_ly,
                 returns_defective_qty_ty, returns_defective_qty_ly,
                 on_hand_qty_ty, on_hand_qty_ly,
                 clearance_on_hand_qty_ty, clearance_on_hand_qty_ly,
                 in_transit_qty_ty, in_transit_qty_ly,
                 in_warehouse_qty_ty, in_warehouse_qty_ly,
                 instock_pct_ty, instock_pct_ly,
                 repl_instock_pct_ty, repl_instock_pct_ly,
                 returns_to_vendor_qty_ty, returns_to_vendor_qty_ly,
                 returns_to_return_center_qty_ty, returns_to_return_center_qty_ly,
                 returns_to_dc_qty_ty, returns_to_dc_qty_ly,
                 returns_recall_qty_ty, returns_recall_qty_ly,
                 inv_adjustment_qty_ty, inv_adjustment_qty_ly,
                 traited_store_count_ty, traited_store_count_ly)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, list(r))


def _parse_item_weekly(ws, headers, period_type=None):
    """Parse Dashboard Report 1 or 2 → walmart_item_weekly."""
    h = {v: i for i, v in enumerate(headers) if v}
    # Detect week column: 'walmart_calendar_week' (Report 2) or 'wm_time_window_week' (Report 1)
    wk_col = "walmart_calendar_week" if "walmart_calendar_week" in h else "wm_time_window_week"
    con = get_db_rw()
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        wk_val = _safe_str(row[h[wk_col]])
        if not wk_val:
            continue
        item = _safe_str(row[h.get("prime_item_number", 3)])
        if not item:
            continue

        # Determine period_type
        pt = period_type or wk_val  # Report 1: wk_val is 'L13W', 'L4W', etc.
        # Normalize LW → L1W (Scintilla uses both)
        if pt == "LW":
            pt = "L1W"
        if pt not in ("weekly", "L1W", "L4W", "L13W", "L26W", "L52W"):
            if wk_val and len(wk_val) >= 6 and wk_val[:4].isdigit():
                pt = "weekly"  # numeric week like 202507
            else:
                pt = wk_val  # keep as-is

        brand = _safe_str(row[h.get("brand_name", 1)])
        division = _brand_to_division(brand)
        sales_type = _safe_str(row[h.get("sales_type_description", 12)])
        adj_code = _safe_str(row[h.get("backroom_inventory_adjustment_type_code", 15)])

        con.execute("""
            INSERT OR IGNORE INTO walmart_item_weekly
                (walmart_week, period_type, brand_name, brand_owner_id,
                 prime_item_number, prime_item_desc, walmart_upc, warehouse_pack_upc,
                 product_description, sales_type, vendor_name, vendor_number,
                 vendor_pack_qty, vendor_pack_length, vendor_pack_height, vendor_pack_width,
                 pos_sales_ty, pos_sales_ly, pos_qty_ty, pos_qty_ly,
                 pos_store_count_ty, pos_store_count_ly,
                 units_per_store_ty, units_per_store_ly,
                 dollars_per_store_ty, dollars_per_store_ly,
                 gross_receipt_qty_ty, gross_receipt_qty_ly,
                 net_receipt_qty_ty, net_receipt_qty_ly,
                 net_receipt_cost_ty, net_receipt_cost_ly,
                 net_receipt_retail_ty, net_receipt_retail_ly,
                 returns_qty_ty, returns_qty_ly,
                 returns_amt_ty, returns_amt_ly,
                 returns_defective_qty_ty, returns_defective_qty_ly,
                 returns_defective_amt_ty, returns_defective_amt_ly,
                 on_hand_qty_ty, on_hand_qty_ly,
                 on_order_qty_ty, on_order_qty_ly,
                 in_transit_qty_ty, in_transit_qty_ly,
                 in_warehouse_qty_ty, in_warehouse_qty_ly,
                 clearance_on_hand_qty_ty, clearance_on_hand_qty_ly,
                 clearance_on_hand_eop_ty, clearance_on_hand_eop_ly,
                 instock_pct_ty, instock_pct_ly,
                 repl_instock_pct_ty, repl_instock_pct_ly,
                 traited_store_count_ty, traited_store_count_ly,
                 valid_store_count_ty, valid_store_count_ly,
                 inv_adjustment_qty_ty, inv_adjustment_qty_ly,
                 backroom_adj_qty, backroom_adj_type_code, backroom_adj_type_desc,
                 division, customer, platform)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, [
            wk_val, pt, brand, _safe_str(row[h.get("brand_owner_id", 2)]),
            item, _safe_str(row[h.get("prime_item_description", 4)], 120),
            _safe_str(row[h.get("walmart_upc_number", 5)]),
            _safe_str(row[h.get("warehouse_pack_upc_number", 6)]),
            _safe_str(row[h.get("product_description", 7)], 120),
            sales_type,
            _safe_str(row[h.get("vendor_name", 13)], 80),
            _safe_str(row[h.get("vendor_number", 14)]),
            _n(row[h.get("vendor_pack_quantity", 8)]) if "vendor_pack_quantity" in h else None,
            _n(row[h.get("vendor_pack_length_quantity", 9)]) if "vendor_pack_length_quantity" in h else None,
            _n(row[h.get("vendor_pack_height_quantity", 10)]) if "vendor_pack_height_quantity" in h else None,
            _n(row[h.get("vendor_pack_width_quantity", 11)]) if "vendor_pack_width_quantity" in h else None,
            # POS
            _n(row[h["pos_sales_this_year"]]) if "pos_sales_this_year" in h else 0,
            _n(row[h["pos_sales_last_year"]]) if "pos_sales_last_year" in h else 0,
            _n(row[h["pos_quantity_this_year"]]) if "pos_quantity_this_year" in h else 0,
            _n(row[h["pos_quantity_last_year"]]) if "pos_quantity_last_year" in h else 0,
            _safe_int(row[h["pos_store_count_this_year"]]) if "pos_store_count_this_year" in h else 0,
            _safe_int(row[h["pos_store_count_last_year"]]) if "pos_store_count_last_year" in h else 0,
            # Velocity
            _n(row[h["units_per_str_with_sales_per_week_or_per_day_ty"]]) if "units_per_str_with_sales_per_week_or_per_day_ty" in h else 0,
            _n(row[h["units_per_str_with_sales_per_week_or_per_day_ly"]]) if "units_per_str_with_sales_per_week_or_per_day_ly" in h else 0,
            _n(row[h["dollar_per_str_with_sales_per_week_or_per_day_ty"]]) if "dollar_per_str_with_sales_per_week_or_per_day_ty" in h else 0,
            _n(row[h["dollar_per_str_with_sales_per_week_or_per_day_ly"]]) if "dollar_per_str_with_sales_per_week_or_per_day_ly" in h else 0,
            # Receipts
            _n(row[h["gross_receipt_quantity_this_year"]]) if "gross_receipt_quantity_this_year" in h else 0,
            _n(row[h["gross_receipt_quantity_last_year"]]) if "gross_receipt_quantity_last_year" in h else 0,
            _n(row[h["net_receipt_quantity_this_year"]]) if "net_receipt_quantity_this_year" in h else 0,
            _n(row[h["net_receipt_quantity_last_year"]]) if "net_receipt_quantity_last_year" in h else 0,
            _n(row[h["net_rcpt_cost_amnt_this_year"]]) if "net_rcpt_cost_amnt_this_year" in h else 0,
            _n(row[h["net_rcpt_cost_amnt_last_year"]]) if "net_rcpt_cost_amnt_last_year" in h else 0,
            _n(row[h["net_receipt_retail_amount_this_year"]]) if "net_receipt_retail_amount_this_year" in h else 0,
            _n(row[h["net_receipt_retail_amount_last_year"]]) if "net_receipt_retail_amount_last_year" in h else 0,
            # Returns
            _n(row[h["total_store_customer_returns_quantity_this_year"]]) if "total_store_customer_returns_quantity_this_year" in h else 0,
            _n(row[h["total_store_customer_returns_quantity_last_year"]]) if "total_store_customer_returns_quantity_last_year" in h else 0,
            _n(row[h["total_store_customer_returns_amount_this_year"]]) if "total_store_customer_returns_amount_this_year" in h else 0,
            _n(row[h["total_store_customer_returns_amount_last_year"]]) if "total_store_customer_returns_amount_last_year" in h else 0,
            _n(row[h["total_store_customer_returns_quantity_defective_this_year"]]) if "total_store_customer_returns_quantity_defective_this_year" in h else 0,
            _n(row[h["total_store_customer_returns_quantity_defective_last_year"]]) if "total_store_customer_returns_quantity_defective_last_year" in h else 0,
            _n(row[h["total_store_customer_returns_amount_defective_this_year"]]) if "total_store_customer_returns_amount_defective_this_year" in h else 0,
            _n(row[h["total_store_customer_returns_amount_defective_last_year"]]) if "total_store_customer_returns_amount_defective_last_year" in h else 0,
            # Inventory
            _n(row[h["store_on_hand_quantity_this_year"]]) if "store_on_hand_quantity_this_year" in h else 0,
            _n(row[h["store_on_hand_quantity_last_year"]]) if "store_on_hand_quantity_last_year" in h else 0,
            _n(row[h["store_on_order_quantity_this_year"]]) if "store_on_order_quantity_this_year" in h else 0,
            _n(row[h["store_on_order_quantity_last_year"]]) if "store_on_order_quantity_last_year" in h else 0,
            _n(row[h["store_in_transit_quantity_this_year"]]) if "store_in_transit_quantity_this_year" in h else 0,
            _n(row[h["store_in_transit_quantity_last_year"]]) if "store_in_transit_quantity_last_year" in h else 0,
            _n(row[h["store_in_warehouse_quantity_this_year"]]) if "store_in_warehouse_quantity_this_year" in h else 0,
            _n(row[h["store_in_warehouse_quantity_last_year"]]) if "store_in_warehouse_quantity_last_year" in h else 0,
            _n(row[h["current_clearance_on_hand_quantity_this_year"]]) if "current_clearance_on_hand_quantity_this_year" in h else 0,
            _n(row[h["current_clearance_on_hand_quantity_last_year"]]) if "current_clearance_on_hand_quantity_last_year" in h else 0,
            _n(row[h["current_clearance_on_hand_quantity_this_year_eop"]]) if "current_clearance_on_hand_quantity_this_year_eop" in h else 0,
            _n(row[h["current_clearance_on_hand_quantity_last_year_eop"]]) if "current_clearance_on_hand_quantity_last_year_eop" in h else 0,
            # Metrics
            _n(row[h["instock_percentage_this_year"]]) if "instock_percentage_this_year" in h else None,
            _n(row[h["instock_percentage_last_year"]]) if "instock_percentage_last_year" in h else None,
            _n(row[h["repl_instock_percentage_this_year"]]) if "repl_instock_percentage_this_year" in h else None,
            _n(row[h["repl_instock_percentage_last_year"]]) if "repl_instock_percentage_last_year" in h else None,
            _safe_int(row[h["traited_store_count_this_year"]]) if "traited_store_count_this_year" in h else 0,
            _safe_int(row[h["traited_store_count_last_year"]]) if "traited_store_count_last_year" in h else 0,
            _safe_int(row[h["valid_store_count_this_year"]]) if "valid_store_count_this_year" in h else 0,
            _safe_int(row[h["valid_store_count_last_year"]]) if "valid_store_count_last_year" in h else 0,
            # Adjustments
            _n(row[h["inventory_adjustment_quantity_this_year"]]) if "inventory_adjustment_quantity_this_year" in h else 0,
            _n(row[h["inventory_adjustment_quantity_last_year"]]) if "inventory_adjustment_quantity_last_year" in h else 0,
            _n(row[h["backroom_adjustment_quantity"]]) if "backroom_adjustment_quantity" in h else 0,
            adj_code,
            _safe_str(row[h.get("backroom_inventory_adjustment_type_description", 16)], 60),
            # Hierarchy
            division, "walmart_stores", "scintilla",
        ])
        count += 1
    con.close()
    return count


def _parse_scorecard(ws, upload_id):
    """Parse Weekly Scorecard (pivoted) → walmart_scorecard.

    Layout:
    Row 1: [None, 'Last Week', None, None, 'Current Month', ...]
    Row 2: [None, '(202552-202552 / ...)', None, None, '(202601-202604 / ...)', ...]
    Row 3: [None, 'TY', 'LY', 'Diff', 'TY', 'LY', 'Diff', ...]
    Row 4+: Section headers and metric rows
    """
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 4:
        return 0

    # Parse period headers from rows 1-2
    period_row = all_rows[0]
    range_row = all_rows[1]
    periods = []
    for i in range(1, len(period_row), 3):
        pname = period_row[i]
        if pname:
            prange = range_row[i] if i < len(range_row) else None
            periods.append((_safe_str(pname, 20), _safe_str(prange, 60), i))

    con = get_db_rw()
    count = 0
    current_vendor = "All Vendors"
    current_group = ""

    # Collect all rows first, then batch insert for reliability
    batch = []
    report_date = datetime.now(CT).date().isoformat()

    for row in all_rows[3:]:
        label = _safe_str(row[0]) if row[0] else None
        if not label:
            continue

        # Detect vendor section header
        if label.startswith("Vendor Scorecard"):
            current_vendor = label.replace("Vendor Scorecard  - ", "").replace("Vendor Scorecard - ", "").strip()
            if len(current_vendor) > 120:
                current_vendor = current_vendor[:120]
            continue

        # Detect metric group
        if label in ("Store Sales", "Store Inventory", "DC Inventory", "Margins and Markdowns",
                      "Store Turns / GMROII", "eCommerce", "Supply Chain"):
            current_group = label
            continue

        # Truncate metric name to fit VARCHAR(60)
        metric_name = label[:60] if len(label) > 60 else label

        # Otherwise it's a metric row
        for pname, prange, col_start in periods:
            ty_val = row[col_start] if col_start < len(row) else None
            ly_val = row[col_start + 1] if col_start + 1 < len(row) else None
            diff_val = row[col_start + 2] if col_start + 2 < len(row) else None

            if ty_val is None and ly_val is None:
                continue

            batch.append([
                current_vendor, current_group, metric_name, pname, prange,
                _n(ty_val) if ty_val is not None else None,
                _n(ly_val) if ly_val is not None else None,
                _n(diff_val) if diff_val is not None else None,
                report_date,
                upload_id,
            ])

    # Batch insert with error handling per row
    for params in batch:
        try:
            con.execute("""
                INSERT OR IGNORE INTO walmart_scorecard
                    (vendor_section, metric_group, metric_name, period, period_range,
                     value_ty, value_ly, value_diff, report_date, upload_id,
                     division, customer, platform)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'golf', 'walmart_stores', 'scintilla')
            """, params)
            count += 1
        except Exception as e:
            logger.error(f"Scorecard insert error: {e} | params={params}")

    con.close()
    return count


def _parse_ecomm(ws, headers):
    """Parse eComm Sales Report → walmart_ecomm_weekly."""
    h = {v: i for i, v in enumerate(headers) if v}
    con = get_db_rw()
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        wk = _safe_str(row[h.get("walmart_calendar_week", 0)])
        item_num = _safe_str(row[h.get("all_links_item_number", 1)])
        if not wk or not item_num:
            continue

        brand = _safe_str(row[h.get("brand_name", 12)])
        division = _brand_to_division(brand)
        vendor_num = _safe_str(row[h.get("vendor_number_9_digit", 4)]) or _safe_str(row[h.get("vendor_number", 92)])

        con.execute("""
            INSERT OR IGNORE INTO walmart_ecomm_weekly
                (walmart_week, all_links_item_number, product_name,
                 base_unit_retail_amount, vendor_number, vendor_name,
                 item_description, fineline_description,
                 walmart_upc, ecomm_upc,
                 omni_category, omni_subcategory, brand_name,
                 dept_description, dept_number,
                 auth_based_qty_ty, auth_based_qty_ly,
                 auth_based_net_sales_ty, auth_based_net_sales_ly,
                 shipped_based_qty_ty, shipped_based_qty_ly,
                 shipped_based_net_sales_ty, shipped_based_net_sales_ly,
                 vendor_pack_cost, vendor_stock_id,
                 season_code, season_description, item_status_code,
                 division, customer, platform)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, [
            wk, item_num,
            _safe_str(row[h.get("product_name", 2)], 200),
            _n(row[h.get("base_unit_retail_amount", 3)]) if "base_unit_retail_amount" in h else None,
            vendor_num,
            _safe_str(row[h.get("vendor_name", 5)], 80),
            _safe_str(row[h.get("all_links_item_description", 6)], 120),
            _safe_str(row[h.get("fineline_description", 7)], 60),
            _safe_str(row[h.get("walmart_upc_number", 8)]),
            _safe_str(row[h.get("ecomm_upc_number", 9)]),
            _safe_str(row[h.get("omni_category_description", 10)], 60),
            _safe_str(row[h.get("omni_subcategory_description", 11)], 60),
            brand,
            _safe_str(row[h.get("accounting_department_description", 13)], 60),
            _safe_str(row[h.get("accounting_department_number", 14)]),
            # Sales metrics
            _n(row[h["auth_based_item_quantity_this_year"]]) if "auth_based_item_quantity_this_year" in h else 0,
            _n(row[h["auth_based_item_quantity_last_year"]]) if "auth_based_item_quantity_last_year" in h else 0,
            _n(row[h["auth_based_net_sales_amount_this_year"]]) if "auth_based_net_sales_amount_this_year" in h else 0,
            _n(row[h["auth_based_net_sales_amount_last_year"]]) if "auth_based_net_sales_amount_last_year" in h else 0,
            _n(row[h["shipped_based_quantity_this_year"]]) if "shipped_based_quantity_this_year" in h else 0,
            _n(row[h["shipped_based_quantity_last_year"]]) if "shipped_based_quantity_last_year" in h else 0,
            _n(row[h["shipped_based_net_sales_amount_this_year"]]) if "shipped_based_net_sales_amount_this_year" in h else 0,
            _n(row[h["shipped_based_net_sales_amount_last_year"]]) if "shipped_based_net_sales_amount_last_year" in h else 0,
            _n(row[h["vendor_pack_cost_amount"]]) if "vendor_pack_cost_amount" in h else None,
            _safe_str(row[h.get("vendor_stock_id", 95)]),
            _safe_str(row[h.get("season_code", 55)]),
            _safe_str(row[h.get("season_description", 56)], 40),
            _safe_str(row[h.get("item_status_code", 40)]),
            division, "walmart_stores", "scintilla",
        ])
        count += 1
    con.close()
    return count


def _parse_order_forecast(ws, headers):
    """Parse Order Forecast → walmart_order_forecast."""
    h = {v: i for i, v in enumerate(headers) if v}
    con = get_db_rw()
    today = datetime.now(CT).date().isoformat()
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        dc = _safe_str(row[h.get("store_dc_nbr", 0)])
        if not dc:
            continue
        con.execute("""
            INSERT OR IGNORE INTO walmart_order_forecast
                (snapshot_date, store_dc_nbr, store_dc_type, vendor_dept_number)
            VALUES (?, ?, ?, ?)
        """, [
            today, dc,
            _safe_str(row[h.get("store_dc_type", 1)]),
            _safe_str(row[h.get("vendor_department_number", 2)]),
        ])
        count += 1
    con.close()
    return count


# ═══════════════════════════════════════════════════════════════════════
#  QUERY ENDPOINTS — serve data to frontend
# ═══════════════════════════════════════════════════════════════════════

@router.get("/api/retail/scorecard")
def get_scorecard(division: str = None, customer: str = None):
    """Return scorecard KPIs grouped by vendor and period."""
    con = get_db()
    from core.hierarchy import hierarchy_filter
    hw, hp = hierarchy_filter(division=division, customer=customer or "walmart_stores")
    rows = con.execute(f"""
        SELECT vendor_section, metric_group, metric_name, period, period_range,
               value_ty, value_ly, value_diff
        FROM walmart_scorecard
        WHERE 1=1 {hw}
        ORDER BY vendor_section, metric_group, metric_name, period
    """, hp).fetchall()
    con.close()
    return {"scorecard": [
        {"vendorSection": r[0], "metricGroup": r[1], "metricName": r[2],
         "period": r[3], "periodRange": r[4],
         "valueTy": _n(r[5]), "valueLy": _n(r[6]), "valueDiff": _n(r[7])}
        for r in rows
    ]}


@router.get("/api/retail/store-performance")
def get_store_performance(
    division: str = None, customer: str = None,
    week: str = None, limit: int = 100, offset: int = 0,
    sort_by: str = "pos_sales_ty", sort_dir: str = "desc",
):
    """Return store-level weekly data for the Store Performance tab."""
    con = get_db()
    from core.hierarchy import hierarchy_filter
    hw, hp = hierarchy_filter(division=division, customer=customer or "walmart_stores")

    where = f"WHERE 1=1 {hw}"
    if week:
        where += " AND walmart_week = ?"
        hp.append(week)

    allowed_sorts = {
        "pos_sales_ty", "pos_qty_ty", "instock_pct_ty", "returns_qty_ty",
        "on_hand_qty_ty", "store_name", "store_number", "walmart_week",
    }
    sb = sort_by if sort_by in allowed_sorts else "pos_sales_ty"
    sd = "ASC" if sort_dir.lower() == "asc" else "DESC"

    rows = con.execute(f"""
        SELECT walmart_week, walmart_year, store_name, store_number,
               pos_qty_ty, pos_qty_ly, pos_sales_ty, pos_sales_ly,
               returns_qty_ty, returns_qty_ly,
               on_hand_qty_ty, on_hand_qty_ly,
               instock_pct_ty, instock_pct_ly,
               repl_instock_pct_ty, repl_instock_pct_ly,
               in_transit_qty_ty, in_warehouse_qty_ty,
               traited_store_count_ty
        FROM walmart_store_weekly
        {where}
        ORDER BY {sb} {sd}
        LIMIT ? OFFSET ?
    """, hp + [limit, offset]).fetchall()

    total = con.execute(f"SELECT COUNT(*) FROM walmart_store_weekly {where}", hp).fetchone()[0]
    con.close()

    # Get distinct weeks for filter dropdown
    con2 = get_db()
    weeks = [r[0] for r in con2.execute(
        "SELECT DISTINCT walmart_week FROM walmart_store_weekly ORDER BY walmart_week DESC"
    ).fetchall()]
    con2.close()

    return {
        "stores": [
            {"week": r[0], "year": r[1], "storeName": r[2], "storeNumber": r[3],
             "posQtyTy": _n(r[4]), "posQtyLy": _n(r[5]),
             "posSalesTy": _n(r[6]), "posSalesLy": _n(r[7]),
             "returnsQtyTy": _n(r[8]), "returnsQtyLy": _n(r[9]),
             "onHandQtyTy": _n(r[10]), "onHandQtyLy": _n(r[11]),
             "instockPctTy": _n(r[12]), "instockPctLy": _n(r[13]),
             "replInstockPctTy": _n(r[14]), "replInstockPctLy": _n(r[15]),
             "inTransitQtyTy": _n(r[16]), "inWarehouseQtyTy": _n(r[17]),
             "traitedStoreCountTy": _n(r[18])}
            for r in rows
        ],
        "total": int(total),
        "weeks": weeks,
    }


@router.get("/api/retail/item-performance")
def get_item_performance(
    division: str = None, customer: str = None,
    period_type: str = "weekly", week: str = None,
    limit: int = 100, offset: int = 0,
    sort_by: str = "pos_sales_ty", sort_dir: str = "desc",
):
    """Return item-level data for the Item Performance tab."""
    con = get_db()
    from core.hierarchy import hierarchy_filter
    hw, hp = hierarchy_filter(division=division, customer=customer or "walmart_stores")

    where = f"WHERE period_type = ? {hw}"
    params = [period_type] + hp
    if week:
        where += " AND walmart_week = ?"
        params.append(week)

    allowed_sorts = {
        "pos_sales_ty", "pos_qty_ty", "units_per_store_ty", "returns_qty_ty",
        "on_hand_qty_ty", "instock_pct_ty", "prime_item_number",
    }
    sb = sort_by if sort_by in allowed_sorts else "pos_sales_ty"
    sd = "ASC" if sort_dir.lower() == "asc" else "DESC"

    rows = con.execute(f"""
        SELECT walmart_week, prime_item_number, prime_item_desc, brand_name,
               sales_type, vendor_name,
               pos_sales_ty, pos_sales_ly, pos_qty_ty, pos_qty_ly,
               pos_store_count_ty, units_per_store_ty, units_per_store_ly,
               dollars_per_store_ty, dollars_per_store_ly,
               returns_qty_ty, returns_qty_ly,
               on_hand_qty_ty, on_hand_qty_ly,
               instock_pct_ty, instock_pct_ly,
               repl_instock_pct_ty, repl_instock_pct_ly,
               net_receipt_qty_ty, net_receipt_qty_ly
        FROM walmart_item_weekly
        {where}
        ORDER BY {sb} {sd}
        LIMIT ? OFFSET ?
    """, params + [limit, offset]).fetchall()

    total = con.execute(f"SELECT COUNT(*) FROM walmart_item_weekly {where}", params).fetchone()[0]
    con.close()

    return {
        "items": [
            {"week": r[0], "itemNumber": r[1], "itemDesc": r[2], "brand": r[3],
             "salesType": r[4], "vendor": r[5],
             "posSalesTy": _n(r[6]), "posSalesLy": _n(r[7]),
             "posQtyTy": _n(r[8]), "posQtyLy": _n(r[9]),
             "posStoreCountTy": _n(r[10]),
             "unitsPerStoreTy": _n(r[11]), "unitsPerStoreLy": _n(r[12]),
             "dollarsPerStoreTy": _n(r[13]), "dollarsPerStoreLy": _n(r[14]),
             "returnsQtyTy": _n(r[15]), "returnsQtyLy": _n(r[16]),
             "onHandQtyTy": _n(r[17]), "onHandQtyLy": _n(r[18]),
             "instockPctTy": _n(r[19]), "instockPctLy": _n(r[20]),
             "replInstockPctTy": _n(r[21]), "replInstockPctLy": _n(r[22]),
             "netReceiptQtyTy": _n(r[23]), "netReceiptQtyLy": _n(r[24])}
            for r in rows
        ],
        "total": int(total),
    }


@router.get("/api/retail/ecomm")
def get_ecomm(
    division: str = None, customer: str = None,
    week: str = None, limit: int = 100, offset: int = 0,
):
    """Return eComm sales data."""
    con = get_db()
    from core.hierarchy import hierarchy_filter
    hw, hp = hierarchy_filter(division=division, customer=customer or "walmart_stores")

    where = f"WHERE 1=1 {hw}"
    if week:
        where += " AND walmart_week = ?"
        hp.append(week)

    rows = con.execute(f"""
        SELECT walmart_week, all_links_item_number, product_name, brand_name,
               base_unit_retail_amount,
               auth_based_qty_ty, auth_based_qty_ly,
               auth_based_net_sales_ty, auth_based_net_sales_ly,
               shipped_based_qty_ty, shipped_based_qty_ly,
               shipped_based_net_sales_ty, shipped_based_net_sales_ly
        FROM walmart_ecomm_weekly
        {where}
        ORDER BY auth_based_net_sales_ty DESC
        LIMIT ? OFFSET ?
    """, hp + [limit, offset]).fetchall()

    total = con.execute(f"SELECT COUNT(*) FROM walmart_ecomm_weekly {where}", hp).fetchone()[0]
    con.close()

    return {
        "items": [
            {"week": r[0], "itemNumber": r[1], "productName": r[2], "brand": r[3],
             "retailAmount": _n(r[4]),
             "authQtyTy": _n(r[5]), "authQtyLy": _n(r[6]),
             "authSalesTy": _n(r[7]), "authSalesLy": _n(r[8]),
             "shippedQtyTy": _n(r[9]), "shippedQtyLy": _n(r[10]),
             "shippedSalesTy": _n(r[11]), "shippedSalesLy": _n(r[12])}
            for r in rows
        ],
        "total": int(total),
    }


@router.get("/api/retail/order-forecast")
def get_order_forecast(division: str = None, customer: str = None):
    """Return latest order forecast snapshot."""
    con = get_db()
    rows = con.execute("""
        SELECT snapshot_date, store_dc_nbr, store_dc_type, vendor_dept_number
        FROM walmart_order_forecast
        ORDER BY snapshot_date DESC, store_dc_nbr
        LIMIT 200
    """).fetchall()
    con.close()
    return {"forecast": [
        {"snapshotDate": str(r[0]), "storeDcNbr": r[1],
         "storeDcType": r[2], "vendorDeptNumber": r[3]}
        for r in rows
    ]}


@router.get("/api/retail/summary")
def get_retail_summary(division: str = None, customer: str = None):
    """Return high-level KPIs for the retail reporting overview."""
    con = get_db()
    from core.hierarchy import hierarchy_filter
    hw, hp = hierarchy_filter(division=division, customer=customer or "walmart_stores")

    # Total rows loaded
    counts = {}
    for tbl in ["walmart_store_weekly", "walmart_item_weekly", "walmart_scorecard",
                "walmart_ecomm_weekly", "walmart_order_forecast"]:
        try:
            r = con.execute(f"SELECT COUNT(*) FROM {tbl}").fetchone()
            counts[tbl] = int(r[0]) if r else 0
        except Exception:
            counts[tbl] = 0

    # Latest week with data
    latest_week = None
    try:
        r = con.execute("SELECT MAX(walmart_week) FROM walmart_store_weekly").fetchone()
        latest_week = r[0] if r else None
    except Exception:
        pass

    # Aggregate POS for latest week
    pos_summary = {"posSalesTy": 0, "posQtyTy": 0, "posSalesLy": 0, "posQtyLy": 0}
    if latest_week:
        try:
            r = con.execute(f"""
                SELECT SUM(pos_sales_ty), SUM(pos_qty_ty), SUM(pos_sales_ly), SUM(pos_qty_ly)
                FROM walmart_store_weekly
                WHERE walmart_week = ? {hw}
            """, [latest_week] + hp).fetchone()
            if r:
                pos_summary = {
                    "posSalesTy": _n(r[0]), "posQtyTy": _n(r[1]),
                    "posSalesLy": _n(r[2]), "posQtyLy": _n(r[3]),
                }
        except Exception:
            pass

    # Upload log
    uploads = []
    try:
        rows = con.execute("""
            SELECT upload_id, filename, report_type, rows_loaded, uploaded_at, status
            FROM retail_upload_log
            ORDER BY uploaded_at DESC
            LIMIT 20
        """).fetchall()
        uploads = [{"uploadId": r[0], "filename": r[1], "reportType": r[2],
                     "rowsLoaded": r[3], "uploadedAt": str(r[4]), "status": r[5]}
                    for r in rows]
    except Exception:
        pass

    con.close()
    return {
        "tableCounts": counts,
        "latestWeek": latest_week,
        "posSummary": pos_summary,
        "uploads": uploads,
    }


# ═══════════════════════════════════════════════════════════════════════
#  WALMART ANALYTICS COMPREHENSIVE ENDPOINT
# ═══════════════════════════════════════════════════════════════════════

@router.get("/api/retail/walmart-analytics")
def get_walmart_analytics(division: str = None, customer: str = None):
    """Return comprehensive data for the Walmart Analytics mockup page.

    This endpoint aggregates data from walmart_item_weekly, walmart_store_weekly,
    walmart_scorecard, walmart_ecomm_weekly, and walmart_order_forecast tables
    to support all 5 sections of the analytics page:
    1. Sales Performance (KPIs, sales by type, item-level, category breakdown)
    2. Inventory Health (KPIs, instock trend, OH per store, instock by type)
    3. Vendor Scorecard (grouped metrics)
    4. eCommerce (KPIs, item-level breakdown)
    5. Order Forecast (KPIs, forecasts)
    """
    con = get_db()
    from core.hierarchy import hierarchy_filter
    hw, hp = hierarchy_filter(division=division, customer=customer or "walmart_stores")

    # Determine if data is available (check multiple tables)
    has_item_data = False
    has_scorecard_data = False
    has_store_data = False
    try:
        total_items = con.execute(
            f"SELECT COUNT(*) FROM walmart_item_weekly WHERE 1=1 {hw}", hp
        ).fetchone()
        has_item_data = bool(total_items and total_items[0] > 0)
    except Exception:
        pass
    try:
        total_sc = con.execute(
            f"SELECT COUNT(*) FROM walmart_scorecard WHERE 1=1 {hw}", hp
        ).fetchone()
        has_scorecard_data = bool(total_sc and total_sc[0] > 0)
    except Exception:
        pass
    try:
        total_store = con.execute(
            f"SELECT COUNT(*) FROM walmart_store_weekly WHERE 1=1 {hw}", hp
        ).fetchone()
        has_store_data = bool(total_store and total_store[0] > 0)
    except Exception:
        pass
    data_available = has_item_data or has_scorecard_data or has_store_data

    # ── 1. SALES PERFORMANCE ─────────────────────────────────────────────

    sales_performance = {
        "kpis": {},
        "salesByType": {},
        "items": [],
        "categories": {},
    }

    if has_item_data:
        # KPIs by period type (L52W, L26W, L13W, L4W, L1W/LW)
        period_types = ["L52W", "L26W", "L13W", "L4W", "L1W"]
        for pt in period_types:
            ui_period = "L52W" if pt == "L52W" else "L26W" if pt == "L26W" else "L13W" if pt == "L13W" else "L4W" if pt == "L4W" else "LW"

            r = con.execute(f"""
                SELECT
                    SUM(pos_sales_ty), SUM(pos_sales_ly),
                    SUM(pos_qty_ty), SUM(pos_qty_ly),
                    SUM(CASE WHEN sales_type IN ('Regular', 'Rollback') THEN pos_sales_ty ELSE 0 END) as regular_ty,
                    SUM(CASE WHEN sales_type IN ('Regular', 'Rollback') THEN pos_sales_ly ELSE 0 END) as regular_ly,
                    SUM(CASE WHEN sales_type = 'Clearance' THEN pos_sales_ty ELSE 0 END) as clearance_ty,
                    SUM(CASE WHEN sales_type = 'Clearance' THEN pos_sales_ly ELSE 0 END) as clearance_ly
                FROM walmart_item_weekly
                WHERE period_type = ? AND 1=1 {hw}
            """, [pt] + hp).fetchone()

            if r:
                sales_performance["kpis"][ui_period] = {
                    "posSalesTy": _n(r[0]),
                    "posSalesLy": _n(r[1]),
                    "posQtyTy": _n(r[2]),
                    "posQtyLy": _n(r[3]),
                    "regularSalesTy": _n(r[4]),
                    "regularSalesLy": _n(r[5]),
                    "clearanceSalesTy": _n(r[6]),
                    "clearanceSalesLy": _n(r[7]),
                }

        # Sales by type breakdown for key periods
        for pt in ["L4W", "L13W", "L52W", "L1W"]:
            ui_period = "L52W" if pt == "L52W" else "L26W" if pt == "L26W" else "L13W" if pt == "L13W" else "L4W" if pt == "L4W" else "LW"

            rows = con.execute(f"""
                SELECT
                    COALESCE(sales_type, 'Other') as st,
                    SUM(pos_sales_ty) as sales_ty,
                    SUM(pos_sales_ly) as sales_ly,
                    SUM(pos_qty_ty) as qty_ty
                FROM walmart_item_weekly
                WHERE period_type = ? AND 1=1 {hw}
                GROUP BY COALESCE(sales_type, 'Other')
                ORDER BY sales_ty DESC
            """, [pt] + hp).fetchall()

            sales_performance["salesByType"][ui_period] = {}
            for row in rows:
                st = row[0] or "Other"
                sales_performance["salesByType"][ui_period][st] = {
                    "salesTy": _n(row[1]),
                    "salesLy": _n(row[2]),
                    "qtyTy": _n(row[3]),
                }

        # Item-level performance: per-period nested data for frontend
        # Frontend expects: { name, lw: {posTy, posLy, qtyTy, qtyLy, ohTy, ohLy, instockPct}, l4w: {...}, ... }
        period_map_items = {"L1W": "lw", "L4W": "l4w", "L13W": "l13w", "L26W": "l26w", "L52W": "l52w"}
        item_rows = con.execute(f"""
            SELECT
                prime_item_desc,
                period_type,
                SUM(pos_sales_ty) as pos_sales_ty,
                SUM(pos_sales_ly) as pos_sales_ly,
                SUM(pos_qty_ty) as pos_qty_ty,
                SUM(pos_qty_ly) as pos_qty_ly,
                SUM(on_hand_qty_ty) as oh_ty,
                SUM(on_hand_qty_ly) as oh_ly,
                AVG(instock_pct_ty) as instock_pct
            FROM walmart_item_weekly
            WHERE prime_item_desc IS NOT NULL AND 1=1 {hw}
            GROUP BY prime_item_desc, period_type
            ORDER BY prime_item_desc, period_type
        """, hp).fetchall()

        # Build nested item dict keyed by item name
        item_dict = {}
        for row in item_rows:
            name = row[0]
            pt = row[1]
            fe_key = period_map_items.get(pt)
            if not fe_key:
                continue
            if name not in item_dict:
                item_dict[name] = {"name": name}
            item_dict[name][fe_key] = {
                "posTy": _n(row[2]),
                "posLy": _n(row[3]),
                "qtyTy": _n(row[4]),
                "qtyLy": _n(row[5]),
                "ohTy": _n(row[6]),
                "ohLy": _n(row[7]),
                "instockPct": _n(row[8]),
            }

        # Sort by L4W posTy descending (fallback to LW), take top 50
        sorted_items = sorted(
            item_dict.values(),
            key=lambda x: (x.get("l4w", {}).get("posTy", 0) or 0),
            reverse=True
        )[:50]
        sales_performance["items"] = sorted_items

        # Category breakdown based on item descriptions
        category_keywords = {
            "Junior Clubs": ["junior", "youth", "kid"],
            "Tee-Up Practice": ["tee", "practice", "range", "training"],
            "Balls": ["ball", "golf ball"],
        }

        category_data = {}
        for item in sales_performance["items"]:
            desc_lower = (item.get("name") or "").lower()
            found_cat = "Other"
            for cat, keywords in category_keywords.items():
                if any(kw in desc_lower for kw in keywords):
                    found_cat = cat
                    break

            if found_cat not in category_data:
                category_data[found_cat] = {"salesTy": 0, "salesLy": 0, "qtyTy": 0, "count": 0}
            # Sum L4W data for category breakdown (fallback to LW)
            pd = item.get("l4w") or item.get("lw") or {}
            category_data[found_cat]["salesTy"] += pd.get("posTy", 0) or 0
            category_data[found_cat]["salesLy"] += pd.get("posLy", 0) or 0
            category_data[found_cat]["qtyTy"] += pd.get("qtyTy", 0) or 0
            category_data[found_cat]["count"] += 1

        sales_performance["categories"] = category_data

    # ── 2. INVENTORY HEALTH ──────────────────────────────────────────────

    inventory_health = {
        "kpis": {},
        "instockTrend": {},
        "ohPerStore": {},
        "instockByType": {},
    }

    if has_item_data:
        # Inventory KPIs
        r = con.execute(f"""
            SELECT
                SUM(on_hand_qty_ty),
                AVG(instock_pct_ty),
                SUM(on_hand_qty_ty * 0.5),
                COUNT(DISTINCT prime_item_desc)
            FROM walmart_item_weekly
            WHERE period_type = 'L4W' AND 1=1 {hw}
        """, hp).fetchone()

        if r:
            inventory_health["kpis"] = {
                "totalStoreOh": _n(r[0]),
                "instockPct": _n(r[1]) if r[1] else 0,
                "retailInventory": _n(r[2]),
                "weeksOfSupply": 4.0,
            }

        # Instock trend by item across windows (L52W → L1W)
        items_trend = con.execute(f"""
            SELECT DISTINCT prime_item_desc
            FROM walmart_item_weekly
            WHERE period_type IN ('L1W', 'L4W', 'L13W', 'L26W', 'L52W') AND 1=1 {hw}
            LIMIT 30
        """, hp).fetchall()

        for item_row in items_trend:
            item_desc = item_row[0]
            if not item_desc:
                continue

            trend_vals = []
            for pt in ["L52W", "L26W", "L13W", "L4W", "L1W"]:
                r = con.execute(f"""
                    SELECT AVG(instock_pct_ty)
                    FROM walmart_item_weekly
                    WHERE period_type = ? AND prime_item_desc = ? AND 1=1 {hw}
                """, [pt, item_desc] + hp).fetchone()
                trend_vals.append(_n(r[0]) if r and r[0] else 0)

            inventory_health["instockTrend"][item_desc] = trend_vals

        # OH per traited store by window
        for pt in ["L1W", "L4W", "L13W", "L52W"]:
            ui_period = "LW" if pt == "L1W" else pt

            rows = con.execute(f"""
                SELECT
                    prime_item_desc,
                    on_hand_qty_ty / NULLIF(traited_store_count_ty, 0) as oh_per_store_ty,
                    on_hand_qty_ly / NULLIF(traited_store_count_ly, 0) as oh_per_store_ly
                FROM walmart_item_weekly
                WHERE period_type = ? AND traited_store_count_ty > 0 AND 1=1 {hw}
                ORDER BY on_hand_qty_ty DESC
                LIMIT 20
            """, [pt] + hp).fetchall()

            if rows:
                inventory_health["ohPerStore"][ui_period] = {
                    "ty": [_n(r[1]) for r in rows],
                    "ly": [_n(r[2]) for r in rows],
                }

        # Instock by sales type
        for pt in ["L4W", "L13W"]:
            rows = con.execute(f"""
                SELECT
                    COALESCE(sales_type, 'Other') as st,
                    AVG(instock_pct_ty) as instock_avg
                FROM walmart_item_weekly
                WHERE period_type = ? AND 1=1 {hw}
                GROUP BY COALESCE(sales_type, 'Other')
            """, [pt] + hp).fetchall()

            inventory_health["instockByType"][pt] = {}
            for row in rows:
                inventory_health["instockByType"][pt][row[0]] = _n(row[1])

    # ── 3. VENDOR SCORECARD ──────────────────────────────────────────────

    scorecard_rows = con.execute(f"""
        SELECT
            vendor_section, metric_group, metric_name, period, period_range,
            value_ty, value_ly, value_diff
        FROM walmart_scorecard
        WHERE 1=1 {hw}
        ORDER BY vendor_section, metric_group, metric_name, period
    """, hp).fetchall()

    scorecard = [
        {
            "vendorSection": r[0],
            "metricGroup": r[1],
            "metricName": r[2],
            "period": r[3],
            "periodRange": r[4],
            "valueTy": _n(r[5]),
            "valueLy": _n(r[6]),
            "valueDiff": _n(r[7]),
        }
        for r in scorecard_rows
    ]

    # ── 4. eCOMMERCE ─────────────────────────────────────────────────────

    ecommerce = {
        "kpis": {},
        "items": [],
    }

    if has_item_data:
        # eComm KPIs
        r = con.execute(f"""
            SELECT
                SUM(auth_based_net_sales_ty),
                SUM(auth_based_net_sales_ly),
                SUM(shipped_based_net_sales_ty),
                SUM(shipped_based_net_sales_ly),
                COUNT(DISTINCT all_links_item_number)
            FROM walmart_ecomm_weekly
            WHERE 1=1 {hw}
        """, hp).fetchone()

        if r:
            ecommerce["kpis"] = {
                "authSalesTy": _n(r[0]),
                "authSalesLy": _n(r[1]),
                "shippedSalesTy": _n(r[2]),
                "shippedSalesLy": _n(r[3]),
                "totalItems": _safe_int(r[4]),
            }

        # Item-level eComm data
        ecomm_items = con.execute(f"""
            SELECT
                product_name,
                brand_name,
                base_unit_retail_amount,
                auth_based_net_sales_ty,
                auth_based_net_sales_ly,
                shipped_based_net_sales_ty,
                shipped_based_net_sales_ly
            FROM walmart_ecomm_weekly
            WHERE 1=1 {hw}
            ORDER BY auth_based_net_sales_ty DESC
            LIMIT 50
        """, hp).fetchall()

        ecommerce["items"] = [
            {
                "productName": r[0],
                "brand": r[1],
                "retailAmount": _n(r[2]),
                "authSalesTy": _n(r[3]),
                "authSalesLy": _n(r[4]),
                "shippedSalesTy": _n(r[5]),
                "shippedSalesLy": _n(r[6]),
            }
            for r in ecomm_items
        ]

    # ── 5. ORDER FORECAST ────────────────────────────────────────────────

    order_forecast = {
        "kpis": {},
        "items": [],
    }

    if has_item_data:
        # Latest snapshot
        r = con.execute("""
            SELECT COUNT(DISTINCT store_dc_nbr) as dc_count
            FROM walmart_order_forecast
            ORDER BY snapshot_date DESC
            LIMIT 1
        """).fetchone()

        if r:
            order_forecast["kpis"] = {
                "dcCount": _safe_int(r[0]),
            }

        # Forecast items
        forecast_items = con.execute("""
            SELECT DISTINCT snapshot_date, store_dc_nbr, store_dc_type
            FROM walmart_order_forecast
            ORDER BY snapshot_date DESC
            LIMIT 100
        """).fetchall()

        order_forecast["items"] = [
            {
                "snapshotDate": str(r[0]),
                "storeDcNbr": r[1],
                "storeDcType": r[2],
            }
            for r in forecast_items
        ]

    # ── Latest week for reference ────────────────────────────────────────
    latest_week = None
    try:
        r = con.execute("SELECT MAX(walmart_week) FROM walmart_item_weekly").fetchone()
        latest_week = r[0] if r else None
    except Exception:
        pass

    con.close()

    return {
        "salesPerformance": sales_performance,
        "inventoryHealth": inventory_health,
        "scorecard": scorecard,
        "ecommerce": ecommerce,
        "orderForecast": order_forecast,
        "latestWeek": latest_week,
        "dataAvailable": data_available,
    }


@router.post("/api/retail/import-scorecard")
async def import_scorecard_json(request_body: dict):
    """Import pre-parsed scorecard data as JSON.

    Accepts: { "rows": [ { "vendor_section", "metric_group", "metric_name",
        "period", "period_range", "value_ty", "value_ly", "value_diff" }, ... ] }
    This bypasses Excel parsing — useful when file upload has issues.
    """
    rows = request_body.get("rows", [])
    if not rows:
        raise HTTPException(400, "No rows provided")

    upload_id = str(uuid.uuid4())[:8]
    report_date = datetime.now(CT).date().isoformat()
    con = get_db_rw()
    count = 0

    for r in rows:
        try:
            con.execute("""
                INSERT OR IGNORE INTO walmart_scorecard
                    (vendor_section, metric_group, metric_name, period, period_range,
                     value_ty, value_ly, value_diff, report_date, upload_id,
                     division, customer, platform)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'golf', 'walmart_stores', 'scintilla')
            """, [
                str(r.get("vendor_section", ""))[:120],
                str(r.get("metric_group", ""))[:40],
                str(r.get("metric_name", ""))[:60],
                str(r.get("period", ""))[:20],
                str(r.get("period_range", ""))[:60] if r.get("period_range") else None,
                _n(r.get("value_ty")),
                _n(r.get("value_ly")),
                _n(r.get("value_diff")),
                report_date,
                upload_id,
            ])
            count += 1
        except Exception as e:
            logger.error(f"Scorecard JSON insert error: {e} | row={r}")

    con.close()
    _log_upload(upload_id, "json_import", "scorecard", count, "SUCCESS", None)

    return {
        "status": "ok",
        "upload_id": upload_id,
        "rows_loaded": count,
        "total_submitted": len(rows),
    }


@router.post("/api/retail/pull-from-drive")
async def pull_from_drive():
    """Pull latest Scintilla reports from Google Drive folder.

    Stub for now — will implement Google Drive integration later.
    """
    return {
        "status": "not_implemented",
        "message": "Google Drive integration coming soon. Use manual upload for now.",
    }


# ════════════════════════════════════════════════════════════════════════════
# NIF (New Item Forecast) Item Master — upload + query
# ════════════════════════════════════════════════════════════════════════════

_NIF_SECTION_KEYWORDS = {
    "new": ["NEW OMNI ITEMS", "NEW ITEMS"],
    "go_forward": ["GO FORWARD OMNI ITEMS", "GO FORWARD ITEMS"],
    "deleted": ["DELETED STORE ITEMS", "DELETED ITEMS"],
    "dotcom": ["GO FORWARD OWNED DOTCOM", "ECOMMERCE ITEMS", "1P OWNED ITEMS",
               "GO FORWARD DOTCOM"],
}


def _classify_section(text: str) -> str | None:
    """Return item_status from a section header row, or None if not a header."""
    if not text:
        return None
    t = text.strip().upper()
    for status, keywords in _NIF_SECTION_KEYWORDS.items():
        for kw in keywords:
            if kw in t:
                return status
    return None


def _find_fcst_extra_cols(ws) -> dict:
    """Scan row 11 headers to find DEXTERITY, CATEGORY, COLOR columns dynamically.
    These shift between NIF file versions (S24/S25 at AZ-BB, S26 at AY-BA)."""
    col_map = {}
    for c in range(37, 60):  # AK through ~BH
        val = ws.cell(11, c).value
        if not val:
            continue
        val_upper = str(val).strip().upper()
        if "DEXTERITY" in val_upper:
            col_map["dexterity"] = c
        elif "CATEGORY" in val_upper:
            col_map["category"] = c
        elif "COLOR" in val_upper:
            col_map["color"] = c
    return col_map


def _parse_nif_fcst(ws, event_year: int) -> list[dict]:
    """Parse FCST sheet → list of item dicts with status classification."""
    items = []
    current_status = "new"  # default until first section header

    # Find Dexterity/Category/Color column positions from row 11 headers
    extra_cols = _find_fcst_extra_cols(ws)

    # Data starts at row 12 (row 11 is headers)
    for r in range(12, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        c3 = ws.cell(r, 3).value  # Description

        # Check if this row is a section header
        section = _classify_section(str(c1) if c1 else "")
        if section:
            current_status = section
            continue

        # Skip empty rows or rows without description
        if not c3 or not str(c3).strip():
            continue

        # Skip if c1 looks like another header (all caps, no data cols)
        c7 = ws.cell(r, 7).value  # UPC
        c8 = ws.cell(r, 8).value  # VSN
        if not c7 and not c8 and not ws.cell(r, 15).value:
            continue

        brand = _safe_str(c1, 80) or ""
        # Clean brand — remove leading category prefix like "PGA TOUR - "
        # but keep it as-is since user wants to see the full brand

        item = {
            "event_year": event_year,
            "item_status": current_status,
            "brand": brand,
            "mod_type": _safe_str(ws.cell(r, 2).value, 40),
            "description": _safe_str(c3, 200),
            "wmt_item_number": _safe_str(ws.cell(r, 4).value, 20),
            "upc": _safe_str(c7, 20),
            "vendor_stock_number": _safe_str(c8, 40),
            "brand_id": _safe_str(ws.cell(r, 9).value, 20),
            "wholesale_cost": _n(ws.cell(r, 15).value),  # WM First Cost Collect Suppliers
            "walmart_retail": _n(ws.cell(r, 21).value),  # Unit Retail SRP
            "vendor_pack": _safe_int(ws.cell(r, 23).value),
            "whse_pack": _safe_int(ws.cell(r, 24).value),
            "old_store_count": _safe_int(ws.cell(r, 25).value),
            "new_store_count": _safe_int(ws.cell(r, 26).value),
            "store_count_diff": _safe_int(ws.cell(r, 27).value),
        }

        # Read Dexterity/Category/Color from FCST columns (found dynamically)
        if "dexterity" in extra_cols:
            dex = _safe_str(ws.cell(r, extra_cols["dexterity"]).value, 60)
            if dex and dex.upper() not in ("NA", "N/A"):
                item["dexterity"] = dex
        if "category" in extra_cols:
            cat = _safe_str(ws.cell(r, extra_cols["category"]).value, 80)
            if cat and cat.upper() not in ("NA", "N/A"):
                item["category"] = cat
        if "color" in extra_cols:
            col = _safe_str(ws.cell(r, extra_cols["color"]).value, 60)
            if col and col.upper() not in ("NA", "N/A"):
                item["color"] = col

        items.append(item)

    return items


def _enrich_dimensions(items: list[dict], ws_setup) -> None:
    """Merge carton dimensions from ITEM SET UP sheet into items (by VSN match)."""
    # Build VSN → dimensions map from ITEM SET UP
    dim_map = {}
    for r in range(3, ws_setup.max_row + 1):
        vsn = ws_setup.cell(r, 3).value  # col 3 = VSN in ITEM SET UP
        if not vsn:
            continue
        vsn = str(vsn).strip()
        length = _n(ws_setup.cell(r, 26).value)
        width = _n(ws_setup.cell(r, 27).value)
        height = _n(ws_setup.cell(r, 28).value)
        if length or width or height:
            # CBM = L*W*H in meters (convert from inches: /39.37 each)
            l_m = length / 39.37 if length else 0
            w_m = width / 39.37 if width else 0
            h_m = height / 39.37 if height else 0
            cbm = l_m * w_m * h_m
            cbf = (length * width * height) / 1728 if (length and width and height) else 0
            dim_map[vsn] = {
                "carton_length": length,
                "carton_width": width,
                "carton_height": height,
                "cbm": round(cbm, 4),
                "cbf": round(cbf, 4),
            }

    for item in items:
        vsn = item.get("vendor_stock_number", "")
        if vsn and vsn in dim_map:
            item.update(dim_map[vsn])


def _enrich_color_dexterity(items: list[dict], ws_commit) -> None:
    """Merge color/dexterity from COMMIT sheet into items (by VSN match)."""
    # Build VSN → color/dex map
    cd_map = {}
    for r in range(3, ws_commit.max_row + 1):
        vsn = ws_commit.cell(r, 17).value  # col 17 = Vendor Stk Nbr
        if not vsn:
            continue
        vsn = str(vsn).strip()
        color = _safe_str(ws_commit.cell(r, 18).value, 60)
        dex = _safe_str(ws_commit.cell(r, 19).value, 60)
        if color and color.upper() != "NA":
            cd_map.setdefault(vsn, {})["color"] = color
        if dex and dex.upper() != "NA":
            cd_map.setdefault(vsn, {})["dexterity"] = dex

    for item in items:
        vsn = item.get("vendor_stock_number", "")
        if vsn and vsn in cd_map:
            if "color" in cd_map[vsn] and not item.get("color"):
                item["color"] = cd_map[vsn]["color"]
            if "dexterity" in cd_map[vsn] and not item.get("dexterity"):
                item["dexterity"] = cd_map[vsn]["dexterity"]


def _detect_event_year(ws) -> int | None:
    """Read event year from cell B3 of FCST sheet."""
    val = ws.cell(3, 2).value
    if val:
        try:
            return int(float(val))
        except (TypeError, ValueError):
            pass
    return None


def _detect_effective_week(ws, filename: str = "") -> int:
    """Read effective week — check filename first (e.g. 'WK4'), then IN-STORE row."""
    import re as _re
    # Check filename first (more reliable — e.g. "GG NIF S26 MOD WK4.xlsx")
    if filename:
        m = _re.search(r'WK\s*(\d+)', filename, _re.IGNORECASE)
        if m:
            return int(m.group(1))
    # Fallback: IN-STORE row (row 8) of FCST sheet
    val = ws.cell(8, 2).value
    if val and "WK" in str(val).upper():
        m = _re.search(r'WK\s*(\d+)', str(val), _re.IGNORECASE)
        if m:
            return int(m.group(1))
    return 7  # default


@router.post("/api/retail/upload-nif")
async def upload_nif(file: UploadFile = File(...)):
    """Parse a NIF (New Item Forecast) Excel file and load into walmart_nif_items."""
    import openpyxl

    try:
        data = await file.read()
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True, keep_links=False)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot read Excel file: {e}")

    if "FCST" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="No FCST sheet found — not a NIF file")

    ws_fcst = wb["FCST"]
    event_year = _detect_event_year(ws_fcst)
    if not event_year:
        raise HTTPException(status_code=400, detail="Cannot detect Event year from cell B3")

    effective_week = _detect_effective_week(ws_fcst, file.filename or "")

    # Parse FCST sheet
    items = _parse_nif_fcst(ws_fcst, event_year)
    if not items:
        raise HTTPException(status_code=400, detail="No items found in FCST sheet")

    # Set effective week on all items
    for item in items:
        item["effective_week"] = effective_week

    # Enrich with carton dimensions from ITEM SET UP
    if "ITEM SET UP" in wb.sheetnames:
        _enrich_dimensions(items, wb["ITEM SET UP"])

    # Fallback: enrich color/dexterity from COMMIT sheet if FCST didn't have them
    if "COMMIT" in wb.sheetnames:
        _enrich_color_dexterity(items, wb["COMMIT"])

    # Fallback: derive category from brand only if FCST didn't provide one
    for item in items:
        if not item.get("category"):
            brand = (item.get("brand") or "").upper()
            if "LPGA" in brand:
                item["category"] = "LPGA"
            elif "G2 SERIES" in brand:
                item["category"] = "G2 Series"
            elif "G1 SERIES" in brand:
                item["category"] = "G1 Series"
            elif "TEE UP" in brand:
                item["category"] = "Tee Up"
            else:
                item["category"] = "Other"

    # Insert into database — clear existing data for this event year first
    con = get_db_rw()
    try:
        con.execute("DELETE FROM walmart_nif_items WHERE event_year = ?", [event_year])

        for item in items:
            con.execute("""
                INSERT INTO walmart_nif_items (
                    event_year, effective_week, item_status, description, brand,
                    wmt_item_number, upc, vendor_stock_number, brand_id,
                    wholesale_cost, walmart_retail, vendor_pack, whse_pack,
                    old_store_count, new_store_count, store_count_diff,
                    carton_length, carton_width, carton_height, cbm, cbf,
                    color, dexterity, category, mod_type,
                    division, customer, platform
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'golf', 'walmart_stores', 'scintilla')
            """, [
                item.get("event_year"), item.get("effective_week"),
                item.get("item_status"), item.get("description"), item.get("brand"),
                item.get("wmt_item_number"), item.get("upc"),
                item.get("vendor_stock_number"), item.get("brand_id"),
                item.get("wholesale_cost", 0), item.get("walmart_retail", 0),
                item.get("vendor_pack", 0), item.get("whse_pack", 0),
                item.get("old_store_count", 0), item.get("new_store_count", 0),
                item.get("store_count_diff", 0),
                item.get("carton_length", 0), item.get("carton_width", 0),
                item.get("carton_height", 0), item.get("cbm", 0), item.get("cbf", 0),
                item.get("color"), item.get("dexterity"),
                item.get("category"), item.get("mod_type"),
            ])
    finally:
        con.close()

    logger.info(f"NIF upload: {file.filename} → {len(items)} items for {event_year}")
    return {
        "status": "success",
        "event_year": event_year,
        "effective_week": effective_week,
        "items_loaded": len(items),
        "filename": file.filename,
        "breakdown": {
            "new": sum(1 for i in items if i["item_status"] == "new"),
            "go_forward": sum(1 for i in items if i["item_status"] == "go_forward"),
            "deleted": sum(1 for i in items if i["item_status"] == "deleted"),
            "dotcom": sum(1 for i in items if i["item_status"] == "dotcom"),
        },
    }


@router.get("/api/retail/nif-items")
async def get_nif_items(year: int = Query(None)):
    """Return NIF item master data, optionally filtered by year."""
    con = get_db()
    try:
        if year:
            rows = con.execute("""
                SELECT id, event_year, effective_week, item_status, description, brand,
                       wmt_item_number, upc, vendor_stock_number, brand_id,
                       wholesale_cost, walmart_retail, vendor_pack, whse_pack,
                       old_store_count, new_store_count, store_count_diff,
                       carton_length, carton_width, carton_height, cbm, cbf,
                       color, dexterity, category, mod_type
                FROM walmart_nif_items
                WHERE event_year = ?
                ORDER BY
                    CASE item_status
                        WHEN 'new' THEN 1
                        WHEN 'go_forward' THEN 2
                        WHEN 'dotcom' THEN 3
                        WHEN 'deleted' THEN 4
                    END,
                    brand, description
            """, [year]).fetchall()
        else:
            rows = con.execute("""
                SELECT id, event_year, effective_week, item_status, description, brand,
                       wmt_item_number, upc, vendor_stock_number, brand_id,
                       wholesale_cost, walmart_retail, vendor_pack, whse_pack,
                       old_store_count, new_store_count, store_count_diff,
                       carton_length, carton_width, carton_height, cbm, cbf,
                       color, dexterity, category, mod_type
                FROM walmart_nif_items
                ORDER BY event_year DESC,
                    CASE item_status
                        WHEN 'new' THEN 1
                        WHEN 'go_forward' THEN 2
                        WHEN 'dotcom' THEN 3
                        WHEN 'deleted' THEN 4
                    END,
                    brand, description
            """).fetchall()
    finally:
        con.close()

    items = []
    for r in rows:
        items.append({
            "id": r[0],
            "eventYear": r[1],
            "effectiveWeek": r[2],
            "itemStatus": r[3],
            "description": r[4],
            "brand": r[5],
            "wmtItemNumber": r[6],
            "upc": r[7],
            "vendorStockNumber": r[8],
            "brandId": r[9],
            "wholesaleCost": _n(r[10]),
            "walmartRetail": _n(r[11]),
            "vendorPack": _safe_int(r[12]),
            "whsePack": _safe_int(r[13]),
            "oldStoreCount": _safe_int(r[14]),
            "newStoreCount": _safe_int(r[15]),
            "storeCountDiff": _safe_int(r[16]),
            "cartonLength": _n(r[17]),
            "cartonWidth": _n(r[18]),
            "cartonHeight": _n(r[19]),
            "cbm": _n(r[20]),
            "cbf": _n(r[21]),
            "color": r[22],
            "dexterity": r[23],
            "category": r[24],
            "modType": r[25],
        })

    # Get available years
    con2 = get_db()
    try:
        year_rows = con2.execute(
            "SELECT DISTINCT event_year FROM walmart_nif_items ORDER BY event_year"
        ).fetchall()
    finally:
        con2.close()

    return {
        "items": items,
        "availableYears": [r[0] for r in year_rows],
        "totalItems": len(items),
    }


# ── NIF Item CRUD — update, delete, add ─────────────────────────────

@router.put("/api/retail/nif-items/{item_id}")
async def update_nif_item(item_id: int, body: dict = Body(...)):
    """Update a single NIF item field."""
    updatable = [
        "description", "brand", "wmt_item_number", "upc", "vendor_stock_number",
        "brand_id", "wholesale_cost", "walmart_retail", "vendor_pack", "whse_pack",
        "old_store_count", "new_store_count", "store_count_diff",
        "carton_length", "carton_width", "carton_height", "cbm", "cbf",
        "item_status", "color", "dexterity", "category",
    ]
    sets, vals = [], []
    for k, v in body.items():
        if k in updatable:
            sets.append(f"{k} = %s")
            vals.append(v)
    if not sets:
        raise HTTPException(status_code=400, detail="No updatable fields provided")
    vals.append(item_id)
    con = get_db_rw()
    try:
        con.execute(f"UPDATE walmart_nif_items SET {', '.join(sets)} WHERE id = %s", vals)
    finally:
        con.close()
    return {"status": "ok", "id": item_id}


@router.delete("/api/retail/nif-items/{item_id}")
async def delete_nif_item(item_id: int):
    """Delete a single NIF item."""
    con = get_db_rw()
    try:
        cur = con.execute("DELETE FROM walmart_nif_items WHERE id = %s RETURNING id", [item_id])
        deleted = cur.fetchone()
        if not deleted:
            raise HTTPException(status_code=404, detail=f"NIF item {item_id} not found")
    finally:
        con.close()
    return {"status": "ok", "id": item_id}


@router.post("/api/retail/nif-items")
async def add_nif_item(body: dict = Body(...)):
    """Add a new NIF item."""
    event_year = body.get("event_year") or body.get("eventYear")
    if not event_year:
        raise HTTPException(status_code=400, detail="event_year is required")
    con = get_db_rw()
    try:
        con.execute("""
            INSERT INTO walmart_nif_items (
                event_year, effective_week, item_status, description, brand,
                wmt_item_number, upc, vendor_stock_number, brand_id,
                wholesale_cost, walmart_retail, vendor_pack, whse_pack,
                old_store_count, new_store_count, store_count_diff,
                carton_length, carton_width, carton_height, cbm, cbf,
                color, dexterity, category, division, customer, platform
            ) VALUES (
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, 'golf', 'walmart_stores', 'scintilla'
            )
        """, [
            int(event_year),
            int(body.get("effective_week") or body.get("effectiveWeek") or 0),
            body.get("item_status") or body.get("itemStatus") or "new",
            body.get("description") or "",
            body.get("brand") or "",
            body.get("wmt_item_number") or body.get("wmtItemNumber") or "",
            body.get("upc") or "",
            body.get("vendor_stock_number") or body.get("vendorStockNumber") or "",
            body.get("brand_id") or body.get("brandId") or "",
            float(body.get("wholesale_cost") or body.get("wholesaleCost") or 0),
            float(body.get("walmart_retail") or body.get("walmartRetail") or 0),
            int(body.get("vendor_pack") or body.get("vendorPack") or 0),
            int(body.get("whse_pack") or body.get("whsePack") or 0),
            int(body.get("old_store_count") or body.get("oldStoreCount") or 0),
            int(body.get("new_store_count") or body.get("newStoreCount") or 0),
            int(body.get("store_count_diff") or body.get("storeCountDiff") or 0),
            float(body.get("carton_length") or body.get("cartonLength") or 0),
            float(body.get("carton_width") or body.get("cartonWidth") or 0),
            float(body.get("carton_height") or body.get("cartonHeight") or 0),
            float(body.get("cbm") or 0),
            float(body.get("cbf") or 0),
            body.get("color") or "",
            body.get("dexterity") or "",
            body.get("category") or "",
        ])
    finally:
        con.close()
    return {"status": "ok", "eventYear": int(event_year)}
