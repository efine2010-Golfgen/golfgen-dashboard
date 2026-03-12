"""Factory PO routes."""
import json
import logging
from pathlib import Path
from datetime import datetime
from zoneinfo import ZoneInfo
from fastapi import APIRouter, Request, UploadFile, File, HTTPException

from core.config import DB_DIR, TIMEZONE
from core.auth import require_auth

logger = logging.getLogger("golfgen")
router = APIRouter()


def _load_factory_po():
    fp = DB_DIR / "factory_po_summary.json"
    if fp.exists():
        with open(fp) as f:
            return json.load(f)
    return {"purchaseOrders": [], "unitsByItem": [], "arrivalSchedule": [], "lastUpload": None}

def _save_factory_po(data):
    fp = DB_DIR / "factory_po_summary.json"
    with open(fp, "w") as f:
        json.dump(data, f, indent=2)

def _load_logistics():
    fp = DB_DIR / "logistics_tracking.json"
    if fp.exists():
        with open(fp) as f:
            return json.load(f)
    return {"shipments": [], "itemsByContainer": [], "lastUpload": None}

@router.get("/api/debug/factory-po-test")
async def debug_factory_po_test():
    """Debug endpoint — no auth, returns factory PO data or error details."""
    try:
        data = _load_factory_po()
        return {"ok": True, "keys": list(data.keys()), "poCount": len(data.get("purchaseOrders", []))}
    except Exception as e:
        import traceback
        return {"ok": False, "error": str(e), "trace": traceback.format_exc()}

@router.get("/api/factory-po")
async def get_factory_po(request: Request):
    require_auth(request)
    data = _load_factory_po()
    # Clean bad rows from old parser output at read time
    if data.get("purchaseOrders"):
        data["purchaseOrders"] = [
            p for p in data["purchaseOrders"]
            if p.get("factory") and not str(p["factory"]).startswith("T-")
            and str(p.get("factory", "")).upper() not in ("TOTAL", "TOTALS", "PO#", "")
            and not ("." in str(p.get("poNumber", "")) and len(str(p.get("poNumber", ""))) > 10)
            and (p.get("units") or 0) > 0
        ]
    if data.get("unitsByItem"):
        data["unitsByItem"] = [
            u for u in data["unitsByItem"]
            if u.get("sku") and str(u["sku"]).upper() not in ("TOTAL", "TOTALS", "SKU", "")
        ]
    if data.get("arrivalSchedule"):
        data["arrivalSchedule"] = [
            a for a in data["arrivalSchedule"]
            if a.get("sku") and str(a["sku"]).upper() not in ("TOTAL", "TOTALS", "SKU", "")
        ]

    # ── Enrich POs with container numbers + status from logistics data ──
    logistics = _load_logistics()
    shipments = logistics.get("shipments", [])
    items_by_container = logistics.get("itemsByContainer", [])

    # Build PO → containers mapping from shipments
    po_containers: dict[str, list[dict]] = {}  # po -> [{containerNumber, status, eta}]
    for s in shipments:
        raw_po = (s.get("poNumber") or "").strip()
        cn = (s.get("containerNumber") or "").strip()
        if not raw_po or not cn:
            continue
        # PO field may contain multiple POs separated by / or ,
        for po_part in raw_po.replace(",", "/").split("/"):
            po_part = po_part.strip()
            if not po_part:
                continue
            po_containers.setdefault(po_part, [])
            # Avoid duplicate containers per PO
            existing_cns = [c["containerNumber"] for c in po_containers[po_part]]
            if cn not in existing_cns:
                po_containers[po_part].append({
                    "containerNumber": cn,
                    "status": s.get("status") or "",
                    "eta": s.get("etaFinal") or s.get("etaDischarge") or "",
                })

    # Enrich each PO with its containers
    for po in data.get("purchaseOrders", []):
        po_num = (po.get("poNumber") or "").strip()
        containers = po_containers.get(po_num, [])
        po["containers"] = containers
        # Derive overall status: if any In Transit → In Transit, all Delivered → Delivered
        statuses = [c["status"] for c in containers if c["status"]]
        if statuses:
            if any("transit" in st.lower() for st in statuses):
                po["containerStatus"] = "In Transit"
            elif any("deliver" in st.lower() for st in statuses):
                po["containerStatus"] = "Delivered"
            elif any("pending" in st.lower() for st in statuses):
                po["containerStatus"] = "Pending"
            else:
                po["containerStatus"] = statuses[0]
        else:
            po["containerStatus"] = ""

    # Build container → status lookup for items
    container_status = {}
    for s in shipments:
        cn = (s.get("containerNumber") or "").strip()
        if cn:
            container_status[cn] = s.get("status") or ""

    # Enrich items by container with status
    for item in items_by_container:
        cn = (item.get("containerNumber") or "").strip()
        item["status"] = container_status.get(cn, "")

    # Also attach items by container to the response for cross-reference
    data["logisticsItems"] = items_by_container

    # ── Enrich unitsByItem with container # and arrival date ──
    # Build SKU → [{containerNumber, eta, qty}] from logistics items
    sku_containers: dict[str, list[dict]] = {}
    for item in items_by_container:
        sku = (item.get("sku") or "").strip()
        cn = (item.get("containerNumber") or "").strip()
        if not sku or not cn:
            continue
        sku_containers.setdefault(sku, [])
        sku_containers[sku].append({
            "containerNumber": cn,
            "eta": item.get("eta") or "",
            "qty": item.get("qty") or 0,
        })

    # Also build container → arrival date from shipments
    container_arrival: dict[str, str] = {}
    for s in shipments:
        cn = (s.get("containerNumber") or "").strip()
        if cn:
            container_arrival[cn] = s.get("deliveryDate") or s.get("etaFinal") or s.get("etaDischarge") or ""

    # Expand unitsByItem: if a SKU has multiple containers, create one row per container
    expanded_units = []
    for u in data.get("unitsByItem", []):
        sku = (u.get("sku") or "").strip()
        containers = sku_containers.get(sku, [])
        if containers:
            for c in containers:
                row = dict(u)
                row["containerNumber"] = c["containerNumber"]
                row["containerQty"] = c["qty"]
                row["arrivalDate"] = container_arrival.get(c["containerNumber"], c["eta"])
                expanded_units.append(row)
        else:
            row = dict(u)
            row["containerNumber"] = ""
            row["containerQty"] = 0
            row["arrivalDate"] = ""
            expanded_units.append(row)
    data["unitsByItem"] = expanded_units

    return data

@router.post("/api/factory-po/upload")
async def upload_factory_po(request: Request, file: UploadFile = File(...)):
    require_auth(request)
    import openpyxl
    from io import BytesIO
    contents = await file.read()
    wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
    if "Factory PO Summary" not in wb.sheetnames:
        raise HTTPException(400, "Sheet 'Factory PO Summary' not found")
    ws = wb["Factory PO Summary"]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        r = {}
        for c in row:
            if hasattr(c, 'column_letter') and c.value is not None:
                r[c.column_letter] = c.value
        rows.append(r)
    def _to_str(dt):
        from datetime import datetime as _dt, date as _d
        if isinstance(dt, (_dt, _d)):
            return dt.strftime("%Y-%m-%d")
        return str(dt) if dt else None
    def _sr(v):
        try: return round(float(v), 2)
        except: return 0
    pos = []
    for i in range(7, len(rows)):
        cells = rows[i]
        factory = cells.get("B")
        if not factory or factory == "TOTALS":
            break
        pos.append({
            "factory": str(factory), "paymentTerms": str(cells.get("D") or ""),
            "poNumber": str(cells.get("E") or ""), "units": cells.get("F") or 0,
            "totalCost": _sr(cells.get("G")), "fobDate": _to_str(cells.get("I")),
            "estArrival": _to_str(cells.get("J")), "cbm": _sr(cells.get("K")),
            "landedCost": _sr(cells.get("U")), "oceanFreight": _sr(cells.get("X")),
            "customs": _sr(cells.get("Y"))
        })
    sku_start = next((i for i, r in enumerate(rows) if "UNITS BY ITEM" in str(r.get("B","")) and "PURCHASE ORDER" in str(r.get("B",""))), None)
    sku_rows = []
    if sku_start:
        for i in range(sku_start + 2, len(rows)):
            cells = rows[i]
            sku = cells.get("B")
            if not sku or sku == "TOTAL" or str(sku).startswith("UNITS BY"):
                break
            sku_rows.append({"sku": str(sku), "description": str(cells.get("C") or ""),
                "byPO": {"T-857500": cells.get("D") or 0, "2100": cells.get("E") or 0,
                    "T-857600": cells.get("F") or 0, "T-870100": cells.get("G") or 0,
                    "T-857700": cells.get("I") or 0, "T-857800": cells.get("J") or 0,
                    "T-857900": cells.get("K") or 0, "T-870200": cells.get("L") or 0,
                    "T-870300": cells.get("M") or 0, "T-870400": cells.get("N") or 0,
                    "T-870500": cells.get("O") or 0},
                "total": cells.get("P") or 0})
    arr_start = next((i for i, r in enumerate(rows) if "EST. ARRIVAL" in str(r.get("B",""))), None)
    arrival_rows = []
    if arr_start:
        for i in range(arr_start + 2, len(rows)):
            cells = rows[i]
            sku = cells.get("B")
            if not sku or sku == "TOTAL": break
            arrival_rows.append({"sku": str(sku), "description": str(cells.get("C") or ""),
                "jan2026": cells.get("D") or 0, "feb2026": cells.get("E") or 0,
                "mar2026": cells.get("F") or 0, "may2026": cells.get("G") or 0,
                "jun2026": cells.get("H") or 0, "total": cells.get("I") or 0})
    from datetime import date as _date
    data = {
        "rateInputs": {"usdRmbRate": 7.1, "customsBrokerFee": 250, "htsCode": "9506.31.0080", "htsDutyRate": 0.046},
        "purchaseOrders": pos, "unitsByItem": sku_rows, "arrivalSchedule": arrival_rows,
        "lastUpload": _date.today().isoformat(), "sourceFile": file.filename
    }
    _save_factory_po(data)
    return {"status": "ok", "pos": len(pos), "skus": len(sku_rows), "arrivals": len(arrival_rows), "lastUpload": data["lastUpload"]}
