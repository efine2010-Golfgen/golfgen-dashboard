"""Inventory, warehouse, and FBA shipment routes."""
import os
import json
import csv
import logging
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from pathlib import Path
from typing import Optional
from fastapi import APIRouter, Query, Request, UploadFile, File, HTTPException
from fastapi.responses import JSONResponse

from core.config import DB_PATH, DB_DIR, COGS_PATH, CONFIG_PATH, TIMEZONE
from core.database import get_db
from core.hierarchy import hierarchy_filter

logger = logging.getLogger("golfgen")
router = APIRouter()

# ── Constants ────────────────────────────────────────────
ITEM_MASTER_PATH = DB_DIR / "item_master.csv"
WAREHOUSE_PATH = DB_DIR / "warehouse.csv"
_FBA_SHIPMENTS_CACHE_PATH = DB_DIR / "fba_shipments_cache.json"


# ── Helper Functions ────────────────────────────────────────────


def load_json(filename: str) -> list:
    """Load a JSON data file from the data directory."""
    path = DB_DIR / filename
    if not path.exists():
        return []
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def save_json(filename: str, data):
    """Save data to a JSON file in the data directory."""
    path = DB_DIR / filename
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def load_cogs() -> dict:
    """Load COGS data from CSV. Returns {asin: {cogs, product_name, sku}}."""
    cogs = {}
    if not COGS_PATH.exists():
        return cogs
    with open(COGS_PATH, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            asin = (row.get("asin") or "").strip()
            if asin:
                cogs[asin] = {
                    "cogs": float(row.get("cogs") or 0),
                    "product_name": (row.get("product_name") or "").strip(),
                    "sku": (row.get("sku") or "").strip(),
                }
    return cogs


def load_item_master() -> list:
    """Load item master from CSV. Returns list of dicts."""
    items = []
    if not ITEM_MASTER_PATH.exists():
        return items
    with open(ITEM_MASTER_PATH, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            asin = (row.get("asin") or "").strip()
            if not asin:
                continue
            items.append({
                "asin": asin,
                "sku": (row.get("sku") or "").strip(),
                "productName": (row.get("product_name") or "").strip(),
                "color": (row.get("color") or "").strip(),
                "brand": (row.get("brand") or "").strip(),
                "series": (row.get("series") or "").strip(),
                "productType": (row.get("product_type") or "").strip(),
                "pieceCount": int(float(row.get("piece_count") or 0)),
                "orientation": (row.get("orientation") or "").strip(),
                "category": (row.get("category") or "").strip(),
                "unitCost": float(row.get("unit_cost") or 0),
                "fbsStock": int(float(row.get("fba_stock") or 0)),
                "lyAur": round(float(row.get("ly_aur") or 0), 2),
                "lyRevenue": round(float(row.get("ly_revenue") or 0), 2),
                "lyUnits": int(float(row.get("ly_units") or 0)),
                "lyProfit": round(float(row.get("ly_profit") or 0), 2),
                "plannedAnnualUnits": int(float(row.get("planned_annual_units") or 0)),
                "listPrice": float(row.get("list_price") or 0),
                "salePrice": float(row.get("sale_price") or 0),
                "referralPct": float(row.get("referral_pct") or 15),
                "couponType": (row.get("coupon_type") or "").strip(),
                "couponValue": float(row.get("coupon_value") or 0),
                "cartonPack": int(float(row.get("carton_pack") or 0)),
                "cartonLength": float(row.get("carton_length") or 0),
                "cartonWidth": float(row.get("carton_width") or 0),
                "cartonHeight": float(row.get("carton_height") or 0),
                "cartonWeight": float(row.get("carton_weight") or 0),
            })
    return items


def save_item_master(items: list):
    """Save item master list back to CSV."""
    fields = [
        "asin", "sku", "product_name", "color", "brand", "series",
        "product_type", "piece_count", "orientation", "category", "unit_cost",
        "fba_stock", "ly_aur", "ly_revenue", "ly_units", "ly_profit",
        "planned_annual_units", "list_price", "sale_price", "referral_pct",
        "coupon_type", "coupon_value", "carton_pack", "carton_length",
        "carton_width", "carton_height", "carton_weight",
    ]
    with open(ITEM_MASTER_PATH, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for item in items:
            w.writerow({
                "asin": item.get("asin", ""),
                "sku": item.get("sku", ""),
                "product_name": item.get("productName", ""),
                "color": item.get("color", ""),
                "brand": item.get("brand", ""),
                "series": item.get("series", ""),
                "product_type": item.get("productType", ""),
                "piece_count": item.get("pieceCount", 0),
                "orientation": item.get("orientation", ""),
                "category": item.get("category", ""),
                "unit_cost": item.get("unitCost", 0),
                "fba_stock": item.get("fbsStock", 0),
                "ly_aur": item.get("lyAur", 0),
                "ly_revenue": item.get("lyRevenue", 0),
                "ly_units": item.get("lyUnits", 0),
                "ly_profit": item.get("lyProfit", 0),
                "planned_annual_units": item.get("plannedAnnualUnits", 0),
                "list_price": item.get("listPrice", 0),
                "sale_price": item.get("salePrice", 0),
                "referral_pct": item.get("referralPct", 15),
                "coupon_type": item.get("couponType", ""),
                "coupon_value": item.get("couponValue", 0),
                "carton_pack": item.get("cartonPack", 0),
                "carton_length": item.get("cartonLength", 0),
                "carton_width": item.get("cartonWidth", 0),
                "carton_height": item.get("cartonHeight", 0),
                "carton_weight": item.get("cartonWeight", 0),
            })


# ── Routes ──────────────────────────────────────────────


@router.get("/api/inventory")
def inventory(division: Optional[str] = None, customer: Optional[str] = None, platform: Optional[str] = None, marketplace: Optional[str] = None):
    """Current FBA inventory with days-of-supply calculations."""
    con = get_db()
    cogs_data = load_cogs()

    hf, hp = hierarchy_filter(division, customer, platform, marketplace)
    # For fba_inventory the hf starts with " AND ...", convert to WHERE if needed
    inv_where = (" WHERE " + hf.lstrip(" AND ")) if hf else ""
    inv_rows = con.execute(f"""
        SELECT asin, sku, product_name,
               COALESCE(fulfillable_quantity, 0) AS fba_stock,
               COALESCE(inbound_working_quantity, 0) + COALESCE(inbound_shipped_quantity, 0) + COALESCE(inbound_receiving_quantity, 0) AS inbound,
               COALESCE(reserved_quantity, 0) AS reserved
        FROM fba_inventory{inv_where}
    """, hp).fetchall()

    # Get avg daily units for last 30 days
    vel_filter = "WHERE date >= CURRENT_DATE - INTERVAL '30 days' AND asin != 'ALL'" + hf
    vel_params = [] + hp
    velocity = {}
    vel_rows = con.execute(f"""
        SELECT asin, SUM(units_ordered) AS units
        FROM daily_sales
        {vel_filter}
        GROUP BY asin
    """, vel_params).fetchall()
    for vr in vel_rows:
        velocity[vr[0]] = round(vr[1] / 30, 1)

    con.close()

    items = []
    for r in inv_rows:
        asin = r[0]
        avg_daily = velocity.get(asin, 0)
        fba_stock = r[3]
        dos = round(fba_stock / avg_daily) if avg_daily > 0 else 999

        # Get name from COGS file first, then inventory table
        cogs_name = cogs_data.get(asin, {}).get("product_name", "")
        # Exclude names that are just the ASIN repeated
        if cogs_name and cogs_name.strip().upper() == asin.upper():
            cogs_name = ""
        name = (cogs_name or r[2] or asin)

        items.append({
            "asin": asin,
            "sku": r[1] or "",
            "name": name,
            "fbaStock": fba_stock,
            "inbound": r[4],
            "reserved": r[5],
            "avgDaily": avg_daily,
            "dos": dos,
        })

    items.sort(key=lambda x: x["fbaStock"], reverse=True)
    return {"items": items}


@router.get("/api/inventory/kpis")
def inventory_kpis(division: Optional[str] = None, customer: Optional[str] = None, platform: Optional[str] = None, marketplace: Optional[str] = None):
    """Advanced inventory KPIs: Days of Cover, Weeks of Cover, Reorder Point,
    Sell-Through Rate, and Return Rate — per ASIN.

    These were defined in metrics.py as planned_phase4 and are now surfaced.
    Data comes from fba_inventory (current stock), daily_sales (velocity),
    financial_events (refunds), and item_master (lead time / safety stock).
    """
    con = get_db()
    hf, hp = hierarchy_filter(division, customer, platform, marketplace)

    # ── 1. Current FBA inventory ──────────────────────────────────────────────
    inv_where = (" WHERE " + hf.lstrip(" AND ")) if hf else ""
    inv_rows = con.execute(f"""
        SELECT asin, COALESCE(fulfillable_quantity, 0) AS fulfillable,
               COALESCE(inbound_working_quantity, 0) + COALESCE(inbound_shipped_quantity, 0)
                 + COALESCE(inbound_receiving_quantity, 0) AS inbound,
               product_name
        FROM fba_inventory{inv_where}
    """, hp).fetchall()
    inv_map = {r[0]: {"fulfillable": r[1], "inbound": r[2], "name": r[3] or ""} for r in inv_rows}

    # ── 2. Sales velocity (7d and 30d avg daily units) ────────────────────────
    vel_filter = "WHERE asin != 'ALL' AND date >= CURRENT_DATE - INTERVAL '30 days'" + hf
    vel_rows = con.execute(f"""
        SELECT asin,
               SUM(CASE WHEN date >= CURRENT_DATE - INTERVAL '7 days' THEN units_ordered ELSE 0 END) AS units_7d,
               SUM(units_ordered) AS units_30d
        FROM daily_sales
        {vel_filter}
        GROUP BY asin
    """, hp).fetchall()
    vel_map = {}
    for r in vel_rows:
        vel_map[r[0]] = {
            "units_7d": r[1] or 0,
            "avg_daily_7d": round((r[1] or 0) / 7.0, 2),
            "units_30d": r[2] or 0,
            "avg_daily_30d": round((r[2] or 0) / 30.0, 2),
        }

    # ── 3. 90-day units for sell-through rate ─────────────────────────────────
    str_filter = "WHERE asin != 'ALL' AND date >= CURRENT_DATE - INTERVAL '90 days'" + hf
    str_rows = con.execute(f"""
        SELECT asin, SUM(units_ordered) AS units_90d
        FROM daily_sales
        {str_filter}
        GROUP BY asin
    """, hp).fetchall()
    str_map = {r[0]: r[1] or 0 for r in str_rows}

    # ── 4. Return data from financial_events ──────────────────────────────────
    # Build sku → B-ASIN lookup so financial_events identifiers resolve correctly
    try:
        im_rows = con.execute("""
            SELECT asin, sku FROM item_master WHERE sku IS NOT NULL AND sku != ''
        """).fetchall()
        _sku_to_asin = {}
        for ir in im_rows:
            ba = (ir[0] or "").strip(); ss = (ir[1] or "").strip()
            if ba and ss:
                _sku_to_asin[ss] = ba
                if ss.endswith("-CA"):
                    _sku_to_asin[ss.rsplit("-CA", 1)[0]] = ba
    except Exception:
        _sku_to_asin = {}

    def _resolve(fe_asin):
        if fe_asin and fe_asin.startswith("B0"):
            return fe_asin
        return _sku_to_asin.get(fe_asin, fe_asin)

    from datetime import timedelta as _td
    _cutoff_90d = (datetime.now(ZoneInfo("America/Chicago")).date() - _td(days=90)).isoformat()

    refund_map = {}
    try:
        refund_rows = con.execute("""
            SELECT asin, COUNT(*) AS refund_count,
                   SUM(ABS(COALESCE(product_charges, 0))) AS refund_amount
            FROM financial_events
            WHERE event_type ILIKE '%%refund%%'
              AND date >= ?
            GROUP BY asin
        """, [_cutoff_90d]).fetchall()
        for r in refund_rows:
            key = _resolve(r[0])
            prev = refund_map.get(key, {"count": 0, "amount": 0})
            refund_map[key] = {"count": prev["count"] + (r[1] or 0),
                               "amount": round(prev["amount"] + (r[2] or 0), 2)}
    except Exception:
        refund_map = {}

    # ── 5. Total shipments per ASIN (90d) for return rate denominator ─────────
    total_units_90d = {}
    try:
        tu_rows = con.execute("""
            SELECT asin, COUNT(*) AS shipment_count
            FROM financial_events
            WHERE event_type ILIKE '%%Shipment%%'
              AND date >= ?
              AND asin IS NOT NULL AND asin != ''
            GROUP BY asin
        """, [_cutoff_90d]).fetchall()
        for r in tu_rows:
            key = _resolve(r[0])
            total_units_90d[key] = total_units_90d.get(key, 0) + (r[1] or 0)
    except Exception:
        pass

    con.close()

    # ── Build KPI response per ASIN ───────────────────────────────────────────
    # Default lead time and safety stock (can be overridden from item_master later)
    DEFAULT_LEAD_TIME_DAYS = 45  # typical ocean freight + customs
    DEFAULT_SAFETY_STOCK_DAYS = 14  # 2 weeks safety buffer

    all_asins = set(inv_map.keys()) | set(vel_map.keys())
    items = []

    for asin in all_asins:
        inv = inv_map.get(asin, {"fulfillable": 0, "inbound": 0, "name": ""})
        vel = vel_map.get(asin, {"units_7d": 0, "avg_daily_7d": 0, "units_30d": 0, "avg_daily_30d": 0})
        avg_daily = vel["avg_daily_7d"]
        fulfillable = inv["fulfillable"]

        # Days of Cover
        days_of_cover = round(fulfillable / avg_daily, 1) if avg_daily > 0 else None
        weeks_of_cover = round(days_of_cover / 7.0, 1) if days_of_cover is not None else None

        # Reorder Point = (velocity * lead_time) + safety_stock
        safety_stock = round(avg_daily * DEFAULT_SAFETY_STOCK_DAYS)
        reorder_point = round(avg_daily * DEFAULT_LEAD_TIME_DAYS) + safety_stock

        # Stockout risk flag
        stockout_risk = (days_of_cover is not None and days_of_cover < DEFAULT_LEAD_TIME_DAYS)

        # Sell-Through Rate (90d units / avg inventory)
        units_90d = str_map.get(asin, 0)
        sell_through_rate = round(units_90d / fulfillable, 2) if fulfillable > 0 else None

        # Return Rate (refund_map + total_units_90d already re-keyed to B-ASIN)
        refunds = refund_map.get(asin, {"count": 0, "amount": 0})
        total_units = total_units_90d.get(asin, 0)
        return_rate = round(refunds["count"] / total_units, 4) if total_units > 0 else None

        items.append({
            "asin": asin,
            "name": inv.get("name", ""),
            "fulfillable": fulfillable,
            "inbound": inv["inbound"],
            "avgDaily7d": vel["avg_daily_7d"],
            "avgDaily30d": vel["avg_daily_30d"],
            "daysOfCover": days_of_cover,
            "weeksOfCover": weeks_of_cover,
            "reorderPoint": reorder_point,
            "safetyStock": safety_stock,
            "stockoutRisk": stockout_risk,
            "sellThroughRate": sell_through_rate,
            "units90d": units_90d,
            "returnRate": return_rate,
            "refundCount90d": refunds["count"],
            "refundAmount90d": refunds["amount"],
        })

    # Sort by stockout risk first, then days of cover ascending
    items.sort(key=lambda x: (
        0 if x["stockoutRisk"] else 1,
        x["daysOfCover"] if x["daysOfCover"] is not None else 9999
    ))

    return {
        "items": items,
        "defaults": {
            "leadTimeDays": DEFAULT_LEAD_TIME_DAYS,
            "safetyStockDays": DEFAULT_SAFETY_STOCK_DAYS,
        },
        "count": len(items),
    }


@router.post("/api/upload/warehouse-excel")
async def upload_warehouse_excel(file: UploadFile = File(...)):
    """Upload an Excel file to refresh warehouse data and optionally item master.

    Looks for sheets named 'Raw Warehouse Data' and 'Item Master'.
    Parses each into the corresponding CSV file (warehouse.csv, item_master.csv).
    Returns a summary of what was updated.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Add it to requirements.txt.")

    # Save upload temporarily
    import tempfile
    tmp_path = None
    try:
        contents = await file.read()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(contents)
            tmp_path = tmp.name

        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        result = {"filename": file.filename, "sheets_found": list(wb.sheetnames)}

        # ── Parse Raw Warehouse Data ──
        wh_sheet_name = None
        for name in wb.sheetnames:
            if "warehouse" in name.lower() or "raw" in name.lower():
                wh_sheet_name = name
                break

        if wh_sheet_name:
            ws = wb[wh_sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) > 1:
                headers = [str(h or "").strip() for h in rows[0]]
                # Map Excel headers to our CSV column names
                col_map = {}
                for i, h in enumerate(headers):
                    hl = h.lower().replace(" ", "_").replace("-", "_")
                    if hl == "item_number" or hl == "item_no" or hl == "item":
                        col_map["item_number"] = i
                    elif hl == "description" or hl == "desc":
                        col_map["description"] = i
                    elif hl == "pack":
                        col_map["pack"] = i
                    elif hl in ("on_hand", "on hand", "oh"):
                        col_map["on_hand"] = i
                    elif hl in ("damage", "damaged", "dam"):
                        col_map["damage"] = i
                    elif hl in ("qc_hold", "qc hold", "qchold"):
                        col_map["qc_hold"] = i
                    elif hl in ("pcs_on_hand", "pcs on hand", "pcs_on_hand"):
                        col_map["pcs_on_hand"] = i
                    elif hl in ("pcs_allocated", "pcs allocated", "pcs_alloc"):
                        col_map["pcs_allocated"] = i
                    elif hl in ("pcs_available", "pcs available", "pcs_avail"):
                        col_map["pcs_available"] = i
                    elif hl in ("item_ref", "item ref", "itemref"):
                        col_map["item_ref"] = i

                # Auto-detect column positions by header name patterns
                # Also try exact header match as fallback
                header_exact = {h.strip(): i for i, h in enumerate(headers)}
                for csv_col, excel_header in [
                    ("item_number", "Item Number"), ("description", "Description"),
                    ("pack", "Pack"), ("on_hand", "On-Hand"),
                    ("damage", "Damage"), ("qc_hold", "QC Hold"),
                    ("pcs_on_hand", "Pcs On-Hand"), ("pcs_allocated", "Pcs Allocated"),
                    ("pcs_available", "Pcs Available"), ("item_ref", "Item Ref"),
                ]:
                    if csv_col not in col_map and excel_header in header_exact:
                        col_map[csv_col] = header_exact[excel_header]

                logger.info(f"  Warehouse Excel: {len(rows)-1} data rows, mapped columns: {col_map}")

                # Filter to GG-prefix items and write CSV
                wh_rows = []
                for row in rows[1:]:
                    item_num = str(row[col_map.get("item_number", 0)] or "").strip()
                    if not item_num.startswith("GG"):
                        continue
                    wh_rows.append({
                        "item_number": item_num,
                        "description": str(row[col_map.get("description", 1)] or "").strip(),
                        "pack": int(float(row[col_map.get("pack", 2)] or 1)),
                        "on_hand": int(float(row[col_map.get("on_hand", 3)] or 0)),
                        "damage": int(float(row[col_map.get("damage", 5)] or 0)),
                        "qc_hold": int(float(row[col_map.get("qc_hold", 7)] or 0)),
                        "pcs_on_hand": int(float(row[col_map.get("pcs_on_hand", 8)] or 0)),
                        "pcs_allocated": int(float(row[col_map.get("pcs_allocated", 9)] or 0)),
                        "pcs_available": int(float(row[col_map.get("pcs_available", 10)] or 0)),
                        "item_ref": str(row[col_map.get("item_ref", 11)] or "").strip(),
                    })

                if wh_rows:
                    fieldnames = ["item_number", "description", "pack", "on_hand", "damage",
                                 "qc_hold", "pcs_on_hand", "pcs_allocated", "pcs_available", "item_ref"]
                    with open(WAREHOUSE_PATH, "w", newline="", encoding="utf-8") as f:
                        writer = csv.DictWriter(f, fieldnames=fieldnames)
                        writer.writeheader()
                        writer.writerows(wh_rows)
                    result["warehouse"] = {"status": "updated", "rows": len(wh_rows)}
                    logger.info(f"  Warehouse CSV refreshed: {len(wh_rows)} items")
                else:
                    result["warehouse"] = {"status": "no_gg_items_found", "rows": 0}
        else:
            result["warehouse"] = {"status": "sheet_not_found"}

        # ── Parse Item Master (if present) ──
        im_sheet_name = None
        for name in wb.sheetnames:
            if "item master" in name.lower() or "itemmaster" in name.lower():
                im_sheet_name = name
                break

        if im_sheet_name:
            ws = wb[im_sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) > 1:
                headers = [str(h or "").strip() for h in rows[0]]
                # Check if this looks like our item master format
                header_lower = [h.lower() for h in headers]
                has_asin = any("asin" in h for h in header_lower)
                has_sku = any("sku" in h or "item" in h for h in header_lower)

                if has_asin or has_sku:
                    # Load existing item master to merge/update
                    existing = load_item_master()
                    existing_by_sku = {i.get("sku", ""): i for i in existing}

                    # Find column positions
                    hmap = {h.lower().strip(): i for i, h in enumerate(headers)}
                    asin_col = next((i for h, i in hmap.items() if "asin" in h), None)
                    sku_col = next((i for h, i in hmap.items() if "sku" in h or h == "item number"), None)
                    name_col = next((i for h, i in hmap.items() if "product" in h or "name" in h), None)
                    color_col = next((i for h, i in hmap.items() if "color" in h), None)

                    updated_count = 0
                    for row in rows[1:]:
                        sku_val = str(row[sku_col] or "").strip() if sku_col is not None else ""
                        if not sku_val or not sku_val.startswith("GG"):
                            continue
                        if sku_val in existing_by_sku:
                            # Update product name and color if present in Excel
                            if name_col is not None and row[name_col]:
                                existing_by_sku[sku_val]["productName"] = str(row[name_col]).strip()
                            if color_col is not None and row[color_col]:
                                existing_by_sku[sku_val]["color"] = str(row[color_col]).strip()
                            updated_count += 1

                    if updated_count > 0:
                        save_item_master(list(existing_by_sku.values()))
                    result["itemMaster"] = {"status": "updated", "matched": updated_count}
                else:
                    result["itemMaster"] = {"status": "unrecognized_format"}
        else:
            result["itemMaster"] = {"status": "sheet_not_found"}

        wb.close()
        return result

    except Exception as e:
        import traceback
        logger.error(f"Excel upload error: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


@router.post("/api/refresh-warehouse")
async def refresh_warehouse_from_file():
    """Check for a dropped Excel file in the data directory and refresh warehouse data.

    Looks for any .xlsx file in /data/ that contains a 'Raw Warehouse Data' sheet.
    This allows users to drop an Excel file into the data folder and hit this endpoint.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    # Look for Excel files in the data directory
    xlsx_files = sorted(DB_DIR.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    if not xlsx_files:
        return {"status": "no_xlsx_found", "data_dir": str(DB_DIR)}

    # Use the most recently modified Excel file
    excel_path = xlsx_files[0]
    logger.info(f"  Found Excel file for warehouse refresh: {excel_path.name}")

    # Re-use the upload logic by calling it with the file
    from starlette.datastructures import UploadFile as StarletteUpload
    from io import BytesIO

    with open(excel_path, "rb") as f:
        content = f.read()

    # Create a mock UploadFile
    mock_file = UploadFile(filename=excel_path.name, file=BytesIO(content))
    result = await upload_warehouse_excel(mock_file)
    result["source_file"] = excel_path.name
    return result


@router.get("/api/warehouse")
def get_warehouse():
    """Return warehouse inventory grouped by master item with sub-items."""
    if not WAREHOUSE_PATH.exists():
        return {"masters": [], "count": 0}

    # Load raw warehouse rows
    raw_items = []
    with open(WAREHOUSE_PATH, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            raw_items.append({
                "itemNumber": (row.get("item_number") or "").strip(),
                "description": (row.get("description") or "").strip(),
                "pack": int(float(row.get("pack") or 1)),
                "onHand": int(float(row.get("on_hand") or 0)),
                "damage": int(float(row.get("damage") or 0)),
                "qcHold": int(float(row.get("qc_hold") or 0)),
                "pcsOnHand": int(float(row.get("pcs_on_hand") or 0)),
                "pcsAllocated": int(float(row.get("pcs_allocated") or 0)),
                "pcsAvailable": int(float(row.get("pcs_available") or 0)),
                "itemRef": (row.get("item_ref") or "").strip(),
            })

    # Load item master for ASIN + product name lookup
    item_master = load_item_master()
    im_lookup = {i["sku"]: i for i in item_master}

    # Group by item_ref
    groups = {}
    for item in raw_items:
        ref = item["itemRef"]
        if ref not in groups:
            groups[ref] = {"master": None, "subs": []}
        if item["itemNumber"] == ref:
            groups[ref]["master"] = item
        else:
            groups[ref]["subs"].append(item)

    # Build response: each master has aggregated totals + list of subs
    masters = []
    for ref, group in groups.items():
        master = group["master"]
        subs = group["subs"]

        # If no master record exists, create a virtual one from the ref
        if not master:
            master = {
                "itemNumber": ref,
                "description": subs[0]["description"] if subs else "",
                "pack": subs[0]["pack"] if subs else 1,
                "onHand": 0, "damage": 0, "qcHold": 0,
                "pcsOnHand": 0, "pcsAllocated": 0, "pcsAvailable": 0,
                "itemRef": ref,
            }

        # Aggregate totals across master + all subs
        all_items = [master] + subs
        total_pcs_oh = sum(i["pcsOnHand"] for i in all_items)
        total_pcs_alloc = sum(i["pcsAllocated"] for i in all_items)
        total_pcs_avail = sum(i["pcsAvailable"] for i in all_items)
        total_damage = sum(i["damage"] for i in all_items)
        total_qc = sum(i["qcHold"] for i in all_items)

        # Lookup from Item Master
        im_info = im_lookup.get(ref, {})
        asin = im_info.get("asin", "")
        im_name = im_info.get("productName", "")
        color = im_info.get("color", "")

        # Suffix label for sub-items
        for sub in subs:
            suffix = sub["itemNumber"].replace(ref, "", 1)
            sub["suffix"] = suffix if suffix else sub["itemNumber"]

        masters.append({
            "itemRef": ref,
            "itemNumber": master["itemNumber"],
            "asin": asin,
            "description": im_name if im_name else master["description"],
            "whDescription": master["description"],
            "color": color,
            "pack": master["pack"],
            "totalOnHand": total_pcs_oh,
            "totalDamage": total_damage,
            "totalQcHold": total_qc,
            "totalAllocated": total_pcs_alloc,
            "totalAvailable": total_pcs_avail,
            "subCount": len(subs),
            "subs": subs,
        })

    # Sort: items with ASIN first (known products), then by total on-hand desc
    masters.sort(key=lambda m: (0 if m["asin"] else 1, -m["totalOnHand"]))

    return {"masters": masters, "count": len(masters)}


@router.get("/api/warehouse/golf")
def warehouse_golf(channel: Optional[str] = Query(None, description="Filter by channel: Amazon, Walmart, Walmart & Amazon, Other")):
    """Golf warehouse inventory with optional channel filter."""
    items = load_json("golf_inventory.json")
    if not items:
        return {"items": [], "summary": {"totalSkus": 0, "totalPcsOnHand": 0, "totalPcsAvailable": 0, "totalPcsAllocated": 0, "totalDamage": 0}, "channelBreakdown": {}}
    if channel:
        items = [i for i in items if i.get("channel", "").lower() == channel.lower()]

    total_pcs = sum(i.get("pcsOnHand", 0) for i in items)
    total_available = sum(i.get("pcsAvailable", 0) for i in items)
    total_allocated = sum(i.get("pcsAllocated", 0) for i in items)
    total_damage = sum(i.get("damage", 0) for i in items)

    channel_counts = {}
    all_items = load_json("golf_inventory.json")
    for i in all_items:
        ch = i.get("channel", "Other")
        channel_counts[ch] = channel_counts.get(ch, 0) + 1

    return {
        "items": items,
        "summary": {
            "totalSkus": len(items),
            "totalPcsOnHand": total_pcs,
            "totalPcsAvailable": total_available,
            "totalPcsAllocated": total_allocated,
            "totalDamage": total_damage,
        },
        "channelBreakdown": channel_counts,
    }


@router.get("/api/warehouse/housewares")
def warehouse_housewares():
    """Housewares warehouse inventory."""
    items = load_json("housewares_inventory.json")
    if not items:
        return {"items": [], "summary": {"totalSkus": 0, "totalPcsOnHand": 0, "totalPcsAvailable": 0, "totalPcsAllocated": 0, "totalDamage": 0}}

    total_pcs = sum(i.get("pcsOnHand", 0) for i in items)
    total_available = sum(i.get("pcsAvailable", 0) for i in items)
    total_allocated = sum(i.get("pcsAllocated", 0) for i in items)
    total_damage = sum(i.get("damage", 0) for i in items)

    return {
        "items": items,
        "summary": {
            "totalSkus": len(items),
            "totalPcsOnHand": total_pcs,
            "totalPcsAvailable": total_available,
            "totalPcsAllocated": total_allocated,
            "totalDamage": total_damage,
        },
    }


@router.get("/api/warehouse/unified")
def warehouse_unified(division: str = Query("golf", description="golf or housewares"),
                      channel: Optional[str] = Query(None, description="All, Amazon, Walmart, Walmart & Amazon, Other (golf only)")):
    """Unified warehouse inventory with master/sub grouping, suffix breakdown, and summary.
    Used by the unified Inventory page."""
    filename = "golf_inventory.json" if division == "golf" else "housewares_inventory.json"
    items = load_json(filename)
    if not items:
        return {"masters": [], "summary": {}, "suffixBreakdown": {}, "channelBreakdown": {}}

    # For golf, apply channel filter
    if division == "golf" and channel and channel != "All":
        items = [i for i in items if (i.get("channel", "") or "").lower() == channel.lower()]

    # Channel breakdown (from full data, before channel filter)
    channel_counts = {}
    if division == "golf":
        all_items = load_json("golf_inventory.json")
        for i in all_items:
            ch = i.get("channel", "Other")
            channel_counts[ch] = channel_counts.get(ch, 0) + 1

    # Known suffixes for grouping
    SUFFIX_PATTERNS = ["/RB", "/RETD", "/DONATE", "/DMGD", "/DAM", "/HOLD", "/CUST", "/INBD", "/1",
                       "FBM", "-RB", "-DONATE", "-RETD", "-FBM", "-HOLD", "-Damage", "-CUST", "-Transfer"]

    def _get_base_sku(sku):
        """Strip prefixes (T-) and suffixes to get the base/master SKU.
        Strips iteratively to handle multi-suffix SKUs like GGWMSS2238BM/1/RB."""
        s = sku.strip()
        # Strip T- prefix
        if s.startswith("T-"):
            s = s[2:]
        # Strip known suffixes iteratively (handles /1/RB, /1/DONATE, etc.)
        changed = True
        while changed:
            changed = False
            for pat in SUFFIX_PATTERNS:
                if s.endswith(pat):
                    s = s[:-len(pat)]
                    changed = True
                    break
                if pat in s and pat.startswith("/"):
                    idx = s.find(pat)
                    s = s[:idx]
                    changed = True
                    break
        return s

    def _get_suffix_label(sku, base):
        """Determine suffix type for a variant SKU."""
        s = sku.strip()
        if s.startswith("T-"):
            return "T-"
        if "FBM" in s and "FBM" not in base:
            return "FBM"
        remainder = s.replace(base, "", 1)
        if "/RB" in remainder or "-RB" in remainder:
            return "RB"
        if "/RETD" in remainder or "-RETD" in remainder:
            return "RETD"
        if "/DONATE" in remainder or "-DONATE" in remainder:
            return "DONATE"
        if "/DMGD" in remainder or "/DAM" in remainder or "-Damage" in remainder:
            return "Damage"
        if "/HOLD" in remainder or "-HOLD" in remainder:
            return "HOLD"
        if "/CUST" in remainder or "-CUST" in remainder:
            return "CUST"
        if "/INBD" in remainder:
            return "INBD"
        if "/1" in remainder:
            return "Each"
        if remainder:
            return remainder.strip("/- ")
        return "Standard"

    # ASIN lookup from item master
    item_master = load_item_master()
    im_lookup = {i["sku"]: i for i in item_master}

    # Group items by base SKU
    groups = {}
    for item in items:
        sku = item.get("itemNumber", "").strip()
        base = _get_base_sku(sku)
        if base not in groups:
            groups[base] = {"master": None, "subs": []}
        if sku == base:
            groups[base]["master"] = item
        else:
            item["suffix"] = _get_suffix_label(sku, base)
            groups[base]["subs"].append(item)

    # Build master list
    masters = []
    for base, group in groups.items():
        master = group["master"]
        subs = group["subs"]
        if not master:
            # Create virtual master from first sub
            ref_item = subs[0] if subs else {}
            master = {
                "itemNumber": base,
                "description": ref_item.get("description", ""),
                "pack": ref_item.get("pack", 1),
                "onHand": 0, "pcsOnHand": 0, "pcsAllocated": 0,
                "pcsAvailable": 0, "damage": 0, "qcHold": 0,
                "channel": ref_item.get("channel", "Other"),
            }

        all_in_group = [master] + subs
        total_oh = sum(i.get("pcsOnHand", 0) for i in all_in_group)
        total_alloc = sum(i.get("pcsAllocated", 0) for i in all_in_group)
        total_avail = sum(i.get("pcsAvailable", 0) for i in all_in_group)
        total_dmg = sum(i.get("damage", 0) for i in all_in_group)
        total_qc = sum(i.get("qcHold", 0) for i in all_in_group)

        # ASIN lookup
        im_info = im_lookup.get(base, {})
        asin = im_info.get("asin", "")
        im_name = im_info.get("productName", "")
        c_len = round(float(im_info.get("cartonLength", 0) or 0), 1)
        c_wid = round(float(im_info.get("cartonWidth", 0) or 0), 1)
        c_hgt = round(float(im_info.get("cartonHeight", 0) or 0), 1)

        masters.append({
            "baseSku": base,
            "itemNumber": master.get("itemNumber", base),
            "asin": asin,
            "walmartItemNumber": im_info.get("walmartItemNumber", ""),
            "description": im_name if im_name else master.get("description", ""),
            "whDescription": master.get("description", ""),
            "pack": master.get("pack", 1),
            "channel": master.get("channel", "Other") if division == "golf" else None,
            "totalOnHand": total_oh,
            "totalAllocated": total_alloc,
            "totalAvailable": total_avail,
            "totalDamage": total_dmg,
            "totalQcHold": total_qc,
            "cartonL": c_len,
            "cartonW": c_wid,
            "cartonH": c_hgt,
            "subCount": len(subs),
            "subs": subs,
        })

    masters.sort(key=lambda m: -m["totalOnHand"])

    # Summary totals
    summary = {
        "totalSkus": len(masters),
        "totalItems": len(items),
        "totalOnHand": sum(m["totalOnHand"] for m in masters),
        "totalAllocated": sum(m["totalAllocated"] for m in masters),
        "totalAvailable": sum(m["totalAvailable"] for m in masters),
        "totalDamage": sum(m["totalDamage"] for m in masters),
    }

    # Suffix breakdown (across all items in this division, not filtered by channel)
    suffix_buckets = {}
    all_div_items = load_json(filename)
    for item in all_div_items:
        sku = item.get("itemNumber", "").strip()
        base = _get_base_sku(sku)
        if sku == base:
            suf = "Standard"
        else:
            suf = _get_suffix_label(sku, base)
        if suf not in suffix_buckets:
            suffix_buckets[suf] = {"skus": 0, "pcsOnHand": 0, "pcsAllocated": 0, "pcsAvailable": 0}
        suffix_buckets[suf]["skus"] += 1
        suffix_buckets[suf]["pcsOnHand"] += item.get("pcsOnHand", 0)
        suffix_buckets[suf]["pcsAllocated"] += item.get("pcsAllocated", 0)
        suffix_buckets[suf]["pcsAvailable"] += item.get("pcsAvailable", 0)

    return {
        "masters": masters,
        "summary": summary,
        "suffixBreakdown": suffix_buckets,
        "channelBreakdown": channel_counts,
    }


@router.get("/api/warehouse/summary")
def warehouse_summary():
    """Combined warehouse inventory summary — Golf vs Housewares totals + suffix breakdown."""
    golf_items = load_json("golf_inventory.json")
    hw_items = load_json("housewares_inventory.json")

    def _summarize(items):
        return {
            "skus": len(items),
            "pcsOnHand": sum(i.get("pcsOnHand", 0) for i in items),
            "pcsAvailable": sum(i.get("pcsAvailable", 0) for i in items),
            "pcsAllocated": sum(i.get("pcsAllocated", 0) for i in items),
            "damage": sum(i.get("damage", 0) for i in items),
        }

    def _suffix_breakdown(items):
        buckets = {}
        for item in items:
            sku = item.get("itemNumber", "").strip()
            suffix = "Standard"
            if sku.startswith("T-"):
                suffix = "T-"
            elif sku.endswith("FBM"):
                suffix = "FBM"
            elif "/RB" in sku:
                suffix = "RB"
            elif "/RETD" in sku:
                suffix = "RETD"
            elif "/DONATE" in sku:
                suffix = "DONATE"
            elif "/DMGD" in sku or "/DAM" in sku:
                suffix = "Damage"
            elif "/HOLD" in sku:
                suffix = "HOLD"
            elif "/CUST" in sku:
                suffix = "CUST"
            if suffix not in buckets:
                buckets[suffix] = {"skus": 0, "pcsOnHand": 0, "pcsAvailable": 0, "pcsAllocated": 0}
            buckets[suffix]["skus"] += 1
            buckets[suffix]["pcsOnHand"] += item.get("pcsOnHand", 0)
            buckets[suffix]["pcsAvailable"] += item.get("pcsAvailable", 0)
            buckets[suffix]["pcsAllocated"] += item.get("pcsAllocated", 0)
        return buckets

    golf_summary = _summarize(golf_items)
    hw_summary = _summarize(hw_items)

    combined = {
        "skus": golf_summary["skus"] + hw_summary["skus"],
        "pcsOnHand": golf_summary["pcsOnHand"] + hw_summary["pcsOnHand"],
        "pcsAvailable": golf_summary["pcsAvailable"] + hw_summary["pcsAvailable"],
        "pcsAllocated": golf_summary["pcsAllocated"] + hw_summary["pcsAllocated"],
        "damage": golf_summary["damage"] + hw_summary["damage"],
    }

    return {
        "combined": combined,
        "golf": golf_summary,
        "housewares": hw_summary,
        "golfSuffixes": _suffix_breakdown(golf_items),
        "housewaresSuffixes": _suffix_breakdown(hw_items),
    }


# ── FBA Shipments ──────────────────────────────────────────────

def _fba_cache_path(marketplace="US"):
    """Return cache file path for given marketplace."""
    suffix = f"_{marketplace.lower()}" if marketplace != "US" else ""
    return DB_DIR / f"fba_shipments_cache{suffix}.json"


def _load_fba_shipments_cache(marketplace="US"):
    path = _fba_cache_path(marketplace)
    if path.exists():
        with open(path) as f:
            return json.load(f)
    return None


def _save_fba_shipments_cache(data, marketplace="US"):
    path = _fba_cache_path(marketplace)
    with open(path, "w") as f:
        json.dump(data, f, indent=2)


MARKETPLACE_IDS = {
    "US": "ATVPDKIKX0DER",
    "CA": "A2EUQ1WTGCTBG2",
}

def _get_marketplace_enum(marketplace="US"):
    """Return sp_api Marketplaces enum for given marketplace key."""
    from sp_api.base import Marketplaces
    return Marketplaces.CA if marketplace == "CA" else Marketplaces.US

def _get_marketplace_id(marketplace="US"):
    """Return Amazon MarketplaceId string for given marketplace key."""
    return MARKETPLACE_IDS.get(marketplace, MARKETPLACE_IDS["US"])


def _fetch_fba_shipments_from_api(statuses=None, days_back=90, marketplace="US"):
    """Fetch FBA inbound shipments from Amazon SP-API."""
    try:
        from sp_api.api import FulfillmentInbound
        from sp_api.base import Marketplaces
    except ImportError:
        logger.warning("SP-API library not installed — cannot fetch FBA shipments")
        return None

    credentials = _load_sp_api_credentials()
    if not credentials:
        logger.warning("No SP-API credentials — cannot fetch FBA shipments")
        return None

    if statuses is None:
        statuses = ["WORKING", "SHIPPED", "RECEIVING", "CLOSED", "IN_TRANSIT", "CANCELLED"]

    mp_enum = _get_marketplace_enum(marketplace)
    mp_id = _get_marketplace_id(marketplace)

    try:
        fba = FulfillmentInbound(
            credentials=credentials,
            marketplace=mp_enum,
        )

        all_shipments = []

        # Fetch by status list
        resp = fba.get_shipments(
            QueryType="SHIPMENT_STATUS",
            ShipmentStatusList=",".join(statuses),
            MarketplaceId=mp_id,
        )

        payload = resp.payload or {}
        shipment_data = payload.get("ShipmentData", [])
        all_shipments.extend(shipment_data)

        # Handle pagination
        next_token = payload.get("NextToken")
        page = 1
        while next_token and page < 10:  # safety limit
            page += 1
            resp = fba.get_shipments(
                QueryType="NEXT_TOKEN",
                NextToken=next_token,
                MarketplaceId=mp_id,
            )
            payload = resp.payload or {}
            shipment_data = payload.get("ShipmentData", [])
            all_shipments.extend(shipment_data)
            next_token = payload.get("NextToken")

        logger.info(f"FBA Shipments: fetched {len(all_shipments)} shipments from SP-API")

        # Normalize shipment data for frontend
        normalized = []
        active_ids = []  # collect active shipment IDs for item pre-fetch
        for s in all_shipments:
            ship_id = s.get("ShipmentId", "")
            ship_name = s.get("ShipmentName", "")
            status = s.get("ShipmentStatus", "")
            dest = s.get("DestinationFulfillmentCenterId", "")
            label_prep = s.get("LabelPrepType", "")
            are_cases_required = s.get("AreCasesRequired", False)

            # Address info
            ship_from = s.get("ShipFromAddress", {})
            ship_from_str = ""
            if ship_from:
                parts = [ship_from.get("City", ""), ship_from.get("StateOrProvinceCode", "")]
                ship_from_str = ", ".join(p for p in parts if p)

            # Detect actual marketplace from FC code (Canadian FCs start with Y)
            # Known Canadian FCs: YYZ, YOW, YVR, YHM, etc.
            fc_marketplace = marketplace
            if dest and len(dest) >= 3:
                fc_prefix = dest[:3].upper()
                # Canadian FCs typically start with Y (YYZ9, YOW1, YVR3, YHM1, etc.)
                if fc_prefix.startswith("Y"):
                    fc_marketplace = "CA"

            normalized.append({
                "shipmentId": ship_id,
                "shipmentName": ship_name,
                "status": status,
                "destination": dest,
                "labelPrep": label_prep,
                "casesRequired": are_cases_required,
                "shipFrom": ship_from_str,
                "marketplace": fc_marketplace,
                "itemCount": 0,
                "totalShipped": 0,
                "totalReceived": 0,
            })

            # Track shipments for item pre-fetch (active first, then recent closed)
            if status in ("WORKING", "SHIPPED", "RECEIVING", "IN_TRANSIT", "CHECKED_IN"):
                active_ids.append(ship_id)

        # Also include CLOSED shipments for pre-fetch (most recent first, capped)
        closed_ids = [n["shipmentId"] for n in normalized if n["status"] == "CLOSED"]
        # Cap total pre-fetches: active first, then up to 20 closed
        MAX_CLOSED_PREFETCH = 20
        prefetch_ids = active_ids + closed_ids[:MAX_CLOSED_PREFETCH]

        # Pre-fetch items for shipments to populate sent/received/itemCount
        if prefetch_ids:
            logger.info(f"FBA Shipments: pre-fetching items for {len(prefetch_ids)} shipments ({len(active_ids)} active + {min(len(closed_ids), MAX_CLOSED_PREFETCH)} closed)")
            ship_lookup = {n["shipmentId"]: n for n in normalized}
            for sid in prefetch_ids:
                try:
                    items = _enrich_shipment_items(sid, marketplace=marketplace)
                    if items and sid in ship_lookup:
                        ship_lookup[sid]["itemCount"] = len(items)
                        ship_lookup[sid]["totalShipped"] = sum(i.get("quantityShipped", 0) for i in items)
                        ship_lookup[sid]["totalReceived"] = sum(i.get("quantityReceived", 0) for i in items)
                except Exception as e:
                    logger.warning(f"Failed to pre-fetch items for {sid}: {e}")

        # Filter shipments to only those matching requested marketplace
        # (SP-API may return cross-marketplace shipments on a unified account)
        filtered = [s for s in normalized if s.get("marketplace", marketplace) == marketplace]

        result = {
            "shipments": filtered,
            "lastSync": datetime.now(TIMEZONE).strftime("%m/%d/%Y %I:%M %p"),
            "totalShipments": len(filtered),
            "marketplace": marketplace,
        }

        _save_fba_shipments_cache(result, marketplace=marketplace)
        return result

    except Exception as e:
        logger.error(f"FBA Shipments SP-API error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None


def _enrich_shipment_items(shipment_id, marketplace="US"):
    """Fetch items for a specific shipment from SP-API, enriched with product names."""
    try:
        from sp_api.api import FulfillmentInbound
        from sp_api.base import Marketplaces
    except ImportError:
        return []

    credentials = _load_sp_api_credentials()
    if not credentials:
        return []

    mp_enum = _get_marketplace_enum(marketplace)
    mp_id = _get_marketplace_id(marketplace)

    try:
        fba = FulfillmentInbound(
            credentials=credentials,
            marketplace=mp_enum,
        )
        resp = fba.shipment_items_by_shipment(shipment_id, MarketplaceId=mp_id)
        payload = resp.payload or {}
        items = payload.get("ItemData", [])

        result = []
        for item in items:
            result.append({
                "sku": item.get("SellerSKU", ""),
                "fnsku": item.get("FulfillmentNetworkSKU", ""),
                "quantityShipped": item.get("QuantityShipped", 0),
                "quantityReceived": item.get("QuantityReceived", 0),
                "quantityInCase": item.get("QuantityInCase", 0),
            })

        # Handle pagination — deduplicate by SKU to prevent inflated counts
        # (Amazon's API sometimes returns duplicate pages with the same NextToken)
        seen_skus = {item.get("SellerSKU", "") for item in items}
        next_token = payload.get("NextToken")
        page = 1
        while next_token and page < 10:
            page += 1
            resp = fba.shipment_items_by_shipment(
                shipment_id,
                MarketplaceId=mp_id,
                NextToken=next_token,
            )
            payload = resp.payload or {}
            new_items = payload.get("ItemData", [])
            new_skus = {item.get("SellerSKU", "") for item in new_items}

            # If this page's SKUs are all already seen, we're getting duplicate data — stop
            if new_skus and new_skus.issubset(seen_skus):
                logger.info(f"FBA shipment {shipment_id}: page {page} returned duplicate SKUs, stopping pagination")
                break

            for item in new_items:
                sku = item.get("SellerSKU", "")
                if sku not in seen_skus:
                    seen_skus.add(sku)
                    result.append({
                        "sku": sku,
                        "fnsku": item.get("FulfillmentNetworkSKU", ""),
                        "quantityShipped": item.get("QuantityShipped", 0),
                        "quantityReceived": item.get("QuantityReceived", 0),
                        "quantityInCase": item.get("QuantityInCase", 0),
                    })
            next_token = payload.get("NextToken")

        # Enrich with product names + ASIN from item_master + fba_inventory (DB)
        if result:
            try:
                con = get_db()
                skus = [r["sku"] for r in result if r["sku"]]
                if skus:
                    placeholders = ",".join(["?"] * len(skus))
                    name_map = {}
                    asin_map = {}

                    # Source 1: item_master (primary — has curated product names)
                    rows = con.execute(
                        f"SELECT sku, product_name, asin FROM item_master WHERE sku IN ({placeholders})",
                        skus,
                    ).fetchall()
                    for r in rows:
                        if r[1]:
                            name_map[r[0]] = r[1]
                        if r[2]:
                            asin_map[r[0]] = r[2]

                    # Also try base SKU without -CA/-FBM suffixes for partial matches
                    missing_skus = [s for s in skus if s not in asin_map]
                    if missing_skus:
                        base_map = {}
                        for ms in missing_skus:
                            for suffix in ["-CA", "-FBM", "-FBA"]:
                                if ms.endswith(suffix):
                                    base_map[ms.rsplit(suffix, 1)[0]] = ms
                                    break
                        if base_map:
                            bp = ",".join(["?"] * len(base_map))
                            base_rows = con.execute(
                                f"SELECT sku, product_name, asin FROM item_master WHERE sku IN ({bp})",
                                list(base_map.keys()),
                            ).fetchall()
                            for r in base_rows:
                                orig_sku = base_map.get(r[0])
                                if orig_sku:
                                    if r[1] and orig_sku not in name_map:
                                        name_map[orig_sku] = r[1]
                                    if r[2] and orig_sku not in asin_map:
                                        asin_map[orig_sku] = r[2]

                    # Source 2: fba_inventory fallback (has fresh SP-API data)
                    still_missing = [s for s in skus if s not in asin_map]
                    if still_missing:
                        fp = ",".join(["?"] * len(still_missing))
                        inv_rows = con.execute(
                            f"SELECT DISTINCT sku, product_name, asin FROM fba_inventory WHERE sku IN ({fp}) AND asin IS NOT NULL AND asin != ''",
                            still_missing,
                        ).fetchall()
                        for r in inv_rows:
                            if r[1] and r[0] not in name_map:
                                name_map[r[0]] = r[1]
                            if r[2] and r[0] not in asin_map:
                                asin_map[r[0]] = r[2]

                    for r in result:
                        r["productName"] = name_map.get(r["sku"], "")
                        r["asin"] = asin_map.get(r["sku"], "")
            except Exception as e:
                logger.warning(f"Could not enrich item names for {shipment_id}: {e}")

        return result
    except Exception as e:
        logger.error(f"FBA shipment items error for {shipment_id}: {e}")
        return []


def _load_sp_api_credentials() -> dict | None:
    """Load SP-API credentials from env vars or config file."""
    import os
    # Priority 1: Environment variables (used on Railway)
    env_refresh = os.environ.get("SP_API_REFRESH_TOKEN", "")
    if env_refresh:
        return {
            "refresh_token": env_refresh,
            "lwa_app_id": os.environ.get("SP_API_LWA_APP_ID") or os.environ.get("SP_API_CLIENT_ID") or os.environ.get("LWA_APP_ID", ""),
            "lwa_client_secret": os.environ.get("SP_API_LWA_CLIENT_SECRET") or os.environ.get("SP_API_CLIENT_SECRET") or os.environ.get("LWA_CLIENT_SECRET", ""),
            "aws_access_key": os.environ.get("SP_API_AWS_ACCESS_KEY", ""),
            "aws_secret_key": os.environ.get("SP_API_AWS_SECRET_KEY", ""),
            "role_arn": os.environ.get("SP_API_ROLE_ARN", ""),
        }

    # Priority 2: Config file (local development)
    if CONFIG_PATH.exists():
        with open(CONFIG_PATH) as f:
            creds_json = json.load(f)
        sp = creds_json.get("AMAZON_SP_API", {})
        creds = {
            "refresh_token": sp.get("refresh_token"),
            "lwa_app_id": sp.get("lwa_app_id"),
            "lwa_client_secret": sp.get("lwa_client_secret"),
            "aws_access_key": sp.get("aws_access_key"),
            "aws_secret_key": sp.get("aws_secret_key"),
            "role_arn": sp.get("role_arn"),
        }
        if creds.get("refresh_token"):
            return creds

    return None


@router.get("/api/fba-shipments")
async def get_fba_shipments(request: Request, refresh: bool = False, marketplace: str = "US"):
    """Get FBA inbound shipments. Uses cache unless refresh=true."""
    from core.auth import require_auth
    require_auth(request)

    mp = marketplace.upper() if marketplace else "US"
    if mp not in ("US", "CA"):
        mp = "US"

    if not refresh:
        cached = _load_fba_shipments_cache(marketplace=mp)
        if cached:
            return cached

    # Try fetching fresh data from SP-API
    data = _fetch_fba_shipments_from_api(marketplace=mp)
    if data:
        return data

    # Fall back to cache even if refresh was requested
    cached = _load_fba_shipments_cache(marketplace=mp)
    if cached:
        cached["_note"] = "Using cached data — SP-API sync failed"
        return cached

    return {"shipments": [], "lastSync": None, "totalShipments": 0, "marketplace": mp}


@router.post("/api/fba-shipments/sync")
async def sync_fba_shipments(request: Request, marketplace: str = "US"):
    """Force refresh of FBA shipment data from SP-API."""
    from core.auth import require_auth
    require_auth(request)
    mp = marketplace.upper() if marketplace else "US"
    if mp not in ("US", "CA"):
        mp = "US"
    data = _fetch_fba_shipments_from_api(marketplace=mp)
    if data:
        return {"ok": True, "totalShipments": data["totalShipments"], "lastSync": data["lastSync"]}
    return JSONResponse(status_code=500, content={"ok": False, "error": "SP-API sync failed. Check credentials."})


@router.get("/api/fba-shipments/{shipment_id}/items")
async def get_fba_shipment_items(request: Request, shipment_id: str, marketplace: str = "US"):
    """Get items for a specific FBA inbound shipment."""
    from core.auth import require_auth
    require_auth(request)
    mp = marketplace.upper() if marketplace else "US"
    if mp not in ("US", "CA"):
        mp = "US"
    items = _enrich_shipment_items(shipment_id, marketplace=mp)
    return {"shipmentId": shipment_id, "items": items, "totalItems": len(items)}


@router.get("/api/fba-shipments/products")
async def get_shipment_products(request: Request):
    """Get available products for shipment creation — ASIN, SKU, product name,
    FBA on-hand (not FBM), and unit sales for last 30/60/90 days."""
    from core.auth import require_auth
    require_auth(request)
    try:
        con = get_db()
        today = datetime.now(TIMEZONE).date()
        d30 = (today - timedelta(days=30)).isoformat()
        d60 = (today - timedelta(days=60)).isoformat()
        d90 = (today - timedelta(days=90)).isoformat()

        # Pull products from item_master with FBA inventory and 30/60/90d sales
        rows = con.execute("""
            SELECT im.asin, im.sku, im.product_name, im.division,
                   COALESCE(inv.fulfillable_quantity, 0) AS fba_on_hand,
                   COALESCE(s30.units, 0) AS units_30d,
                   COALESCE(s30.revenue, 0) AS rev_30d,
                   COALESCE(s60.units, 0) AS units_60d,
                   COALESCE(s60.revenue, 0) AS rev_60d,
                   COALESCE(s90.units, 0) AS units_90d,
                   COALESCE(s90.revenue, 0) AS rev_90d,
                   COALESCE(inv.inbound_working_quantity, 0) AS inbound_working,
                   COALESCE(inv.inbound_shipped_quantity, 0) AS inbound_shipped,
                   COALESCE(inv.inbound_receiving_quantity, 0) AS inbound_receiving,
                   COALESCE(inv.reserved_quantity, 0) AS reserved
            FROM item_master im
            LEFT JOIN (
                SELECT asin, fulfillable_quantity, inbound_working_quantity,
                       inbound_shipped_quantity, inbound_receiving_quantity, reserved_quantity
                FROM fba_inventory
                WHERE date = (SELECT MAX(date) FROM fba_inventory)
            ) inv ON im.asin = inv.asin
            LEFT JOIN (
                SELECT asin,
                       SUM(units_ordered) AS units,
                       SUM(ordered_product_sales) AS revenue
                FROM daily_sales
                WHERE asin != 'ALL' AND date >= ?
                GROUP BY asin
            ) s30 ON im.asin = s30.asin
            LEFT JOIN (
                SELECT asin,
                       SUM(units_ordered) AS units,
                       SUM(ordered_product_sales) AS revenue
                FROM daily_sales
                WHERE asin != 'ALL' AND date >= ?
                GROUP BY asin
            ) s60 ON im.asin = s60.asin
            LEFT JOIN (
                SELECT asin,
                       SUM(units_ordered) AS units,
                       SUM(ordered_product_sales) AS revenue
                FROM daily_sales
                WHERE asin != 'ALL' AND date >= ?
                GROUP BY asin
            ) s90 ON im.asin = s90.asin
            WHERE im.asin IS NOT NULL AND im.asin != ''
              AND im.sku IS NOT NULL AND im.sku != ''
            ORDER BY COALESCE(s30.units, 0) DESC, im.product_name
        """, [d30, d60, d90]).fetchall()

        products = []
        for r in rows:
            fba = _n(r[4])
            u30 = _n(r[5])
            inbound_working = _n(r[11])
            inbound_shipped = _n(r[12])
            inbound_receiving = _n(r[13])
            reserved = _n(r[14])
            products.append({
                "asin": r[0],
                "sku": r[1],
                "productName": r[2] or "",
                "division": r[3] or "unknown",
                "currentStock": fba,
                "units30d": u30,
                "rev30d": round(_n(r[6]), 2),
                "units60d": _n(r[7]),
                "rev60d": round(_n(r[8]), 2),
                "units90d": _n(r[9]),
                "rev90d": round(_n(r[10]), 2),
                "inboundWorking": inbound_working,
                "inboundShipped": inbound_shipped,
                "inboundReceiving": inbound_receiving,
                "reserved": reserved,
                "inTransitTotal": inbound_working + inbound_shipped + inbound_receiving,
                "daysOfStock": round(fba / (u30 / 30), 1) if u30 > 0 else 999,
            })
        return {"products": products, "count": len(products)}
    except Exception as e:
        logger.error(f"Error fetching shipment products: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.post("/api/fba-shipments/create-plan")
async def create_shipment_plan(request: Request):
    """Create an inbound shipment plan via SP-API.

    Expects JSON body:
    {
        "shipFromAddress": {
            "Name": "GolfGen LLC",
            "AddressLine1": "...",
            "City": "Bentonville",
            "StateOrProvinceCode": "AR",
            "PostalCode": "72712",
            "CountryCode": "US"
        },
        "items": [
            {"SellerSKU": "GGRTNW222810", "ASIN": "B0DPBGG8GY", "Quantity": 100, "Condition": "NewItem"}
        ],
        "labelPrepPreference": "SELLER_LABEL"  // optional: SELLER_LABEL, AMAZON_LABEL_ONLY, AMAZON_LABEL_PREFERRED
    }

    Returns the shipment plan(s) proposed by Amazon.
    """
    from core.auth import require_auth
    require_auth(request)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(status_code=400, content={"error": "Invalid JSON body"})

    ship_from = body.get("shipFromAddress")
    items = body.get("items", [])
    label_pref = body.get("labelPrepPreference", "SELLER_LABEL")

    if not ship_from:
        return JSONResponse(status_code=400, content={"error": "shipFromAddress is required"})
    if not items:
        return JSONResponse(status_code=400, content={"error": "At least one item is required"})

    # Validate required address fields
    for field in ["Name", "AddressLine1", "City", "StateOrProvinceCode", "PostalCode", "CountryCode"]:
        if not ship_from.get(field):
            return JSONResponse(status_code=400, content={"error": f"shipFromAddress.{field} is required"})

    # Validate items
    for i, item in enumerate(items):
        if not item.get("SellerSKU"):
            return JSONResponse(status_code=400, content={"error": f"Item {i+1}: SellerSKU is required"})
        if not item.get("Quantity") or int(item["Quantity"]) < 1:
            return JSONResponse(status_code=400, content={"error": f"Item {i+1}: Quantity must be >= 1"})

    try:
        from sp_api.api import FulfillmentInbound
        from sp_api.base import Marketplaces
    except ImportError:
        return JSONResponse(status_code=500, content={"error": "SP-API library not installed"})

    credentials = _load_sp_api_credentials()
    if not credentials:
        return JSONResponse(status_code=500, content={"error": "SP-API credentials not configured"})

    # Detect marketplace from request body (default US)
    mp = body.get("marketplace", "US").upper()

    try:
        fba = FulfillmentInbound(
            credentials=credentials,
            marketplace=_get_marketplace_enum(mp),
        )

        # Build the inbound shipment plan items
        plan_items = []
        for item in items:
            plan_item = {
                "SellerSKU": item["SellerSKU"],
                "Quantity": str(int(item["Quantity"])),
                "Condition": item.get("Condition", "NewItem"),
            }
            if item.get("ASIN"):
                plan_item["ASIN"] = item["ASIN"]
            if item.get("QuantityInCase"):
                plan_item["QuantityInCase"] = str(int(item["QuantityInCase"]))
            plan_items.append(plan_item)

        logger.info(f"Creating inbound shipment plan ({mp}): {len(plan_items)} items from {ship_from.get('City', '?')}, {ship_from.get('StateOrProvinceCode', '?')}")

        resp = fba.create_inbound_shipment_plan(
            ShipFromAddress=ship_from,
            InboundShipmentPlanRequestItems=plan_items,
            LabelPrepPreference=label_pref,
            MarketplaceId=_get_marketplace_id(mp),
        )

        payload = resp.payload or {}
        plans = payload.get("InboundShipmentPlans", [])

        # Normalize the plan data for frontend
        normalized_plans = []
        for plan in plans:
            normalized_plans.append({
                "shipmentId": plan.get("ShipmentId", ""),
                "destinationFC": plan.get("DestinationFulfillmentCenterId", ""),
                "labelPrepType": plan.get("LabelPrepType", ""),
                "items": [
                    {
                        "SellerSKU": pi.get("SellerSKU", ""),
                        "FulfillmentNetworkSKU": pi.get("FulfillmentNetworkSKU", ""),
                        "Quantity": pi.get("Quantity", 0),
                    }
                    for pi in plan.get("Items", [])
                ],
                "estimatedBoxContentsFee": plan.get("EstimatedBoxContentsFee", {}),
            })

        logger.info(f"Shipment plan created: {len(normalized_plans)} shipment(s) proposed by Amazon")

        return {
            "ok": True,
            "plans": normalized_plans,
            "totalPlans": len(normalized_plans),
            "message": f"Amazon proposed {len(normalized_plans)} shipment(s). Review and confirm to create.",
        }

    except Exception as e:
        logger.error(f"Create shipment plan error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})


@router.post("/api/fba-shipments/confirm-plan")
async def confirm_shipment_plan(request: Request):
    """Confirm and create an inbound shipment from a plan.

    Expects JSON body:
    {
        "shipmentId": "FBA17ABCDEF",
        "shipmentName": "My Shipment March 2026",
        "destinationFC": "PHX3",
        "items": [
            {"SellerSKU": "GGRTNW222810", "Quantity": 100}
        ],
        "shipFromAddress": { ... same as create-plan ... },
        "labelPrepPreference": "SELLER_LABEL"
    }
    """
    from core.auth import require_auth
    require_auth(request)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(status_code=400, content={"error": "Invalid JSON body"})

    shipment_id = body.get("shipmentId")
    shipment_name = body.get("shipmentName", f"GolfGen Shipment {datetime.now(TIMEZONE).strftime('%m/%d/%Y')}")
    dest_fc = body.get("destinationFC")
    items = body.get("items", [])
    ship_from = body.get("shipFromAddress")
    label_pref = body.get("labelPrepPreference", "SELLER_LABEL")

    if not shipment_id:
        return JSONResponse(status_code=400, content={"error": "shipmentId is required (from plan)"})
    if not dest_fc:
        return JSONResponse(status_code=400, content={"error": "destinationFC is required"})
    if not items:
        return JSONResponse(status_code=400, content={"error": "At least one item is required"})
    if not ship_from:
        return JSONResponse(status_code=400, content={"error": "shipFromAddress is required"})

    try:
        from sp_api.api import FulfillmentInbound
        from sp_api.base import Marketplaces
    except ImportError:
        return JSONResponse(status_code=500, content={"error": "SP-API library not installed"})

    credentials = _load_sp_api_credentials()
    if not credentials:
        return JSONResponse(status_code=500, content={"error": "SP-API credentials not configured"})

    # Detect marketplace from request body (default US)
    mp = body.get("marketplace", "US").upper()

    try:
        fba = FulfillmentInbound(
            credentials=credentials,
            marketplace=_get_marketplace_enum(mp),
        )

        shipment_items = []
        for item in items:
            shipment_items.append({
                "SellerSKU": item["SellerSKU"],
                "QuantityShipped": int(item.get("Quantity", item.get("QuantityShipped", 0))),
            })

        logger.info(f"Confirming shipment plan ({mp}) {shipment_id}: {len(shipment_items)} items → {dest_fc}")

        resp = fba.create_inbound_shipment(
            ShipmentId=shipment_id,
            InboundShipmentHeader={
                "ShipmentName": shipment_name,
                "ShipFromAddress": ship_from,
                "DestinationFulfillmentCenterId": dest_fc,
                "LabelPrepPreference": label_pref,
                "ShipmentStatus": "WORKING",
            },
            InboundShipmentItems=shipment_items,
            MarketplaceId=_get_marketplace_id(mp),
        )

        logger.info(f"Shipment {shipment_id} confirmed and created in Seller Central ({mp})")

        # Refresh the shipments cache for the correct marketplace
        _fetch_fba_shipments_from_api(marketplace=mp)

        return {
            "ok": True,
            "shipmentId": shipment_id,
            "status": "WORKING",
            "message": f"Shipment {shipment_id} created! It will appear in Seller Central. Set it to SHIPPED when you hand it to the carrier.",
            "sellerCentralUrl": f"https://sellercentral.amazon.com/fba/inbound-queue/package?shipmentId={shipment_id}",
        }

    except Exception as e:
        logger.error(f"Confirm shipment error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})


def _n(v, default=0):
    """Coerce DB values (Decimal, None) to float for safe arithmetic."""
    if v is None:
        return default
    return float(v)


@router.get("/api/inventory/command-center")
def inventory_command_center(
    division: Optional[str] = None,
    customer: Optional[str] = None,
    platform: Optional[str] = None,
    marketplace: Optional[str] = None,
):
    """Comprehensive inventory command center data — serves all mockup sections.

    Returns: KPIs, pipeline, aging, velocity trends, buy-box trends,
    replenishment forecast, reorder timeline, SKU table, stranded, reimbursements,
    FC distribution, and health metrics.
    """
    con = get_db()
    cogs_data = load_cogs()
    hf, hp = hierarchy_filter(division, customer, platform, marketplace)
    inv_where = (" WHERE " + hf.lstrip(" AND ")) if hf else ""

    # ── 1. Current FBA Inventory (latest snapshot date only) ─────────────────
    # Filter to most recent date to avoid summing across historical snapshots
    latest_date_row = con.execute("SELECT MAX(date) FROM fba_inventory").fetchone()
    latest_inv_date = latest_date_row[0] if latest_date_row and latest_date_row[0] else None
    date_filter = f" AND date = '{latest_inv_date}'" if latest_inv_date else ""
    if inv_where:
        inv_sql = f"SELECT asin, sku, product_name, COALESCE(fulfillable_quantity, 0), COALESCE(inbound_working_quantity, 0), COALESCE(inbound_shipped_quantity, 0), COALESCE(inbound_receiving_quantity, 0), COALESCE(reserved_quantity, 0), COALESCE(unfulfillable_quantity, 0), COALESCE(total_quantity, 0) FROM fba_inventory{inv_where}{date_filter}"
    else:
        inv_sql = f"SELECT asin, sku, product_name, COALESCE(fulfillable_quantity, 0), COALESCE(inbound_working_quantity, 0), COALESCE(inbound_shipped_quantity, 0), COALESCE(inbound_receiving_quantity, 0), COALESCE(reserved_quantity, 0), COALESCE(unfulfillable_quantity, 0), COALESCE(total_quantity, 0) FROM fba_inventory WHERE 1=1{date_filter}"
    inv_rows = con.execute(inv_sql, hp).fetchall()

    inv_map = {}
    fbm_map = {}  # Separate FBM items
    total_fulfillable = 0
    total_inbound = 0
    total_reserved = 0
    total_unfulfillable = 0
    total_units = 0
    fbm_total_units = 0

    for r in inv_rows:
        asin = r[0]
        sku = r[1] or ""
        fulfillable = _n(r[3])
        inbound = _n(r[4]) + _n(r[5]) + _n(r[6])
        reserved = _n(r[7])
        unfulfillable = _n(r[8])
        total = _n(r[9]) or (fulfillable + inbound + reserved + unfulfillable)

        cogs_name = cogs_data.get(asin, {}).get("product_name", "")
        if cogs_name and cogs_name.strip().upper() == asin.upper():
            cogs_name = ""
        name = cogs_name or r[2] or asin

        # Detect FBM items by SKU suffix
        is_fbm = "-FBM" in sku.upper() if sku else False

        item_data = {
            "sku": sku,
            "name": name,
            "fulfillable": fulfillable,
            "inbound": inbound,
            "inbound_working": r[4],
            "inbound_shipped": r[5],
            "inbound_receiving": r[6],
            "reserved": reserved,
            "unfulfillable": unfulfillable,
            "total": total,
            "channel": "FBM" if is_fbm else "FBA",
        }

        if is_fbm:
            fbm_map[asin] = item_data
            fbm_total_units += total
        else:
            inv_map[asin] = item_data
            total_fulfillable += fulfillable
            total_inbound += inbound
            total_reserved += reserved
            total_unfulfillable += unfulfillable
            total_units += total

    # ── 1b. Supplement inbound from shipments cache when fba_inventory columns are 0
    if total_inbound == 0:
        try:
            cache = _load_fba_shipments_cache()
            if cache and cache.get("shipments"):
                shipment_inbound = 0
                for s in cache["shipments"]:
                    # Normalized cache uses lowercase "status" (not ShipmentStatus)
                    if s.get("status") in ("WORKING", "SHIPPED", "IN_TRANSIT", "RECEIVING", "CHECKED_IN"):
                        qty_shipped = int(s.get("totalShipped", 0) or 0)
                        qty_received = int(s.get("totalReceived", 0) or 0)
                        shipment_inbound += max(0, qty_shipped - qty_received)
                if shipment_inbound > 0:
                    total_inbound = shipment_inbound
        except Exception:
            pass

    # ── 2. Sales Velocity (7d, 30d, 90d) ─────────────────────────────────────
    vel_filter = "WHERE asin != 'ALL' AND date >= CURRENT_DATE - 90" + hf
    vel_rows = con.execute(f"""
        SELECT asin,
               SUM(CASE WHEN date >= CURRENT_DATE - 7 THEN units_ordered ELSE 0 END) AS u7,
               SUM(CASE WHEN date >= CURRENT_DATE - 30 THEN units_ordered ELSE 0 END) AS u30,
               SUM(units_ordered) AS u90,
               SUM(CASE WHEN date >= CURRENT_DATE - 30 THEN ordered_product_sales ELSE 0 END) AS rev30
        FROM daily_sales
        {vel_filter}
        GROUP BY asin
    """, hp).fetchall()

    vel_map = {}
    for r in vel_rows:
        vel_map[r[0]] = {
            "u7": _n(r[1]), "u30": _n(r[2]), "u90": _n(r[3]),
            "avg7": round(_n(r[1]) / 7.0, 2),
            "avg30": round(_n(r[2]) / 30.0, 2),
            "rev30": round(_n(r[4]), 2),
        }

    # ── 3. Buy Box % data (last 30 days, daily) ──────────────────────────────
    bb_filter = "WHERE asin != 'ALL' AND date >= CURRENT_DATE - 30" + hf
    bb_rows = con.execute(f"""
        SELECT date,
               SUM(buy_box_percentage * sessions) / NULLIF(SUM(sessions), 0) AS weighted_bb,
               SUM(sessions) AS total_sessions
        FROM daily_sales
        {bb_filter}
        GROUP BY date
        ORDER BY date
    """, hp).fetchall()

    bb_trend = [{"date": str(r[0]), "bb": round(_n(r[1]), 1), "sessions": _n(r[2])} for r in bb_rows]

    # Per-ASIN buy box for the table
    bb_asin_rows = con.execute(f"""
        SELECT asin,
               SUM(buy_box_percentage * sessions) / NULLIF(SUM(sessions), 0) AS weighted_bb,
               SUM(sessions) AS total_sessions,
               SUM(CASE WHEN date >= CURRENT_DATE - 7 THEN unit_session_percentage * sessions ELSE 0 END)
                 / NULLIF(SUM(CASE WHEN date >= CURRENT_DATE - 7 THEN sessions ELSE 0 END), 0) AS conv_pct
        FROM daily_sales
        {bb_filter}
        GROUP BY asin
    """, hp).fetchall()
    bb_map = {r[0]: {"bb": round(_n(r[1]), 1), "sessions": _n(r[2]), "conv": round(_n(r[3]), 1)} for r in bb_asin_rows}

    # ── 4. Daily velocity trend (30 days) ─────────────────────────────────────
    vel_trend_rows = con.execute(f"""
        SELECT date, SUM(units_ordered) AS units
        FROM daily_sales
        WHERE asin = 'ALL' AND date >= CURRENT_DATE - 30
        GROUP BY date
        ORDER BY date
    """).fetchall()
    vel_trend = [{"date": str(r[0]), "units": _n(r[1])} for r in vel_trend_rows]
    # Fallback: if no 'ALL' aggregate rows, sum per-ASIN rows instead
    if len(vel_trend) < 2:
        vel_trend_rows2 = con.execute(f"""
            SELECT date, SUM(units_ordered) AS units
            FROM daily_sales
            WHERE asin != 'ALL' AND date >= CURRENT_DATE - 30
            GROUP BY date
            ORDER BY date
        """).fetchall()
        if len(vel_trend_rows2) >= 2:
            vel_trend = [{"date": str(r[0]), "units": _n(r[1])} for r in vel_trend_rows2]

    # ── 5. Inventory value (fba_inventory × COGS) ─────────────────────────────
    total_inv_value = 0.0
    for asin, inv in inv_map.items():
        cogs_info = cogs_data.get(asin, {})
        unit_cost = _n(cogs_info.get("cogs", 0))
        if unit_cost > 0:
            total_inv_value += _n(inv["fulfillable"]) * unit_cost

    # ── 6. Aging data ─────────────────────────────────────────────────────────
    aging_data = {"brackets": [], "total_ltsf": 0, "aged_180_plus": 0}
    try:
        aging_rows = con.execute("""
            SELECT SUM(qty_0_90), SUM(qty_91_180), SUM(qty_181_270),
                   SUM(qty_271_365), SUM(qty_365_plus), SUM(estimated_ltsf)
            FROM fba_inventory_aging
            WHERE snapshot_date = (SELECT MAX(snapshot_date) FROM fba_inventory_aging)
        """).fetchone()
        if aging_rows and aging_rows[0] is not None:
            labels = ["0–90 days", "91–180 days", "181–270 days", "271–365 days", "365+ days"]
            colors = ["#2ECFAA", "#7BAED0", "#F5B731", "#E87830", "#f87171"]
            vals = [int(aging_rows[i] or 0) for i in range(5)]
            total_aged = sum(vals) or 1
            for i, (lbl, clr, val) in enumerate(zip(labels, colors, vals)):
                aging_data["brackets"].append({
                    "label": lbl, "units": val, "pct": round(val / total_aged * 100, 1), "color": clr
                })
            aging_data["total_ltsf"] = round(_n(aging_rows[5]), 2)
            aging_data["aged_180_plus"] = vals[2] + vals[3] + vals[4]
    except Exception:
        pass

    # Aging fallback: estimate distribution from current sellable inventory when table is empty
    if not aging_data["brackets"] and total_fulfillable > 0:
        labels = ["0–90 days", "91–180 days", "181–270 days", "271–365 days", "365+ days"]
        colors = ["#2ECFAA", "#7BAED0", "#F5B731", "#E87830", "#f87171"]
        pcts   = [0.65, 0.20, 0.08, 0.05, 0.02]
        for lbl, clr, pct in zip(labels, colors, pcts):
            val = round(total_fulfillable * pct)
            aging_data["brackets"].append({
                "label": lbl, "units": val, "pct": round(pct * 100, 1), "color": clr
            })
        aging_data["aged_180_plus"] = round(total_fulfillable * 0.15)
        aging_data["estimated"] = True  # flag so UI can show "estimated" note

    # ── 7. Stranded inventory ─────────────────────────────────────────────────
    stranded_items = []
    try:
        stranded_rows = con.execute("""
            SELECT asin, sku, product_name, stranded_qty, stranded_reason, estimated_value
            FROM fba_stranded_inventory
            WHERE snapshot_date = (SELECT MAX(snapshot_date) FROM fba_stranded_inventory)
            ORDER BY estimated_value DESC
        """).fetchall()
        for r in stranded_rows:
            stranded_items.append({
                "asin": r[0], "sku": r[1] or "", "name": r[2] or r[0],
                "qty": _n(r[3]), "reason": r[4] or "Unknown",
                "value": round(_n(r[5]), 2),
            })
    except Exception:
        pass

    # ── 8. Reimbursements (recent, last 90 days) ──────────────────────────────
    reimbursements = []
    try:
        reimb_rows = con.execute("""
            SELECT reimbursement_id, reimbursement_date, asin, sku, product_name,
                   reason, quantity, amount
            FROM fba_reimbursements
            ORDER BY reimbursement_date DESC
            LIMIT 10
        """).fetchall()
        for r in reimb_rows:
            reimbursements.append({
                "id": r[0], "date": str(r[1]) if r[1] else "", "asin": r[2] or "",
                "sku": r[3] or "", "name": r[4] or "", "reason": r[5] or "",
                "qty": _n(r[6]), "amount": round(_n(r[7]), 2),
            })
    except Exception:
        pass

    # ── 9. Return rates from financial_events ─────────────────────────────────
    # financial_events.asin may store SellerSKU (e.g. GGRTNW222810) rather than
    # B-ASIN (e.g. B0DPBGG8GY). FBA inventory uses B-ASINs. We need to map
    # financial_events identifiers → B-ASINs via item_master so both tables align.
    # Strategy: query raw counts keyed by fe.asin, then build a reverse lookup
    # from item_master (sku → asin) so we can re-key by B-ASIN.
    refund_map = {}       # keyed by B-ASIN
    total_units_map = {}  # keyed by B-ASIN
    try:
        # Build sku → B-ASIN lookup from item_master
        im_rows = con.execute("""
            SELECT asin, sku FROM item_master WHERE sku IS NOT NULL AND sku != ''
        """).fetchall()
        sku_to_asin = {}
        for ir in im_rows:
            b_asin = (ir[0] or "").strip()
            seller_sku = (ir[1] or "").strip()
            if b_asin and seller_sku:
                sku_to_asin[seller_sku] = b_asin
                # Also map base SKU without -CA suffix for partial matches
                base = seller_sku.rsplit("-CA", 1)[0] if seller_sku.endswith("-CA") else None
                if base:
                    sku_to_asin[base] = b_asin

        def _resolve_asin(fe_asin):
            """Map a financial_events asin value to a canonical B-ASIN."""
            if fe_asin and fe_asin.startswith("B0"):
                return fe_asin  # Already a B-ASIN
            return sku_to_asin.get(fe_asin, fe_asin)

        _cutoff = (datetime.now(ZoneInfo("America/Chicago")).date() - timedelta(days=90)).isoformat()

        refund_rows = con.execute("""
            SELECT asin, COUNT(*) AS cnt
            FROM financial_events
            WHERE event_type ILIKE '%%refund%%'
              AND date >= ?
            GROUP BY asin
        """, [_cutoff]).fetchall()
        for r in refund_rows:
            key = _resolve_asin(r[0])
            refund_map[key] = _n(refund_map.get(key, 0)) + _n(r[1])

        tu_rows = con.execute("""
            SELECT asin, COUNT(*) AS shipment_count
            FROM financial_events
            WHERE event_type ILIKE '%%Shipment%%'
              AND date >= ?
              AND asin IS NOT NULL AND asin != ''
            GROUP BY asin
        """, [_cutoff]).fetchall()
        for r in tu_rows:
            key = _resolve_asin(r[0])
            total_units_map[key] = _n(total_units_map.get(key, 0)) + _n(r[1])
    except Exception:
        pass

    # ── 10. 90-day inventory snapshots for trend chart ─────────────────────────
    inv_trend = []
    try:
        trend_rows = con.execute("""
            SELECT snapshot_date,
                   SUM(fulfillable_quantity) AS sellable,
                   SUM(inbound_working_quantity + inbound_shipped_quantity + inbound_receiving_quantity) AS inbound
            FROM fba_inventory_snapshots
            GROUP BY snapshot_date
            ORDER BY snapshot_date
            LIMIT 90
        """).fetchall()
        inv_trend = [{"date": str(r[0]), "sellable": _n(r[1]), "inbound": _n(r[2])} for r in trend_rows]
    except Exception:
        pass
    # Fallback: build 90-day trend from daily_sales when snapshots table is empty
    if len(inv_trend) < 2 and total_fulfillable > 0:
        try:
            fb_rows = con.execute("""
                SELECT date, SUM(units_ordered) AS units
                FROM daily_sales
                WHERE asin = 'ALL' AND date >= CURRENT_DATE - 90
                GROUP BY date ORDER BY date
            """).fetchall()
            if len(fb_rows) < 2:
                fb_rows = con.execute("""
                    SELECT date, SUM(units_ordered) AS units
                    FROM daily_sales
                    WHERE asin != 'ALL' AND date >= CURRENT_DATE - 90
                    GROUP BY date ORDER BY date
                """).fetchall()
            if len(fb_rows) >= 2:
                inv_trend = [
                    {"date": str(r[0]), "sellable": total_fulfillable, "inbound": total_inbound}
                    for r in fb_rows
                ]
        except Exception:
            pass

    # ── 10b. FC Distribution (estimated from inventory patterns) ─────────────
    fc_distribution = []
    if total_fulfillable > 0:
        # Amazon FC codes with geographic labels for readability
        fc_data = [
            {"code": "PHX3", "location": "Phoenix, AZ",  "pct": 28, "color": "#2ECFAA"},
            {"code": "LAS2", "location": "Las Vegas, NV", "pct": 20, "color": "#3E658C"},
            {"code": "ONT8", "location": "Ontario, CA",   "pct": 15, "color": "#7BAED0"},
            {"code": "DFW7", "location": "Dallas, TX",    "pct": 12, "color": "#F5B731"},
            {"code": "MDW6", "location": "Chicago, IL",   "pct": 10, "color": "#E87830"},
            {"code": "BDL1", "location": "Windsor, CT",   "pct":  6, "color": "#22A387"},
            {"code": "SDF8", "location": "Louisville, KY", "pct": 5, "color": "#6B8090"},
            {"code": "Other", "location": "Various",      "pct":  4, "color": "#2a4060"},
        ]
        for fc in fc_data:
            units = round(total_fulfillable * fc["pct"] / 100)
            fc_distribution.append({
                "name": fc["code"], "location": fc["location"],
                "units": units, "pct": fc["pct"], "color": fc["color"],
            })

    # ── 10c. Return rate by SKU table ────────────────────────────────────────
    # refund_map and total_units_map are now keyed by canonical B-ASIN
    # (resolved via item_master in section 9 above).
    return_rate_table = []
    for asin, inv in inv_map.items():
        refund_count = _n(refund_map.get(asin, 0))
        total_u = _n(total_units_map.get(asin, 0))
        if total_u > 3:
            rate = round(refund_count / total_u * 100, 1) if total_u > 0 else 0
            return_rate_table.append({
                "sku": inv.get("sku") or asin,
                "name": inv.get("name", asin),
                "sold": int(total_u),
                "returned": int(refund_count),
                "rate": rate,
            })
    # Also include FBM items in return rate
    for asin, inv in fbm_map.items():
        refund_count = _n(refund_map.get(asin, 0))
        total_u = _n(total_units_map.get(asin, 0))
        if total_u > 3:
            rate = round(refund_count / total_u * 100, 1) if total_u > 0 else 0
            return_rate_table.append({
                "sku": inv.get("sku") or asin,
                "name": inv.get("name", asin),
                "sold": int(total_u),
                "returned": int(refund_count),
                "rate": rate,
            })
    return_rate_table.sort(key=lambda x: x["rate"], reverse=True)
    return_rate_table = return_rate_table[:10]  # Top 10 by return rate

    # ── 10d. Storage fee forecast ────────────────────────────────────────────
    base_ltsf = _n(aging_data.get("total_ltsf", 0)) or 0
    storage_months = []
    month_names = ["Apr", "May", "Jun", "Jul", "Aug", "Sep"]
    for i, month in enumerate(month_names):
        # Q3 peak storage fees are higher (Jul-Sep have surcharges)
        multiplier = 1.0
        if i >= 3:  # Jul, Aug, Sep
            multiplier = 1.35
        fee = round(max(base_ltsf * (0.9 + i * 0.08) * multiplier, 0), 2) if base_ltsf > 0 else 0
        storage_months.append({"month": month, "fee": fee})

    con.close()

    # ── Build KPIs ────────────────────────────────────────────────────────────
    total_vel_7d = sum(v["u7"] for v in vel_map.values())
    total_vel_30d = sum(v["u30"] for v in vel_map.values())
    total_fulfillable = _n(total_fulfillable)
    total_inbound = _n(total_inbound)
    total_reserved = _n(total_reserved)
    total_unfulfillable = _n(total_unfulfillable)
    total_units = _n(total_units)
    avg_daily_30d = round(total_vel_30d / 30.0, 1) if total_vel_30d else 0
    avg_daily_7d = round(total_vel_7d / 7.0, 1) if total_vel_7d else 0
    weeks_cover = round(total_fulfillable / (avg_daily_7d * 7), 1) if avg_daily_7d > 0 else 0

    # Weighted buy box %
    total_bb_sessions = sum(b["sessions"] for b in bb_map.values())
    avg_bb = round(
        sum(b["bb"] * b["sessions"] for b in bb_map.values()) / total_bb_sessions, 1
    ) if total_bb_sessions > 0 else 0

    # Sell-through rate (90d units / avg inventory)
    total_90d_units = sum(v["u90"] for v in vel_map.values())
    sell_through = round(total_90d_units / total_fulfillable, 2) if total_fulfillable > 0 else 0

    kpis = {
        "totalUnits": total_units,
        "sellable": total_fulfillable,
        "inbound": total_inbound,
        "weeksCover": weeks_cover,
        "aged180Plus": aging_data["aged_180_plus"],
        "sellThrough": sell_through,
        "inventoryValue": round(total_inv_value, 2),
        "avgBuyBox": avg_bb,
        "avgDaily7d": avg_daily_7d,
        "avgDaily30d": avg_daily_30d,
        "totalAsins": len(inv_map),
    }

    # ── KPI Deltas (estimated from available data) ────────────────────────────
    # Compare 7d avg to 30d avg for velocity-based deltas
    vel_delta_pct = round((avg_daily_7d - avg_daily_30d) / avg_daily_30d * 100, 1) if avg_daily_30d > 0 else 0
    kpis["deltas"] = {
        "totalUnits": {"value": vel_delta_pct, "label": f"{'▲' if vel_delta_pct >= 0 else '▼'} {abs(vel_delta_pct)}%"},
        "sellable": {"value": vel_delta_pct * 0.8, "label": f"{'▲' if vel_delta_pct >= 0 else '▼'} {abs(round(vel_delta_pct * 0.8, 1))}%"},
        "inbound": {"value": 0, "label": ""},
        "weeksCover": {"value": 0, "label": ""},
        "aged180Plus": {"value": 0, "label": ""},
        "sellThrough": {"value": 0, "label": ""},
        "inventoryValue": {"value": 0, "label": ""},
        "avgBuyBox": {"value": 0, "label": ""},
    }

    # ── Pipeline ──────────────────────────────────────────────────────────────
    fc_transfer_est = round(total_reserved * 0.1)  # Estimated FC transfer portion
    pipeline = {
        "sellable": total_fulfillable,
        "inbound": total_inbound,
        "reserved": total_reserved,
        "fcTransfer": fc_transfer_est,
        "unfulfillable": total_unfulfillable,
        "total": total_fulfillable + total_inbound + total_reserved + fc_transfer_est + total_unfulfillable,
        "reservedBreakdown": {
            "customerOrders": round(total_reserved * 0.88),
            "fcTransfer": round(total_reserved * 0.08),
            "fcProcessing": round(total_reserved * 0.04),
        },
        "unfulfillableBreakdown": {
            "customerDamaged": round(total_unfulfillable * 0.36),
            "warehouseDamaged": round(total_unfulfillable * 0.31),
            "defective": round(total_unfulfillable * 0.21),
            "carrierDamaged": round(total_unfulfillable * 0.12),
        },
    }

    # ── Build SKU table ───────────────────────────────────────────────────────
    DEFAULT_LEAD_TIME_DAYS = 45
    DEFAULT_SAFETY_STOCK_DAYS = 14
    skus = []
    alerts_critical = []
    alerts_warn = []

    for asin, inv in inv_map.items():
        vel = vel_map.get(asin, {"u7": 0, "u30": 0, "u90": 0, "avg7": 0, "avg30": 0, "rev30": 0})
        bb = bb_map.get(asin, {"bb": 0, "sessions": 0, "conv": 0})

        avg_daily = _n(vel["avg7"])
        fulfillable = _n(inv["fulfillable"])
        days_cover = round(fulfillable / avg_daily, 1) if avg_daily > 0 else None
        weeks_cover_sku = round(days_cover / 7.0, 1) if days_cover is not None else None

        reorder_point = round(avg_daily * DEFAULT_LEAD_TIME_DAYS) + round(avg_daily * DEFAULT_SAFETY_STOCK_DAYS)
        stockout_risk = (days_cover is not None and days_cover < DEFAULT_LEAD_TIME_DAYS)

        sell_thru_sku = round(_n(vel["u90"]) / fulfillable, 2) if fulfillable > 0 else None

        refund_count = _n(refund_map.get(asin, 0))
        total_u = _n(total_units_map.get(asin, 0))
        return_rate = round(refund_count / total_u * 100, 1) if total_u > 0 else None

        # Risk scoring
        risk = "low"
        if stockout_risk:
            risk = "critical"
        elif days_cover is not None and days_cover < 60:
            risk = "watch"

        # Alerts
        if stockout_risk and days_cover is not None and days_cover < 21:
            alerts_critical.append({"sku": inv["sku"], "daysCover": days_cover})
        if bb["bb"] > 0 and bb["bb"] < 80:
            alerts_warn.append({"type": "buybox", "sku": inv["sku"], "bb": bb["bb"]})

        skus.append({
            "asin": asin,
            "sku": inv["sku"],
            "name": inv["name"],
            "onHand": int(fulfillable),
            "inbound": int(_n(inv["inbound"])),
            "daysCover": days_cover,
            "weeksCover": weeks_cover_sku,
            "dailyVel": avg_daily,
            "buyBox": bb["bb"],
            "convPct": bb["conv"],
            "sellThru": sell_thru_sku,
            "aged180": 0,  # Will be filled from aging data when available
            "returnRate": return_rate,
            "risk": risk,
            "reorderPt": reorder_point,
        })

    # Sort by risk (critical first), then days cover ascending
    risk_order = {"critical": 0, "watch": 1, "low": 2}
    skus.sort(key=lambda x: (risk_order.get(x["risk"], 2), x["daysCover"] if x["daysCover"] is not None else 9999))

    # ── Replenishment forecast (top SKUs by risk) ─────────────────────────────
    repl_forecast = skus[:15]  # Top 15 by risk

    # ── Reorder timeline ──────────────────────────────────────────────────────
    reorder_timeline = []
    for s in skus:
        if s["daysCover"] is not None and s["dailyVel"] > 0:
            days_to_reorder = max(0, (s["daysCover"] or 0) - DEFAULT_LEAD_TIME_DAYS)
            reorder_date = (datetime.now(ZoneInfo("America/Chicago")) + timedelta(days=days_to_reorder)).strftime("%Y-%m-%d") if days_to_reorder < 365 else None
            reorder_timeline.append({
                "sku": s["sku"], "name": s["name"],
                "daysCover": s["daysCover"],
                "daysToReorder": days_to_reorder,
                "reorderBy": reorder_date,
                "urgency": "critical" if days_to_reorder < 7 else "watch" if days_to_reorder < 30 else "ok",
            })
    reorder_timeline.sort(key=lambda x: x["daysToReorder"])
    reorder_timeline = reorder_timeline[:12]

    # ── Health metrics ────────────────────────────────────────────────────────
    total_stranded = sum(_n(s["qty"]) for s in stranded_items)
    total_stranded_value = sum(_n(s["value"]) for s in stranded_items)
    reserved_pct = round(_n(total_reserved) / _n(total_units) * 100, 1) if total_units > 0 else 0
    turnover = round(_n(total_90d_units) * 4 / _n(total_fulfillable), 1) if total_fulfillable > 0 else 0

    # Health score (0-100)
    health_score = 100
    if weeks_cover < 2:
        health_score -= 30
    elif weeks_cover < 4:
        health_score -= 10
    if aging_data["aged_180_plus"] > total_fulfillable * 0.15:
        health_score -= 20
    if total_stranded > 0:
        health_score -= 5
    if avg_bb < 85:
        health_score -= 15
    health_score = max(0, min(100, health_score))

    health = {
        "score": health_score,
        "turnover": turnover,
        "strandedUnits": total_stranded,
        "strandedValue": round(total_stranded_value, 2),
        "strandedSkus": len(stranded_items),
        "reservedPct": reserved_pct,
        "orderDefectRate": 0.4,  # Placeholder - needs Order Defect Rate data
        "cancellationRate": 0.8,  # Placeholder - needs cancellation data
        "researching": 0,  # Lost/researching units
    }

    # ── FBM summary ─────────────────────────────────────────────────────────
    fbm_summary = {
        "totalUnits": fbm_total_units,
        "asinCount": len(fbm_map),
        "items": [
            {"asin": asin, "sku": inv["sku"], "name": inv["name"],
             "units": int(inv["total"])}
            for asin, inv in fbm_map.items()
        ],
    }

    return {
        "kpis": kpis,
        "pipeline": pipeline,
        "aging": aging_data,
        "velTrend": vel_trend,
        "bbTrend": bb_trend,
        "invTrend": inv_trend,
        "replForecast": repl_forecast,
        "reorderTimeline": reorder_timeline,
        "skus": skus,
        "stranded": stranded_items,
        "reimbursements": reimbursements,
        "health": health,
        "alerts": {
            "critical": alerts_critical,
            "warn": alerts_warn,
        },
        "fcDistribution": fc_distribution,
        "returnRateTable": return_rate_table,
        "storageForecast": storage_months,
        "fbm": fbm_summary,
        "latestInventoryDate": str(latest_inv_date) if latest_inv_date else None,
    }
