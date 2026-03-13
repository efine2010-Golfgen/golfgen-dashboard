"""
Supply Chain Report — Unified PO / OTW / Invoice views.

Data Store: JSON file (supply_chain_store.json) with flat denormalized records.
Each record = one SKU × PO × Container × Invoice row.

Upload: Single Excel file with 2 tabs:
  Tab 1 "PO & Invoice Data" — PO#, Factory, Invoice#, Container#, SKU, Desc, Units, CBM, Est.Arrival, Status
  Tab 2 "Shipment Data"     — Container#, PO#, Shipper, HB#, Type, Vessel, ETD, ETD Port, ETA, ETA Port,
                               ETA Delivery, Delivery Location, Actual Delivery

Three GET views derive from the same store:
  /api/supply-chain/otw      — container-level OTW summary
  /api/supply-chain/po       — PO-level summary with ordered vs shipped
  /api/supply-chain/invoices — invoice-level summary
  /api/supply-chain/upload   — single-file upload (POST)
"""

from __future__ import annotations
import json, os, re
from datetime import datetime, date
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo
from fastapi import APIRouter, UploadFile, File, Depends
import openpyxl

from core.config import DB_DIR
from core.auth import require_auth

router = APIRouter(prefix="/api/supply-chain", tags=["supply-chain"])

STORE_PATH = os.path.join(DB_DIR, "supply_chain_store.json")
TZ = ZoneInfo("America/Chicago")


# ── helpers ──────────────────────────────────────────────────────────────

def _load_store() -> dict:
    """Load the data store. Returns {"records": [...], "lastUpload": ..., "sourceFile": ...}"""
    if os.path.exists(STORE_PATH):
        with open(STORE_PATH, "r") as f:
            return json.load(f)
    return {"records": [], "lastUpload": None, "sourceFile": None}


def _save_store(data: dict):
    with open(STORE_PATH, "w") as f:
        json.dump(data, f, default=str)


def _dash(v):
    """Return '—' for empty/None values."""
    if v is None or (isinstance(v, str) and v.strip() in ("", "—", "-", "N/A", "n/a", "(blank)")):
        return "—"
    return v


def _parse_date(v):
    """Parse various date formats to ISO string or '—'."""
    if v is None:
        return "—"
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        v = v.strip()
        if v in ("—", "-", "", "(blank)", "N/A"):
            return "—"
        # Try common formats
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d", "%d-%b-%Y", "%b %d, %Y"):
            try:
                return datetime.strptime(v, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return v  # return as-is if can't parse
    return "—"


def _derive_eta_month(eta_str: str) -> str:
    """Derive 'Jan 2026' style month from an ISO date string."""
    if eta_str == "—" or not eta_str:
        return "—"
    try:
        dt = datetime.strptime(eta_str, "%Y-%m-%d")
        return dt.strftime("%b %Y")
    except (ValueError, TypeError):
        return "—"


def _cell_val(ws, row, col):
    """Get cell value from openpyxl worksheet."""
    cell = ws.cell(row=row, column=col)
    return cell.value


def _str(v):
    if v is None:
        return ""
    return str(v).strip()


def _float_or_zero(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _int_or_zero(v):
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return 0


# ── Upload endpoint ──────────────────────────────────────────────────────

@router.post("/upload")
async def upload_supply_chain(file: UploadFile = File(...), _user=Depends(require_auth)):
    """
    Upload a single Excel file with up to 3 tabs:
      - PO Summary: PO-level data (Factory, PO#, Units, FOB Date, Status, financials)
      - Logistics Tracking: Container-level shipping data (Container#, PO#, ETD, ETA, etc.)
      - Invoice Summary: Invoice-level data (PO#, Container#, Invoice#, Date)

    Records are keyed by {PO#}_{Container#} for container-level, or {PO#}_PO for PO-only.
    Handles both SKU-level and PO-level spreadsheets.
    """
    import tempfile
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    content = await file.read()
    tmp.write(content)
    tmp.close()

    try:
        wb = openpyxl.load_workbook(tmp.name, data_only=True)
    except Exception as e:
        os.unlink(tmp.name)
        return {"status": "error", "message": f"Cannot read Excel file: {e}"}

    try:
        return _parse_workbook(wb, file.filename)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"status": "error", "message": f"Parse error: {e}"}
    finally:
        os.unlink(tmp.name)


def _normalize_status(raw: str) -> str:
    """Normalize status text to one of: Delivered, In Transit, Open, Closed, or —."""
    if raw == "—":
        return "—"
    sl = raw.lower().strip()
    if "deliver" in sl:
        return "Delivered"
    elif "transit" in sl or "ship" in sl or "sail" in sl:
        return "In Transit"
    elif "close" in sl:
        return "Closed"
    elif "open" in sl or "pending" in sl or "order" in sl:
        return "Open"
    return raw


def _find_header_row(ws, required_keywords: list[str], max_scan: int = 15) -> tuple:
    """Scan first N rows for a header row containing required keywords. Returns (row_number, col_map) or (None, {})."""
    for r in range(1, min(ws.max_row + 1, max_scan)):
        row_vals = [_str(_cell_val(ws, r, c)).lower() for c in range(1, ws.max_column + 1)]
        joined = " ".join(row_vals)
        if all(kw in joined for kw in required_keywords):
            return r, row_vals
    return None, []


def _map_columns(ws, header_row: int, mappings: dict) -> dict:
    """Build column map from header row. mappings = {field_name: [list of possible header substrings]}."""
    col_map = {}
    for c in range(1, ws.max_column + 1):
        hdr = _str(_cell_val(ws, header_row, c)).lower()
        if not hdr:
            continue
        for field, matchers in mappings.items():
            if field in col_map:
                continue  # already mapped, don't overwrite
            for m in matchers:
                if callable(m):
                    if m(hdr):
                        col_map[field] = c
                        break
                elif m == hdr:
                    col_map[field] = c
                    break
    return col_map


def _parse_workbook(wb, filename: str) -> dict:
    """Core parser: reads PO Summary, Logistics Tracking, and Invoice Summary sheets."""

    # ── Detect sheets ──
    po_sheet = ship_sheet = inv_sheet = None
    for name in wb.sheetnames:
        lower = name.lower().strip()
        if lower in ("po summary", "po_summary", "po data", "purchase orders", "factory po summary"):
            po_sheet = wb[name]
        elif "po" in lower and ("invoice" in lower or "data" in lower):
            if po_sheet is None:
                po_sheet = wb[name]
        elif "logistics" in lower or "shipment" in lower or "otw" in lower or "shipping" in lower:
            ship_sheet = wb[name]
        elif "invoice" in lower and "summary" in lower:
            inv_sheet = wb[name]

    if po_sheet is None and ship_sheet is None:
        return {"status": "error", "message": "Could not find PO Summary or Logistics Tracking sheet in the workbook."}

    records = {}  # keyed by record_id
    now_iso = datetime.now(TZ).isoformat()

    def _make_record(po_num, container="—"):
        container_key = container if container != "—" else "PO"
        rid = f"{po_num}_{container_key}"
        if rid not in records:
            records[rid] = {
                "record_id": rid, "po_number": po_num, "factory": "—",
                "payment_terms": "—", "invoice_number": "—", "invoice_date": "—",
                "container_number": container, "sku": "—", "description": "—",
                "units": 0, "cbm": 0.0, "x_factory_date": "—",
                "container_eta": "—", "eta_month": "—", "status": "—",
                "shipper": "—", "hb_number": "—", "container_type": "—",
                "vessel": "—", "etd": "—", "etd_port": "—", "eta_port": "—",
                "eta_delivery": "—", "delivery_location": "—", "actual_delivery": "—",
                "total_cost": 0.0, "landed_cost": 0.0, "freight_cost": 0.0,
                "duty_cost": 0.0, "last_updated": now_iso,
            }
        return records[rid]

    # ── 1. Parse Logistics Tracking (container × PO level) ──
    if ship_sheet is not None:
        hr, _ = _find_header_row(ship_sheet, ["po"])
        if hr is None:
            hr, _ = _find_header_row(ship_sheet, ["container"])
        if hr is not None:
            sc = _map_columns(ship_sheet, hr, {
                "container_number": [lambda h: ("container" in h and "#" in h) or h in ("cntr#", "cntr #", "container#")],
                "po_number": [lambda h: "po" in h and "invoice" not in h and "#" in h or h in ("po#", "po #", "po")],
                "shipper": [lambda h: "shipper" in h],
                "hb_number": [lambda h: "hb" in h or "hbl" in h],
                "container_type": [lambda h: "mode" in h or ("type" in h and "delivery" not in h)],
                "vessel": [lambda h: "vessel" in h],
                "etd": [lambda h: ("etd" in h and "port" not in h) or h == "etd origin"],
                "etd_port": [lambda h: ("departure" in h and "port" in h)],
                "eta_discharge": [lambda h: ("eta" in h and "discharge" in h) or (h == "eta discharge")],
                "eta_port": [lambda h: "arrival" in h and "port" in h],
                "eta_delivery": [lambda h: "eta" in h and ("final" in h or "deliv" in h or "dest" in h)],
                "delivery_location": [lambda h: "final" in h and "location" in h],
                "actual_delivery": [lambda h: "delivery" in h and "date" in h],
                "invoice_number": [lambda h: "invoice" in h],
                "units": [lambda h: h in ("units", "qty", "quantity") or (h == "unit")],
                "cbm": [lambda h: "cbm" in h],
                "status": [lambda h: h == "status"],
                "freight_cost": [lambda h: "freight" in h and "cost" in h],
                "duty_cost": [lambda h: "duty" in h and "cost" in h],
            })

            for r in range(hr + 1, ship_sheet.max_row + 1):
                cntr = _str(_cell_val(ship_sheet, r, sc.get("container_number", 1)))
                po = _str(_cell_val(ship_sheet, r, sc.get("po_number", 2)))
                if not cntr and not po:
                    continue
                if not po or po.lower() in ("total", "subtotal", "grand total"):
                    continue
                if not cntr:
                    cntr = "—"

                rec = _make_record(po, cntr)
                rec["shipper"] = _dash(_str(_cell_val(ship_sheet, r, sc.get("shipper", 0)))) if sc.get("shipper") else rec["shipper"]
                rec["hb_number"] = _dash(_str(_cell_val(ship_sheet, r, sc.get("hb_number", 0)))) if sc.get("hb_number") else rec["hb_number"]
                rec["container_type"] = _dash(_str(_cell_val(ship_sheet, r, sc.get("container_type", 0)))) if sc.get("container_type") else rec["container_type"]
                rec["vessel"] = _dash(_str(_cell_val(ship_sheet, r, sc.get("vessel", 0)))) if sc.get("vessel") else rec["vessel"]
                rec["etd"] = _parse_date(_cell_val(ship_sheet, r, sc["etd"])) if sc.get("etd") else rec["etd"]
                rec["etd_port"] = _dash(_str(_cell_val(ship_sheet, r, sc["etd_port"]))) if sc.get("etd_port") else rec["etd_port"]
                rec["eta_port"] = _dash(_str(_cell_val(ship_sheet, r, sc["eta_port"]))) if sc.get("eta_port") else rec["eta_port"]
                rec["eta_delivery"] = _parse_date(_cell_val(ship_sheet, r, sc["eta_delivery"])) if sc.get("eta_delivery") else rec["eta_delivery"]
                rec["delivery_location"] = _dash(_str(_cell_val(ship_sheet, r, sc["delivery_location"]))) if sc.get("delivery_location") else rec["delivery_location"]
                rec["actual_delivery"] = _parse_date(_cell_val(ship_sheet, r, sc["actual_delivery"])) if sc.get("actual_delivery") else rec["actual_delivery"]
                # ETA (discharge / arrival)
                eta_val = _parse_date(_cell_val(ship_sheet, r, sc["eta_discharge"])) if sc.get("eta_discharge") else "—"
                if eta_val != "—":
                    rec["container_eta"] = eta_val
                    rec["eta_month"] = _derive_eta_month(eta_val)
                # Invoice from logistics
                inv = _dash(_str(_cell_val(ship_sheet, r, sc["invoice_number"]))) if sc.get("invoice_number") else "—"
                if inv != "—":
                    rec["invoice_number"] = inv
                # Units and CBM
                if sc.get("units"):
                    rec["units"] = _int_or_zero(_cell_val(ship_sheet, r, sc["units"]))
                if sc.get("cbm"):
                    rec["cbm"] = _float_or_zero(_cell_val(ship_sheet, r, sc["cbm"]))
                # Status
                if sc.get("status"):
                    rec["status"] = _normalize_status(_dash(_str(_cell_val(ship_sheet, r, sc["status"]))))
                # Freight / duty costs
                if sc.get("freight_cost"):
                    rec["freight_cost"] = _float_or_zero(_cell_val(ship_sheet, r, sc["freight_cost"]))
                if sc.get("duty_cost"):
                    rec["duty_cost"] = _float_or_zero(_cell_val(ship_sheet, r, sc["duty_cost"]))

    # ── 2. Parse PO Summary (PO-level data) ──
    if po_sheet is not None:
        hr, _ = _find_header_row(po_sheet, ["po", "unit"])
        if hr is None:
            hr, _ = _find_header_row(po_sheet, ["po#"])
        if hr is not None:
            pc = _map_columns(po_sheet, hr, {
                "po_number": [lambda h: ("po" in h and "#" in h) or h in ("po#", "po #", "po number")],
                "factory": [lambda h: "factory" in h or "supplier" in h],
                "payment_terms": [lambda h: "payment" in h and "term" in h],
                "sku": [lambda h: h in ("sku", "item #", "item#", "item number", "item no")],
                "description": [lambda h: "description" in h or "desc" in h],
                "units": [lambda h: h in ("units", "qty", "quantity") or (h == "unit")],
                "cbm": [lambda h: h == "cbm"],
                "fob_date": [lambda h: "fob" in h and ("date" in h or h == "fob")],
                "est_arrival": [lambda h: ("est" in h and "arrival" in h) or h == "est arrival"],
                "status": [lambda h: h == "status"],
                "total_cost": [lambda h: "total" in h and "$" in h],
                "landed_cost": [lambda h: "landed" in h and "$" in h],
                "ocean_freight": [lambda h: "ocean" in h and "frt" in h],
                "customs": [lambda h: "customs" in h and "$" in h],
            })

            has_sku = "sku" in pc

            for r in range(hr + 1, po_sheet.max_row + 1):
                po_num = _str(_cell_val(po_sheet, r, pc.get("po_number", 1)))
                if not po_num or po_num.lower() in ("total", "subtotal", "grand total"):
                    continue

                # Determine if we already have container-level records for this PO
                existing_po_recs = [rid for rid in records if rid.startswith(f"{po_num}_") and not rid.endswith("_PO")]

                if has_sku:
                    # SKU-level sheet: create individual records
                    sku = _str(_cell_val(po_sheet, r, pc["sku"]))
                    container = _dash(_str(_cell_val(po_sheet, r, pc.get("container_number", 0)))) if pc.get("container_number") else "—"
                    container_key = container if container != "—" else "PO"
                    rid = f"{po_num}_{container_key}_{sku}" if sku else f"{po_num}_{container_key}"
                    if rid not in records:
                        records[rid] = _make_record(po_num, container).copy()
                        records[rid]["record_id"] = rid
                    rec = records[rid]
                    rec["sku"] = _dash(sku)
                    rec["description"] = _dash(_str(_cell_val(po_sheet, r, pc["description"]))) if pc.get("description") else "—"
                else:
                    # PO-level sheet (no SKU): enrich existing container records, or create PO-level record
                    if existing_po_recs:
                        # Enrich all container-level records for this PO with PO-level fields
                        for rid in existing_po_recs:
                            rec = records[rid]
                            rec["factory"] = _dash(_str(_cell_val(po_sheet, r, pc["factory"]))) if pc.get("factory") else rec.get("factory", "—")
                            rec["payment_terms"] = _dash(_str(_cell_val(po_sheet, r, pc["payment_terms"]))) if pc.get("payment_terms") else rec.get("payment_terms", "—")
                            if pc.get("units"):
                                u = _int_or_zero(_cell_val(po_sheet, r, pc["units"]))
                                if u > 0 and rec["units"] == 0:
                                    rec["units"] = u
                            if pc.get("cbm"):
                                c = _float_or_zero(_cell_val(po_sheet, r, pc["cbm"]))
                                if c > 0 and rec["cbm"] == 0.0:
                                    rec["cbm"] = c
                            rec["x_factory_date"] = _parse_date(_cell_val(po_sheet, r, pc["fob_date"])) if pc.get("fob_date") else rec.get("x_factory_date", "—")
                            if pc.get("est_arrival"):
                                ea = _parse_date(_cell_val(po_sheet, r, pc["est_arrival"]))
                                if ea != "—" and rec.get("container_eta", "—") == "—":
                                    rec["container_eta"] = ea
                                    rec["eta_month"] = _derive_eta_month(ea)
                            if pc.get("status"):
                                s = _normalize_status(_dash(_str(_cell_val(po_sheet, r, pc["status"]))))
                                if s != "—" and rec.get("status", "—") == "—":
                                    rec["status"] = s
                            if pc.get("total_cost"):
                                rec["total_cost"] = _float_or_zero(_cell_val(po_sheet, r, pc["total_cost"]))
                            if pc.get("landed_cost"):
                                rec["landed_cost"] = _float_or_zero(_cell_val(po_sheet, r, pc["landed_cost"]))
                            if pc.get("ocean_freight"):
                                rec["freight_cost"] = _float_or_zero(_cell_val(po_sheet, r, pc["ocean_freight"]))
                            if pc.get("customs"):
                                rec["duty_cost"] = _float_or_zero(_cell_val(po_sheet, r, pc["customs"]))
                        continue  # don't create a duplicate PO-level record
                    else:
                        rec = _make_record(po_num, "—")

                # Set PO-level fields
                rec["factory"] = _dash(_str(_cell_val(po_sheet, r, pc["factory"]))) if pc.get("factory") else rec.get("factory", "—")
                rec["payment_terms"] = _dash(_str(_cell_val(po_sheet, r, pc["payment_terms"]))) if pc.get("payment_terms") else rec.get("payment_terms", "—")
                if pc.get("units"):
                    u = _int_or_zero(_cell_val(po_sheet, r, pc["units"]))
                    if u > 0:
                        rec["units"] = u
                if pc.get("cbm"):
                    c = _float_or_zero(_cell_val(po_sheet, r, pc["cbm"]))
                    if c > 0:
                        rec["cbm"] = c
                rec["x_factory_date"] = _parse_date(_cell_val(po_sheet, r, pc["fob_date"])) if pc.get("fob_date") else rec.get("x_factory_date", "—")
                if pc.get("est_arrival"):
                    ea = _parse_date(_cell_val(po_sheet, r, pc["est_arrival"]))
                    if ea != "—":
                        rec["container_eta"] = ea
                        rec["eta_month"] = _derive_eta_month(ea)
                if pc.get("status"):
                    s = _normalize_status(_dash(_str(_cell_val(po_sheet, r, pc["status"]))))
                    if s != "—":
                        rec["status"] = s
                if pc.get("total_cost"):
                    rec["total_cost"] = _float_or_zero(_cell_val(po_sheet, r, pc["total_cost"]))
                if pc.get("landed_cost"):
                    rec["landed_cost"] = _float_or_zero(_cell_val(po_sheet, r, pc["landed_cost"]))
                if pc.get("ocean_freight"):
                    rec["freight_cost"] = _float_or_zero(_cell_val(po_sheet, r, pc["ocean_freight"]))
                if pc.get("customs"):
                    rec["duty_cost"] = _float_or_zero(_cell_val(po_sheet, r, pc["customs"]))

            # For PO-level sheets without SKU, enrich existing container records with PO-level data
            if not has_sku:
                # Build PO -> fields lookup from PO-only records
                po_fields = {}
                for rid, rec in records.items():
                    if rid.endswith("_PO"):
                        po_fields[rec["po_number"]] = rec
                # Propagate PO-level fields to container records
                for rid, rec in records.items():
                    if not rid.endswith("_PO") and rec["po_number"] in po_fields:
                        po_rec = po_fields[rec["po_number"]]
                        for f in ("factory", "payment_terms", "x_factory_date", "total_cost", "landed_cost"):
                            if rec.get(f, "—") == "—" and po_rec.get(f, "—") != "—":
                                rec[f] = po_rec[f]
                        # Update status from PO if container status is missing
                        if rec.get("status", "—") == "—" and po_rec.get("status", "—") != "—":
                            rec["status"] = po_rec["status"]
                        # Update ETA from PO if missing
                        if rec.get("container_eta", "—") == "—" and po_rec.get("container_eta", "—") != "—":
                            rec["container_eta"] = po_rec["container_eta"]
                            rec["eta_month"] = po_rec["eta_month"]

                # Remove PO-level records that are duplicated by container records
                po_nums_with_containers = set()
                for rid in records:
                    if not rid.endswith("_PO"):
                        po_num = rid.rsplit("_", 1)[0] if "_" in rid else rid
                        # Extract PO number (first part before _)
                        po_nums_with_containers.add(records[rid]["po_number"])
                to_remove = [rid for rid, rec in records.items()
                             if rid.endswith("_PO") and rec["po_number"] in po_nums_with_containers]
                for rid in to_remove:
                    del records[rid]

    # ── 3. Parse Invoice Summary (invoice-level enrichment) ──
    if inv_sheet is not None:
        hr, _ = _find_header_row(inv_sheet, ["invoice"])
        if hr is not None:
            ic = _map_columns(inv_sheet, hr, {
                "po_number": [lambda h: ("po" in h and "#" in h) or h in ("po#", "po #", "po number")],
                "container_number": [lambda h: "container" in h and ("#" in h or "number" in h)],
                "invoice_number": [lambda h: "invoice" in h and ("#" in h or "number" in h or h == "invoice #")],
                "invoice_date": [lambda h: "invoice" in h and "date" in h],
                "inv_units": [lambda h: "unit" in h or "qty" in h],
                "inv_cbm": [lambda h: "cbm" in h],
                "invoice_shipper": [lambda h: "shipper" in h],
                "sold_to": [lambda h: "sold" in h],
            })

            # Build lookup tables
            inv_by_po_cntr = {}  # (po, container) -> {invoice_date, invoice_number, units, cbm, shipper, sold_to}
            inv_by_po = {}       # po -> [{invoice_number, invoice_date, ...}]

            for r in range(hr + 1, inv_sheet.max_row + 1):
                po = _str(_cell_val(inv_sheet, r, ic.get("po_number", 1)))
                cntr = _str(_cell_val(inv_sheet, r, ic.get("container_number", 2)))
                inv_num = _dash(_str(_cell_val(inv_sheet, r, ic.get("invoice_number", 3))))
                inv_date = _parse_date(_cell_val(inv_sheet, r, ic.get("invoice_date", 4))) if ic.get("invoice_date") else "—"
                inv_units = _int_or_zero(_cell_val(inv_sheet, r, ic.get("inv_units", 5))) if ic.get("inv_units") else 0
                inv_cbm = _float_or_zero(_cell_val(inv_sheet, r, ic.get("inv_cbm", 6))) if ic.get("inv_cbm") else 0.0
                shipper = _dash(_str(_cell_val(inv_sheet, r, ic.get("invoice_shipper", 7)))) if ic.get("invoice_shipper") else "—"
                sold_to = _dash(_str(_cell_val(inv_sheet, r, ic.get("sold_to", 8)))) if ic.get("sold_to") else "—"

                if not po or po.lower() in ("total", "subtotal", "grand total"):
                    continue

                inv_data = {"invoice_number": inv_num, "invoice_date": inv_date,
                            "inv_units": inv_units, "inv_cbm": inv_cbm,
                            "invoice_shipper": shipper, "sold_to": sold_to}

                if cntr:
                    inv_by_po_cntr[(po, cntr)] = inv_data
                if po not in inv_by_po:
                    inv_by_po[po] = []
                inv_by_po[po].append(inv_data)

            # Enrich records with invoice data
            for rid, rec in records.items():
                po = rec["po_number"]
                cntr = rec["container_number"]
                inv_data = None
                if cntr != "—":
                    inv_data = inv_by_po_cntr.get((po, cntr))
                if inv_data is None and po in inv_by_po:
                    # Take first matching invoice for this PO
                    inv_data = inv_by_po[po][0]
                if inv_data:
                    if inv_data.get("invoice_date", "—") != "—":
                        rec["invoice_date"] = inv_data["invoice_date"]
                    if inv_data.get("invoice_number", "—") != "—" and rec["invoice_number"] == "—":
                        rec["invoice_number"] = inv_data["invoice_number"]

    # ── Save to store ──
    all_records = list(records.values())
    store = _load_store()
    existing = {r["record_id"]: r for r in store.get("records", [])}
    for rec in all_records:
        existing[rec["record_id"]] = rec
    store["records"] = list(existing.values())
    store["lastUpload"] = datetime.now(TZ).isoformat()
    store["sourceFile"] = filename
    _save_store(store)

    return {
        "status": "ok",
        "records_parsed": len(all_records),
        "total_records": len(store["records"]),
        "lastUpload": store["lastUpload"],
        "sourceFile": store["sourceFile"],
    }


# ── Also support legacy upload endpoints (backward compat) ──

@router.post("/v2/upload")
async def upload_v2(file: UploadFile = File(...), _user=Depends(require_auth)):
    """Alias for the main upload endpoint."""
    return await upload_supply_chain(file, _user)


# ── GET endpoints ────────────────────────────────────────────────────────

@router.get("/otw")
async def get_otw_summary(_user=Depends(require_auth)):
    """
    View 1: OTW Summary — container-level.
    One row per Container × PO. Leads with shipping/logistics data.
    """
    store = _load_store()
    records = store.get("records", [])

    # Group by (container_number, po_number) for container-level view
    container_groups: dict[str, dict] = {}
    item_detail: list[dict] = []  # for item pivot table

    for rec in records:
        cntr = rec.get("container_number", "—")
        po = rec.get("po_number", "—")
        key = f"{cntr}|{po}"

        if key not in container_groups:
            container_groups[key] = {
                "container_number": cntr,
                "hb_number": rec.get("hb_number", "—"),
                "vessel": rec.get("vessel", "—"),
                "etd": rec.get("etd", "—"),
                "etd_port": rec.get("etd_port", "—"),
                "container_eta": rec.get("container_eta", "—"),
                "eta_port": rec.get("eta_port", "—"),
                "eta_delivery": rec.get("eta_delivery", "—"),
                "delivery_location": rec.get("delivery_location", "—"),
                "actual_delivery": rec.get("actual_delivery", "—"),
                "units": 0,
                "cbm": 0.0,
                "po_number": po,
                "invoice_number": rec.get("invoice_number", "—"),
                "status": rec.get("status", "—"),
                "eta_month": rec.get("eta_month", "—"),
                "container_type": rec.get("container_type", "—"),
                "shipper": rec.get("shipper", "—"),
                "x_factory_date": rec.get("x_factory_date", "—"),
            }

        container_groups[key]["units"] += rec.get("units", 0)
        container_groups[key]["cbm"] += rec.get("cbm", 0.0)

        # Collect item detail
        item_detail.append({
            "sku": rec.get("sku", "—"),
            "description": rec.get("description", "—"),
            "container_number": cntr,
            "units": rec.get("units", 0),
        })

    rows = list(container_groups.values())
    # Round CBM
    for r in rows:
        r["cbm"] = round(r["cbm"], 1)

    # Sort: Delivered first, then In Transit, then Open; within each by ETA
    status_order = {"Delivered": 0, "In Transit": 1, "Open": 2, "—": 3}
    rows.sort(key=lambda x: (status_order.get(x["status"], 3), x.get("container_eta", "zzz")))

    # KPI summary
    total_containers = len(set(r["container_number"] for r in rows if r["container_number"] != "—"))
    total_units = sum(r["units"] for r in rows)
    total_cbm = round(sum(r["cbm"] for r in rows), 1)
    delivered = len([r for r in rows if r["status"] == "Delivered"])
    in_transit = len([r for r in rows if r["status"] == "In Transit"])
    open_count = len([r for r in rows if r["status"] == "Open"])

    # Build item pivot: rows = SKUs, columns = container numbers
    all_containers = list(dict.fromkeys(r["container_number"] for r in rows))
    sku_map: dict[str, dict] = {}
    for item in item_detail:
        sku = item["sku"]
        if sku not in sku_map:
            sku_map[sku] = {"sku": sku, "description": item["description"], "containers": {}, "total": 0}
        cntr = item["container_number"]
        sku_map[sku]["containers"][cntr] = sku_map[sku]["containers"].get(cntr, 0) + item["units"]
        sku_map[sku]["total"] += item["units"]

    item_pivot = list(sku_map.values())
    item_pivot.sort(key=lambda x: x["sku"])

    return {
        "rows": rows,
        "summary": {
            "totalContainers": total_containers,
            "totalUnits": total_units,
            "totalCBM": total_cbm,
            "delivered": delivered,
            "inTransit": in_transit,
            "open": open_count,
        },
        "itemPivot": item_pivot,
        "allContainers": all_containers,
        "lastUpload": store.get("lastUpload"),
        "sourceFile": store.get("sourceFile"),
    }


@router.get("/po")
async def get_po_summary(_user=Depends(require_auth)):
    """
    View 2: PO Summary — PO-level.
    One row per PO × Container/Invoice. Shows ordered vs shipped.
    """
    store = _load_store()
    records = store.get("records", [])

    # Group by PO to calculate totals, then expand per container
    po_totals: dict[str, int] = {}  # po_number -> total units ordered
    for rec in records:
        po = rec.get("po_number", "—")
        po_totals[po] = po_totals.get(po, 0) + rec.get("units", 0)

    # Group by (po_number, container_number) for row-level
    po_container_groups: dict[str, dict] = {}
    item_detail: list[dict] = []

    for rec in records:
        po = rec.get("po_number", "—")
        cntr = rec.get("container_number", "—")
        key = f"{po}|{cntr}"

        if key not in po_container_groups:
            po_container_groups[key] = {
                "po_number": po,
                "factory": rec.get("factory", "—"),
                "units_ordered": 0,  # will set on first row of PO
                "units_shipped": 0,
                "difference": 0,
                "cbm": 0.0,
                "container_number": cntr,
                "invoice_number": rec.get("invoice_number", "—"),
                "payment_terms": rec.get("payment_terms", "—"),
                "x_factory_date": rec.get("x_factory_date", "—"),
                "etd": rec.get("etd", "—"),
                "etd_port": rec.get("etd_port", "—"),
                "eta_month": rec.get("eta_month", "—"),
                "container_eta": rec.get("container_eta", "—"),
                "eta_port": rec.get("eta_port", "—"),
                "status": rec.get("status", "—"),
            }

        po_container_groups[key]["units_shipped"] += rec.get("units", 0)
        po_container_groups[key]["cbm"] += rec.get("cbm", 0.0)

        item_detail.append({
            "sku": rec.get("sku", "—"),
            "description": rec.get("description", "—"),
            "po_number": po,
            "units": rec.get("units", 0),
        })

    rows = list(po_container_groups.values())
    for r in rows:
        r["cbm"] = round(r["cbm"], 1)

    # Set units_ordered on first row of each PO, '—' on subsequent
    seen_pos = set()
    rows.sort(key=lambda x: (x["po_number"], x["container_number"]))
    for r in rows:
        po = r["po_number"]
        if po not in seen_pos:
            r["units_ordered"] = po_totals.get(po, 0)
            seen_pos.add(po)
        else:
            r["units_ordered"] = "—"

    # Compute difference for each PO (first row only)
    po_shipped: dict[str, int] = {}
    for r in rows:
        po = r["po_number"]
        if r["status"] in ("Delivered", "In Transit"):
            po_shipped[po] = po_shipped.get(po, 0) + r["units_shipped"]

    for r in rows:
        if r["units_ordered"] != "—":
            r["difference"] = r["units_ordered"] - po_shipped.get(r["po_number"], 0)

    # KPI summary
    unique_pos = set(r["po_number"] for r in rows)
    total_ordered = sum(po_totals.values())
    total_shipped = sum(r["units_shipped"] for r in rows if r["status"] in ("Delivered", "In Transit"))
    status_counts = {}
    for r in rows:
        s = r["status"]
        status_counts[s] = status_counts.get(s, 0) + 1

    # Build item pivot: rows = SKUs, columns = PO numbers
    all_pos = list(dict.fromkeys(r["po_number"] for r in rows))
    sku_map: dict[str, dict] = {}
    for item in item_detail:
        sku = item["sku"]
        if sku not in sku_map:
            sku_map[sku] = {"sku": sku, "description": item["description"], "pos": {}, "total": 0}
        po = item["po_number"]
        sku_map[sku]["pos"][po] = sku_map[sku]["pos"].get(po, 0) + item["units"]
        sku_map[sku]["total"] += item["units"]

    item_pivot = list(sku_map.values())
    item_pivot.sort(key=lambda x: x["sku"])

    return {
        "rows": rows,
        "summary": {
            "totalPOs": len(unique_pos),
            "unitsOrdered": total_ordered,
            "unitsShipped": total_shipped,
            "delivered": status_counts.get("Delivered", 0),
            "inTransit": status_counts.get("In Transit", 0),
            "open": status_counts.get("Open", 0),
        },
        "itemPivot": item_pivot,
        "allPOs": all_pos,
        "lastUpload": store.get("lastUpload"),
        "sourceFile": store.get("sourceFile"),
    }


@router.get("/invoices")
async def get_invoice_summary(_user=Depends(require_auth)):
    """
    View 3: Invoice Summary — invoice-level.
    One row per Invoice × PO × Container.
    """
    store = _load_store()
    records = store.get("records", [])

    # Group by (invoice_number, po_number, container_number)
    inv_groups: dict[str, dict] = {}
    item_detail: list[dict] = []

    for rec in records:
        inv = rec.get("invoice_number", "—")
        po = rec.get("po_number", "—")
        cntr = rec.get("container_number", "—")
        key = f"{inv}|{po}|{cntr}"

        if key not in inv_groups:
            inv_groups[key] = {
                "invoice_number": inv,
                "invoice_date": rec.get("invoice_date", "—"),
                "total_units": 0,
                "cbm": 0.0,
                "po_number": po,
                "container_number": cntr,
                "container_eta": rec.get("container_eta", "—"),
                "eta_port": rec.get("eta_port", "—"),
                "factory": rec.get("factory", "—"),
                "status": rec.get("status", "—"),
                "eta_month": rec.get("eta_month", "—"),
            }

        inv_groups[key]["total_units"] += rec.get("units", 0)
        inv_groups[key]["cbm"] += rec.get("cbm", 0.0)

        item_detail.append({
            "sku": rec.get("sku", "—"),
            "description": rec.get("description", "—"),
            "invoice_number": inv,
            "units": rec.get("units", 0),
        })

    rows = list(inv_groups.values())
    for r in rows:
        r["cbm"] = round(r["cbm"], 1)

    # Sort by status then ETA
    status_order = {"Delivered": 0, "In Transit": 1, "Open": 2, "—": 3}
    rows.sort(key=lambda x: (status_order.get(x["status"], 3), x.get("container_eta", "zzz")))

    # KPI
    unique_invoices = set(r["invoice_number"] for r in rows if r["invoice_number"] != "—")
    total_units = sum(r["total_units"] for r in rows)
    total_cbm = round(sum(r["cbm"] for r in rows), 1)
    status_counts = {}
    for r in rows:
        s = r["status"]
        status_counts[s] = status_counts.get(s, 0) + 1

    # Build item pivot: rows = SKUs, columns = invoice numbers
    all_invoices = list(dict.fromkeys(r["invoice_number"] for r in rows))
    sku_map: dict[str, dict] = {}
    for item in item_detail:
        sku = item["sku"]
        if sku not in sku_map:
            sku_map[sku] = {"sku": sku, "description": item["description"], "invoices": {}, "total": 0}
        inv = item["invoice_number"]
        sku_map[sku]["invoices"][inv] = sku_map[sku]["invoices"].get(inv, 0) + item["units"]
        sku_map[sku]["total"] += item["units"]

    item_pivot = list(sku_map.values())
    item_pivot.sort(key=lambda x: x["sku"])

    return {
        "rows": rows,
        "summary": {
            "totalInvoices": len(unique_invoices),
            "totalUnits": total_units,
            "totalCBM": total_cbm,
            "delivered": status_counts.get("Delivered", 0),
            "inTransit": status_counts.get("In Transit", 0),
            "open": status_counts.get("Open", 0),
        },
        "itemPivot": item_pivot,
        "allInvoices": all_invoices,
        "lastUpload": store.get("lastUpload"),
        "sourceFile": store.get("sourceFile"),
    }


@router.get("/store")
async def get_raw_store(_user=Depends(require_auth)):
    """Debug endpoint: return the raw data store."""
    store = _load_store()
    return {
        "total_records": len(store.get("records", [])),
        "lastUpload": store.get("lastUpload"),
        "sourceFile": store.get("sourceFile"),
        "records": store.get("records", [])[:200],  # cap at 200 for debug
    }
